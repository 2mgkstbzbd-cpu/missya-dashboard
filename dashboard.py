import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import io
import time
import os
import socket
import hashlib
import re
import json
import zipfile
from copy import copy
from datetime import datetime
import html as _html
from PIL import Image, ImageDraw, ImageFont
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode, JsCode
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
def _is_nan(x):
    try:
        return x != x
    except Exception:
        return False

def fmt_num(x, na="—"):
    if x is None or _is_nan(x):
        return na
    try:
        s = f"{float(x):,.2f}"
    except Exception:
        return str(x)
    s = s.rstrip("0").rstrip(".")
    return s

def fmt_num_fixed(x, decimals: int = 1, na: str = "—"):
    if x is None or _is_nan(x):
        return na
    try:
        return f"{float(x):,.{int(decimals)}f}"
    except Exception:
        return str(x)

def sanitize_filename(name: str, default: str = "export"):
    s = str(name or "").strip()
    if not s:
        s = default
    s = re.sub(r"[\\\\/:*?\"<>|]+", "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:120] if len(s) > 120 else s

def _df_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str,
    title_lines: list[str] | None = None,
    number_headers: set[str] | None = None,
    number_formats: dict[str, str] | None = None,
    trend_type_header: str | None = None,
    percent_headers: set[str] | None = None,
    percent_formats: dict[str, str] | None = None,
    store_type_header: str | None = None,
    group_headers: bool = False,
):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb[sheet_name]

    title_lines = title_lines or []
    insert_n = len(title_lines) + (1 if title_lines else 0)
    if insert_n:
        ws.insert_rows(1, amount=insert_n)
        end_col = get_column_letter(ws.max_column)
        for i, line in enumerate(title_lines, start=1):
            ws.merge_cells(f"A{i}:{end_col}{i}")
            c = ws[f"A{i}"]
            c.value = str(line)
            c.font = Font(bold=True, size=12, color="111827")
            c.alignment = Alignment(horizontal="left", vertical="center")

    header_row = (len(title_lines) + 2) if title_lines else 1

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _get_headers(row_idx: int):
        hs = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=col).value
            hs.append(str(v) if v is not None else "")
        return hs

    group_row = None
    subheader_row = header_row
    if group_headers:
        group_row = header_row
        ws.insert_rows(group_row, amount=1)
        subheader_row = group_row + 1

    ws.freeze_panes = f"A{subheader_row + 1}"

    headers = _get_headers(subheader_row)

    if group_headers:
        seg_specs = [
            ("段粉", "段粉-", "DCFCE7", "166534"),
            ("雅系列", "雅系列-", "DBEAFE", "1D4ED8"),
            ("中老年（提）", "中老年-", "FEF9C3", "A16207"),
        ]
        grouped_cols: set[int] = set()
        for label, prefix, bg, fg in seg_specs:
            idxs = [i + 1 for i, h in enumerate(headers) if str(h).startswith(prefix)]
            if not idxs:
                continue
            idxs = sorted(set(idxs))
            blocks = []
            start = idxs[0]
            prev = idxs[0]
            for x in idxs[1:]:
                if x == prev + 1:
                    prev = x
                else:
                    blocks.append((start, prev))
                    start = x
                    prev = x
            blocks.append((start, prev))

            for c1, c2 in blocks:
                grouped_cols.update(range(c1, c2 + 1))
                top_cell = ws.cell(row=group_row, column=c1)
                top_cell.value = label
                top_cell.fill = PatternFill("solid", fgColor=bg)
                top_cell.font = Font(bold=True, color=fg)
                top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                top_cell.border = border
                if c2 > c1:
                    ws.merge_cells(start_row=group_row, start_column=c1, end_row=group_row, end_column=c2)

                for col in range(c1, c2 + 1):
                    sh = ws.cell(row=subheader_row, column=col)
                    full = headers[col - 1]
                    sub = full.split("-", 1)[1] if "-" in full else full
                    sub = sub.replace("(提)", "").strip()
                    sh.value = sub
                    sh.fill = PatternFill("solid", fgColor=bg)
                    sh.font = Font(bold=True, color=fg)
                    sh.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    sh.border = border

        if not grouped_cols:
            def _sub_style(s: str):
                k = str(s or "").strip()
                if k in ("A", "A数量"):
                    return "DCFCE7", "166534"
                if k in ("B", "B数量"):
                    return "DBEAFE", "1D4ED8"
                if k in ("C", "C数量"):
                    return "FEF9C3", "A16207"
                if k in ("D", "D数量"):
                    return "FEE2E2", "991B1B"
                if k == "持续升级":
                    return "DCFCE7", "166534"
                if k == "持续降级":
                    return "FEE2E2", "991B1B"
                if k == "先降级后升级":
                    return "DBEAFE", "1D4ED8"
                if k == "先升级后降级":
                    return "FEF9C3", "A16207"
                if k == "持续持平":
                    return "E5E7EB", "374151"
                if k in ("持平升级", "升级持平"):
                    return "DCFCE7", "166534"
                if k in ("持平降级", "降级持平"):
                    return "FEE2E2", "991B1B"
                return None, None

            group_map: dict[str, list[int]] = {}
            for i, h in enumerate(headers, start=1):
                hs = str(h or "")
                if "-" not in hs:
                    continue
                prefix, _ = hs.split("-", 1)
                prefix = str(prefix).strip()
                if not prefix:
                    continue
                group_map.setdefault(prefix, []).append(i)

            for label, idxs in group_map.items():
                idxs = sorted(set(idxs))
                blocks = []
                start = idxs[0]
                prev = idxs[0]
                for x in idxs[1:]:
                    if x == prev + 1:
                        prev = x
                    else:
                        blocks.append((start, prev))
                        start = x
                        prev = x
                blocks.append((start, prev))

                for c1, c2 in blocks:
                    grouped_cols.update(range(c1, c2 + 1))
                    top_cell = ws.cell(row=group_row, column=c1)
                    top_cell.value = label
                    top_cell.fill = header_fill
                    top_cell.font = header_font
                    top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    top_cell.border = border
                    if c2 > c1:
                        ws.merge_cells(start_row=group_row, start_column=c1, end_row=group_row, end_column=c2)

                    for col in range(c1, c2 + 1):
                        sh = ws.cell(row=subheader_row, column=col)
                        full = headers[col - 1]
                        sub = full.split("-", 1)[1] if "-" in full else full
                        sub = str(sub).strip()
                        if sub in ("A", "B", "C", "D"):
                            sub = f"{sub}数量"
                        sh.value = sub
                        bg, fg = _sub_style(sub)
                        if bg and fg:
                            sh.fill = PatternFill("solid", fgColor=bg)
                            sh.font = Font(bold=True, color=fg)
                        else:
                            sh.fill = header_fill
                            sh.font = header_font
                        sh.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        sh.border = border

        for col in range(1, ws.max_column + 1):
            if col in grouped_cols:
                continue
            v = ws.cell(row=subheader_row, column=col).value
            c = ws.cell(row=group_row, column=col)
            c.value = v
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
            ws.merge_cells(start_row=group_row, start_column=col, end_row=subheader_row, end_column=col)

        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=group_row, column=col)
            if c.value is None:
                continue
            c.border = border
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=subheader_row, column=col)
            if c.value is None:
                continue
            c.border = border
    else:
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=subheader_row, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

    trend_col_idx = None
    if trend_type_header:
        try:
            trend_col_idx = headers.index(trend_type_header) + 1
        except ValueError:
            trend_col_idx = None
    trend_col_idxs = set()
    if trend_col_idx:
        trend_col_idxs.add(trend_col_idx)
    for i, h in enumerate(headers, start=1):
        hs = str(h or "").strip()
        if hs in ("趋势类型", "近三周期变化") or hs.endswith("趋势类型") or hs.endswith("变化类型"):
            trend_col_idxs.add(i)

    store_type_col_idx = None
    if store_type_header:
        try:
            store_type_col_idx = headers.index(store_type_header) + 1
        except ValueError:
            store_type_col_idx = None
    store_type_col_idxs = set()
    if store_type_col_idx:
        store_type_col_idxs.add(store_type_col_idx)
    for i, h in enumerate(headers, start=1):
        hs = str(h or "").strip()
        if "门店类型" in hs or hs in ("等级",):
            store_type_col_idxs.add(i)

    max_rows_for_full_style = 12000
    apply_full = ws.max_row <= max_rows_for_full_style
    num_fmt = "0.#"
    pct_fmt = "0.#%"
    number_headers = number_headers or set()
    percent_headers = percent_headers or set()
    number_formats = number_formats or {}
    percent_formats = percent_formats or {}

    even_fill = PatternFill("solid", fgColor="F8FAFC")
    total_fill = None

    for r in range(subheader_row + 1, ws.max_row + 1):
        is_total_row = False
        for c0 in range(1, min(4, ws.max_column) + 1):
            if str(ws.cell(row=r, column=c0).value or "").strip() == "合计":
                is_total_row = True
                break
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if apply_full:
                cell.border = border
                if (r - subheader_row) % 2 == 0:
                    cell.fill = even_fill
                if is_total_row:
                    cell.font = Font(bold=True, color="A16207")
            fmt_override = number_formats.get(headers[col - 1])
            if fmt_override:
                cell.number_format = fmt_override
            elif headers[col - 1] in number_headers:
                cell.number_format = num_fmt
            pct_override = percent_formats.get(headers[col - 1])
            if pct_override:
                cell.number_format = pct_override
            elif headers[col - 1] in percent_headers:
                cell.number_format = pct_fmt
            if trend_col_idxs and col in trend_col_idxs:
                v = str(cell.value or "").strip()
                if v == "持续增长":
                    cell.fill = PatternFill("solid", fgColor="DCFCE7")
                    cell.font = Font(bold=True, color="166534")
                elif v == "持续下滑":
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
                    cell.font = Font(bold=True, color="991B1B")
                elif v == "先下滑后增长":
                    cell.fill = PatternFill("solid", fgColor="DBEAFE")
                    cell.font = Font(bold=True, color="1D4ED8")
                elif v == "先增长后下滑":
                    cell.fill = PatternFill("solid", fgColor="FEF9C3")
                    cell.font = Font(bold=True, color="A16207")
                elif v == "持续升级":
                    cell.fill = PatternFill("solid", fgColor="DCFCE7")
                    cell.font = Font(bold=True, color="166534")
                elif v == "持续降级":
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
                    cell.font = Font(bold=True, color="991B1B")
                elif v == "先降级后升级":
                    cell.fill = PatternFill("solid", fgColor="DBEAFE")
                    cell.font = Font(bold=True, color="1D4ED8")
                elif v == "先升级后降级":
                    cell.fill = PatternFill("solid", fgColor="FEF9C3")
                    cell.font = Font(bold=True, color="A16207")
                elif v == "持续持平":
                    cell.fill = PatternFill("solid", fgColor="E5E7EB")
                    cell.font = Font(bold=True, color="374151")
                elif v in ("持平升级", "升级持平"):
                    cell.fill = PatternFill("solid", fgColor="DCFCE7")
                    cell.font = Font(bold=True, color="166534")
                elif v in ("持平降级", "降级持平"):
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
                    cell.font = Font(bold=True, color="991B1B")
            if store_type_col_idxs and col in store_type_col_idxs:
                v = str(cell.value or "").strip().upper()
                if "A" in v:
                    cell.fill = PatternFill("solid", fgColor="DCFCE7")
                    cell.font = Font(bold=True, color="166534")
                elif "B" in v:
                    cell.fill = PatternFill("solid", fgColor="DBEAFE")
                    cell.font = Font(bold=True, color="1D4ED8")
                elif "C" in v:
                    cell.fill = PatternFill("solid", fgColor="FEF9C3")
                    cell.font = Font(bold=True, color="A16207")
                elif "D" in v:
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
                    cell.font = Font(bold=True, color="991B1B")

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for r in range(1, min(ws.max_row, header_row + 2000) + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = max(10, min(36, max_len + 2))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def _merge_single_sheet_workbooks(items: list[tuple[bytes, str]]):
    wb_out = Workbook()
    if wb_out.worksheets:
        wb_out.remove(wb_out.worksheets[0])

    for xlsx_bytes, sheet_name in items:
        src_wb = load_workbook(io.BytesIO(xlsx_bytes))
        src_ws = src_wb.active

        tgt_ws = wb_out.create_sheet(title=str(sheet_name))
        tgt_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
        tgt_ws.freeze_panes = src_ws.freeze_panes

        for k, dim in src_ws.column_dimensions.items():
            tgt_ws.column_dimensions[k].width = dim.width
            tgt_ws.column_dimensions[k].hidden = dim.hidden

        for k, dim in src_ws.row_dimensions.items():
            tgt_ws.row_dimensions[k].height = dim.height
            tgt_ws.row_dimensions[k].hidden = dim.hidden

        for rng in list(src_ws.merged_cells.ranges):
            tgt_ws.merge_cells(str(rng))

        for row in src_ws.iter_rows(values_only=False):
            for cell in row:
                tgt = tgt_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                tgt.number_format = cell.number_format
                tgt.font = copy(cell.font)
                tgt.fill = copy(cell.fill)
                tgt.border = copy(cell.border)
                tgt.alignment = copy(cell.alignment)
                tgt.protection = copy(cell.protection)

    out = io.BytesIO()
    wb_out.save(out)
    return out.getvalue()

def fmt_pct_ratio(r, na="—", decimals=1):
    if r is None or _is_nan(r):
        return na
    v = float(r) * 100.0
    s = f"{v:.{decimals}f}".rstrip("0").rstrip(".")
    return f"{s}%"

def fmt_pct_value(p, na="—", decimals=1):
    if p is None or _is_nan(p):
        return na
    v = float(p)
    sign = "+" if v > 0 else ("-" if v < 0 else "")
    s = f"{abs(v):.{decimals}f}".rstrip("0").rstrip(".")
    return f"{sign}{s}%"

def _pil_load_font(size: int, bold: bool = False):
    candidates = [
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\msyhbd.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\simsun.ttc",
        r"C:\Windows\Fonts\arial.ttf",
    ]
    if bold:
        bold_first = [r"C:\Windows\Fonts\msyhbd.ttc", r"C:\Windows\Fonts\arialbd.ttf"]
        candidates = bold_first + [c for c in candidates if c not in bold_first]
    for p in candidates:
        try:
            return ImageFont.truetype(p, size=size)
        except Exception:
            continue
    return ImageFont.load_default()

def _pil_table_png(df: pd.DataFrame, title_lines: list[str], font_size: int = 16, col_types: dict | None = None):
    d = df.copy().fillna("")
    font = _pil_load_font(font_size, bold=False)
    font_b = _pil_load_font(font_size + 2, bold=True)
    pad_x = 14
    pad_y = 10
    line_spacing = 2
    grid_color = (220, 223, 230)
    text_color = (31, 35, 40)
    header_bg = (245, 247, 250)
    title_bg = (255, 255, 255)
    row_bg_odd = (255, 255, 255)
    row_bg_even = (252, 253, 255)

    headers = d.columns.tolist()
    rows = d.values.tolist()
    n_rows = len(rows)
    n_cols = len(headers)
    col_types = col_types or {}

    def _text_wh(txt, fnt):
        if txt is None:
            return 0, 0
        s = str(txt)
        if not s:
            return 0, 0
        lines = s.splitlines() or [s]
        widths = []
        heights = []
        for line in lines:
            t = str(line)
            if t:
                bbox = fnt.getbbox(t)
                widths.append(bbox[2] - bbox[0])
                heights.append(bbox[3] - bbox[1])
            else:
                bbox = fnt.getbbox("Ag")
                widths.append(0)
                heights.append(bbox[3] - bbox[1])
        w = max(widths) if widths else 0
        h = int(sum(heights) + max(0, len(lines) - 1) * line_spacing)
        return int(w), int(h)

    def _cell_text(v, t):
        if t == "num":
            return fmt_num_fixed(v, decimals=1, na="0")
        if t == "pct":
            try:
                return fmt_pct_ratio(v, na="—", decimals=1)
            except Exception:
                return ""
        if v is None:
            return ""
        return str(v)

    fixed_w = {"spark": 160, "tag": 160}
    col_widths = []
    for j in range(n_cols):
        name = headers[j]
        t = col_types.get(name, "text")
        if t in fixed_w:
            col_widths.append(fixed_w[t])
            continue
        w0, _ = _text_wh(name, font_b)
        w = w0
        for i in range(min(n_rows, 120)):
            txt = _cell_text(rows[i][j], t)
            w1, _ = _text_wh(txt, font)
            if w1 > w:
                w = w1
        col_widths.append(int(w + pad_x * 2))

    row_h = max(_text_wh("Ag", font)[1] + pad_y * 2, font_size + pad_y * 2 + 6)
    has_groups = any("\n" in str(h) for h in headers)
    if has_groups:
        group_labels = []
        sub_labels = []
        for h in headers:
            s = str(h)
            parts = s.split("\n", 1)
            if len(parts) == 2:
                group_labels.append(parts[0])
                sub_labels.append(parts[1])
            else:
                group_labels.append("")
                sub_labels.append(s)
        group_h = max(_text_wh("Ag", font_b)[1] + pad_y * 2, max((_text_wh(x, font_b)[1] for x in group_labels), default=0) + pad_y * 2)
        sub_h = max(_text_wh("Ag", font_b)[1] + pad_y * 2, max((_text_wh(x, font_b)[1] for x in sub_labels), default=0) + pad_y * 2)
        header_h = int(group_h + sub_h)
    else:
        header_h = max(_text_wh("Ag", font_b)[1] + pad_y * 2, font_size + pad_y * 2 + 10)
        try:
            header_h = max(header_h, max(_text_wh(h, font_b)[1] for h in headers) + pad_y * 2)
        except Exception:
            pass

    title_lines = [str(x) for x in title_lines if str(x).strip()]
    title_h = 0
    if title_lines:
        title_h = (row_h * len(title_lines)) + 10

    table_w = int(sum(col_widths) + 1)
    table_h = int(header_h + (n_rows * row_h) + 1)
    img_w = table_w
    img_h = int(title_h + table_h)

    img = Image.new("RGB", (img_w, img_h), title_bg)
    draw = ImageDraw.Draw(img)

    y = 0
    if title_lines:
        draw.rectangle([0, 0, img_w, title_h], fill=title_bg)
        ty = 6
        for k, line in enumerate(title_lines):
            fnt = font_b if k == 0 else font
            draw.text((pad_x, ty), line, fill=text_color, font=fnt)
            ty += row_h
        y = title_h

    x = 0
    draw.rectangle([0, y, img_w, y + header_h], fill=header_bg)
    if has_groups:
        y_top = y
        y_sub = y + group_h
        gx = 0
        cur = group_labels[0] if group_labels else ""
        start_x = 0
        for j in range(n_cols + 1):
            if j == n_cols or group_labels[j] != cur:
                end_x = gx
                draw.rectangle([start_x, y_top, end_x, y_sub], outline=grid_color, width=1)
                if cur:
                    tw, th = _text_wh(cur, font_b)
                    draw.multiline_text(
                        (start_x + (end_x - start_x - tw) / 2, y_top + (group_h - th) / 2),
                        cur,
                        fill=text_color,
                        font=font_b,
                        align="center",
                        spacing=line_spacing,
                    )
                if j < n_cols:
                    start_x = gx
                    cur = group_labels[j]
            if j < n_cols:
                gx += col_widths[j]

        x = 0
        for j in range(n_cols):
            w = col_widths[j]
            draw.rectangle([x, y_sub, x + w, y + header_h], outline=grid_color, width=1)
            txt = sub_labels[j]
            tw, th = _text_wh(txt, font_b)
            draw.multiline_text(
                (x + (w - tw) / 2, y_sub + (sub_h - th) / 2),
                txt,
                fill=text_color,
                font=font_b,
                align="center",
                spacing=line_spacing,
            )
            x += w
    else:
        for j in range(n_cols):
            w = col_widths[j]
            draw.rectangle([x, y, x + w, y + header_h], outline=grid_color, width=1)
            txt = headers[j]
            tw, th = _text_wh(txt, font_b)
            draw.multiline_text(
                (x + (w - tw) / 2, y + (header_h - th) / 2),
                txt,
                fill=text_color,
                font=font_b,
                align="center",
                spacing=line_spacing,
            )
            x += w

    def _draw_sparkline(cell_box, v):
        try:
            vals = json.loads(v) if isinstance(v, str) else list(v)
        except Exception:
            vals = []
        try:
            vals = [float(x) for x in vals if x is not None]
        except Exception:
            vals = []
        if len(vals) < 2:
            return
        left, top, right, bottom = cell_box
        inner_l = left + pad_x
        inner_r = right - pad_x
        inner_t = top + pad_y
        inner_b = bottom - pad_y
        if inner_r - inner_l < 10 or inner_b - inner_t < 10:
            return
        vmin = min(vals)
        vmax = max(vals)
        if vmax == vmin:
            vmax = vmin + 1.0
        n = len(vals)
        pts = []
        for i, val in enumerate(vals):
            px = inner_l + (inner_r - inner_l) * (i / (n - 1))
            py = inner_b - (inner_b - inner_t) * ((val - vmin) / (vmax - vmin))
            pts.append((px, py))
        line_color = (255, 112, 0)
        draw.line(pts, fill=line_color, width=2, joint="curve")

    def _tag_style(tag: str):
        t = (tag or "").strip()
        m = re.search(r"[A-Da-d]", t)
        if m:
            k = m.group(0).upper()
            if k == "A":
                return (219, 246, 229), (15, 81, 50)
            if k == "B":
                return (219, 234, 254), (30, 64, 175)
            if k == "C":
                return (254, 243, 199), (146, 64, 14)
            if k == "D":
                return (254, 226, 226), (127, 29, 29)
        if t == "持续增长":
            return (219, 246, 229), (15, 81, 50)
        if t == "持续下滑":
            return (254, 226, 226), (127, 29, 29)
        if t == "先下滑后增长":
            return (219, 234, 254), (30, 64, 175)
        if t == "先增长后下滑":
            return (254, 243, 199), (146, 64, 14)
        if t in ("持续升级", "持平升级", "升级持平"):
            return (219, 246, 229), (15, 81, 50)
        if t in ("持续降级", "持平降级", "降级持平"):
            return (254, 226, 226), (127, 29, 29)
        if t == "先降级后升级":
            return (219, 234, 254), (30, 64, 175)
        if t == "先升级后降级":
            return (254, 243, 199), (146, 64, 14)
        if t == "持续持平":
            return (229, 231, 235), (55, 65, 81)
        return (229, 231, 235), (55, 65, 81)

    def _draw_tag(cell_box, tag: str):
        left, top, right, bottom = cell_box
        bg, fg = _tag_style(tag)
        rect_l = left + pad_x
        rect_r = right - pad_x
        rect_t = top + int(pad_y * 0.7)
        rect_b = bottom - int(pad_y * 0.7)
        if rect_r - rect_l < 10 or rect_b - rect_t < 10:
            return
        try:
            draw.rounded_rectangle([rect_l, rect_t, rect_r, rect_b], radius=14, fill=bg, outline=grid_color, width=1)
        except Exception:
            draw.rectangle([rect_l, rect_t, rect_r, rect_b], fill=bg, outline=grid_color, width=1)
        txt = (tag or "").strip()
        tw, th = _text_wh(txt, font)
        draw.text((rect_l + (rect_r - rect_l - tw) / 2, rect_t + (rect_b - rect_t - th) / 2), txt, fill=fg, font=font)

    for i in range(n_rows):
        x = 0
        yy = y + header_h + i * row_h
        row_bg = row_bg_even if (i % 2 == 1) else row_bg_odd
        for j in range(n_cols):
            w = col_widths[j]
            left = x
            right = x + w
            top = yy
            bottom = yy + row_h
            draw.rectangle([left, top, right, bottom], fill=row_bg, outline=grid_color, width=1)
            name = headers[j]
            t = col_types.get(name, "text")
            v = rows[i][j]
            if t == "spark":
                _draw_sparkline((left, top, right, bottom), v)
            elif t == "tag":
                _draw_tag((left, top, right, bottom), str(v))
            else:
                txt = _cell_text(v, t)
                tw, th = _text_wh(txt, font)
                if "\n" in str(txt):
                    draw.multiline_text(
                        (left + (w - tw) / 2, top + (row_h - th) / 2),
                        txt,
                        fill=text_color,
                        font=font,
                        align="center",
                        spacing=line_spacing,
                    )
                else:
                    draw.text((left + (w - tw) / 2, top + (row_h - th) / 2), txt, fill=text_color, font=font)
            x += w

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def _pil_line_png(x_labels: list[str], y_vals: list[float], title_lines: list[str], color: tuple[int, int, int] = (124, 58, 237)):
    title_lines = [str(x) for x in (title_lines or []) if str(x).strip()]
    x_labels = [str(x) for x in (x_labels or [])]
    y_vals = [0.0 if (v is None or _is_nan(v)) else float(v) for v in (y_vals or [])]

    font = _pil_load_font(18, bold=False)
    font_b = _pil_load_font(20, bold=True)

    w = 1400
    h = 720 if title_lines else 620
    pad_l = 90
    pad_r = 40
    pad_t = 90 if title_lines else 50
    pad_b = 90
    bg = (255, 255, 255)
    grid = (229, 231, 235)
    axis = (17, 24, 39)
    txt = (31, 41, 55)

    img = Image.new("RGB", (w, h), bg)
    draw = ImageDraw.Draw(img)

    y = 20
    for i, line in enumerate(title_lines):
        f = font_b if i == 0 else font
        draw.text((pad_l, y), str(line), fill=axis, font=f)
        y += 34

    plot_top = pad_t
    plot_left = pad_l
    plot_right = w - pad_r
    plot_bottom = h - pad_b

    n = len(y_vals)
    if n <= 0:
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()

    y_max = max(y_vals) if y_vals else 0.0
    y_max = max(1.0, float(y_max))
    y_min = 0.0

    def _xy(i: int, v: float):
        if n == 1:
            x = (plot_left + plot_right) / 2
        else:
            x = plot_left + (plot_right - plot_left) * (i / (n - 1))
        yy = plot_bottom - (plot_bottom - plot_top) * ((v - y_min) / (y_max - y_min)) if y_max > y_min else plot_bottom
        return float(x), float(yy)

    steps = 5
    for k in range(steps + 1):
        v = y_min + (y_max - y_min) * (k / steps)
        yy = plot_bottom - (plot_bottom - plot_top) * (k / steps)
        draw.line([(plot_left, yy), (plot_right, yy)], fill=grid, width=1)
        s = f"{v:.1f}"
        bbox = font.getbbox(s)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        draw.text((plot_left - tw - 12, yy - th / 2), s, fill=txt, font=font)

    draw.line([(plot_left, plot_top), (plot_left, plot_bottom)], fill=axis, width=2)
    draw.line([(plot_left, plot_bottom), (plot_right, plot_bottom)], fill=axis, width=2)

    pts = [_xy(i, y_vals[i]) for i in range(n)]
    for i in range(1, n):
        draw.line([pts[i - 1], pts[i]], fill=color, width=4)

    r = 7
    for i, (x, yy) in enumerate(pts):
        draw.ellipse([x - r, yy - r, x + r, yy + r], fill=color, outline=color)
        v = y_vals[i]
        s = f"{v:.1f}"
        bbox = font_b.getbbox(s)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        draw.text((x - tw / 2, yy - th - 14), s, fill=axis, font=font_b)

    for i, (x, yy) in enumerate(pts):
        lab = x_labels[i] if i < len(x_labels) else ""
        if not lab:
            continue
        bbox = font.getbbox(lab)
        tw = bbox[2] - bbox[0]
        draw.text((x - tw / 2, plot_bottom + 14), lab, fill=txt, font=font)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

_COORD_NUM_RE = re.compile(r"[-+]?\d+(?:\.\d+)?")

def _parse_lon_lat(v):
    if v is None:
        return None, None
    s = str(v).strip()
    if not s or s.lower() in {"nan", "none"}:
        return None, None
    nums = _COORD_NUM_RE.findall(s)
    if len(nums) < 2:
        return None, None
    a = float(nums[0])
    b = float(nums[1])

    def _is_lon(x): return 70 <= x <= 140
    def _is_lat(x): return 0 <= x <= 60

    if _is_lon(a) and _is_lat(b):
        lon, lat = a, b
    elif _is_lon(b) and _is_lat(a):
        lon, lat = b, a
    else:
        lon, lat = (a, b) if abs(a) >= abs(b) else (b, a)
    if not _is_lon(lon) or not _is_lat(lat):
        return None, None
    return lon, lat

def get_host_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = '127.0.0.1'
    finally:
        s.close()
    return ip

# -----------------------------------------------------------------------------
# 1. Page Config
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="美思雅数据分析系统",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)
import streamlit.components.v1 as components

components.html(
    """
<script>
(() => {
  const doc = window.parent && window.parent.document ? window.parent.document : document;
  const ensure = () => {
    if (!doc || !doc.documentElement) return;

    doc.documentElement.setAttribute("translate", "no");
    doc.documentElement.classList.add("notranslate");

    if (doc.body) {
      doc.body.setAttribute("translate", "no");
      doc.body.classList.add("notranslate");
    }

    const head = doc.head || (doc.getElementsByTagName("head")[0] || null);
    if (head && !head.querySelector('meta[name="google"][content="notranslate"]')) {
      const meta = doc.createElement("meta");
      meta.name = "google";
      meta.content = "notranslate";
      head.appendChild(meta);
    }
  };

  ensure();
  const t = setInterval(() => {
    ensure();
    if (doc && doc.body) clearInterval(t);
  }, 100);
})();
</script>
""",
    height=0,
)

_required_password = os.getenv("DASHBOARD_PASSWORD", "").strip()
if _required_password:
    if not st.session_state.get("_authed", False):
        st.markdown("### 🔒 访问验证")
        _pwd = st.text_input("请输入访问密码", type="password")
        if st.button("验证", type="primary"):
            if _pwd == _required_password:
                st.session_state["_authed"] = True
                st.rerun()
            else:
                st.error("密码错误")
        st.stop()

if 'drill_level' not in st.session_state:
    st.session_state.drill_level = 1
if 'selected_prov' not in st.session_state:
    st.session_state.selected_prov = None
if 'selected_dist' not in st.session_state:
    st.session_state.selected_dist = None
if 'perf_time_mode' not in st.session_state:
    st.session_state.perf_time_mode = '近12个月'
if 'perf_provs' not in st.session_state:
    st.session_state.perf_provs = []
if 'perf_cats' not in st.session_state:
    st.session_state.perf_cats = []

# -----------------------------------------------------------------------------
# 2. Custom CSS
# -----------------------------------------------------------------------------
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
    :root {
        --bg-1: #F5F5F7;
        --bg-2: #ECECEC;
        --bg-3: #E0E0E0;
        --panel: #FFFFFF;
        --panel-2: rgba(255, 255, 255, 0.85);
        --stroke: #E0E0E0;
        --stroke-strong: #D1D1D1;
        --text: #1B1530;
        --text-muted: rgba(27, 21, 48, 0.7);
        --primary: #5B2EA6;
        --primary-2: #6A3AD0;
        --accent: #FFC400;
        --accent-2: #FFB000;
        --danger: #E5484D;
        --success: #2FBF71;
        --shadow: 0 10px 26px rgba(0, 0, 0, 0.08);
        --shadow-soft: 0 4px 12px rgba(0, 0, 0, 0.05);
        --radius: 12px;
        --radius-sm: 10px;
        --transition: 240ms cubic-bezier(.2,.8,.2,1);
        --focus: 0 0 0 3px rgba(91, 46, 166, 0.2);
        --tbl-header-bg: #4285F4;
        --tbl-header-bg-hover: #2F76E4;
        --tbl-header-border: #2B63C4;
        --tbl-header-fg: #A16207;
        --tbl-header-icon: #A16207;
        --tbl-header-shadow: 0 6px 16px rgba(0, 0, 0, 0.16);
        --tbl-header-font-size: 15px;
        --tbl-header-font-weight: 800;
        --tbl-cell-font-size: 13px;
    }
    
    @media (prefers-color-scheme: dark) {
        :root {
            --tbl-header-bg: #2B66D9;
            --tbl-header-bg-hover: #2358C2;
            --tbl-header-border: #1B46A0;
            --tbl-header-fg: #A16207;
            --tbl-header-icon: #A16207;
            --tbl-header-shadow: 0 10px 22px rgba(0, 0, 0, 0.32);
        }
    }

    html, body, [class*="css"] {
        font-family: 'Inter', 'Microsoft YaHei', sans-serif;
        color: var(--text);
    }

    .stApp {
        background: #F5F5F7;
    }

    [data-testid="stSidebar"], [data-testid="collapsedControl"] {
        display: none !important;
    }

    div[data-testid="stDataFrame"] thead tr th,
    div[data-testid="stTable"] thead tr th {
        background: var(--tbl-header-bg) !important;
        color: var(--tbl-header-fg) !important;
        font-weight: var(--tbl-header-font-weight) !important;
        font-size: var(--tbl-header-font-size) !important;
        border-bottom: 1px solid var(--tbl-header-border) !important;
    }

    div[data-testid="stDataFrame"] thead tr th:hover,
    div[data-testid="stTable"] thead tr th:hover {
        background: var(--tbl-header-bg-hover) !important;
    }

    div[data-testid="stDataFrame"] thead tr th:active,
    div[data-testid="stTable"] thead tr th:active {
        box-shadow: var(--tbl-header-shadow) !important;
    }

    .out-kpi-card {
        background: linear-gradient(180deg, rgba(66,133,244,0.08) 0%, rgba(255,255,255,0.92) 60%, #FFFFFF 100%);
        border-radius: 14px;
        padding: 16px 16px 14px;
        border: 1px solid rgba(66,133,244,0.22);
        box-shadow: 0 10px 26px rgba(0,0,0,0.06);
        margin-bottom: 10px;
        position: relative;
        overflow: hidden;
    }
    .out-kpi-bar {
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 3px;
        background: linear-gradient(90deg, var(--tbl-header-bg) 0%, var(--success) 60%, var(--accent) 100%);
        opacity: 0.9;
    }
    .out-kpi-head { display:flex; align-items:center; gap:10px; margin-bottom: 10px; }
    .out-kpi-ico {
        width: 34px;
        height: 34px;
        border-radius: 10px;
        display:flex;
        justify-content:center;
        align-items:center;
        background: rgba(66,133,244,0.16);
        border: 1px solid rgba(66,133,244,0.28);
        color: var(--tbl-header-bg);
        font-weight: 900;
        font-size: 18px;
    }
    .out-kpi-title { font-size: 15px; color: rgba(27,21,48,0.78); font-weight: 800; letter-spacing: 0.2px; }
    .out-kpi-val { font-size: 26px; font-weight: 900; color: #1B1530; margin-bottom: 4px; }
    .out-kpi-sub { font-size: 13px; display:flex; justify-content:space-between; align-items:center; color: rgba(27,21,48,0.72); }
    .out-kpi-sub2 { font-size: 12px; display:flex; justify-content:space-between; align-items:center; color: rgba(27,21,48,0.62); margin-top: 4px; }
    .out-kpi-progress { background: rgba(27,21,48,0.10); border-radius: 999px; height: 6px; width: 100%; overflow: hidden; }
    .out-kpi-progress-bar { height: 100%; border-radius: 999px; }
    .trend-up { color: var(--success); font-weight: 800; }
    .trend-down { color: var(--danger); font-weight: 800; }
    .trend-neutral { color: rgba(27,21,48,0.72); font-weight: 800; }
    @media (max-width: 768px) {
        .out-kpi-card { padding: 14px 14px 12px; }
        .out-kpi-val { font-size: 22px; }
        .out-kpi-title { font-size: 14px; }
    }

    h1, h2, h3, h4, h5, h6 {
        color: var(--text);
        letter-spacing: 0.2px;
        text-shadow: none;
    }

    [data-testid="stAppViewContainer"] {
        color: var(--text);
    }

    /* Reset global text visibility */
    .stMarkdown, .stText, p, li, span, label {
        color: var(--text) !important;
        text-shadow: none;
    }
    
    /* Caption specific */
    .stCaption {
        color: var(--text-muted) !important;
    }

    /* --- SIDEBAR STYLING REMOVED TO RESTORE VISIBILITY --- */
    
    /* Only keep global metric styling that doesn't affect visibility */
    div[data-testid="stMetric"] {
        background: var(--panel);
        border: 1px solid var(--stroke);
        border-radius: var(--radius);
        box-shadow: var(--shadow-soft);
        padding: 18px;
    }

    div[data-testid="stMetric"] * {
        color: var(--text) !important;
    }
    
    div[data-testid="stMetric"] label {
        color: var(--text-muted) !important;
    }

    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        color: var(--primary) !important;
    }

    div[data-testid="stMetric"] div[data-testid="stMetricDelta"] {
        color: var(--accent-2) !important;
    }
    
    /* Buttons */
    div.stButton > button {
        border-radius: var(--radius-sm);
    }
    
    /* Analysis Button Customization */
    div.stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #FFC400 0%, #FFB000 100%) !important;
        border: 1px solid rgba(255, 176, 0, 0.4) !important;
        color: #5B2EA6 !important;
        font-weight: 700 !important;
        text-shadow: none !important;
        box-shadow: 0 4px 12px rgba(255, 196, 0, 0.25) !important;
        transition: all 0.2s ease !important;
    }

    div.stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #FFD54F 0%, #FFC107 100%) !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 16px rgba(255, 196, 0, 0.35) !important;
        border-color: rgba(255, 176, 0, 0.6) !important;
    }
    
    div.stButton > button[kind="primary"]:active {
        transform: translateY(1px) !important;
        box-shadow: 0 2px 8px rgba(255, 196, 0, 0.2) !important;
    }
    
    /* Tabs styling kept simple */
    .stTabs [data-baseweb="tab-list"] {
        gap: 10px;
    }

    /* Outbound subtabs (radio styled as tabs) */
    div[data-testid="stRadio"] .out-subtab-hint {display:none;}
    div[data-testid="stRadio"] [data-baseweb="radio"] > div {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        padding: 0 !important;
        margin: 0 !important;
    }
    div[data-testid="stRadio"] [data-baseweb="radio"] input {
        position: absolute !important;
        opacity: 0 !important;
        pointer-events: none !important;
    }
    div[data-testid="stRadio"] [data-baseweb="radio"] div[role="radio"] {
        display: none !important;
    }
    div[data-testid="stRadio"] [data-baseweb="radio"] span {
        font-weight: 600 !important;
        color: rgba(27, 21, 48, 0.75) !important;
    }
    div[data-testid="stRadio"] [data-baseweb="radio"] input:checked ~ div span {
        color: rgba(27, 21, 48, 0.95) !important;
    }
    div[data-testid="stRadio"] [data-testid="stRadio"] > div[role="radiogroup"] {
        border-bottom: 1px solid rgba(0, 0, 0, 0.08) !important;
        padding-bottom: 6px !important;
        gap: 10px !important;
    }
    div[data-testid="stRadio"] [data-baseweb="radio"] {
        position: relative !important;
        padding: 8px 0 10px 0 !important;
        margin-right: 14px !important;
    }
    div[data-testid="stRadio"] [data-baseweb="radio"] input:checked ~ div::after {
        content: "" !important;
        position: absolute !important;
        left: 0 !important;
        right: 0 !important;
        bottom: -7px !important;
        height: 2px !important;
        background: #E5484D !important;
        border-radius: 2px !important;
        transition: all 0.2s ease !important;
    }

    .out-subtab-content {
        animation: outFadeUp 240ms cubic-bezier(.2,.8,.2,1);
    }
    @keyframes outFadeUp {
        from { opacity: 0; transform: translateY(8px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    /* Ensure DataFrame styling is applied even if internal structure varies */
    [data-testid="stDataFrame"] {
        background: var(--panel);
        border: 1px solid var(--stroke);
        border-radius: var(--radius);
        box-shadow: var(--shadow-soft);
        overflow: hidden;
    }
    
    /* Target all possible table cells within the dataframe container */
    [data-testid="stDataFrame"] td, 
    [data-testid="stDataFrame"] th,
    [data-testid="stDataFrame"] [role="gridcell"],
    [data-testid="stDataFrame"] [role="columnheader"],
    [data-testid="stDataFrame"] div[data-testid="stDataFrameResizable"] {
        text-align: center !important;
        vertical-align: middle !important;
        color: var(--text) !important;
        display: flex;
        justify-content: center;
        align-items: center;
    }
    
    /* Force header content center */
    [data-testid="stDataFrame"] [role="columnheader"] > div {
        justify-content: center !important;
        text-align: center !important;
        width: 100%;
        display: flex;
    }
    
    /* Force cell content center */
    [data-testid="stDataFrame"] [role="gridcell"] > div {
        justify-content: center !important;
        text-align: center !important;
        width: 100%;
        display: flex;
    }

    /* Essential Visibility Controls */
    button[kind="header"], [data-testid="collapsedControl"] {
        visibility: visible !important;
        z-index: 999999 !important;
    }

    header {visibility: visible !important;}
    [data-testid="stSidebarNav"] {display: block !important;}
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    .stDeployButton {display:none;}
    [data-testid="stStatusWidget"] {visibility: hidden;}
    [data-testid="stToolbar"] {display: none !important;}
    [data-testid="stHeader"] {background: transparent !important;}
    [data-testid="stHeader"] a {display:none !important;}
    [data-testid="stViewerBadge"] {display:none !important;}
    [data-testid="stGitHubLink"] {display:none !important;}

    /* Sidebar explicit frosted light scheme to ensure readability */
    [data-testid="stSidebar"] {
        background: rgba(255, 255, 255, 0.88) !important;
        backdrop-filter: blur(10px) !important;
        border-right: 1px solid rgba(0, 0, 0, 0.08) !important;
    }
    [data-testid="stSidebar"] * {
        color: #333333 !important;
    }
    [data-testid="stSidebar"] summary svg,
    [data-testid="stSidebar"] svg {
        fill: #333333 !important;
    }
    [data-testid="stSidebar"] details[data-testid="stExpander"] > summary:hover {
        background: rgba(0,0,0,0.05) !important;
    }

    /* Sidebar inputs and selects */
    [data-testid="stSidebar"] input,
    [data-testid="stSidebar"] div[data-baseweb="select"] > div {
        background: rgba(255,255,255,0.95) !important;
        border: 1px solid rgba(0,0,0,0.15) !important;
        color: #333333 !important;
    }
    [data-testid="stSidebar"] input::placeholder {
        color: rgba(0,0,0,0.55) !important;
    }
    [data-testid="stSidebar"] div[data-baseweb="select"] > div:hover,
    [data-testid="stSidebar"] input:hover {
        border-color: rgba(91,46,166,0.6) !important;
    }

    /* File uploader dropzone */
    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
        background: rgba(255,255,255,0.92) !important;
        border: 1px dashed rgba(0,0,0,0.18) !important;
        color: #333333 !important;
    }
    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] span {
        color: #333333 !important;
    }
    [data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] {
        background: rgba(91,46,166,0.10) !important;
        color: #333333 !important;
        border: 1px solid rgba(91,46,166,0.35) !important;
    }

    /* Collapsed control arrow visibility and contrast */
    [data-testid="collapsedControl"] {
        color: #333333 !important;
        background: rgba(255,255,255,0.55) !important;
        border: 1px solid rgba(0,0,0,0.12) !important;
        border-radius: 6px !important;
        top: 56px !important;
    }

    @media (max-width: 768px) {
        div[data-testid="stMetric"] {
            padding: 14px;
        }
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
    div[data-testid="stDataFrame"] div[role="gridcell"] { display: flex; align-items: center; }
    div[data-testid="stDataFrame"] div[role="columnheader"] { display: flex; align-items: center; justify-content: center; }
    .msy-table-wrap {
        width: 100%;
        overflow-x: auto;
        border-radius: 12px;
        border: 1px solid rgba(0,0,0,0.08);
        background: rgba(255,255,255,0.9);
        box-shadow: 0 6px 18px rgba(0,0,0,0.06);
    }
    table.msy-table {
        width: 100%;
        border-collapse: collapse;
        table-layout: auto;
        font-size: 14px;
        line-height: 1.45;
    }
    table.msy-table thead th {
        position: sticky;
        top: 0;
        background: #1F2937;
        color: #F9FAFB;
        font-weight: 700;
        padding: 10px 12px;
        border: 1px solid rgba(255,255,255,0.12);
        text-align: center;
        vertical-align: middle;
        white-space: nowrap;
    }
    table.msy-table tbody td {
        padding: 10px 12px;
        border: 1px solid rgba(0,0,0,0.08);
        text-align: center;
        vertical-align: middle;
        font-variant-numeric: tabular-nums;
        white-space: nowrap;
    }
    table.msy-table tbody tr:nth-child(even) td {
        background: #F8FAFC;
    }
    table.msy-table tbody tr:hover td {
        background: #EEF2FF;
    }
</style>
""", unsafe_allow_html=True)

def _format_cell(v):
    if v is None or pd.isna(v):
        return ""
    if isinstance(v, (int, float, np.integer, np.floating)):
        return fmt_num(v, na="")
    return str(v)


# -----------------------------------------------------------------------------
# AgGrid Helper
# -----------------------------------------------------------------------------
JS_COLOR_CONDITIONAL = JsCode("""
function(params) {
    if (params.value > 0) {
        return {'color': '#28A745', 'textAlign': 'center', 'fontWeight': 'bold'};
    } else if (params.value < 0) {
        return {'color': '#DC3545', 'textAlign': 'center', 'fontWeight': 'bold'};
    }
    return {'textAlign': 'center'};
}
""")

JS_CENTER = JsCode("""
function(params) {
    return {'textAlign': 'center'};
}
""")

JS_FMT_NUM = JsCode("""
function(params) {
    const v = params.value;
    if (v === null || v === undefined) return '';
    const n = Number(v);
    if (!isFinite(n)) return '';
    const isInt = Math.abs(n - Math.round(n)) < 1e-9;
    const minF = isInt ? 0 : 1;
    const maxF = isInt ? 0 : 1;
    return n.toLocaleString('zh-CN', {minimumFractionDigits: minF, maximumFractionDigits: maxF});
}
""")

JS_FMT_NUM_1DP = JsCode("""
function(params) {
    const v = params.value;
    if (v === null || v === undefined) return '';
    const n = Number(v);
    if (!isFinite(n)) return '';
    return n.toLocaleString('zh-CN', {minimumFractionDigits: 1, maximumFractionDigits: 1});
}
""")

JS_FMT_PCT_RATIO = JsCode("""
function(params) {
    const v = params.value;
    if (v === null || v === undefined) return '';
    const n = Number(v);
    if (!isFinite(n)) return '';
    const p = n * 100;
    const isInt = Math.abs(p - Math.round(p)) < 1e-9;
    const minF = isInt ? 0 : 1;
    const maxF = isInt ? 0 : 1;
    return p.toLocaleString('zh-CN', {minimumFractionDigits: minF, maximumFractionDigits: maxF}) + '%';
}
""")

JS_FMT_PCT_RATIO_1DP = JsCode("""
function(params) {
    const v = params.value;
    if (v === null || v === undefined) return '';
    const n = Number(v);
    if (!isFinite(n)) return '';
    const p = n * 100;
    return p.toLocaleString('zh-CN', {minimumFractionDigits: 1, maximumFractionDigits: 1}) + '%';
}
""")

# Custom Cell Renderer for Progress Bar (Mockup using HTML)
JS_PROGRESS_BAR = JsCode("""
class ProgressBarRenderer {
    init(params) {
        this.eGui = document.createElement('div');
        this.eGui.style.width = '100%';
        this.eGui.style.height = '100%';
        this.eGui.style.display = 'flex';
        this.eGui.style.alignItems = 'center';
        
        const fmt1 = (v) => {
            if (v === null || v === undefined) return '';
            const n = Number(v);
            if (!isFinite(n)) return '';
            if (Math.abs(n - Math.round(n)) < 1e-9) return Math.round(n).toLocaleString('zh-CN');
            return n.toLocaleString('zh-CN', {minimumFractionDigits: 1, maximumFractionDigits: 1});
        };

        let value = params.value;
        if (value === null || value === undefined) value = 0;

        let maxValue = 0;
        if (params.colDef && params.colDef.cellRendererParams && params.colDef.cellRendererParams.maxValue !== undefined) {
            maxValue = Number(params.colDef.cellRendererParams.maxValue) || 0;
        }

        const percent = maxValue > 0 ? Math.min(Math.max((Number(value) / maxValue) * 100, 0), 100) : 0;

        let color = '#007bff';
        if (percent >= 100) color = '#28a745';
        else if (percent < 60) color = '#dc3545';
        else color = '#ffc107';
        
        this.eGui.innerHTML = `
            <div style="width: 100%; height: 20px; background-color: #e9ecef; border-radius: 3px; position: relative;">
                <div style="width: ${percent}%; height: 100%; background-color: ${color}; border-radius: 3px; transition: width 0.5s;"></div>
                <span style="position: absolute; top: 0; left: 0; width: 100%; height: 100%; text-align: center; line-height: 20px; font-size: 12px; color: #000;">${fmt1(value)}</span>
            </div>
        `;
    }
    getGui() {
        return this.eGui;
    }
}
""")

# Custom Cell Renderer for Count (No %)
JS_PROGRESS_BAR_COUNT = JsCode("""
class ProgressBarCountRenderer {
    init(params) {
        this.eGui = document.createElement('div');
        this.eGui.style.width = '100%';
        this.eGui.style.height = '100%';
        this.eGui.style.display = 'flex';
        this.eGui.style.alignItems = 'center';
        
        const fmt1 = (v) => {
            if (v === null || v === undefined) return '';
            const n = Number(v);
            if (!isFinite(n)) return '';
            if (Math.abs(n - Math.round(n)) < 1e-9) return Math.round(n).toLocaleString('zh-CN');
            return n.toLocaleString('zh-CN', {minimumFractionDigits: 1, maximumFractionDigits: 1});
        };

        let value = params.value;
        if (value === null || value === undefined) value = 0;

        let maxValue = 0;
        if (params.colDef && params.colDef.cellRendererParams && params.colDef.cellRendererParams.maxValue !== undefined) {
            maxValue = Number(params.colDef.cellRendererParams.maxValue) || 0;
        }

        const percent = maxValue > 0 ? Math.min(Math.max((Number(value) / maxValue) * 100, 0), 100) : 0;

        let color = '#28a745';
        if (percent > 0) color = '#ffc107';
        if (percent >= 60) color = '#dc3545';
        
        this.eGui.innerHTML = `
            <div style="width: 100%; height: 20px; background-color: #e9ecef; border-radius: 3px; position: relative;">
                <div style="width: ${percent}%; height: 100%; background-color: ${color}; border-radius: 3px; transition: width 0.5s;"></div>
                <span style="position: absolute; top: 0; left: 0; width: 100%; height: 100%; text-align: center; line-height: 20px; font-size: 12px; color: #000;">${fmt1(value)}</span>
            </div>
        `;
    }
    getGui() {
        return this.eGui;
    }
}
""")

JS_SPARKLINE = JsCode("""
class SparklineRenderer {
    init(params) {
        this.eGui = document.createElement('div');
        this.eGui.style.width = '100%';
        this.eGui.style.height = '100%';
        this.eGui.style.display = 'flex';
        this.eGui.style.alignItems = 'center';
        this.eGui.style.justifyContent = 'center';

        let raw = params.value;
        let arr = [];
        try {
            if (Array.isArray(raw)) arr = raw;
            else if (typeof raw === 'string') arr = JSON.parse(raw);
            else arr = [];
        } catch (e) {
            arr = [];
        }
        arr = (arr || []).map(v => Number(v) || 0);
        const w = (params.colDef && params.colDef.cellRendererParams && params.colDef.cellRendererParams.width) ? params.colDef.cellRendererParams.width : 120;
        const h = (params.colDef && params.colDef.cellRendererParams && params.colDef.cellRendererParams.height) ? params.colDef.cellRendererParams.height : 28;
        const pad = 3;
        const n = arr.length;
        if (!n) {
            this.eGui.innerHTML = '';
            return;
        }
        let minV = Math.min(...arr);
        let maxV = Math.max(...arr);
        const range = (maxV - minV) || 1;
        const denom = Math.max(1, n - 1);
        const pts = arr.map((v, i) => {
            const x = pad + (i * (w - pad * 2)) / denom;
            const y = (h - pad) - ((v - minV) / range) * (h - pad * 2);
            return `${x},${y}`;
        }).join(' ');
        this.eGui.innerHTML = `
            <svg width="${w}" height="${h}" viewBox="0 0 ${w} ${h}" style="display:block;">
                <polyline points="${pts}" fill="none" stroke="#f97316" stroke-width="2" stroke-linejoin="round" stroke-linecap="round"></polyline>
            </svg>
        `;
    }
    getGui() {
        return this.eGui;
    }
}
""")

JS_TREND_TAG = JsCode("""
class TrendTagRenderer {
    init(params) {
        this.eGui = document.createElement('div');
        this.eGui.style.width = '100%';
        this.eGui.style.height = '100%';
        this.eGui.style.display = 'flex';
        this.eGui.style.alignItems = 'center';
        this.eGui.style.justifyContent = 'center';
        const v = (params.value === null || params.value === undefined) ? '' : String(params.value);
        let bg = '#f3f4f6';
        let fg = '#374151';
        if (v.indexOf('持续增长') >= 0 || v.indexOf('持续升级') >= 0) { bg = '#dcfce7'; fg = '#166534'; }
        else if (v.indexOf('持续下滑') >= 0 || v.indexOf('持续降级') >= 0) { bg = '#fee2e2'; fg = '#991b1b'; }
        else if (v.indexOf('先下滑后增长') >= 0 || v.indexOf('先降级后升级') >= 0) { bg = '#dbeafe'; fg = '#1d4ed8'; }
        else if (v.indexOf('先增长后下滑') >= 0 || v.indexOf('先升级后降级') >= 0) { bg = '#fef9c3'; fg = '#a16207'; }
        else if (v.indexOf('持续持平') >= 0) { bg = '#e5e7eb'; fg = '#374151'; }
        else if (v.indexOf('持平升级') >= 0 || v.indexOf('升级持平') >= 0) { bg = '#dcfce7'; fg = '#166534'; }
        else if (v.indexOf('持平降级') >= 0 || v.indexOf('降级持平') >= 0) { bg = '#fee2e2'; fg = '#991b1b'; }
        this.eGui.innerHTML = `
            <span style="display:inline-flex; align-items:center; justify-content:center; padding:2px 10px; border-radius:999px; background:${bg}; color:${fg}; font-size:12px; font-weight:800; border:1px solid rgba(0,0,0,0.06);">
                ${v || '—'}
            </span>
        `;
    }
    getGui() {
        return this.eGui;
    }
}
""")

JS_STORE_TAG = JsCode("""
class StoreTagRenderer {
    init(params) {
        this.eGui = document.createElement('div');
        this.eGui.style.width = '100%';
        this.eGui.style.height = '100%';
        this.eGui.style.display = 'flex';
        this.eGui.style.alignItems = 'center';
        this.eGui.style.justifyContent = 'center';
        const raw = (params.value === null || params.value === undefined) ? '' : String(params.value);
        const v = raw.trim();
        let bg = '#f3f4f6';
        let fg = '#374151';
        let letter = '';
        const m = v.match(/[A-D]/i);
        if (m && m[0]) letter = m[0].toUpperCase();
        if (letter === 'A') { bg = '#dcfce7'; fg = '#166534'; }
        else if (letter === 'B') { bg = '#dbeafe'; fg = '#1d4ed8'; }
        else if (letter === 'C') { bg = '#fef9c3'; fg = '#a16207'; }
        else if (letter === 'D') { bg = '#fee2e2'; fg = '#991b1b'; }
        this.eGui.innerHTML = `
            <span style="display:inline-flex; align-items:center; justify-content:center; padding:2px 10px; border-radius:999px; background:${bg}; color:${fg}; font-size:12px; font-weight:800; border:1px solid rgba(0,0,0,0.06);">
                ${v || '—'}
            </span>
        `;
    }
    getGui() {
        return this.eGui;
    }
}
""")

def show_aggrid_table(df, height=None, key=None, on_row_selected=None, 
                      columns_props=None, 
                      column_defs=None,
                      grid_options_overrides=None,
                      auto_height_limit=2000):
    """
    Standardized AgGrid Table
    :param df: DataFrame to display
    :param height: Fixed height (optional)
    :param key: Unique key
    :param on_row_selected: 'single' or 'multiple' or None
    :param columns_props: Dict of col_name -> {type: 'percent'|'money'|'growth'|'bar', ...}
    :param auto_height_limit: Max height for auto calculation
    """
    if df is None or df.empty:
        # Custom Empty State
        st.markdown("""
            <div style="text-align: center; padding: 40px; background: #f8f9fa; border-radius: 8px; border: 1px dashed #d9d9d9;">
                <div style="font-size: 24px; margin-bottom: 10px;">📭</div>
                <div style="color: #666; font-size: 14px;">暂无数据</div>
            </div>
        """, unsafe_allow_html=True)
        return None

    # Inject CSS for Custom AgGrid Styling
    st.markdown("""
        <style>
        /* --- 1. Header Styling --- */
        .ag-header {
            background-color: var(--tbl-header-bg) !important;
            border-bottom: 1px solid var(--tbl-header-border) !important;
        }
        .ag-header-row,
        .ag-header-group-cell,
        .ag-header-cell {
            background-color: var(--tbl-header-bg) !important;
        }
        .ag-header-group-cell:hover,
        .ag-header-cell:hover {
            background-color: var(--tbl-header-bg-hover) !important;
        }
        .ag-header-group-cell:active,
        .ag-header-cell:active {
            box-shadow: var(--tbl-header-shadow) !important;
        }
        .ag-header-cell {
            color: var(--tbl-header-fg) !important;
            font-family: 'Inter', 'Microsoft YaHei', sans-serif !important;
            font-size: var(--tbl-header-font-size) !important;
            font-weight: var(--tbl-header-font-weight) !important;
            padding: 0 12px !important;
        }
        .ag-header-group-cell {
            color: var(--tbl-header-fg) !important;
            font-family: 'Inter', 'Microsoft YaHei', sans-serif !important;
            font-size: var(--tbl-header-font-size) !important;
            font-weight: var(--tbl-header-font-weight) !important;
        }
        .ag-header-cell .ag-icon,
        .ag-header-group-cell .ag-icon,
        .ag-sort-indicator-icon,
        .ag-icon-asc,
        .ag-icon-desc,
        .ag-icon-menu {
            color: var(--tbl-header-icon) !important;
            fill: var(--tbl-header-icon) !important;
            opacity: 1 !important;
        }
        /* Strict Centering for Header */
        .ag-header-cell-label {
            display: flex !important;
            justify-content: center !important;
            align-items: center !important;
            text-align: center !important;
            width: 100% !important;
        }
        .ag-header-cell-label, .ag-header-cell-text {
            white-space: normal !important;
            overflow: visible !important;
            text-overflow: clip !important;
            line-height: 1.2 !important;
        }
        .ag-header-cell-text {
            font-size: 12px !important;
        }

        .ag-header-cell.hdr-a,
        .ag-header-group-cell.hdr-a {
            background-color: #16A34A !important;
            color: #FFFFFF !important;
        }
        .ag-header-cell.hdr-b,
        .ag-header-group-cell.hdr-b {
            background-color: #2563EB !important;
            color: #FFFFFF !important;
        }
        .ag-header-cell.hdr-c,
        .ag-header-group-cell.hdr-c {
            background-color: #F59E0B !important;
            color: #111827 !important;
        }
        .ag-header-cell.hdr-d,
        .ag-header-group-cell.hdr-d {
            background-color: #DC2626 !important;
            color: #FFFFFF !important;
        }
        .ag-header-cell.hdr-up,
        .ag-header-group-cell.hdr-up {
            background-color: #16A34A !important;
            color: #FFFFFF !important;
        }
        .ag-header-cell.hdr-du,
        .ag-header-group-cell.hdr-du {
            background-color: #2563EB !important;
            color: #FFFFFF !important;
        }
        .ag-header-cell.hdr-ud,
        .ag-header-group-cell.hdr-ud {
            background-color: #F59E0B !important;
            color: #111827 !important;
        }
        .ag-header-cell.hdr-down,
        .ag-header-group-cell.hdr-down {
            background-color: #DC2626 !important;
            color: #FFFFFF !important;
        }
        .ag-header-cell.hdr-flat,
        .ag-header-group-cell.hdr-flat {
            background-color: #E5E7EB !important;
            color: #111827 !important;
        }
        .ag-header-cell.hdr-a .ag-header-cell-text,
        .ag-header-cell.hdr-b .ag-header-cell-text,
        .ag-header-cell.hdr-d .ag-header-cell-text,
        .ag-header-cell.hdr-up .ag-header-cell-text,
        .ag-header-cell.hdr-du .ag-header-cell-text,
        .ag-header-cell.hdr-down .ag-header-cell-text,
        .ag-header-group-cell.hdr-a .ag-header-cell-text,
        .ag-header-group-cell.hdr-b .ag-header-cell-text,
        .ag-header-group-cell.hdr-d .ag-header-cell-text,
        .ag-header-group-cell.hdr-up .ag-header-cell-text,
        .ag-header-group-cell.hdr-du .ag-header-cell-text,
        .ag-header-group-cell.hdr-down .ag-header-cell-text {
            color: #FFFFFF !important;
        }
        .ag-header-cell.hdr-c .ag-header-cell-text,
        .ag-header-cell.hdr-ud .ag-header-cell-text,
        .ag-header-cell.hdr-flat .ag-header-cell-text,
        .ag-header-group-cell.hdr-c .ag-header-cell-text,
        .ag-header-group-cell.hdr-ud .ag-header-cell-text,
        .ag-header-group-cell.hdr-flat .ag-header-cell-text {
            color: #111827 !important;
        }
        .ag-header-cell.hdr-a .ag-icon,
        .ag-header-cell.hdr-b .ag-icon,
        .ag-header-cell.hdr-d .ag-icon,
        .ag-header-cell.hdr-up .ag-icon,
        .ag-header-cell.hdr-du .ag-icon,
        .ag-header-cell.hdr-down .ag-icon,
        .ag-header-group-cell.hdr-a .ag-icon,
        .ag-header-group-cell.hdr-b .ag-icon,
        .ag-header-group-cell.hdr-d .ag-icon,
        .ag-header-group-cell.hdr-up .ag-icon,
        .ag-header-group-cell.hdr-du .ag-icon,
        .ag-header-group-cell.hdr-down .ag-icon {
            color: #FFFFFF !important;
            fill: #FFFFFF !important;
        }
        .ag-header-cell.hdr-c .ag-icon,
        .ag-header-cell.hdr-ud .ag-icon,
        .ag-header-cell.hdr-flat .ag-icon,
        .ag-header-group-cell.hdr-c .ag-icon,
        .ag-header-group-cell.hdr-ud .ag-icon,
        .ag-header-group-cell.hdr-flat .ag-icon {
            color: #111827 !important;
            fill: #111827 !important;
        }

        /* Strict Centering for Cells */
        .ag-cell, .ag-cell-value {
            display: flex !important;
            justify-content: center !important;
            align-items: center !important;
            text-align: center !important;
        }
        
        /* Remove default separator bars in header */
        .ag-header-cell::after, .ag-header-group-cell::after {
            display: none !important;
        }

        /* --- 2. Row & Cell Styling --- */
        .ag-row {
            font-family: 'Inter', 'Microsoft YaHei', sans-serif !important;
            font-size: var(--tbl-cell-font-size) !important;
            color: #333333 !important;
            border-bottom-color: #f0f0f0 !important;
        }
        .ag-row-odd {
            background-color: #f8f9fa !important;
        }
        .ag-row-even {
            background-color: #ffffff !important;
        }
        .ag-row-hover {
            background-color: #f0f7ff !important;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important;
            z-index: 5;
        }
        .ag-row-selected {
            background-color: #e6f7ff !important;
            border-left: 2px solid #4096ff !important; /* Left highlight */
        }
        
        /* Removed duplicate .ag-cell rule, handled above */

        /* Selected Row Text */
        .ag-row-selected .ag-cell {
            font-weight: 500 !important;
        }

        .ag-row.ag-row-pinned,
        .ag-row.ag-row-pinned-bottom {
            background-color: var(--tbl-header-bg) !important;
        }
        .ag-row-pinned .ag-cell,
        .ag-row-pinned-bottom .ag-cell {
            color: var(--tbl-header-fg) !important;
            font-weight: 900 !important;
            border-top: 1px solid var(--tbl-header-border) !important;
        }
        .ag-row-pinned .ag-cell .ag-cell-value,
        .ag-row-pinned-bottom .ag-cell .ag-cell-value {
            color: var(--tbl-header-fg) !important;
            font-weight: 900 !important;
        }
        .ag-row-pinned .ag-cell .ag-icon,
        .ag-row-pinned-bottom .ag-cell .ag-icon {
            color: var(--tbl-header-icon) !important;
            fill: var(--tbl-header-icon) !important;
        }
        
        /* --- 3. Container & Borders --- */
        .ag-root-wrapper {
            border: 1px solid #e5e6eb !important;
            border-radius: 4px !important;
            overflow: hidden !important; /* For radius */
        }
        
        /* --- 4. Scrollbars (Optional, for better look) --- */
        .ag-body-viewport::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }
        .ag-body-viewport::-webkit-scrollbar-thumb {
            background: #ccc;
            border-radius: 4px;
        }
        .ag-body-viewport::-webkit-scrollbar-track {
            background: #f1f1f1;
        }

        /* --- 5. Mobile Optimization --- */
        @media (max-width: 768px) {
            .ag-header-cell {
                font-size: 13px !important;
                padding: 0 4px !important;
            }
            .ag-header-group-cell {
                font-size: 13px !important;
            }
            .ag-cell {
                font-size: 12px !important;
                padding: 0 4px !important;
            }
            .ag-header-group-cell:active,
            .ag-header-cell:active {
                box-shadow: none !important;
            }
        }
        </style>
    """, unsafe_allow_html=True)

    gb = GridOptionsBuilder.from_dataframe(df)

    percent_cols = set()
    if columns_props:
        for col, props in columns_props.items():
            c_type = (props or {}).get('type')
            if c_type in ['percent', 'growth']:
                percent_cols.add(col)
    for col in df.columns:
        if ('同比' in str(col)) or ('增长' in str(col)) or ('达成率' in str(col)) or (str(col).endswith('率')):
            percent_cols.add(col)
    
    total_row = {c: None for c in df.columns}
    if len(df.columns) > 0:
        total_row[df.columns[0]] = '合计'
    yoy_cols = [c for c in df.columns if ('同比' in str(c)) or (str(c) == '同比增长')]

    for c in df.columns:
        if c == df.columns[0]:
            continue
        if c in percent_cols:
            continue
        s = pd.to_numeric(df[c], errors='coerce')
        if s.notna().sum() == 0:
            continue
        total_row[c] = float(s.fillna(0).sum())

    def _infer_yoy_pair(yoy_col: str):
        if yoy_col not in df.columns:
            return None

        for old, new in [
            ('同比(箱)', '箱数'),
            ('同比（箱）', '箱数'),
            ('同比(门店)', '门店数'),
            ('同比（门店）', '门店数'),
        ]:
            if old in str(yoy_col):
                cur = str(yoy_col).replace(old, new)
                last = str(yoy_col).replace(old, '同期(箱数)' if '箱' in old else '同期(门店数)')
                if cur in df.columns and last in df.columns:
                    return cur, last

        if str(yoy_col) == '同比增长':
            for cur, last in [
                ('本月', '同期'),
                ('本月业绩', '同期业绩'),
                ('本月(万)', '同期(万)'),
                ('本月业绩(万)', '同期业绩(万)'),
                ('实际', '同期'),
            ]:
                if cur in df.columns and last in df.columns:
                    return cur, last

        base = (
            str(yoy_col)
            .replace('同比增长', '')
            .replace('同比', '')
            .replace('增长', '')
            .strip()
        )
        if not base:
            return None
        last_candidates = [c for c in df.columns if ('同期' in str(c) or '去年' in str(c)) and base in str(c)]
        cur_candidates = [c for c in df.columns if ('同期' not in str(c) and '去年' not in str(c) and '同比' not in str(c) and '增长' not in str(c)) and base in str(c)]
        if len(cur_candidates) == 1 and len(last_candidates) == 1:
            return cur_candidates[0], last_candidates[0]
        return None

    for c in yoy_cols:
        pair = _infer_yoy_pair(c)
        if not pair:
            continue
        cur_col, last_col = pair
        try:
            cur_sum = float(pd.to_numeric(df[cur_col], errors='coerce').fillna(0).sum())
            last_sum = float(pd.to_numeric(df[last_col], errors='coerce').fillna(0).sum())
            total_row[c] = (cur_sum - last_sum) / last_sum if last_sum > 0 else None
        except Exception:
            total_row[c] = None
    
    # Configure General Options
    gb.configure_grid_options(
        rowHeight=40, # increased for padding
        headerHeight=60,
        animateRows=True,
        suppressCellFocus=True, # remove blue outline on click
        enableCellTextSelection=True,
        suppressDragLeaveHidesColumns=True,
        sideBar={
            "toolPanels": [
                {
                    "id": "columns",
                    "labelDefault": "列",
                    "iconKey": "columns",
                    "toolPanel": "agColumnsToolPanel",
                    "toolPanelParams": {
                        "suppressRowGroups": True,
                        "suppressValues": True,
                        "suppressPivots": True,
                        "suppressPivotMode": True
                    }
                }
            ],
            "defaultToolPanel": None
        }
    )
    
    # Default Config: Centered, Resizable, Sortable, Filterable
    gb.configure_default_column(
        resizable=True,
        filterable=True,
        sortable=True,
        cellStyle=JS_CENTER,
        headerClass='ag-header-center',
        headerStyle={'textAlign': 'center', 'justifyContent': 'center'},
        wrapHeaderText=True,
        autoHeaderHeight=True,
        minWidth=70,
        flex=1
    )
    
    configured_cols = set()

    # Apply Column Specific Props
    if columns_props:
        for col, props in columns_props.items():
            if col not in df.columns:
                continue
            
            c_type = props.get('type')
            max_value = None
            if c_type in ("bar", "bar_count"):
                s = pd.to_numeric(df[col], errors='coerce')
                max_value = float(s.max()) if len(s) and pd.notna(s.max()) else 0.0
            
            if c_type == 'growth':
                gb.configure_column(col, 
                                    cellStyle=JS_COLOR_CONDITIONAL, 
                                    type=["numericColumn", "numberColumnFilter"], 
                                    valueFormatter=JS_FMT_PCT_RATIO,
                                    minWidth=70,
                                    flex=1)
                configured_cols.add(col)
            elif c_type == 'percent':
                 gb.configure_column(col, 
                                    type=["numericColumn", "numberColumnFilter"], 
                                    valueFormatter=JS_FMT_PCT_RATIO,
                                    minWidth=70,
                                    flex=1)
                 configured_cols.add(col)
            elif c_type == 'money':
                gb.configure_column(col, 
                                    type=["numericColumn", "numberColumnFilter"], 
                                    valueFormatter=JS_FMT_NUM,
                                    minWidth=70,
                                    flex=1)
                configured_cols.add(col)
            elif c_type == 'bar':
                # Use custom renderer
                gb.configure_column(col, 
                                    cellRenderer=JS_PROGRESS_BAR,
                                    cellRendererParams={'maxValue': max_value},
                                    type=["numericColumn", "numberColumnFilter"],
                                    valueFormatter=JS_FMT_NUM,
                                    minWidth=70,
                                    flex=1)
                configured_cols.add(col)
            elif c_type == 'bar_count':
                # Use custom renderer for count
                gb.configure_column(col, 
                                    cellRenderer=JS_PROGRESS_BAR_COUNT,
                                    cellRendererParams={'maxValue': max_value},
                                    type=["numericColumn", "numberColumnFilter"],
                                    valueFormatter=JS_FMT_NUM,
                                    minWidth=70,
                                    flex=1)
                configured_cols.add(col)
                
    # Generic Auto-Type Logic (Fallbacks)
    for col in df.columns:
        if col in configured_cols:
            continue
        
        # Check if column has 'growth' or '同比' -> Growth Color
        if '同比' in col or '增长' in col:
            gb.configure_column(col, 
                                cellStyle=JS_COLOR_CONDITIONAL, 
                                type=["numericColumn", "numberColumnFilter"], 
                                valueFormatter=JS_FMT_PCT_RATIO,
                                minWidth=70,
                                flex=1)
        
        # Check if '达成率' or '率' -> Percent
        elif '达成率' in col or '占比' in col or str(col).endswith('率'):
            gb.configure_column(col, 
                                type=["numericColumn", "numberColumnFilter"], 
                                valueFormatter=JS_FMT_PCT_RATIO,
                                minWidth=70,
                                flex=1)
            
            # Optional: Add Data Bar style for '达成率' if requested
            if '达成率' in col:
                 gb.configure_column(col,
                    cellStyle=JsCode("""
                        function(params) {
                            let ratio = params.value;
                            if (ratio === null || isNaN(ratio)) return {'textAlign': 'center'};
                            let percent = ratio * 100;
                             let color = '#28a745'; // Green
                             if (percent < 100) color = '#ffc107'; // Yellow
                             if (percent < 60) color = '#dc3545'; // Red
                             return {
                                 'textAlign': 'center', 
                                 'background': `linear-gradient(90deg, ${color} ${Math.min(percent, 100)}%, transparent ${Math.min(percent, 100)}%)`
                             };
                        }
                    """),
                    valueFormatter=JS_FMT_PCT_RATIO,
                    minWidth=70,
                    flex=1
                 )

        # Money/Number
        elif pd.api.types.is_numeric_dtype(df[col]):
            gb.configure_column(col, 
                                type=["numericColumn", "numberColumnFilter"], 
                                valueFormatter=JS_FMT_NUM,
                                minWidth=70,
                                flex=1)
        else:
            if col == df.columns[0]:
                gb.configure_column(col, minWidth=95, flex=1.2, tooltipField=col)
            else:
                gb.configure_column(col, minWidth=100, flex=1.2, tooltipField=col)

    # Selection
    if on_row_selected:
        gb.configure_selection('single', use_checkbox=False)
        
    gridOptions = gb.build()
    gridOptions['pinnedBottomRowData'] = [total_row]
    if column_defs:
        gridOptions['columnDefs'] = column_defs
        gridOptions['groupHeaderHeight'] = 40
        gridOptions['headerHeight'] = 46
    if grid_options_overrides:
        gridOptions.update(grid_options_overrides)
    
    # --- Auto Height & Pagination Logic ---
    # 1. Calculate ideal height for all rows
    n_rows = len(df)
    row_h = 40  # consistent with configure_grid_options rowHeight
    header_h = 60 # consistent with configure_grid_options headerHeight
    padding = 20
    
    calc_full_height = header_h + (n_rows * row_h) + padding + 40 # +40 buffer for potential horizontal scrollbar/total row
    
    # 2. Thresholds
    MAX_HEIGHT_NO_SCROLL = 600  # If content < 600px, show full height (no scroll/pagination)
    PAGE_SIZE = 20              # If content > 600px, use pagination with 20 rows/page
    
    # 3. Determine Mode
    # If explicit height provided, use it (and scroll if needed)
    # Else, apply auto-logic
    if height:
        final_height = height
        # If explicitly short height, maybe enable pagination? No, trust caller or AgGrid default scroll.
    else:
        if calc_full_height <= MAX_HEIGHT_NO_SCROLL:
            final_height = max(150, calc_full_height) # At least 150px
            # No pagination needed
            gridOptions['pagination'] = False
        else:
            # Content too long -> Use Pagination
            gridOptions['pagination'] = True
            gridOptions['paginationPageSize'] = PAGE_SIZE
            # Height fits PageSize rows + Header + PaginationPanel
            # PageSize * RowHeight + Header + PagerPanel(~50px)
            final_height = (PAGE_SIZE * row_h) + header_h + 50 + padding
    
    # Enable SideBar for Columns Tool Panel (Optional, user asked for "Drop down menu for each column")
    # AgGrid default filter menu is on column header. 
    
    # --- Responsive & Horizontal Scroll Logic ---
    # If too many columns, disable 'fit_columns_on_grid_load' to allow horizontal scroll
    # Heuristic: > 8-10 columns or if we suspect wide content
    should_fit_columns = True
    if len(df.columns) > 10:
        should_fit_columns = False
    
    return AgGrid(
        df,
        gridOptions=gridOptions,
        height=final_height,
        width='100%',
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        update_mode=GridUpdateMode.SELECTION_CHANGED | GridUpdateMode.VALUE_CHANGED,
        fit_columns_on_grid_load=should_fit_columns,
        allow_unsafe_jscode=True, 
        theme='streamlit', 
        key=key
    )

# -----------------------------------------------------------------------------
# 3. Data Logic
# -----------------------------------------------------------------------------

@st.cache_data(show_spinner=False)
def load_data_v2(file_bytes: bytes, file_name: str):
    debug_logs = []
    try:
        file_name_lower = (file_name or "").lower()
        bio = io.BytesIO(file_bytes)
        if file_name_lower.endswith('.csv'):
            df = pd.read_csv(bio, encoding='gb18030')
            df_stock = None
            df_q4_raw = None
            df_perf_raw = None
        else:
            xl = pd.ExcelFile(bio)
            df = xl.parse(0)
            df_stock = xl.parse(1) if len(xl.sheet_names) > 1 else None
            df_q4_raw = xl.parse(2) if len(xl.sheet_names) > 2 else None
            df_perf_raw = None
            
            debug_logs.append(f"Total Sheets: {len(xl.sheet_names)} | Names: {xl.sheet_names}")

            # Sheet 4 Detection Logic (Robust)
            if len(xl.sheet_names) > 3:
                preferred = next((s for s in xl.sheet_names if 'sheet4' in str(s).strip().lower()), None)
                candidate_names = [preferred] if preferred else []
                candidate_names += [s for s in xl.sheet_names if s not in candidate_names]
                
                for sname in candidate_names:
                    try:
                        # Optimization: Read only header first (0 rows) to check columns
                        tmp_header = xl.parse(sname, nrows=0)
                        cols = [str(c).strip() for c in tmp_header.columns]
                    except Exception as e:
                        debug_logs.append(f"Error parsing header of {sname}: {str(e)}")
                        continue
                    
                    # Fuzzy match for keys
                    key_hits = sum(1 for k in ['年份', '月份', '省区'] if any(k in c for c in cols))
                    signal_hits = sum(1 for k in ['发货仓', '原价金额', '基本数量', '大分类', '月分析', '客户简称'] if any(k in c for c in cols))
                    
                    debug_logs.append(f"Checking '{sname}': keys={key_hits}, signals={signal_hits}")
                    
                    if key_hits >= 2 and signal_hits >= 1:
                        # Found it! Now read the full sheet
                        try:
                            df_perf_raw = xl.parse(sname)
                            debug_logs.append(f"-> MATCHED Sheet4: {sname}")
                            break
                        except Exception as e:
                            debug_logs.append(f"Error reading body of {sname}: {e}")
            else:
                 debug_logs.append("Warning: Less than 4 sheets found.")
            
        # --- Process Sheet 1 (Sales) ---
        # Ensure column names are clean
        df.columns = [str(c).strip() for c in df.columns]
        
        # --- Handle Long Format (Rows) -> Wide Format (Columns) ---
        # User indicates Time (Month) is in Column F (index 5)
        # Potential Columns: F=Time, I=Prov, J=Dist, K=Qty (based on user info)
        is_long_format = False
        time_col = None
        
        # Check if Column F exists and looks like Month
        if len(df.columns) > 5:
            col_f = df.columns[5]
            # Check a sample of values in Col F for "月" or date-like
            sample_vals = df[col_f].dropna().head(10).astype(str).tolist()
            if any('月' in v for v in sample_vals):
                is_long_format = True
                time_col = col_f
        
        if is_long_format:
            # Identify Key Columns for Pivot
            # Try to map by name or index
            # User hints: Prov(I=8), Dist(J=9), Qty(K=10)
            
            col_prov = df.columns[8] if len(df.columns) > 8 else None
            col_dist = df.columns[9] if len(df.columns) > 9 else None
            col_qty = df.columns[10] if len(df.columns) > 10 else None
            
            # Fallback: Search by name
            if col_prov is None: col_prov = next((c for c in df.columns if '省' in c), None)
            if col_dist is None: col_dist = next((c for c in df.columns if '经销' in c or '客户' in c), None)
            if col_qty is None: col_qty = next((c for c in df.columns if '数' in c or 'Qty' in c or '箱' in c), None)
            
            # Store Column? If not found, default to Dist or blank
            col_store = next((c for c in df.columns if '门店' in c), None)
            
            if col_prov and col_dist and col_qty and time_col:
                # Prepare for Pivot
                pivot_index = [col_prov, col_dist]
                if col_store:
                    pivot_index.append(col_store)
                
                # Pivot
                # Ensure Qty is numeric
                df[col_qty] = pd.to_numeric(df[col_qty], errors='coerce').fillna(0)
                
                df_wide = df.pivot_table(
                    index=pivot_index,
                    columns=time_col,
                    values=col_qty,
                    aggfunc='sum'
                ).reset_index()
                
                # Handle Missing Store Column if needed
                if not col_store:
                    df_wide['门店名称'] = df_wide[col_dist] # Use Dist as Store if missing
                    
                df = df_wide
                # Reset clean columns
                df.columns = [str(c).strip() for c in df.columns]
                
        # Identify Month Columns (Assume '1月', '2月', etc. or columns 4 onwards if strict structure)
        # Based on user requirement: Col 1-3 are dimensions, 4+ are months.
        # Let's try to detect "X月" pattern first, fallback to index.
        month_cols = [c for c in df.columns if '月' in c and c not in ['品牌省区名称', '经销商名称', '门店名称']]
        
        # If headers are standard: 品牌省区名称, 经销商名称, 门店名称
        # Normalize dimension columns
        rename_map = {}
        if '品牌省区名称' in df.columns: rename_map['品牌省区名称'] = '省区'
        if '经销商名称' not in df.columns and len(df.columns) > 1: rename_map[df.columns[1]] = '经销商名称'
        if '门店名称' not in df.columns and len(df.columns) > 2: rename_map[df.columns[2]] = '门店名称'
        
        df = df.rename(columns=rename_map)
        
        # Validate critical columns
        required = ['省区', '经销商名称', '门店名称']
        for req in required:
            if req not in df.columns:
                # Fallback: Assume positional 0, 1, 2
                if len(df.columns) >= 3:
                    df.columns.values[0] = '省区'
                    df.columns.values[1] = '经销商名称'
                    df.columns.values[2] = '门店名称'
                else:
                    st.error(f"数据格式错误：缺失列 {req}")
                    return None, None, None, None

        # Re-identify month cols after rename
        month_cols = [c for c in df.columns if c not in required]
        
        # Ensure numeric
        for col in month_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
        # --- Core Metric Calculation ---
        # 1. Total Shipment
        df['总出库数'] = df[month_cols].sum(axis=1)
        
        # 2. Effective Months (Count where Shipment > 0)
        df['有效月份数'] = df[month_cols].gt(0).sum(axis=1).astype(int)
        
        # 3. Avg Monthly Shipment
        # Optimized: Vectorized calculation instead of apply
        df['平均每月出库数'] = np.where(df['有效月份数'] > 0, df['总出库数'] / df['有效月份数'], 0.0)
        
        # Classification
        # Optimized: Vectorized select
        conditions = [
            df['平均每月出库数'] >= 4,
            (df['平均每月出库数'] >= 2) & (df['平均每月出库数'] < 4),
            (df['平均每月出库数'] >= 1) & (df['平均每月出库数'] < 2)
        ]
        choices = ['A类门店 (>=4)', 'B类门店 (2-4)', 'C类门店 (1-2)']
        df['门店分类'] = np.select(conditions, choices, default='D类门店 (<1)')
        
        # --- Process Sheet 2 (Stock) ---
        if df_stock is not None:
            # Clean columns
            df_stock.columns = [str(c).strip() for c in df_stock.columns]
            
            # Validate Stock Columns (A-L strict structure check or Name check)
            # User defined: 经销商编码(A), 经销商名称(B), 产品编码(C), 产品名称(D), 库存数量(E), 箱数(F), 省区(G), 客户简称(H), 大类(I), 小类(J), 重量(K), 规格(L)
            # We map by index to be safe if names vary slightly, or by expected names.
            # Let's use expected names map based on index to standardize.
            # UPDATE: Use '客户简称' (H列, index 7) as the primary '经销商名称' for analysis.
            # Rename original '经销商名称' (B列, index 1) to '经销商全称' for reference.
            stock_cols_map = {
                0: '经销商编码', 1: '经销商全称', 2: '产品编码', 3: '产品名称', 
                4: '库存数量(听/盒)', 5: '箱数', 6: '省区名称', 7: '经销商名称', # Map '客户简称' to '经销商名称'
                8: '产品大类', 9: '产品小类', 10: '重量', 11: '规格'
            }
            
            if len(df_stock.columns) >= 12:
                # Rename columns by index to ensure standard access
                new_cols = list(df_stock.columns)
                for idx, name in stock_cols_map.items():
                    new_cols[idx] = name
                df_stock.columns = new_cols
                
                # Ensure numeric '箱数'
                df_stock['箱数'] = pd.to_numeric(df_stock['箱数'], errors='coerce').fillna(0)
                
                # Clean Distributor Name (客户简称)
                df_stock['经销商名称'] = df_stock['经销商名称'].astype(str).str.strip()
                
                # Fix PyArrow mixed type error for mixed columns
                df_stock['重量'] = df_stock['重量'].astype(str)
                df_stock['规格'] = df_stock['规格'].astype(str)
                
                # --- Smart Classification Logic (Specific Category) ---
                # Rules:
                # - 雅系列：仅当产品名称包含「雅赋/雅耀/雅舒/雅护」之一时命中
                # - 分段：仅在「产品大类=美思雅段粉」范围内，且产品名称包含「1段/2段/3段」之一时命中
                
                # Optimized: Vectorized Logic using np.select and str.contains
                # Pre-calculate boolean masks
                name_series = df_stock['产品名称'].astype(str)
                cat_series = df_stock['产品大类'].astype(str)
                
                mask_ya = name_series.str.contains('雅赋|雅耀|雅舒|雅护', regex=True)
                mask_seg_cat = cat_series == '美思雅段粉'
                
                # For segments, we need to extract which segment it is. 
                # Since we need the specific string ('1段' etc), np.select is good but we need to know WHICH one.
                # Let's use extraction for segments.
                seg_extract = name_series.str.extract(r'(1段|2段|3段)')[0]
                
                # Logic:
                # 1. If '雅系列' keyword -> return keyword. (Need to extract which one? Old logic returned the keyword itself e.g. '雅赋')
                # 2. If '美思雅段粉' and has segment -> return segment.
                # 3. Else '其他'
                
                # Extract Ya keyword
                ya_extract = name_series.str.extract(r'(雅赋|雅耀|雅舒|雅护)')[0]
                
                # Construct final series
                # Priority: Ya > Segment (if logic follows original sequence, Ya was checked first)
                
                df_stock['具体分类'] = np.where(
                    mask_ya, ya_extract,
                    np.where(
                        mask_seg_cat & seg_extract.notna(), seg_extract,
                        '其他'
                    )
                )
                df_stock['具体分类'] = df_stock['具体分类'].fillna('其他').astype(str)
                 
            else:
                st.warning("库存表 (Sheet2) 列数不足 12 列，无法进行库存分析。")
                df_stock = None

        # --- Process Sheet 3 (Outbound Base Table) ---
        if df_q4_raw is not None:
            df_q4_raw.columns = [str(c).strip() for c in df_q4_raw.columns]

            df_out = df_q4_raw.copy()

            month_src = df_out.columns[5] if len(df_out.columns) > 5 else None
            prov_src = df_out.columns[8] if len(df_out.columns) > 8 else None
            dist_src = df_out.columns[9] if len(df_out.columns) > 9 else None
            qty_src = df_out.columns[10] if len(df_out.columns) > 10 else None

            rename_map = {}
            if month_src: rename_map[month_src] = '月份'
            if prov_src: rename_map[prov_src] = '省区'
            if dist_src: rename_map[dist_src] = '经销商名称'
            if qty_src: rename_map[qty_src] = '数量(箱)'

            cat_src = next((c for c in df_out.columns if '产品大类' in str(c)), None)
            if cat_src is None:
                cat_src = next((c for c in df_out.columns if ('大类' in str(c)) and ('省区' not in str(c))), None)
            sub_src = next((c for c in df_out.columns if '产品小类' in str(c)), None)
            if sub_src is None:
                sub_src = next((c for c in df_out.columns if ('小类' in str(c)) and ('产品' in str(c))), None)
            if cat_src is None and len(df_out.columns) > 11:
                cat_src = df_out.columns[11]
            if sub_src is None and len(df_out.columns) > 12:
                sub_src = df_out.columns[12]

            if cat_src: rename_map[cat_src] = '产品大类'
            if sub_src: rename_map[sub_src] = '产品小类'

            df_out = df_out.rename(columns=rename_map)
            df_out = df_out.loc[:, ~df_out.columns.duplicated()]

            if '经销商名称' in df_out.columns:
                df_out['经销商名称'] = df_out['经销商名称'].astype(str).str.strip()
            if '数量(箱)' in df_out.columns:
                df_out['数量(箱)'] = pd.to_numeric(df_out['数量(箱)'], errors='coerce').fillna(0)
            if '产品大类' in df_out.columns:
                df_out['产品大类'] = df_out['产品大类'].astype(str).str.strip()
            if '产品小类' in df_out.columns:
                df_out['产品小类'] = df_out['产品小类'].astype(str).str.strip()

            df_q4_raw = df_out

        # --- Process Sheet 4 (Performance / Shipment) ---
        if df_perf_raw is not None:
            df_perf_raw.columns = [str(c).strip() for c in df_perf_raw.columns]
            df_perf = df_perf_raw.copy()

            col_year = next((c for c in df_perf.columns if c == '年份' or '年' in c), None)
            col_month = next((c for c in df_perf.columns if c == '月份' or '月' in c), None)
            col_prov = next((c for c in df_perf.columns if c == '省区' or '省区' in c), None)
            col_dist = next((c for c in df_perf.columns if c == '经销商名称' or c == '客户简称' or '客户简称' in c), None)
            col_qty = next((c for c in df_perf.columns if c == '箱数' or c == '基本数量' or '数量' in c), None)
            col_amt = next((c for c in df_perf.columns if c == '发货金额' or c == '原价金额' or '金额' in c), None)
            col_wh = next((c for c in df_perf.columns if c == '发货仓' or '发货仓' in c), None)
            col_mid = next((c for c in df_perf.columns if c == '中类' or '中类' in c), None)
            col_grp = next((c for c in df_perf.columns if c == '归类' or '归类' in c), None)
            col_bigcat = next((c for c in df_perf.columns if c == '大分类' or '大分类' in c), None)
            col_big = next((c for c in df_perf.columns if c == '大类' or '大类' in c), None)
            col_small = next((c for c in df_perf.columns if c == '小类' or '小类' in c), None)
            col_cat = next((c for c in df_perf.columns if c == '月分析' or '月分析' in c), None)

            if len(df_perf.columns) > 24:
                col_qty = df_perf.columns[24]

            rename_perf = {}
            if col_year: rename_perf[col_year] = '年份'
            if col_month: rename_perf[col_month] = '月份'
            if col_prov: rename_perf[col_prov] = '省区'
            if col_dist: rename_perf[col_dist] = '经销商名称'
            if col_qty: rename_perf[col_qty] = '发货箱数'
            if col_amt: rename_perf[col_amt] = '发货金额'
            if col_wh: rename_perf[col_wh] = '发货仓'
            if col_mid: rename_perf[col_mid] = '中类'
            if col_grp: rename_perf[col_grp] = '归类'
            if col_bigcat:
                rename_perf[col_bigcat] = '大分类'
            elif col_cat:
                rename_perf[col_cat] = '大分类'
            if col_big: rename_perf[col_big] = '大类'
            if col_small: rename_perf[col_small] = '小类'

            df_perf = df_perf.rename(columns=rename_perf)

            for c in ['省区', '经销商名称', '发货仓', '中类', '归类', '大分类', '大类', '小类']:
                if c in df_perf.columns:
                    df_perf[c] = df_perf[c].fillna('').astype(str).str.strip()
            
            # --- FIX: Ensure '经销商名称' exists ---
            if '经销商名称' not in df_perf.columns:
                # Try to find alias
                alt_dist = next((c for c in df_perf.columns if '客户' in c or '经销' in c), None)
                if alt_dist:
                    df_perf = df_perf.rename(columns={alt_dist: '经销商名称'})
                else:
                    # Fallback: Create empty if absolutely necessary (but better to warn)
                    df_perf['经销商名称'] = '未知经销商'
            # --------------------------------------

            if '大分类' in df_perf.columns and '类目' not in df_perf.columns:
                df_perf['类目'] = df_perf['大分类']

            if '年份' in df_perf.columns:
                # Handle "25年" or "2025" strings by extracting digits
                # NOTE: Use regex extraction to handle "25年" -> "25"
                df_perf['年份'] = df_perf['年份'].astype(str).str.extract(r'(\d+)')[0].astype(float).fillna(0).astype(int)
                # Normalize 2-digit years to 4-digit (e.g. 25 -> 2025)
                df_perf['年份'] = df_perf['年份'].apply(lambda y: y + 2000 if 0 < y < 100 else y)

            if '月份' in df_perf.columns:
                 # Handle "1月" or "01" strings
                df_perf['月份'] = df_perf['月份'].astype(str).str.extract(r'(\d+)')[0].astype(float).fillna(0).astype(int)
            if '发货箱数' in df_perf.columns:
                df_perf['发货箱数'] = pd.to_numeric(df_perf['发货箱数'], errors='coerce').fillna(0)
            if '发货金额' in df_perf.columns:
                df_perf['发货金额'] = pd.to_numeric(df_perf['发货金额'], errors='coerce').fillna(0)

            if '年份' in df_perf.columns and '月份' in df_perf.columns:
                df_perf = df_perf[(df_perf['年份'] > 0) & (df_perf['月份'].between(1, 12))]
                df_perf['年月'] = pd.to_datetime(df_perf['年份'].astype(str) + '-' + df_perf['月份'].astype(str).str.zfill(2) + '-01')
            else:
                df_perf['年月'] = pd.NaT

            df_perf_raw = df_perf

        # --- Process Sheet 5 (Target) ---
        df_target_raw = None
        try:
            if len(xl.sheet_names) > 4:
                df_target_raw = xl.parse(4)
                df_target_raw.columns = [str(c).strip() for c in df_target_raw.columns]
                
                # Expected Cols: D(品类), E(月份), F(任务量) -> Index 3, 4, 5
                # Rename by index to be safe
                rename_target = {}
                if len(df_target_raw.columns) > 3: rename_target[df_target_raw.columns[3]] = '品类'
                if len(df_target_raw.columns) > 4: rename_target[df_target_raw.columns[4]] = '月份'
                if len(df_target_raw.columns) > 5: rename_target[df_target_raw.columns[5]] = '任务量'
                
                df_target_raw = df_target_raw.rename(columns=rename_target)
                
                # Basic Cleaning
                if '月份' in df_target_raw.columns:
                     # Handle "1月" or "01" strings
                    df_target_raw['月份'] = df_target_raw['月份'].astype(str).str.extract(r'(\d+)')[0].astype(float).fillna(0).astype(int)
                if '任务量' in df_target_raw.columns:
                    df_target_raw['任务量'] = pd.to_numeric(df_target_raw['任务量'], errors='coerce').fillna(0)
            else:
                 debug_logs.append("Warning: Sheet5 (Target) not found.")
        except Exception as e:
            debug_logs.append(f"Error parsing Sheet5: {e}")
            df_target_raw = None

        return df, month_cols, df_stock, df_q4_raw, df_perf_raw, df_target_raw, debug_logs
        
    except Exception as e:
        st.error(f"数据加载失败: {str(e)}")
        return None, None, None, None, None, None, [str(e)]

@st.cache_data(ttl=3600)
def load_data_v3(file_bytes: bytes, file_name: str):
    debug_logs = []
    try:
        file_name_lower = (file_name or "").lower()
        bio = io.BytesIO(file_bytes)
        
        # Init Returns
        df = None
        month_cols = []
        df_stock = None
        df_q4_raw = None
        df_perf_raw = None
        df_target_raw = None
        df_scan_raw = None
        df_newcust_raw = None

        if file_name_lower.endswith('.csv'):
            df = pd.read_csv(bio, encoding='gb18030')
        else:
            xl = pd.ExcelFile(bio)
            debug_logs.append(f"Sheet Names: {xl.sheet_names}")

            def _pick_sheet_by_name(names, keywords, default_index=None):
                candidates = [str(s) for s in names if any(k in str(s) for k in keywords)]
                if not candidates:
                    if default_index is not None and len(names) > int(default_index):
                        return names[int(default_index)]
                    return None

                def _name_score(value: str):
                    text = str(value)
                    nums = re.findall(r"(\d{4,8})", text)
                    if nums:
                        return max(int(x) for x in nums)
                    nums = re.findall(r"(\d{3,4})", text)
                    if nums:
                        return max(int(x) for x in nums)
                    return -1

                candidates.sort(key=lambda s: (_name_score(s), str(s)), reverse=True)
                return candidates[0]
            
            # Sheet 1: Sales
            if len(xl.sheet_names) > 0: df = xl.parse(0)
            
            # Sheet 2: Stock
            if len(xl.sheet_names) > 1: df_stock = xl.parse(1)
            
            # Sheet 3: Outbound (Q4)
            if len(xl.sheet_names) > 2: df_q4_raw = xl.parse(2)
            
            # Sheet 4: Performance
            if len(xl.sheet_names) > 3:
                preferred = next((s for s in xl.sheet_names if any(k in str(s) for k in ["发货", "业绩", "Performance", "perf", "Perf"])), None)
                if preferred is None:
                    preferred = next((s for s in xl.sheet_names if 'sheet4' in str(s).strip().lower()), None)
                candidate_names = [preferred] if preferred else []
                candidate_names += [s for s in xl.sheet_names if s not in candidate_names]
                for sname in candidate_names:
                    try:
                        tmp_header = xl.parse(sname, nrows=0)
                        cols = [str(c).strip() for c in tmp_header.columns]
                        key_hits = sum(1 for k in ['年份', '月份', '省区'] if any(k in c for c in cols))
                        signal_hits = sum(1 for k in ['发货仓', '原价金额', '基本数量', '大分类', '月分析', '客户简称'] if any(k in c for c in cols))
                        if key_hits >= 2 and signal_hits >= 1:
                            df_perf_raw = xl.parse(sname)
                            debug_logs.append(f"-> MATCHED Sheet4: {sname}")
                            break
                    except: continue
                if df_perf_raw is None:
                    try:
                        tmp_header = xl.parse(3, nrows=0)
                        cols = [str(c).strip() for c in tmp_header.columns]
                        signal_hits = sum(1 for k in ['日期', '业务部', '原价金额', '基本数量', '客户简称', '客户名称', '大类'] if any(k in c for c in cols))
                        if signal_hits >= 3:
                            df_perf_raw = xl.parse(3)
                            debug_logs.append(f"-> FALLBACK Sheet4(idx3): {xl.sheet_names[3]}")
                    except Exception:
                        pass
            
            # Sheet 5: Target
            if len(xl.sheet_names) > 4: df_target_raw = xl.parse(4)

            # Sheet 6: Scan Data (用于扫码相关分析)
            scan_sheet_name = _pick_sheet_by_name(xl.sheet_names, ["扫码", "扫描"], default_index=5)
            if scan_sheet_name is not None:
                try:
                    df_scan_raw = xl.parse(scan_sheet_name)
                    debug_logs.append(f"-> Scan Data: {scan_sheet_name}")
                except Exception:
                    df_scan_raw = None

            if len(xl.sheet_names) > 7:
                try:
                    _sname_nc = xl.sheet_names[7]
                    _raw_nc = xl.parse(7, header=None)
                    _best_i = None
                    _best_score = -1
                    _kw = ["省区", "门店", "新客", "经销", "客户", "时间", "日期", "月份", "年月"]
                    for _i in range(min(20, int(_raw_nc.shape[0]))):
                        try:
                            vs = [str(x or "").strip() for x in _raw_nc.iloc[_i].tolist()]
                            txt = " ".join(vs)
                            score = sum(1 for k in _kw if k in txt)
                            if score > _best_score:
                                _best_score = score
                                _best_i = _i
                        except Exception:
                            continue
                    if _best_i is not None and _best_score >= 2:
                        df_newcust_raw = xl.parse(7, header=int(_best_i))
                        debug_logs.append(f"-> New Customers from Sheet8: {_sname_nc} (header row={int(_best_i)+1})")
                    else:
                        df_newcust_raw = xl.parse(7)
                        debug_logs.append(f"-> New Customers from Sheet8: {_sname_nc}")
                except Exception:
                    df_newcust_raw = None

            if df_newcust_raw is None and len(xl.sheet_names) > 0:
                def _pick_newcust_sheet(_xl):
                    best = None
                    best_score = -1
                    for sname in _xl.sheet_names:
                        try:
                            h = _xl.parse(sname, nrows=0)
                            cols = [str(c).strip() for c in h.columns]
                            score = 0
                            score += 2 if any("新客" in c for c in cols) else 0
                            score += 1 if any("门店" in c for c in cols) else 0
                            score += 1 if any("省区" in c for c in cols) else 0
                            score += 1 if any(("经销商" in c) or ("客户简称" in c) or ("客户" in c) for c in cols) else 0
                            score += 1 if any(("时间" in c) or ("日期" in c) for c in cols) else 0
                            if score > best_score:
                                best_score = score
                                best = sname
                        except Exception:
                            continue
                    return best if best_score >= 3 else None

                chosen_newcust = _pick_newcust_sheet(xl)
                if chosen_newcust:
                    try:
                        _raw_nc = xl.parse(chosen_newcust, header=None)
                        _best_i = None
                        _best_score = -1
                        _kw = ["省区", "门店", "新客", "经销", "客户", "时间", "日期", "月份", "年月"]
                        for _i in range(min(20, int(_raw_nc.shape[0]))):
                            try:
                                vs = [str(x or "").strip() for x in _raw_nc.iloc[_i].tolist()]
                                txt = " ".join(vs)
                                score = sum(1 for k in _kw if k in txt)
                                if score > _best_score:
                                    _best_score = score
                                    _best_i = _i
                            except Exception:
                                continue
                        if _best_i is not None and _best_score >= 2:
                            df_newcust_raw = xl.parse(chosen_newcust, header=int(_best_i))
                            debug_logs.append(f"-> New Customers matched: {chosen_newcust} (header row={int(_best_i)+1})")
                        else:
                            df_newcust_raw = xl.parse(chosen_newcust)
                            debug_logs.append(f"-> New Customers matched: {chosen_newcust}")
                    except Exception:
                        df_newcust_raw = None

        # --- Process Sheet 1 (Sales) ---
        if df is not None:
            df.columns = [str(c).strip() for c in df.columns]
            
            # Identify Month Columns
            is_long_format = False
            time_col = None
            if len(df.columns) > 5:
                col_f = df.columns[5]
                sample_vals = df[col_f].dropna().head(10).astype(str).tolist()
                if any('月' in v for v in sample_vals):
                    is_long_format = True
                    time_col = col_f
            
            if is_long_format:
                col_prov = df.columns[8] if len(df.columns) > 8 else None
                col_dist = df.columns[9] if len(df.columns) > 9 else None
                col_qty = df.columns[10] if len(df.columns) > 10 else None
                
                if col_prov is None: col_prov = next((c for c in df.columns if '省' in c), None)
                if col_dist is None: col_dist = next((c for c in df.columns if '经销' in c or '客户' in c), None)
                if col_qty is None: col_qty = next((c for c in df.columns if '数' in c or 'Qty' in c or '箱' in c), None)
                col_store = next((c for c in df.columns if '门店' in c), None)
                
                if col_prov and col_dist and col_qty and time_col:
                    df[col_qty] = pd.to_numeric(df[col_qty], errors='coerce').fillna(0)
                    pivot_index = [col_prov, col_dist]
                    if col_store: pivot_index.append(col_store)
                    df_wide = df.pivot_table(index=pivot_index, columns=time_col, values=col_qty, aggfunc='sum').reset_index()
                    if not col_store: df_wide['门店名称'] = df_wide[col_dist]
                    df = df_wide
                    df.columns = [str(c).strip() for c in df.columns]
            
            rename_map = {}
            if '品牌省区名称' in df.columns: rename_map['品牌省区名称'] = '省区'
            if '经销商名称' not in df.columns and len(df.columns) > 1: rename_map[df.columns[1]] = '经销商名称'
            if '门店名称' not in df.columns and len(df.columns) > 2: rename_map[df.columns[2]] = '门店名称'
            df = df.rename(columns=rename_map)
            
            required = ['省区', '经销商名称', '门店名称']
            for req in required:
                if req not in df.columns:
                    if len(df.columns) >= 3:
                        df.columns.values[0] = '省区'
                        df.columns.values[1] = '经销商名称'
                        df.columns.values[2] = '门店名称'
            
            month_cols = [c for c in df.columns if '月' in c and c not in required]
            for col in month_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                
            df['总出库数'] = df[month_cols].sum(axis=1)
            df['有效月份数'] = df[month_cols].gt(0).sum(axis=1).astype(int)
            df['平均每月出库数'] = np.where(df['有效月份数'] > 0, df['总出库数'] / df['有效月份数'], 0.0)
            
            conditions = [df['平均每月出库数'] >= 4, (df['平均每月出库数'] >= 2) & (df['平均每月出库数'] < 4), (df['平均每月出库数'] >= 1) & (df['平均每月出库数'] < 2)]
            choices = ['A类门店 (>=4)', 'B类门店 (2-4)', 'C类门店 (1-2)']
            df['门店分类'] = np.select(conditions, choices, default='D类门店 (<1)')

        # --- Process Sheet 2 (Stock) ---
        if df_stock is not None:
            df_stock.columns = [str(c).strip() for c in df_stock.columns]
            df_cols = list(df_stock.columns)
            def _col_by_name(names: list[str]):
                for n in names:
                    if n in df_stock.columns:
                        return n
                for c in df_stock.columns:
                    s = str(c).strip()
                    if any(n in s for n in names):
                        return c
                return None

            col_dist_abbr = _col_by_name(["客户简称"])
            col_dist_full = _col_by_name(["经销商名称", "经销商全称"])
            col_dist_code = _col_by_name(["经销商编码"])
            col_prod_code = _col_by_name(["产品编码"])
            col_prod_name = _col_by_name(["产品名称"])
            col_qty = _col_by_name(["库存数量", "库存数量(听/盒)"])
            col_box = _col_by_name(["箱数", "箱"])
            col_prov = _col_by_name(["省区名称", "省区"])
            col_big = _col_by_name(["产品大类", "大类"])
            col_small = _col_by_name(["产品小类", "小类"])
            col_weight = _col_by_name(["重量"])
            col_spec = _col_by_name(["规格"])
            col_batch = _col_by_name(["批次号", "批次", "批号", "LOT", "lot"])

            if col_prod_name is None and len(df_cols) >= 4:
                col_prod_name = df_cols[3]
            if col_box is None and len(df_cols) >= 6:
                col_box = df_cols[5]
            if col_prov is None and len(df_cols) >= 7:
                col_prov = df_cols[6]
            if col_dist_abbr is None and len(df_cols) >= 8:
                col_dist_abbr = df_cols[7]
            if col_big is None and len(df_cols) >= 10:
                col_big = df_cols[9]
            if col_small is None and len(df_cols) >= 11:
                col_small = df_cols[10]
            if col_weight is None and len(df_cols) >= 12:
                col_weight = df_cols[11] if col_spec is not None else None
            if col_batch is None and len(df_cols) >= 13:
                col_batch = df_cols[12]

            df_stock = pd.DataFrame({
                "经销商编码": df_stock[col_dist_code] if col_dist_code is not None else (df_stock[df_cols[0]] if len(df_cols) > 0 else pd.Series([], dtype=object)),
                "经销商全称": df_stock[col_dist_full] if col_dist_full is not None else (df_stock[df_cols[1]] if len(df_cols) > 1 else pd.Series([], dtype=object)),
                "产品编码": df_stock[col_prod_code] if col_prod_code is not None else (df_stock[df_cols[2]] if len(df_cols) > 2 else pd.Series([], dtype=object)),
                "产品名称": df_stock[col_prod_name] if col_prod_name is not None else pd.Series([], dtype=object),
                "库存数量(听/盒)": df_stock[col_qty] if col_qty is not None else (df_stock[df_cols[4]] if len(df_cols) > 4 else pd.Series([], dtype=object)),
                "箱数": df_stock[col_box] if col_box is not None else pd.Series([], dtype=object),
                "省区": df_stock[col_prov] if col_prov is not None else pd.Series([], dtype=object),
                "经销商名称": df_stock[col_dist_abbr] if col_dist_abbr is not None else (df_stock[col_dist_full] if col_dist_full is not None else pd.Series([], dtype=object)),
                "产品大类": df_stock[col_big] if col_big is not None else pd.Series([], dtype=object),
                "产品小类": df_stock[col_small] if col_small is not None else pd.Series([], dtype=object),
                "重量": df_stock[col_weight] if col_weight is not None else pd.Series([], dtype=object),
                "规格": df_stock[col_spec] if col_spec is not None else pd.Series([], dtype=object),
                "批次号": df_stock[col_batch] if col_batch is not None else pd.Series([], dtype=object),
            })

            df_stock["箱数"] = pd.to_numeric(df_stock["箱数"], errors="coerce").fillna(0.0)
            df_stock["库存数量(听/盒)"] = pd.to_numeric(df_stock["库存数量(听/盒)"], errors="coerce").fillna(0.0)
            for _c in ["省区", "产品大类", "产品小类", "重量", "规格", "批次号"]:
                df_stock[_c] = df_stock[_c].fillna("").astype(str).str.strip()
            if "批次号" in df_stock.columns:
                df_stock["批次号"] = df_stock["批次号"].astype(str).str.extract(r"(\d{8})")[0].fillna("").astype(str).str.strip()
            df_stock["经销商名称"] = df_stock["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
            df_stock["经销商全称"] = df_stock["经销商全称"].fillna("").astype(str).str.strip()
            df_stock["产品名称"] = df_stock["产品名称"].fillna("").astype(str).str.strip()

            name_series = df_stock["产品名称"].astype(str)
            mask_ya = name_series.str.contains("雅赋|雅耀|雅舒|雅护", regex=True)
            mask_seg_cat = df_stock["产品大类"].astype(str) == "美思雅段粉"
            seg_extract = name_series.str.extract(r"(1段|2段|3段)")[0]
            ya_extract = name_series.str.extract(r"(雅赋|雅耀|雅舒|雅护)")[0]
            df_stock["具体分类"] = np.where(mask_ya, ya_extract, np.where(mask_seg_cat & seg_extract.notna(), seg_extract, "其他"))
            df_stock["具体分类"] = df_stock["具体分类"].fillna("其他").astype(str)

        # --- Process Sheet 3 (Outbound) FIX ---
        if df_q4_raw is not None:
            # Deduplicate
            cols = pd.Series(df_q4_raw.columns)
            for dup in cols[cols.duplicated()].unique(): 
                cols[cols[cols == dup].index.values.tolist()] = [dup + '.' + str(i) if i != 0 else dup for i in range(sum(cols == dup))]
            df_q4_raw.columns = cols
            
            df_out = df_q4_raw.copy()
            
            # Map Indices (User Requirement: +8 shift)
            # M(12)=Year, N(13)=Month, Q(16)=Prov, R(17)=Dist(CustomerAbbr), S(18)=Qty, U(20)=SubCat
            idx_map = {
                4: '门店编号',
                12: '年份',
                13: '月份',
                16: '省区',
                17: '经销商名称',
                18: '数量(箱)',
                24: '门店状态',
                20: '产品小类',
                19: '产品大类'
            }
            curr_cols = list(df_out.columns)
            
            # Avoid Name Collision: Rename existing columns that clash with target names
            target_names = list(idx_map.values())
            for i, c in enumerate(curr_cols):
                if c in target_names and i not in idx_map:
                    new_n = f"{c}_old_{i}"
                    df_out.rename(columns={c: new_n}, inplace=True)
                    debug_logs.append(f"Renamed collision '{c}' -> '{new_n}'")
            
            # Refresh columns after collision avoidance
            curr_cols = list(df_out.columns)
            
            for idx, name in idx_map.items():
                if idx < len(curr_cols):
                    df_out.rename(columns={curr_cols[idx]: name}, inplace=True)

            
            # Clean Dist Name
            if '经销商名称' in df_out.columns:
                 df_out['经销商名称'] = df_out['经销商名称'].astype(str).str.replace(r'\s+', '', regex=True)
                 debug_logs.append(f"Sheet3 Dist Sample: {df_out['经销商名称'].head(3).tolist()}")

            if '数量(箱)' in df_out.columns:
                df_out['数量(箱)'] = pd.to_numeric(df_out['数量(箱)'], errors='coerce').fillna(0)
            
            if '产品大类' in df_out.columns: df_out['产品大类'] = df_out['产品大类'].astype(str).str.strip()
            if '产品小类' in df_out.columns: df_out['产品小类'] = df_out['产品小类'].astype(str).str.strip()
            
            # Clean Year
            if '年份' in df_out.columns:
                # Extract digits and normalize
                df_out['年份'] = df_out['年份'].astype(str).str.extract(r'(\d+)')[0].astype(float).fillna(0).astype(int)
                # Normalize 25 -> 2025
                df_out['年份'] = df_out['年份'].apply(lambda y: y + 2000 if 20 <= y < 100 else y)

            df_q4_raw = df_out

        # --- Process Sheet 4 (Perf) ---
        if df_perf_raw is not None:
            df_perf_raw.columns = [str(c).strip() for c in df_perf_raw.columns]
            df_perf = df_perf_raw.copy()
            _perf_cols = list(df_perf.columns)
            def _col_by_idx(i: int):
                try:
                    return _perf_cols[i] if i < len(_perf_cols) else None
                except Exception:
                    return None

            col_year = next((c for c in df_perf.columns if str(c).strip() in ('年份', 'Year', 'year') or str(c).strip() == '年'), None)
            col_month = next((c for c in df_perf.columns if str(c).strip() in ('月份', 'Month', 'month') or str(c).strip() == '月'), None)
            col_date = next((c for c in df_perf.columns if ('日期' in str(c)) or (str(c).strip() in ('date', 'Date'))), None)
            col_prov = next((c for c in df_perf.columns if str(c).strip() == '省区' or ('省区' in str(c)) or (str(c).strip() == '省')), None)
            col_dist = next((c for c in df_perf.columns if str(c).strip() == '客户简称' or ('客户简称' in str(c))), None)
            if col_dist is None:
                col_dist = next((c for c in df_perf.columns if str(c).strip() == '经销商名称' or ('经销商' in str(c))), None)
            if col_dist is None:
                col_dist = next((c for c in df_perf.columns if str(c).strip() == '客户名称' or ('客户名称' in str(c))), None)
            col_qty = next((c for c in df_perf.columns if str(c).strip() == '发货箱数' or str(c).strip() == '箱数' or str(c).strip() == '基本数量' or ('数量' in str(c)) or ('箱' in str(c)) or ('件数' in str(c)) or (str(c).strip().endswith('件'))), None)
            col_amt = next((c for c in df_perf.columns if str(c).strip() == '发货金额' or str(c).strip() == '原价金额' or ('金额' in str(c))), None)
            col_wh = next((c for c in df_perf.columns if str(c).strip() == '发货仓' or ('发货仓' in str(c))), None)
            col_mid = next((c for c in df_perf.columns if str(c).strip() == '中类' or ('中类' in str(c))), None)
            col_grp = next((c for c in df_perf.columns if str(c).strip() == '归类' or ('归类' in str(c))), None)
            col_bigcat = next((c for c in df_perf.columns if str(c).strip() == '大分类' or ('大分类' in str(c))), None)
            col_big = next((c for c in df_perf.columns if str(c).strip() == '大类' or ('大类' in str(c))), None)
            col_small = next((c for c in df_perf.columns if str(c).strip() == '小类' or ('小类' in str(c))), None)
            col_small_code = next((c for c in df_perf.columns if any(k in str(c) for k in ["产品小类码", "产品小类编码", "小类码", "小类编码"])), None)
            col_weight = next((c for c in df_perf.columns if "重量" in str(c).replace(" ", "")), None)
            col_cat = next((c for c in df_perf.columns if str(c).strip() == '月分析' or ('月分析' in str(c))), None)

            if col_prov is None:
                col_prov = _col_by_idx(5)  # F列：省区
            if col_date is None:
                col_date = _col_by_idx(6)  # G列：日期
            _y_qty = _col_by_idx(24)  # Y列：箱数
            if _y_qty is not None:
                col_qty = _y_qty
            elif col_qty is None:
                col_qty = _col_by_idx(9)  # J列：发货件数（兜底）
            if col_amt is None:
                col_amt = _col_by_idx(10)  # K列：发货额
            if col_big is None:
                col_big = _col_by_idx(13)  # N列：产品大类
            if col_small is None:
                col_small = _col_by_idx(18)  # S列：产品小类
            if col_small is None:
                col_small = _col_by_idx(14)  # O列：产品小类
            if col_small_code is None:
                col_small_code = _col_by_idx(18)
            if col_weight is None:
                col_weight = _col_by_idx(14)
            if col_dist is None:
                col_dist = _col_by_idx(19)  # T列：客户简称（用于匹配经销商）

            rename_perf = {}
            if col_year: rename_perf[col_year] = '年份'
            if col_month: rename_perf[col_month] = '月份'
            if col_date: rename_perf[col_date] = '日期'
            if col_prov: rename_perf[col_prov] = '省区'
            if col_dist: rename_perf[col_dist] = '经销商名称'
            if col_qty: rename_perf[col_qty] = '发货箱数'
            if col_amt: rename_perf[col_amt] = '发货金额'
            if col_wh: rename_perf[col_wh] = '发货仓'
            if col_mid: rename_perf[col_mid] = '中类'
            if col_grp: rename_perf[col_grp] = '归类'
            if col_bigcat:
                rename_perf[col_bigcat] = '大分类'
            elif col_cat:
                rename_perf[col_cat] = '大分类'
            if col_big: rename_perf[col_big] = '大类'
            if col_small: rename_perf[col_small] = '小类'
            if col_small_code and col_small_code != col_small:
                rename_perf[col_small_code] = '小类码'
            if col_weight:
                rename_perf[col_weight] = '重量'

            df_perf = df_perf.rename(columns=rename_perf)

            if '经销商名称' not in df_perf.columns:
                alt_dist = next((c for c in df_perf.columns if ('客户' in str(c)) or ('经销' in str(c))), None)
                if alt_dist:
                    df_perf = df_perf.rename(columns={alt_dist: '经销商名称'})
                else:
                    df_perf['经销商名称'] = ''
                    debug_logs.append("Warning: Sheet4 missing distributor column; set '经销商名称' to empty.")

            for c in ['省区', '经销商名称', '发货仓', '中类', '归类', '大分类', '大类', '小类']:
                if c in df_perf.columns:
                    if c == "经销商名称":
                        df_perf[c] = df_perf[c].fillna('').astype(str).str.replace(r'\s+', '', regex=True)
                    else:
                        df_perf[c] = df_perf[c].fillna('').astype(str).str.strip()

            if '年份' in df_perf.columns:
                df_perf['年份'] = df_perf['年份'].astype(str).str.extract(r'(\d+)')[0].astype(float).fillna(0).astype(int)
                df_perf['年份'] = df_perf['年份'].apply(lambda y: y + 2000 if 0 < y < 100 else y)
            if '月份' in df_perf.columns:
                df_perf['月份'] = df_perf['月份'].astype(str).str.extract(r'(\d+)')[0].astype(float).fillna(0).astype(int)
            if '发货箱数' in df_perf.columns:
                df_perf['发货箱数'] = pd.to_numeric(df_perf['发货箱数'], errors='coerce').fillna(0)
            if '发货金额' in df_perf.columns:
                df_perf['发货金额'] = pd.to_numeric(df_perf['发货金额'], errors='coerce').fillna(0)

            # Fallback: if Year/Month exist but look invalid (mostly 0), try Date
            year_valid_ratio = 0.0
            month_valid_ratio = 0.0
            if '年份' in df_perf.columns:
                year_valid_ratio = (df_perf['年份'] > 2000).mean()
            if '月份' in df_perf.columns:
                month_valid_ratio = (df_perf['月份'].between(1, 12)).mean()

            if (year_valid_ratio < 0.1 or month_valid_ratio < 0.1) and '日期' in df_perf.columns:
                debug_logs.append(f"Fallback to Date column because Year/Month valid ratio is low (Y:{year_valid_ratio:.2f}, M:{month_valid_ratio:.2f})")
                dt = pd.to_datetime(df_perf['日期'], errors='coerce')
                if dt.isna().mean() > 0.8:
                    try:
                        dt2 = pd.to_datetime(pd.to_numeric(df_perf['日期'], errors='coerce'), unit='D', origin='1899-12-30', errors='coerce')
                        if dt2.isna().mean() < dt.isna().mean():
                            dt = dt2
                    except Exception:
                        pass
                
                # Override if we got valid dates
                if dt.notna().mean() > 0.1:
                    df_perf['年份'] = pd.to_numeric(dt.dt.year, errors='coerce').fillna(0).astype(int)
                    df_perf['月份'] = pd.to_numeric(dt.dt.month, errors='coerce').fillna(0).astype(int)
                    debug_logs.append("Overridden Year/Month from Date column")

            if ('年份' not in df_perf.columns or '月份' not in df_perf.columns) and '日期' in df_perf.columns:
                dt = pd.to_datetime(df_perf['日期'], errors='coerce')
                if dt.isna().mean() > 0.8:
                    try:
                        dt2 = pd.to_datetime(pd.to_numeric(df_perf['日期'], errors='coerce'), unit='D', origin='1899-12-30', errors='coerce')
                        if dt2.isna().mean() < dt.isna().mean():
                            dt = dt2
                    except Exception:
                        pass
                
                # Only update if we found valid dates
                if dt.notna().mean() > 0.1:
                    df_perf['年份'] = pd.to_numeric(dt.dt.year, errors='coerce').fillna(0).astype(int)
                    df_perf['月份'] = pd.to_numeric(dt.dt.month, errors='coerce').fillna(0).astype(int)


            if '年份' in df_perf.columns and '月份' in df_perf.columns:
                df_perf = df_perf[(df_perf['年份'] > 0) & (df_perf['月份'].between(1, 12))]
                df_perf['年月'] = pd.to_datetime(
                    df_perf['年份'].astype(str) + '-' + df_perf['月份'].astype(str).str.zfill(2) + '-01',
                    errors='coerce'
                )
            else:
                df_perf['年月'] = pd.NaT
            df_perf_raw = df_perf

        # --- Process Sheet 5 (Target) ---
        if df_target_raw is not None:
            df_target_raw.columns = [str(c).strip() for c in df_target_raw.columns]
            rename_target = {}
            if len(df_target_raw.columns) > 3: rename_target[df_target_raw.columns[3]] = '品类'
            if len(df_target_raw.columns) > 4: rename_target[df_target_raw.columns[4]] = '月份'
            if len(df_target_raw.columns) > 5: rename_target[df_target_raw.columns[5]] = '任务量'
            df_target_raw = df_target_raw.rename(columns=rename_target)
            if '月份' in df_target_raw.columns:
                df_target_raw['月份'] = df_target_raw['月份'].astype(str).str.extract(r'(\d+)')[0].astype(float).fillna(0).astype(int)
            if '任务量' in df_target_raw.columns:
                df_target_raw['任务量'] = pd.to_numeric(df_target_raw['任务量'], errors='coerce').fillna(0)

        # --- Process Sheet 6 (Scan Data) ---
        if df_scan_raw is not None:
            df0 = df_scan_raw

            def _pick_scan_sheet(xl_obj: pd.ExcelFile):
                candidates = [str(s) for s in xl_obj.sheet_names if any(k in str(s) for k in ["扫码", "扫描"])]
                if not candidates:
                    return None
                best = None
                best_key = -1
                for sname in candidates:
                    try:
                        tmp = xl_obj.parse(sname)
                        cols = [str(c).strip() for c in tmp.columns]
                        y_col = next((c for c in cols if c in ["年", "年份"]), None)
                        m_col = next((c for c in cols if c in ["月", "月份"]), None)
                        d_col = next((c for c in cols if c in ["日", "天"]), None)
                        if y_col is None or m_col is None or d_col is None:
                            continue
                        yy = pd.to_numeric(tmp[y_col].astype(str).str.extract(r"(\d+)")[0], errors="coerce").fillna(0).astype(int)
                        yy = yy.apply(lambda v: v + 2000 if 0 < v < 100 else v)
                        mm = pd.to_numeric(tmp[m_col].astype(str).str.extract(r"(\d+)")[0], errors="coerce").fillna(0).astype(int)
                        dd = pd.to_numeric(tmp[d_col].astype(str).str.extract(r"(\d+)")[0], errors="coerce").fillna(0).astype(int)
                        key = (yy * 10000 + mm * 100 + dd).max()
                        key = int(key) if pd.notna(key) else -1
                        if key > best_key:
                            best_key = key
                            best = sname
                    except Exception:
                        continue
                return best

            chosen_scan_sheet = None

            cols = [str(c).strip() for c in df0.columns]
            def _col_by_name(names: list[str]):
                for n in names:
                    if n in df0.columns:
                        return n
                for c in df0.columns:
                    s = str(c).strip()
                    if any(n in s for n in names):
                        return c
                return None

            store_col = _col_by_name(["门店名称"])
            dist_col = _col_by_name(["经销商名称", "客户简称"])
            prov_col = _col_by_name(["省区"])
            cat_col = _col_by_name(["产品大类", "大类"])
            small_col = _col_by_name(["产品小类", "产品名称", "重量"])
            coord_col = _col_by_name(["经纬度", "GPS位置"])
            y_col = _col_by_name(["年份", "年"])
            m_col = _col_by_name(["月份", "月"])
            d_col = _col_by_name(["日"])

            def _col(idx: int):
                if idx < df0.shape[1]:
                    return df0.iloc[:, idx]
                return pd.Series([None] * len(df0))

            store_src = df0[store_col] if store_col is not None else _col(1)
            dist_src = df0[dist_col] if dist_col is not None else _col(18)
            prov_src = df0[prov_col] if prov_col is not None else _col(17)
            cat_src = df0[cat_col] if cat_col is not None else _col(19)
            small_src = df0[small_col] if small_col is not None else _col(20)
            y_src = df0[y_col] if y_col is not None else _col(13)
            m_src = df0[m_col] if m_col is not None else _col(14)
            d_src = df0[d_col] if d_col is not None else _col(15)

            if df0.shape[1] > 18:
                _s_store = _col(1).fillna("").astype(str)
                _s_dist = _col(18).fillna("").astype(str)
                if (_s_store.str.strip() != "").mean() >= 0.05:
                    store_src = _col(1)
                if (_s_dist.str.strip() != "").mean() >= 0.05:
                    dist_src = _col(18)

            df_scan_raw = pd.DataFrame({
                "门店名称": store_src,
                "经销商名称": dist_src,
                "客户简称": dist_src,
                "省区": prov_src,
                "产品大类": cat_src,
                "产品小类": small_src,
                "经纬度": df0[coord_col] if coord_col is not None else _col(12),
                "年份": y_src,
                "月份": m_src,
                "日": d_src,
            })

            df_scan_raw["年份"] = df_scan_raw["年份"].astype(str).str.extract(r"(\d+)")[0].astype(float).fillna(0).astype(int)
            df_scan_raw["年份"] = df_scan_raw["年份"].apply(lambda y: y + 2000 if 0 < y < 100 else y)
            df_scan_raw["月份"] = df_scan_raw["月份"].astype(str).str.extract(r"(\d+)")[0].astype(float).fillna(0).astype(int)
            df_scan_raw["日"] = df_scan_raw["日"].astype(str).str.extract(r"(\d+)")[0].astype(float).fillna(0).astype(int)

            for c in ["门店名称", "省区", "经销商名称", "客户简称", "产品大类", "产品小类"]:
                df_scan_raw[c] = df_scan_raw[c].fillna("").astype(str).str.strip()
            df_scan_raw["门店名称"] = df_scan_raw["门店名称"].astype(str).str.replace(r"\s+", "", regex=True)
            df_scan_raw["经销商名称"] = df_scan_raw["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True)
            df_scan_raw["客户简称"] = df_scan_raw["客户简称"].astype(str).str.replace(r"\s+", "", regex=True)

            coords = df_scan_raw["经纬度"].apply(_parse_lon_lat)
            df_scan_raw["经度"] = coords.apply(lambda x: x[0])
            df_scan_raw["纬度"] = coords.apply(lambda x: x[1])

        if df_newcust_raw is not None and not getattr(df_newcust_raw, "empty", True):
            try:
                df0 = df_newcust_raw.copy()
                df0.columns = [str(c).strip() for c in df0.columns]
                def _col_by_name(names: list[str]):
                    for n in names:
                        if n in df0.columns:
                            return n
                    for c in df0.columns:
                        s = str(c).strip()
                        if any(n in s for n in names):
                            return c
                    return None

                def _col(idx: int):
                    if idx < df0.shape[1]:
                        return df0.iloc[:, idx]
                    return pd.Series([None] * len(df0))

                # Sheet8 has a fixed layout from the user base table: A省区, E门店, F日期, G新客数, J客户名称.
                col_prov = 0 if df0.shape[1] > 0 else _col_by_name(["省区"])
                col_store = 4 if df0.shape[1] > 4 else _col_by_name(["门店名称", "门店"])
                col_time = 5 if df0.shape[1] > 5 else _col_by_name(["时间", "日期"])
                col_val = 6 if df0.shape[1] > 6 else _col_by_name(["新客数", "新客数量", "新客"])
                col_dist = 9 if df0.shape[1] > 9 else _col_by_name(["经销商名称", "客户简称", "客户名称", "客户"])
                col_year = _col_by_name(["年份", "年"])
                col_month = _col_by_name(["月份", "月"])
                col_ym = _col_by_name(["年月", "年/月", "年-月", "月度"])

                prov_s = df0.iloc[:, col_prov] if isinstance(col_prov, int) and col_prov < df0.shape[1] else (df0[col_prov] if col_prov is not None else _col(0))
                store_s = df0.iloc[:, col_store] if isinstance(col_store, int) and col_store < df0.shape[1] else (df0[col_store] if col_store is not None else _col(4))
                time_s = df0.iloc[:, col_time] if isinstance(col_time, int) and col_time < df0.shape[1] else (df0[col_time] if col_time is not None else _col(5))
                val_s = df0.iloc[:, col_val] if isinstance(col_val, int) and col_val < df0.shape[1] else (df0[col_val] if col_val is not None else _col(6))
                dist_s = df0.iloc[:, col_dist] if isinstance(col_dist, int) and col_dist < df0.shape[1] else (df0[col_dist] if col_dist is not None else _col(9))
                df_newcust_raw = pd.DataFrame({
                    "省区": prov_s,
                    "门店名称": store_s,
                    "经销商名称": dist_s,
                    "时间": time_s,
                    "新客数": val_s,
                })
                for c in ["省区", "门店名称", "经销商名称", "时间"]:
                    df_newcust_raw[c] = df_newcust_raw[c].fillna("").astype(str).str.strip()
                df_newcust_raw["经销商名称"] = df_newcust_raw["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True)
                df_newcust_raw["新客数"] = pd.to_numeric(df_newcust_raw["新客数"], errors="coerce").fillna(0.0)
                yy = None
                mm = None
                if col_year is not None and col_month is not None:
                    yy = pd.to_numeric(df0[col_year].astype(str).str.extract(r"(\d+)")[0], errors="coerce").fillna(0).astype(int)
                    mm = pd.to_numeric(df0[col_month].astype(str).str.extract(r"(\d+)")[0], errors="coerce").fillna(0).astype(int)
                    yy = yy.apply(lambda y: y + 2000 if 0 < y < 100 else y)
                if (yy is None or mm is None or (yy.eq(0).mean() > 0.5)) and (col_ym is not None and col_ym in df0.columns):
                    s_ym = df0[col_ym]
                    s_ym_str = s_ym.fillna("").astype(str)
                    m1 = s_ym_str.str.extract(r"(\d{4})\D{0,3}(\d{1,2})")
                    yy1 = pd.to_numeric(m1[0], errors="coerce").fillna(0).astype(int)
                    mm1 = pd.to_numeric(m1[1], errors="coerce").fillna(0).astype(int)
                    m2 = s_ym_str.str.extract(r"(\d{6})")[0]
                    yyyymm = pd.to_numeric(m2, errors="coerce").fillna(0).astype(int)
                    yy2 = (yyyymm // 100).astype(int)
                    mm2 = (yyyymm % 100).astype(int)
                    yy = yy1.where(yy1 != 0, yy2)
                    mm = mm1.where(mm1 != 0, mm2)
                if yy is None or mm is None or (yy.eq(0).mean() > 0.5):
                    t0 = df_newcust_raw["时间"]
                    t_str = t0.fillna("").astype(str)
                    ym = t_str.str.extract(r"(\d{4})\D{0,3}(\d{1,2})")
                    yy = pd.to_numeric(ym[0], errors="coerce").fillna(0).astype(int)
                    mm = pd.to_numeric(ym[1], errors="coerce").fillna(0).astype(int)
                    if (yy.eq(0).mean() > 0.5) or (mm.eq(0).mean() > 0.5):
                        t_num = pd.to_numeric(t0, errors="coerce")
                        dt = pd.to_datetime(t0, errors="coerce")
                        dt2 = pd.to_datetime(t_num, unit="D", origin="1899-12-30", errors="coerce")
                        dt = dt.fillna(dt2)
                        yy = dt.dt.year.fillna(0).astype(int)
                        mm = dt.dt.month.fillna(0).astype(int)

                yy = pd.to_numeric(yy, errors="coerce").fillna(0).astype(int) if yy is not None else pd.Series([0] * len(df_newcust_raw))
                mm = pd.to_numeric(mm, errors="coerce").fillna(0).astype(int) if mm is not None else pd.Series([0] * len(df_newcust_raw))
                if yy.eq(0).mean() > 0.5 and (mm.ne(0).mean() > 0.2):
                    anchor_year = 0
                    try:
                        if df_q4_raw is not None and not getattr(df_q4_raw, "empty", True) and "年份" in df_q4_raw.columns:
                            anchor_year = int(pd.to_numeric(df_q4_raw["年份"], errors="coerce").fillna(0).max())
                    except Exception:
                        anchor_year = 0
                    try:
                        if (anchor_year < 2000) and df_perf_raw is not None and not getattr(df_perf_raw, "empty", True) and "年份" in df_perf_raw.columns:
                            anchor_year = int(pd.to_numeric(df_perf_raw["年份"], errors="coerce").fillna(0).max())
                    except Exception:
                        pass
                    if anchor_year < 2000:
                        anchor_year = int(datetime.now().year)
                    yy = yy.where(yy != 0, anchor_year)

                df_newcust_raw["_ym"] = (yy * 100 + mm).astype(int)
                df_newcust_raw = df_newcust_raw[df_newcust_raw["_ym"].between(200001, 209912)].copy()
            except Exception:
                df_newcust_raw = None

        return df, month_cols, df_stock, df_q4_raw, df_perf_raw, df_target_raw, df_scan_raw, df_newcust_raw, debug_logs
        
    except Exception as e:
        import traceback
        return None, None, None, None, None, None, None, None, [f"Error: {str(e)}", traceback.format_exc()]

@st.cache_data(ttl=3600)
def load_builtin_perf_2025():
    base_dir = os.path.dirname(__file__) if "__file__" in globals() else os.getcwd()
    candidate_dirs = [base_dir, os.path.join(base_dir, "builtin_data")]
    split_paths_by_dir = {}
    for d in candidate_dirs:
        if os.path.isdir(d):
            for f in os.listdir(d):
                if f.startswith("perf_2025_part") and (f.endswith(".csv") or f.endswith(".csv.gz")):
                    split_paths_by_dir.setdefault(d, []).append(os.path.join(d, f))
    split_paths = []
    preferred_dir = os.path.join(base_dir, "builtin_data")
    if split_paths_by_dir.get(preferred_dir):
        split_paths = sorted(split_paths_by_dir[preferred_dir])
    elif split_paths_by_dir.get(base_dir):
        split_paths = sorted(split_paths_by_dir[base_dir])

    df0 = None
    if split_paths:
        try:
            dfs = []
            for p in split_paths:
                dfs.append(pd.read_csv(p))
            if dfs:
                df0 = pd.concat(dfs, ignore_index=True)
        except Exception:
            df0 = None

    if df0 is None:
        path = os.path.join(base_dir, "分析底表0115.xlsx")
        if not os.path.exists(path):
            return pd.DataFrame()
        try:
            xl = pd.ExcelFile(path)
            sheet_name = next((s for s in xl.sheet_names if "发货" in str(s)), None)
            if sheet_name is None and len(xl.sheet_names) > 3:
                sheet_name = xl.sheet_names[3]
            if sheet_name is None:
                return pd.DataFrame()
            df0 = xl.parse(sheet_name)
        except Exception:
            return pd.DataFrame()

    df0.columns = [str(c).strip() for c in df0.columns]
    col_year = next((c for c in df0.columns if str(c).strip() == "年份" or "年" in str(c)), None)
    col_month = next((c for c in df0.columns if str(c).strip() == "月份" or "月" in str(c)), None)
    col_prov = next((c for c in df0.columns if "省区" in str(c)), None)
    col_dist = next((c for c in df0.columns if "客户简称" in str(c)), None) or next((c for c in df0.columns if "购货单位" in str(c)), None)
    col_qty = next((c for c in df0.columns if "基本数量" in str(c)), None) or next((c for c in df0.columns if "箱" in str(c) or "数量" in str(c)), None)
    col_amt = next((c for c in df0.columns if "原价金额" in str(c)), None) or next((c for c in df0.columns if "金额" in str(c)), None)
    col_wh = next((c for c in df0.columns if "发货仓" in str(c)), None)
    col_grp = next((c for c in df0.columns if "归类" in str(c)), None)
    col_bigcat = next((c for c in df0.columns if str(c).strip() == "大分类"), None) or next((c for c in df0.columns if "月分析" in str(c)), None)
    col_big = next((c for c in df0.columns if str(c).strip() == "大类"), None)
    col_mid = next((c for c in df0.columns if str(c).strip() == "中类"), None)
    col_small = next((c for c in df0.columns if str(c).strip() == "小类"), None)

    df = pd.DataFrame()
    if col_year is not None: df["年份"] = df0[col_year]
    if col_month is not None: df["月份"] = df0[col_month]
    if col_prov is not None: df["省区"] = df0[col_prov]
    if col_dist is not None: df["经销商名称"] = df0[col_dist]
    if col_qty is not None: df["发货箱数"] = df0[col_qty]
    if col_amt is not None: df["发货金额"] = df0[col_amt]
    if col_wh is not None: df["发货仓"] = df0[col_wh]
    if col_mid is not None: df["中类"] = df0[col_mid]
    if col_grp is not None: df["归类"] = df0[col_grp]
    if col_bigcat is not None: df["大分类"] = df0[col_bigcat]
    if col_big is not None: df["大类"] = df0[col_big]
    if col_small is not None: df["小类"] = df0[col_small]

    for c in ["省区", "经销商名称", "发货仓", "中类", "归类", "大分类", "大类", "小类"]:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str).str.strip()
    if "年份" in df.columns:
        df["年份"] = df["年份"].astype(str).str.extract(r"(\d+)")[0].astype(float).fillna(0).astype(int)
        df["年份"] = df["年份"].apply(lambda y: y + 2000 if 0 < y < 100 else y)
    if "月份" in df.columns:
        df["月份"] = df["月份"].astype(str).str.extract(r"(\d+)")[0].astype(float).fillna(0).astype(int)
    if "发货箱数" in df.columns:
        df["发货箱数"] = pd.to_numeric(df["发货箱数"], errors="coerce").fillna(0)
    if "发货金额" in df.columns:
        df["发货金额"] = pd.to_numeric(df["发货金额"], errors="coerce").fillna(0)
    if "年份" in df.columns and "月份" in df.columns:
        df = df[(df["年份"] == 2025) & (df["月份"].between(1, 12))]
        df["年月"] = pd.to_datetime(df["年份"].astype(str) + "-" + df["月份"].astype(str).str.zfill(2) + "-01", errors="coerce")
    else:
        return pd.DataFrame()
    return df

@st.cache_data(ttl=3600)
def load_builtin_scan_2025():
    base_dir = os.path.dirname(__file__) if "__file__" in globals() else os.getcwd()
    candidate_dirs = [base_dir, os.path.join(base_dir, "builtin_data")]
    split_paths_by_dir = {}
    for d in candidate_dirs:
        if os.path.isdir(d):
            for f in os.listdir(d):
                if f.startswith("scan_2025_part") and (f.endswith(".csv") or f.endswith(".csv.gz")):
                    split_paths_by_dir.setdefault(d, []).append(os.path.join(d, f))
    split_paths = []
    preferred_dir = os.path.join(base_dir, "builtin_data")
    if split_paths_by_dir.get(preferred_dir):
        split_paths = sorted(split_paths_by_dir[preferred_dir])
    elif split_paths_by_dir.get(base_dir):
        split_paths = sorted(split_paths_by_dir[base_dir])

    df0 = None
    if split_paths:
        try:
            dfs = []
            for p in split_paths:
                dfs.append(pd.read_csv(p))
            if dfs:
                df0 = pd.concat(dfs, ignore_index=True)
        except Exception:
            df0 = None

    if df0 is None:
        path = os.path.join(base_dir, "分析底表0115.xlsx")
        if not os.path.exists(path):
            return pd.DataFrame()
        try:
            xl = pd.ExcelFile(path)
            if len(xl.sheet_names) <= 5:
                return pd.DataFrame()
            df0 = xl.parse(5)
        except Exception:
            return pd.DataFrame()

    if df0 is None or df0.empty:
        return pd.DataFrame()

    def _col(idx: int):
        if idx < df0.shape[1]:
            return df0.iloc[:, idx]
        return pd.Series([None] * len(df0))

    df = pd.DataFrame({
        "门店名称": _col(1),
        "经销商名称": _col(18),
        "省区": _col(17),
        "产品大类": _col(19),
        "产品小类": _col(20),
        "经纬度": _col(12),
        "年份": _col(13),
        "月份": _col(14),
        "日": _col(15),
    })

    df["年份"] = df["年份"].astype(str).str.extract(r"(\d+)")[0].astype(float).fillna(0).astype(int)
    df["年份"] = df["年份"].apply(lambda y: y + 2000 if 0 < y < 100 else y)
    df["月份"] = df["月份"].astype(str).str.extract(r"(\d+)")[0].astype(float).fillna(0).astype(int)
    df["日"] = df["日"].astype(str).str.extract(r"(\d+)")[0].astype(float).fillna(0).astype(int)

    for c in ["门店名称", "省区", "经销商名称", "产品大类", "产品小类"]:
        df[c] = df[c].fillna("").astype(str).str.strip()

    coords = df["经纬度"].apply(_parse_lon_lat)
    df["经度"] = coords.apply(lambda x: x[0])
    df["纬度"] = coords.apply(lambda x: x[1])

    df = df[df["年份"] == 2025]
    return df

@st.cache_data(show_spinner=False, ttl=3600)
def load_project_targets_sheet(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    try:
        bio = io.BytesIO(file_bytes)
        xl = pd.ExcelFile(bio)
        names = [str(s) for s in xl.sheet_names]
        preferred = [s for s in names if any(k in s for k in ["专案", "项目", "专案数据"])]
        candidates = preferred + [s for s in names if s not in preferred]

        def _has_store_type(cols) -> bool:
            return any("门店类型" in str(c) for c in cols)

        for sname in candidates:
            try:
                tmp = xl.parse(sname, nrows=0)
                cols = [str(c).strip() for c in tmp.columns]
                if _has_store_type(cols):
                    df = xl.parse(sname)
                    df.columns = [str(c).strip() for c in df.columns]
                    df = df.dropna(how="all")
                    return df
            except Exception:
                continue

        for sname in candidates:
            try:
                raw = xl.parse(sname, header=None)
                if raw is None or raw.empty:
                    continue
                header_row = None
                for r in range(min(80, len(raw))):
                    row = raw.iloc[r].tolist()
                    if any("门店类型" in str(x) for x in row if x is not None):
                        header_row = r
                        break
                if header_row is None:
                    continue
                cols = [str(c).strip() if c is not None else "" for c in raw.iloc[header_row].tolist()]
                df = raw.iloc[header_row + 1 :].copy()
                df.columns = cols
                df = df.dropna(how="all")
                return df
            except Exception:
                continue

        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()

def _first_col_contains(df: pd.DataFrame, keywords: list[str]):
    if df is None or df.empty:
        return None
    cols = [str(c).strip() for c in df.columns]
    for c in cols:
        ok = True
        for k in keywords:
            if k not in c:
                ok = False
                break
        if ok:
            return c
    return None

def _build_project_tracking_store_df(
    df_out_base: pd.DataFrame,
    df_proj_raw: pd.DataFrame,
    year: int,
    month: int | None,
):
    if df_out_base is None or df_out_base.empty:
        return pd.DataFrame(), ["未检测到出库数据"]
    if df_proj_raw is None or df_proj_raw.empty:
        return pd.DataFrame(), ["未检测到专案数据（请确认第7个sheet包含“门店类型”列）"]

    df_proj = df_proj_raw.copy()
    df_proj.columns = [str(c).strip() for c in df_proj.columns]

    prov_col = _first_col_contains(df_proj, ["省区"]) or _first_col_contains(df_proj, ["省区名称"])
    dist_col = _first_col_contains(df_proj, ["经销商"]) or _first_col_contains(df_proj, ["客户简称"]) or _first_col_contains(df_proj, ["客户"])
    store_type_col = _first_col_contains(df_proj, ["门店类型"])
    store_col = None
    for c in df_proj.columns:
        s = str(c)
        if ("门店" in s) and ("门店类型" not in s) and ("编码" not in s):
            store_col = c
            break

    if store_type_col is None and len(df_proj.columns) > 4:
        store_type_col = df_proj.columns[4]

    logs = []
    missing = []
    for k, v in [("省区", prov_col), ("经销商", dist_col), ("门店", store_col), ("门店类型", store_type_col)]:
        if v is None:
            missing.append(k)
    if missing:
        return pd.DataFrame(), [f"专案数据缺少关键列：{', '.join(missing)}"]

    seg_map = {"美思雅段粉": "段粉", "雅系列": "雅系列", "中老年": "中老年"}

    seg_target_candidates = {
        "段粉": [
            _first_col_contains(df_proj, ["段粉", "目标"]),
            _first_col_contains(df_proj, ["段粉", "任务"]),
            _first_col_contains(df_proj, ["段粉"]),
        ],
        "雅系列": [
            _first_col_contains(df_proj, ["雅系列", "目标"]),
            _first_col_contains(df_proj, ["雅系列", "任务"]),
            _first_col_contains(df_proj, ["雅系列"]),
        ],
        "中老年": [
            _first_col_contains(df_proj, ["中老年", "目标"]),
            _first_col_contains(df_proj, ["中老年", "任务"]),
            _first_col_contains(df_proj, ["中老年"]),
        ],
    }
    seg_target_col = {k: next((c for c in v if c), None) for k, v in seg_target_candidates.items()}

    if any(seg_target_col[k] is None for k in ["段粉", "雅系列", "中老年"]):
        miss2 = [k for k in ["段粉", "雅系列", "中老年"] if seg_target_col[k] is None]
        return pd.DataFrame(), [f"专案数据未找到目标列：{', '.join(miss2)}（请确保列名含“段粉/雅系列/中老年”）"]

    mid_target_is_ti = "提" in str(seg_target_col["中老年"])

    df_tgt = df_proj[[prov_col, dist_col, store_col, store_type_col, seg_target_col["段粉"], seg_target_col["雅系列"], seg_target_col["中老年"]]].copy()
    df_tgt.columns = ["省区", "经销商名称", "门店名称", "门店类型", "段粉_目标", "雅系列_目标", "中老年_目标"]
    for c in ["省区", "经销商名称", "门店名称", "门店类型"]:
        df_tgt[c] = df_tgt[c].fillna("").astype(str).str.strip()
        if c in ("经销商名称", "门店名称"):
            df_tgt[c] = df_tgt[c].str.replace(r"\s+", "", regex=True)
    for c in ["段粉_目标", "雅系列_目标", "中老年_目标"]:
        df_tgt[c] = pd.to_numeric(df_tgt[c], errors="coerce").fillna(0.0)
    if not mid_target_is_ti:
        df_tgt["中老年_目标"] = df_tgt["中老年_目标"] * 3.0

    df_out_all = df_out_base.copy()
    if "_年" in df_out_all.columns:
        df_out_all = df_out_all[df_out_all["_年"].astype(int) == int(year)].copy()

    cat_col = "产品大类" if "产品大类" in df_out_all.columns else ("_模块大类" if "_模块大类" in df_out_all.columns else None)
    if cat_col is None:
        return pd.DataFrame(), ["出库数据缺少“大类/产品大类”字段"]

    weight_col = None
    if "重量" in df_out_all.columns:
        weight_col = "重量"
    elif "产品小类" in df_out_all.columns:
        weight_col = "产品小类"
    elif "_模块小类" in df_out_all.columns:
        weight_col = "_模块小类"

    if "_门店名" not in df_out_all.columns:
        store_name_col = None
        for c in df_out_all.columns:
            if "门店" in str(c) and "类型" not in str(c):
                store_name_col = c
                break
        if store_name_col:
            df_out_all["_门店名"] = df_out_all[store_name_col].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
        else:
            df_out_all["_门店名"] = pd.NA

    for c in ["省区", "经销商名称", "_门店名", cat_col]:
        if c in df_out_all.columns:
            df_out_all[c] = df_out_all[c].fillna("").astype(str).str.strip()
    if "经销商名称" in df_out_all.columns:
        df_out_all["经销商名称"] = df_out_all["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True)
    if "_门店名" in df_out_all.columns:
        df_out_all["_门店名"] = df_out_all["_门店名"].astype(str).str.replace(r"\s+", "", regex=True)
    if weight_col is not None and weight_col in df_out_all.columns:
        df_out_all[weight_col] = df_out_all[weight_col].fillna("").astype(str).str.strip()

    proj_dists = sorted([x for x in df_tgt["经销商名称"].dropna().astype(str).unique().tolist() if x and x.lower() not in ("nan", "none", "null")])
    if proj_dists:
        df_out_all = df_out_all[df_out_all["经销商名称"].astype(str).isin(proj_dists)].copy()

    df_today_base = df_out_all.copy()
    try:
        df_today_base = df_today_base[df_today_base[cat_col].astype(str).isin(list(seg_map.keys()))].copy()
        if weight_col is not None and weight_col in df_today_base.columns:
            w_digits = df_today_base[weight_col].astype(str).str.extract(r"(\d+)")[0].fillna("")
            is_800 = w_digits.astype(str).eq("800")
            df_today_base = df_today_base[~((df_today_base[cat_col].astype(str) == "美思雅段粉") & (~is_800))].copy()
    except Exception:
        df_today_base = df_out_all.copy()

    today_m, today_d = None, None
    if {"_年", "_月", "_日"}.issubset(set(df_today_base.columns)):
        try:
            yy = pd.to_numeric(df_today_base["_年"], errors="coerce").fillna(0).astype(int)
            mm = pd.to_numeric(df_today_base["_月"], errors="coerce")
            dd = pd.to_numeric(df_today_base["_日"], errors="coerce")
            tmp = pd.DataFrame({"_年": yy, "_月": mm, "_日": dd})
            tmp = tmp[(tmp["_年"] == int(year)) & tmp["_月"].between(1, 12) & tmp["_日"].between(1, 31)]
            if not tmp.empty:
                key = (tmp["_年"].astype(int) * 10000 + tmp["_月"].astype(int) * 100 + tmp["_日"].astype(int)).max()
                key = int(key)
                today_m = int((key % 10000) // 100)
                today_d = int(key % 100)
        except Exception:
            today_m, today_d = None, None
    if today_m is not None and today_d is not None:
        logs.append(f"今日出库日期：{today_m:02d}-{today_d:02d}")

    df_out = df_out_all.copy()
    if month is not None and "_月" in df_out.columns:
        df_out = df_out[df_out["_月"].astype(int) == int(month)].copy()
    df_out_today = df_today_base.copy()
    if today_m is not None and today_d is not None and {"_月", "_日"}.issubset(set(df_out_today.columns)):
        df_out_today = df_out_today[
            (pd.to_numeric(df_out_today["_月"], errors="coerce").astype("Int64") == int(today_m))
            & (pd.to_numeric(df_out_today["_日"], errors="coerce").astype("Int64") == int(today_d))
        ].copy()
    else:
        df_out_today = df_out_today.iloc[0:0].copy()

    def _prep_out(df_in: pd.DataFrame) -> pd.DataFrame:
        if df_in is None or df_in.empty:
            return df_in.iloc[0:0].copy()
        d = df_in.copy()
        d = d[d[cat_col].astype(str).isin(list(seg_map.keys()))].copy()
        if weight_col is not None and weight_col in d.columns:
            w_digits = d[weight_col].astype(str).str.extract(r"(\d+)")[0].fillna("")
            is_800 = w_digits.astype(str).eq("800")
            d = d[~((d[cat_col].astype(str) == "美思雅段粉") & (~is_800))].copy()
        d["数量(箱)"] = pd.to_numeric(d.get("数量(箱)", 0.0), errors="coerce").fillna(0.0)
        d["_seg"] = d[cat_col].map(seg_map).fillna("")
        return d

    df_out = _prep_out(df_out)
    df_out_today = _prep_out(df_out_today)

    agg_out = (
        df_out.groupby(["省区", "经销商名称", "_门店名", "_seg"], as_index=False)["数量(箱)"]
        .sum()
        .rename(columns={"_门店名": "门店名称"})
    )
    agg_out.loc[agg_out["_seg"] == "中老年", "数量(箱)"] = agg_out.loc[agg_out["_seg"] == "中老年", "数量(箱)"] * 3.0

    out_pv = agg_out.pivot_table(index=["省区", "经销商名称", "门店名称"], columns="_seg", values="数量(箱)", aggfunc="sum").fillna(0.0).reset_index()
    for seg in ["段粉", "雅系列", "中老年"]:
        if seg not in out_pv.columns:
            out_pv[seg] = 0.0
    out_pv = out_pv.rename(columns={"段粉": "段粉_出库", "雅系列": "雅系列_出库", "中老年": "中老年_出库"})

    agg_today = (
        df_out_today.groupby(["省区", "经销商名称", "_门店名", "_seg"], as_index=False)["数量(箱)"]
        .sum()
        .rename(columns={"_门店名": "门店名称"})
    )
    agg_today.loc[agg_today["_seg"] == "中老年", "数量(箱)"] = agg_today.loc[agg_today["_seg"] == "中老年", "数量(箱)"] * 3.0
    today_pv = agg_today.pivot_table(index=["省区", "经销商名称", "门店名称"], columns="_seg", values="数量(箱)", aggfunc="sum").fillna(0.0).reset_index()
    for seg in ["段粉", "雅系列", "中老年"]:
        if seg not in today_pv.columns:
            today_pv[seg] = 0.0
    today_pv = today_pv.rename(columns={"段粉": "段粉_今日出库", "雅系列": "雅系列_今日出库", "中老年": "中老年_今日出库"})

    store_df = df_tgt.merge(out_pv, on=["省区", "经销商名称", "门店名称"], how="left")
    store_df = store_df.merge(today_pv, on=["省区", "经销商名称", "门店名称"], how="left")
    for c in ["段粉_出库", "雅系列_出库", "中老年_出库", "段粉_今日出库", "雅系列_今日出库", "中老年_今日出库"]:
        store_df[c] = pd.to_numeric(store_df.get(c, 0.0), errors="coerce").fillna(0.0)

    def _rate(out_v, tgt_v):
        try:
            t = float(tgt_v or 0.0)
            if t <= 0:
                return None
            return float(out_v or 0.0) / t
        except Exception:
            return None

    store_df["段粉-目标值"] = store_df["段粉_目标"].astype(float)
    store_df["段粉-出库值"] = store_df["段粉_出库"].astype(float)
    store_df["段粉-完成率"] = store_df.apply(lambda r: _rate(r["段粉-出库值"], r["段粉-目标值"]), axis=1)
    store_df["段粉-今日出库"] = store_df["段粉_今日出库"].astype(float)

    store_df["雅系列-目标值"] = store_df["雅系列_目标"].astype(float)
    store_df["雅系列-出库值"] = store_df["雅系列_出库"].astype(float)
    store_df["雅系列-完成率"] = store_df.apply(lambda r: _rate(r["雅系列-出库值"], r["雅系列-目标值"]), axis=1)
    store_df["雅系列-今日出库"] = store_df["雅系列_今日出库"].astype(float)

    store_df["中老年-目标值(提)"] = store_df["中老年_目标"].astype(float)
    store_df["中老年-出库值(提)"] = store_df["中老年_出库"].astype(float)
    store_df["中老年-完成率"] = store_df.apply(lambda r: _rate(r["中老年-出库值(提)"], r["中老年-目标值(提)"]), axis=1)
    store_df["中老年-今日出库(提)"] = store_df["中老年_今日出库"].astype(float)

    store_df["库存"] = 0.0
    store_df["本月新客"] = 0.0
    store_df["近三月新客"] = 0.0
    store_df["累计新客"] = 0.0
    store_df["本月扫码"] = 0.0
    store_df["本月扫码率"] = 0.0

    anchor_ym = None
    if month is not None:
        anchor_ym = int(int(year) * 100 + int(month))
    else:
        try:
            if "_月" in df_out_all.columns:
                mm = pd.to_numeric(df_out_all["_月"], errors="coerce").fillna(0).astype(int)
                mm = mm[(mm >= 1) & (mm <= 12)]
                if not mm.empty:
                    anchor_ym = int(int(year) * 100 + int(mm.max()))
        except Exception:
            anchor_ym = None

    if anchor_ym is None:
        try:
            _scan = globals().get("df_scan_raw")
            if _scan is not None and not getattr(_scan, "empty", True):
                _s0 = _scan.copy()
                _s0["_ym"] = (pd.to_numeric(_s0.get("年份", 0), errors="coerce").fillna(0).astype(int) * 100 + pd.to_numeric(_s0.get("月份", 0), errors="coerce").fillna(0).astype(int)).astype(int)
                _s0 = _s0[_s0["_ym"].between(200001, 209912)]
                _s0 = _s0[_s0["_ym"].astype(int).astype(str).str.startswith(str(int(year)))]
                if not _s0.empty:
                    anchor_ym = int(pd.to_numeric(_s0["_ym"], errors="coerce").fillna(0).max())
        except Exception:
            anchor_ym = None

    if anchor_ym is not None:
        y = int(anchor_ym // 100)
        m = int(anchor_ym % 100)
        prev3 = []
        y2, m2 = y, m
        for _ in range(3):
            m2 -= 1
            if m2 <= 0:
                y2 -= 1
                m2 += 12
            prev3.append(int(y2 * 100 + m2))

        nc0 = globals().get("df_newcust_raw")
        if nc0 is not None and not getattr(nc0, "empty", True) and "_ym" in nc0.columns:
            nc = nc0.copy()
            if "省区" in nc.columns:
                nc["省区"] = nc["省区"].fillna("").astype(str).str.strip()
            if "经销商名称" in nc.columns:
                nc["经销商名称"] = nc["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
            if "门店名称" in nc.columns:
                nc["门店名称"] = nc["门店名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
            if "新客数" in nc.columns:
                nc["新客数"] = pd.to_numeric(nc["新客数"], errors="coerce").fillna(0.0)
            nc["_ym"] = pd.to_numeric(nc["_ym"], errors="coerce").fillna(0).astype(int)
            nc = nc[nc["_ym"].between(200001, 209912)].copy()
            nc = nc[nc["_ym"] <= anchor_ym].copy()
            cur = nc[nc["_ym"] == anchor_ym].groupby(["省区", "经销商名称", "门店名称"], as_index=False)["新客数"].sum().rename(columns={"新客数": "本月新客"})
            p3 = nc[nc["_ym"].isin(prev3)].groupby(["省区", "经销商名称", "门店名称"], as_index=False)["新客数"].sum().rename(columns={"新客数": "近三月新客"})
            cum = nc[(nc["_ym"] >= 202501) & (nc["_ym"] <= anchor_ym)].groupby(["省区", "经销商名称", "门店名称"], as_index=False)["新客数"].sum().rename(columns={"新客数": "累计新客"})
            for _df in (cur, p3, cum):
                _df["省区"] = _df["省区"].fillna("").astype(str).str.strip()
                _df["经销商名称"] = _df["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                _df["门店名称"] = _df["门店名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
            store_df = store_df.merge(cur, on=["省区", "经销商名称", "门店名称"], how="left")
            store_df = store_df.merge(p3, on=["省区", "经销商名称", "门店名称"], how="left")
            store_df = store_df.merge(cum, on=["省区", "经销商名称", "门店名称"], how="left")
            for _c in ["本月新客", "近三月新客", "累计新客"]:
                if _c in store_df.columns:
                    store_df[_c] = pd.to_numeric(store_df[_c], errors="coerce").fillna(0.0)
                else:
                    store_df[_c] = 0.0

        st0 = globals().get("df_stock_raw")
        if st0 is not None and not getattr(st0, "empty", True):
            _s = st0.copy()
            if "省区" not in _s.columns and "省区名称" in _s.columns:
                _s["省区"] = _s["省区名称"]
            for _c in ["省区", "经销商名称", "产品大类", "产品名称", "重量"]:
                if _c in _s.columns:
                    if _c == "经销商名称":
                        _s[_c] = _s[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                    else:
                        _s[_c] = _s[_c].fillna("").astype(str).str.strip()
            if "箱数" in _s.columns:
                _s["箱数"] = pd.to_numeric(_s["箱数"], errors="coerce").fillna(0.0)
            else:
                _s["箱数"] = 0.0
            _mask_seg = pd.Series(False, index=_s.index)
            if "产品大类" in _s.columns:
                _mask_seg |= _s["产品大类"].astype(str).str.strip().eq("美思雅段粉")
                _mask_seg |= _s["产品大类"].astype(str).str.contains("中老年", regex=False)
            if "产品名称" in _s.columns:
                _mask_seg |= _s["产品名称"].astype(str).str.contains(r"(雅赋|雅耀|雅舒|雅护)", regex=True)
            if _mask_seg.any():
                _s = _s[_mask_seg].copy()
            if "产品大类" in _s.columns and "重量" in _s.columns:
                w_digits = _s["重量"].astype(str).str.extract(r"(\d{3})")[0].fillna("")
                _s = _s[~((_s["产品大类"].astype(str).str.strip() == "美思雅段粉") & (w_digits.astype(str) != "800"))].copy()
            inv = _s.groupby(["省区", "经销商名称"], as_index=False)["箱数"].sum().rename(columns={"箱数": "库存"})
            for _c in ["省区", "经销商名称"]:
                inv[_c] = inv[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True) if _c == "经销商名称" else inv[_c].fillna("").astype(str).str.strip()
            store_df = store_df.merge(inv, on=["省区", "经销商名称"], how="left", suffixes=("", "_y"))
            if "库存_y" in store_df.columns:
                store_df.drop(columns=["库存"], inplace=True, errors="ignore")
                store_df.rename(columns={"库存_y": "库存"}, inplace=True)
            store_df["库存"] = pd.to_numeric(store_df.get("库存", 0), errors="coerce").fillna(0.0)

        sc0 = globals().get("df_scan_raw")
        if sc0 is not None and not getattr(sc0, "empty", True):
            s = sc0.copy()
            for _c in ["省区", "经销商名称", "门店名称", "产品大类", "产品小类"]:
                if _c in s.columns:
                    if _c in ("经销商名称", "门店名称"):
                        s[_c] = s[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                    else:
                        s[_c] = s[_c].fillna("").astype(str).str.strip()
            s["年份"] = pd.to_numeric(s.get("年份", 0), errors="coerce").fillna(0).astype(int)
            s["月份"] = pd.to_numeric(s.get("月份", 0), errors="coerce").fillna(0).astype(int)
            s = s[(s["年份"] > 0) & (s["月份"].between(1, 12))].copy()
            s["_ym"] = (s["年份"] * 100 + s["月份"]).astype(int)
            s = s[s["_ym"] == anchor_ym].copy()
            if not s.empty:
                _mask = pd.Series(False, index=s.index)
                if "产品大类" in s.columns:
                    _mask |= s["产品大类"].astype(str).str.strip().eq("美思雅段粉")
                    _mask |= s["产品大类"].astype(str).str.contains("中老年", regex=False)
                if "产品小类" in s.columns:
                    _mask |= s["产品小类"].astype(str).str.contains(r"(雅赋|雅耀|雅舒|雅护)", regex=True)
                s = s[_mask].copy()
                if not s.empty and "产品大类" in s.columns and "产品小类" in s.columns:
                    w_digits = s["产品小类"].astype(str).str.extract(r"(\d{3})")[0].fillna("")
                    s = s[~((s["产品大类"].astype(str).str.strip() == "美思雅段粉") & (w_digits.astype(str) != "800"))].copy()
                if not s.empty:
                    scan_agg = (
                        s.groupby(["省区", "经销商名称", "门店名称"], as_index=False)
                        .size()
                        .rename(columns={"size": "_扫码听数"})
                    )
                    scan_agg["本月扫码"] = pd.to_numeric(scan_agg["_扫码听数"], errors="coerce").fillna(0.0) / 6.0
                    scan_agg.drop(columns=["_扫码听数"], inplace=True, errors="ignore")
                    store_df = store_df.merge(scan_agg, on=["省区", "经销商名称", "门店名称"], how="left", suffixes=("", "_y"))
                    if "本月扫码_y" in store_df.columns:
                        store_df.drop(columns=["本月扫码"], inplace=True, errors="ignore")
                        store_df.rename(columns={"本月扫码_y": "本月扫码"}, inplace=True)
                    store_df["本月扫码"] = pd.to_numeric(store_df.get("本月扫码", 0), errors="coerce").fillna(0.0)

        out_box = (
            pd.to_numeric(store_df.get("段粉-出库值", 0), errors="coerce").fillna(0.0)
            + pd.to_numeric(store_df.get("雅系列-出库值", 0), errors="coerce").fillna(0.0)
            + (pd.to_numeric(store_df.get("中老年-出库值(提)", 0), errors="coerce").fillna(0.0) / 3.0)
        )
        scan_box = pd.to_numeric(store_df.get("本月扫码", 0), errors="coerce").fillna(0.0)
        store_df["本月扫码率"] = np.where(out_box > 0, scan_box / out_box, 0.0)
        store_df["本月扫码率"] = pd.to_numeric(store_df.get("本月扫码率", 0), errors="coerce").fillna(0.0)

    keep = [
        "省区",
        "经销商名称",
        "门店名称",
        "门店类型",
        "段粉-目标值",
        "段粉-出库值",
        "段粉-完成率",
        "段粉-今日出库",
        "雅系列-目标值",
        "雅系列-出库值",
        "雅系列-完成率",
        "雅系列-今日出库",
        "中老年-目标值(提)",
        "中老年-出库值(提)",
        "中老年-完成率",
        "中老年-今日出库(提)",
        "库存",
        "本月新客",
        "近三月新客",
        "累计新客",
        "本月扫码",
        "本月扫码率",
    ]
    store_df = store_df[keep].copy()
    return store_df, logs

# -----------------------------------------------------------------------------
# 4. Layout
# -----------------------------------------------------------------------------

st.markdown("## 🛠️ 数据控制台")

if 'hc_mode' not in st.session_state:
    st.session_state.hc_mode = False

st.toggle("高对比模式", key="hc_mode")

if st.session_state.get("hc_mode"):
    st.markdown("""
    <style>
      :root {
        --tbl-header-bg: #0B57D0;
        --tbl-header-bg-hover: #0846AB;
        --tbl-header-border: #06357F;
        --tbl-header-fg: #FFFFFF;
        --tbl-header-icon: #FFFFFF;
        --tbl-header-shadow: 0 10px 22px rgba(0, 0, 0, 0.38);
      }
    </style>
    """, unsafe_allow_html=True)

if 'exp_upload' not in st.session_state:
    st.session_state.exp_upload = True
if 'exp_filter' not in st.session_state:
    st.session_state.exp_filter = True

with st.expander("📥 数据导入", expanded=st.session_state.exp_upload):
    uploaded_file = st.file_uploader("导入数据表 (Excel/CSV)", type=['xlsx', 'xls', 'csv'], key="main_uploader")
    c_u1, c_u2 = st.columns([1, 3])
    with c_u1:
        if st.button("🔄 清缓存重解析", key="clear_cache_reparse_btn"):
            for k in [
                "main_uploader",
                "_uploaded_bytes",
                "_uploaded_sig",
                "_uploaded_name",
                "_active_file_sig",
                "_parsed_cache",
                "out_subtab_cache",
                "out_m_excel_cache",
                "out_m_zip_cache",
                "out_m_month_cols",
                "out_m_drill_level",
                "out_m_selected_prov",
                "out_m_selected_dist",
            ]:
                st.session_state.pop(k, None)
            try:
                st.cache_data.clear()
            except Exception:
                pass
            st.rerun()
    with c_u2:
        st.caption("如果上传后仍看不到新客列，点一次这里可强制清理解析/页面缓存。")

if uploaded_file is None:
    st.markdown(
        """
        <div style='text-align: center; padding: 60px 20px; background-color: white; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.05); margin-bottom: 40px;'>
            <h1 style='color: #4096ff; margin-bottom: 16px;'>👋 欢迎使用美思雅数据分析系统</h1>
            <p style='color: #666; font-size: 16px; margin-bottom: 0;'>请上传 Excel 数据文件以解锁完整分析面板</p>
        </div>
        """,
        unsafe_allow_html=True
    )
    # st.stop()  # Streamlit Cloud Health Check Fix: Avoid blocking the app here

# Main Logic
if uploaded_file:
    uploaded_name = uploaded_file.name
    cached_bytes = st.session_state.get("_uploaded_bytes")
    cached_sig = st.session_state.get("_uploaded_sig")
    cached_name = st.session_state.get("_uploaded_name")

    if cached_bytes is None or cached_sig is None or cached_name != uploaded_name:
        cached_bytes = uploaded_file.getvalue()
        cached_sig = hashlib.md5(cached_bytes).hexdigest()
        st.session_state["_uploaded_bytes"] = cached_bytes
        st.session_state["_uploaded_sig"] = cached_sig
        st.session_state["_uploaded_name"] = uploaded_name

    if st.session_state.get("_active_file_sig") != cached_sig:
        st.session_state["_active_file_sig"] = cached_sig
        st.session_state["run_analysis"] = False
        st.session_state.pop("out_subtab_cache", None)
        st.session_state.pop("out_m_excel_cache", None)
        st.session_state.pop("out_m_zip_cache", None)

    parsed_cache = st.session_state.get("_parsed_cache", {})
    if cached_sig in parsed_cache:
        # Check cache format
        cache_val = parsed_cache[cached_sig]
        if len(cache_val) == 9:
            df_raw, month_cols, df_stock_raw, df_q4_raw, df_perf_raw, df_target_raw, df_scan_raw, df_newcust_raw, debug_logs = cache_val
        elif len(cache_val) == 8:
            df_raw, month_cols, df_stock_raw, df_q4_raw, df_perf_raw, df_target_raw, df_scan_raw, df_newcust_raw, debug_logs = load_data_v3(cached_bytes, uploaded_name)
            parsed_cache[cached_sig] = (df_raw, month_cols, df_stock_raw, df_q4_raw, df_perf_raw, df_target_raw, df_scan_raw, df_newcust_raw, debug_logs)
            st.session_state["_parsed_cache"] = parsed_cache
        else:
            df_raw, month_cols, df_stock_raw, df_q4_raw, df_perf_raw, df_target_raw, df_scan_raw, df_newcust_raw, debug_logs = load_data_v3(cached_bytes, uploaded_name)
            parsed_cache[cached_sig] = (df_raw, month_cols, df_stock_raw, df_q4_raw, df_perf_raw, df_target_raw, df_scan_raw, df_newcust_raw, debug_logs)
            st.session_state["_parsed_cache"] = parsed_cache
    else:
        df_raw, month_cols, df_stock_raw, df_q4_raw, df_perf_raw, df_target_raw, df_scan_raw, df_newcust_raw, debug_logs = load_data_v3(cached_bytes, uploaded_name)
        parsed_cache[cached_sig] = (df_raw, month_cols, df_stock_raw, df_q4_raw, df_perf_raw, df_target_raw, df_scan_raw, df_newcust_raw, debug_logs)
        if len(parsed_cache) > 2:
            for k in list(parsed_cache.keys())[:-2]:
                parsed_cache.pop(k, None)
        st.session_state["_parsed_cache"] = parsed_cache

    df_perf_2025 = load_builtin_perf_2025()
    if df_perf_2025 is not None and not df_perf_2025.empty:
        if df_perf_raw is None or getattr(df_perf_raw, "empty", True):
            df_perf_raw = df_perf_2025.copy()
        else:
            # Ensure Year type consistency
            if "年份" in df_perf_raw.columns:
                df_perf_raw["年份"] = pd.to_numeric(df_perf_raw["年份"], errors="coerce")
            if "年份" in df_perf_2025.columns:
                df_perf_2025["年份"] = pd.to_numeric(df_perf_2025["年份"], errors="coerce")
                
            years = df_perf_raw.get("年份", pd.Series(dtype=float))
            # Only append if 2025 is NOT already in the uploaded file
            if not bool((years == 2025).any()):
                df_perf_raw = pd.concat([df_perf_2025, df_perf_raw], ignore_index=True, sort=False)
                
    df_scan_2025 = load_builtin_scan_2025()
    if df_scan_2025 is not None and not df_scan_2025.empty:
        if df_scan_raw is None or getattr(df_scan_raw, "empty", True):
            df_scan_raw = df_scan_2025.copy()
        else:
            # Ensure Year type consistency
            if "年份" in df_scan_raw.columns:
                df_scan_raw["年份"] = pd.to_numeric(df_scan_raw["年份"], errors="coerce")
            if "年份" in df_scan_2025.columns:
                df_scan_2025["年份"] = pd.to_numeric(df_scan_2025["年份"], errors="coerce")
                
            years_s = df_scan_raw.get("年份", pd.Series(dtype=float))
            # Only append if 2025 is NOT already in the uploaded file
            if not bool((years_s == 2025).any()):
                df_scan_raw = pd.concat([df_scan_2025, df_scan_raw], ignore_index=True, sort=False)

    if df_raw is None and debug_logs:
        st.error("数据加载失败。详细日志如下：")
        st.text("\n".join(debug_logs))

    df_filter_src = None
    for _df in (df_raw, df_q4_raw, df_stock_raw, df_perf_raw, df_scan_raw, df_newcust_raw):
        if _df is None or getattr(_df, "empty", True):
            continue
        if ("省区" in _df.columns) and ("经销商名称" in _df.columns):
            df_filter_src = _df
            break

    if df_filter_src is None:
        st.error("未找到包含「省区」「经销商名称」字段的数据表（第1个sheet已删除也可用其它sheet，但需包含这两列）。")
        st.stop()
    else:
        nc_ok = df_newcust_raw is not None and not getattr(df_newcust_raw, "empty", True) and ("_ym" in df_newcust_raw.columns)
        if nc_ok:
            try:
                _yms = sorted([int(x) for x in pd.to_numeric(df_newcust_raw["_ym"], errors="coerce").dropna().astype(int).unique().tolist() if 200001 <= int(x) <= 209912])
            except Exception:
                _yms = []
            st.caption(f"新客表：已读取（行数={int(len(df_newcust_raw))}，月份={('、'.join([str(x) for x in _yms[:12]]) + ('…' if len(_yms) > 12 else '')) if _yms else '未识别到月份'}）")
        else:
            st.caption("新客表：未读取（请确认底表第8个sheet包含新客数据，并点「清缓存重解析」后重新上传）")
        def _col_series(_df: pd.DataFrame, _name: str) -> pd.Series:
            if _df is None or getattr(_df, "empty", True) or (_name not in _df.columns):
                return pd.Series([], dtype=object)
            _v = _df[_name]
            if isinstance(_v, pd.DataFrame):
                if _v.shape[1] >= 1:
                    return _v.iloc[:, 0]
                return pd.Series([""] * int(len(_df)), index=_df.index, dtype=object)
            return _v

        store_geo_df = None
        try:
            if df_raw is not None and not getattr(df_raw, "empty", True) and int(df_raw.shape[1]) >= 9:
                _store = df_raw.iloc[:, 8].fillna("").astype(str).str.strip()
                _city = df_raw.iloc[:, 2].fillna("").astype(str).str.strip() if int(df_raw.shape[1]) >= 3 else pd.Series([""] * len(df_raw))
                _dist = df_raw.iloc[:, 3].fillna("").astype(str).str.strip() if int(df_raw.shape[1]) >= 4 else pd.Series([""] * len(df_raw))
                _status = df_raw.iloc[:, 17].fillna("").astype(str).str.strip() if int(df_raw.shape[1]) >= 18 else pd.Series([""] * len(df_raw))
                _m0 = pd.DataFrame({"门店名称": _store, "市": _city, "区/县": _dist, "门店状态": _status})
                _m0 = _m0[_m0["门店名称"].fillna("").astype(str).str.strip() != ""].copy()
                if not _m0.empty:
                    _m0["_k_store_geo"] = _m0["门店名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)

                    def _first_non_empty(vs):
                        for x in vs.tolist():
                            s = str(x or "").strip()
                            if s and s.lower() not in ("nan", "none", "null"):
                                return s
                        return ""

                    store_geo_df = (
                        _m0.groupby(["_k_store_geo"], as_index=False)
                        .agg({"市": _first_non_empty, "区/县": _first_non_empty, "门店状态": _first_non_empty})
                    )
        except Exception:
            store_geo_df = None

        # --- Filters Area ---
        with st.expander("🔎 筛选搜索", expanded=st.session_state.exp_filter):
            # Province Filter
            _prov_s = _col_series(df_filter_src, "省区").dropna().astype(str).str.strip()
            provinces = ["全部"] + sorted([x for x in _prov_s.unique().tolist() if x])
            sel_prov = st.selectbox("选择省区 (Province)", provinces)
            
            # Distributor Filter
            if sel_prov != '全部':
                _m = _col_series(df_filter_src, "省区").fillna("").astype(str).str.strip() == str(sel_prov).strip()
                _dist_s = _col_series(df_filter_src.loc[_m], "经销商名称").dropna().astype(str).str.strip()
                dist_options = ["全部"] + sorted([x for x in _dist_s.unique().tolist() if x])
            else:
                _dist_s = _col_series(df_filter_src, "经销商名称").dropna().astype(str).str.strip()
                dist_options = ["全部"] + sorted([x for x in _dist_s.unique().tolist() if x])
            sel_dist = st.selectbox("选择经销商 (Distributor)", dist_options)

            cat_set = set()
            for _df, _col in [
                (df_perf_raw, '大分类'),
                (df_perf_raw, '产品大类'),
                (df_q4_raw, '产品大类'),
                (df_stock_raw, '产品大类'),
                (df_scan_raw, '产品大类'),
            ]:
                if _df is not None and not getattr(_df, "empty", True) and _col in _df.columns:
                    cat_set |= set(_df[_col].fillna('').astype(str).str.strip().tolist())
            cat_options = ['全部'] + sorted([x for x in cat_set if x])
            sel_cat = st.selectbox("选择产品大类 (Category)", cat_options, key="main_sel_cat")
        
        # Apply Filters
        df = df_filter_src.copy()
        if sel_prov != '全部':
            _m = _col_series(df, "省区").fillna("").astype(str).str.strip() == str(sel_prov).strip()
            df = df.loc[_m].copy()
        if sel_dist != '全部':
            _m = _col_series(df, "经销商名称").fillna("").astype(str).str.strip() == str(sel_dist).strip()
            df = df.loc[_m].copy()
            
        if not st.session_state.get('run_analysis', False):
            st.markdown("### ✅ 数据已加载")
            st.caption("点击「开始分析 🚀」进入分析页面。")
            if st.button("开始分析 🚀", type="primary", key="main_start_analysis"):
                st.session_state['run_analysis'] = True

        # Share / external-access UI intentionally removed
            
        if st.session_state.get('run_analysis', False):
            
            # --- Header ---
            st.title("📈 美思雅数据分析系统")
            st.markdown(f"当前数据范围: **{sel_prov}** / **{sel_dist}** | 包含 **{len(df)}** 家门店")
            
            main_tab_options = ["📊 核心概览", "🚀 业绩分析", "📦 库存分析", "🚚 出库分析", "📱 扫码分析"]
            _default_nav = st.session_state.get("main_nav", "📊 核心概览") if isinstance(st.session_state.get("main_nav", None), str) else "📊 核心概览"
            if _default_nav not in main_tab_options:
                _default_nav = "📊 核心概览"
            main_tab = st.segmented_control(
                "主导航",
                options=main_tab_options,
                default=_default_nav,
                key="main_nav",
                label_visibility="collapsed",
            )
            
            # === TAB 1: OVERVIEW ===
            if main_tab == "📊 核心概览":
                st.caption(f"筛选口径：省区={sel_prov}｜经销商={sel_dist}｜产品大类={st.session_state.get('main_sel_cat', '全部')}")

                # --- Common Helpers for Tab 1 ---
                def _fmt_wan(x): return fmt_num((x or 0) / 10000)
                def _fmt_pct(x): return fmt_pct_ratio(x) if x is not None else "—"
                def _arrow(x): return "↑" if x and x>0 else ("↓" if x and x<0 else "")
                def _trend_cls(x): return "trend-up" if x and x > 0 else ("trend-down" if x and x < 0 else "trend-neutral")

                # Card Renderer for Performance (Tab 7 Style)
                def _render_perf_card(title, icon, val_wan, target_wan, rate, yoy_val_wan, yoy_pct):
                    trend_cls = _trend_cls(yoy_pct)
                    arrow = _arrow(yoy_pct)
                    rate_txt = _fmt_pct(rate)
                    yoy_txt = _fmt_pct(yoy_pct)
                    pct_val = min(max(rate * 100 if rate else 0, 0), 100)
                    prog_color = "#28A745" if rate and rate >= 1.0 else ("#FFC107" if rate and rate >= 0.8 else "#DC3545")

                    st.markdown(f"""
                    <div class="out-kpi-card">
                        <div class="out-kpi-bar"></div>
                        <div class="out-kpi-head">
                            <div class="out-kpi-ico">{icon}</div>
                            <div class="out-kpi-title">{title}</div>
                        </div>
                        <div class="out-kpi-val">¥ {val_wan}万</div>
                        <div class="out-kpi-sub2" style="margin-top:8px;">
                            <span>达成率</span>
                            <span style="font-weight:800; color:{prog_color}">{rate_txt}</span>
                        </div>
                        <div class="out-kpi-progress" style="margin-top:6px;">
                            <div class="out-kpi-progress-bar" style="background:{prog_color}; width:{pct_val}%;"></div>
                        </div>
                        <div class="out-kpi-sub2" style="margin-top:10px;">
                            <span>目标</span>
                            <span>{target_wan}万</span>
                        </div>
                        <div class="out-kpi-sub2">
                            <span>同期</span>
                            <span>{yoy_val_wan}万</span>
                        </div>
                        <div class="out-kpi-sub2">
                            <span>同比</span>
                            <span class="{trend_cls}">{arrow} {yoy_txt}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                # Card Renderer for Outbound/Scan (General Style)
                def _render_general_card(title, icon, main_val, sub_items):
                    # sub_items: list of (label, value_html)
                    rows_html = ""
                    for label, val_html in sub_items:
                        rows_html += f'<div class="out-kpi-sub2"><span>{label}</span><span>{val_html}</span></div>'
                    
                    st.markdown(f"""
                    <div class="out-kpi-card">
                        <div class="out-kpi-bar"></div>
                        <div class="out-kpi-head">
                            <div class="out-kpi-ico">{icon}</div>
                            <div class="out-kpi-title">{title}</div>
                        </div>
                        <div class="out-kpi-val">{main_val}</div>
                        <div style="margin-top:10px;">{rows_html}</div>
                    </div>
                    """, unsafe_allow_html=True)

                sel_bigcat = st.session_state.get("main_sel_cat", "全部")

                def _filter_common(_df):
                    if _df is None or getattr(_df, "empty", True):
                        return pd.DataFrame()
                    d = _df.copy()
                    for c in ['省区', '经销商名称', '产品大类', '大分类']:
                        if c in d.columns:
                            d[c] = d[c].fillna('').astype(str).str.strip()
                    if sel_prov != '全部' and '省区' in d.columns:
                        d = d[d['省区'] == sel_prov]
                    if sel_dist != '全部' and '经销商名称' in d.columns:
                        d = d[d['经销商名称'] == sel_dist]
                    if sel_bigcat != '全部':
                        if '产品大类' in d.columns:
                            d = d[d['产品大类'] == sel_bigcat]
                        elif '大分类' in d.columns:
                            d = d[d['大分类'] == sel_bigcat]
                    return d

                # ---------------------------------------------------------
                # 1. 核心业绩指标 (From Tab 7)
                # ---------------------------------------------------------
                st.markdown("### 🚀 核心业绩指标")
                df_perf = _filter_common(df_perf_raw)
                if not df_perf.empty:
                    # Data Prep
                    if '年份' in df_perf.columns:
                        df_perf['年份'] = pd.to_numeric(df_perf['年份'], errors='coerce').fillna(0).astype(int)
                    if '月份' in df_perf.columns:
                        df_perf['月份'] = pd.to_numeric(df_perf['月份'], errors='coerce').fillna(0).astype(int)
                    amt_col = '发货金额' if '发货金额' in df_perf.columns else None
                    if amt_col:
                        df_perf[amt_col] = pd.to_numeric(df_perf[amt_col], errors='coerce').fillna(0)
                    
                    years_avail = sorted([y for y in df_perf['年份'].unique().tolist() if y > 2000])
                    perf_y = max(years_avail) if years_avail else 2025
                    months_avail = sorted([m for m in df_perf[df_perf['年份'] == perf_y]['月份'].unique().tolist() if 1 <= m <= 12])
                    perf_m = max(months_avail) if months_avail else 1
                    last_y = perf_y - 1

                    # Actuals
                    cur_m_amt = df_perf[(df_perf['年份'] == perf_y) & (df_perf['月份'] == perf_m)][amt_col].sum() if amt_col else 0
                    last_m_amt = df_perf[(df_perf['年份'] == last_y) & (df_perf['月份'] == perf_m)][amt_col].sum() if amt_col else 0
                    cur_y_amt = df_perf[df_perf['年份'] == perf_y][amt_col].sum() if amt_col else 0
                    last_y_amt = df_perf[df_perf['年份'] == last_y][amt_col].sum() if amt_col else 0

                    yoy_m = (cur_m_amt - last_m_amt) / last_m_amt if last_m_amt > 0 else 0
                    yoy_y = (cur_y_amt - last_y_amt) / last_y_amt if last_y_amt > 0 else 0

                    # Targets
                    t_cur_m = 0.0
                    t_cur_y = 0.0
                    if df_target_raw is not None and not getattr(df_target_raw, "empty", True):
                        df_t = df_target_raw.copy()
                        for c in ['省区', '品类']:
                            if c in df_t.columns: df_t[c] = df_t[c].fillna('').astype(str).str.strip()
                        if '月份' in df_t.columns: df_t['月份'] = pd.to_numeric(df_t['月份'], errors='coerce').fillna(0).astype(int)
                        if '任务量' in df_t.columns: df_t['任务量'] = pd.to_numeric(df_t['任务量'], errors='coerce').fillna(0)
                        
                        if sel_prov != '全部' and '省区' in df_t.columns:
                            df_t = df_t[df_t['省区'] == sel_prov]
                        # Target usually doesn't filter by Distributor, but filters by Category
                        if sel_bigcat != '全部' and '品类' in df_t.columns:
                            df_t = df_t[df_t['品类'] == sel_bigcat]
                        
                        t_cur_m = df_t[df_t['月份'] == perf_m]['任务量'].sum()
                        t_cur_y = df_t['任务量'].sum() # Total Year Target

                    rate_m = (cur_m_amt / t_cur_m) if t_cur_m > 0 else None
                    rate_y = (cur_y_amt / t_cur_y) if t_cur_y > 0 else None

                    c1, c2 = st.columns(2)
                    with c1:
                        _render_perf_card(f"本月业绩（{perf_m}月）", "📅", _fmt_wan(cur_m_amt), _fmt_wan(t_cur_m), rate_m, _fmt_wan(last_m_amt), yoy_m)
                    with c2:
                        _render_perf_card(f"年度累计业绩（{perf_y}年）", "🏆", _fmt_wan(cur_y_amt), _fmt_wan(t_cur_y), rate_y, _fmt_wan(last_y_amt), yoy_y)
                else:
                    st.info("业绩数据为空或不含匹配字段")

                st.markdown("---")

                # ---------------------------------------------------------
                # 2. 库存关键指标 (From Tab 6)
                # ---------------------------------------------------------
                st.markdown("### 📦 库存关键指标")
                df_stock = _filter_common(df_stock_raw)
                if not df_stock.empty:
                    # Prepare Data for Metrics
                    stock_box_col = '箱数' if '箱数' in df_stock.columns else next((c for c in df_stock.columns if '箱' in str(c)), None)
                    stock_boxes = float(pd.to_numeric(df_stock[stock_box_col], errors='coerce').fillna(0).sum()) if stock_box_col else 0.0
                    
                    # Q4 Avg Sales (Need logic from Tab 6)
                    total_q4_avg = 0.0
                    if df_q4_raw is not None and not getattr(df_q4_raw, "empty", True):
                        # Simple estimation: Filter Q4 raw by current filters -> Sum Q4 months -> Divide by 3
                        # Tab 6 logic is more complex (Distributor based), but for Overview Total, simple sum is close enough.
                        # However, let's try to match Tab 6 logic: Sum 'Q4_Avg' of relevant distributors.
                        
                        # 1. Get filtered distributors
                        valid_dists = df_stock['经销商名称'].unique()
                        
                        # 2. Calculate Q4 Sales for these distributors
                        df_q4_f = df_q4_raw.copy()
                        if '年份' in df_q4_f.columns: df_q4_f = df_q4_f[df_q4_f['年份'] == 2025] # Q4 assumption
                        if '经销商名称' in df_q4_f.columns:
                            df_q4_f = df_q4_f[df_q4_f['经销商名称'].isin(valid_dists)]
                        
                        # Filter for Oct, Nov, Dec
                        if '月份' in df_q4_f.columns:
                            df_q4_f['月份'] = pd.to_numeric(df_q4_f['月份'], errors='coerce').fillna(0).astype(int)
                            df_q4_f = df_q4_f[df_q4_f['月份'].isin([10, 11, 12])]
                        
                        qty_col = '数量(箱)' if '数量(箱)' in df_q4_f.columns else next((c for c in df_q4_f.columns if '数量' in str(c)), None)
                        if qty_col:
                            total_q4_sales = pd.to_numeric(df_q4_f[qty_col], errors='coerce').sum()
                            total_q4_avg = total_q4_sales / 3.0

                    dos = stock_boxes / total_q4_avg if total_q4_avg > 0 else 0.0
                    
                    # Abnormal Count (Simplify for Overview)
                    # Tab 6 calculates per distributor. Here we just show global metrics.
                    
                    m1, m2, m3 = st.columns(3)
                    m1.metric("📦 总库存 (箱)", fmt_num(stock_boxes))
                    m2.metric("📉 Q4月均销", fmt_num(total_q4_avg))
                    m3.metric("📅 整体可销月 (DOS)", fmt_num(dos))
                else:
                    st.info("库存数据为空")

                st.markdown("---")

                # ---------------------------------------------------------
                # 3. 出库关键指标 (From Tab Out)
                # ---------------------------------------------------------
                st.markdown("### 🚚 出库关键指标")
                df_out = _filter_common(df_q4_raw)
                if not df_out.empty:
                    # Date Prep
                    tmp = df_out.copy()
                    for c in ['年份', '月份']: 
                        if c in tmp.columns: tmp[c] = pd.to_numeric(tmp[c], errors='coerce').fillna(0).astype(int)
                    if '日' in tmp.columns: tmp['日'] = pd.to_numeric(tmp['日'], errors='coerce').fillna(0).astype(int)
                    qty_col = '数量(箱)' if '数量(箱)' in tmp.columns else next((c for c in tmp.columns if '数量' in str(c) or '箱' in str(c)), None)
                    if qty_col:
                        tmp['数量(箱)'] = pd.to_numeric(tmp[qty_col], errors='coerce').fillna(0)
                        tmp = tmp[tmp['年份'] > 0]
                        
                        oy = int(tmp['年份'].max())
                        om = int(tmp[tmp['年份'] == oy]['月份'].max())
                        od = int(tmp[(tmp['年份'] == oy) & (tmp['月份'] == om)]['日'].max())
                        
                        # Current
                        today_boxes = tmp[(tmp['年份'] == oy) & (tmp['月份'] == om) & (tmp['日'] == od)]['数量(箱)'].sum()
                        month_boxes = tmp[(tmp['年份'] == oy) & (tmp['月份'] == om)]['数量(箱)'].sum()
                        year_boxes = tmp[tmp['年份'] == oy]['数量(箱)'].sum()
                        
                        # Last Year
                        ly = oy - 1
                        l_today_boxes = tmp[(tmp['年份'] == ly) & (tmp['月份'] == om) & (tmp['日'] == od)]['数量(箱)'].sum()
                        l_month_boxes = tmp[(tmp['年份'] == ly) & (tmp['月份'] == om)]['数量(箱)'].sum()
                        l_year_boxes = tmp[tmp['年份'] == ly]['数量(箱)'].sum()
                        
                        # YoY
                        yoy_d = (today_boxes - l_today_boxes) / l_today_boxes if l_today_boxes > 0 else 0
                        yoy_m = (month_boxes - l_month_boxes) / l_month_boxes if l_month_boxes > 0 else 0
                        yoy_y = (year_boxes - l_year_boxes) / l_year_boxes if l_year_boxes > 0 else 0
                        
                        k1, k2, k3 = st.columns(3)
                        with k1:
                            trend = _trend_cls(yoy_d)
                            arr = _arrow(yoy_d)
                            _render_general_card("本日出库", "🚚", f"{fmt_num(today_boxes)} 箱", [
                                ("同期", f"{fmt_num(l_today_boxes)} 箱"),
                                ("同比", f'<span class="{trend}">{arr} {_fmt_pct(yoy_d)}</span>')
                            ])
                        with k2:
                            trend = _trend_cls(yoy_m)
                            arr = _arrow(yoy_m)
                            _render_general_card(f"本月累计出库（{om}月）", "📦", f"{fmt_num(month_boxes)} 箱", [
                                ("同期", f"{fmt_num(l_month_boxes)} 箱"),
                                ("同比", f'<span class="{trend}">{arr} {_fmt_pct(yoy_m)}</span>')
                            ])
                        with k3:
                            trend = _trend_cls(yoy_y)
                            arr = _arrow(yoy_y)
                            _render_general_card(f"本年累计出库（{oy}年）", "🧾", f"{fmt_num(year_boxes)} 箱", [
                                ("同期", f"{fmt_num(l_year_boxes)} 箱"),
                                ("同比", f'<span class="{trend}">{arr} {_fmt_pct(yoy_y)}</span>')
                            ])
                else:
                    st.info("出库数据为空")

                st.markdown("---")

                # ---------------------------------------------------------
                # 4. 扫码率概览 (From Tab Scan)
                # ---------------------------------------------------------
                st.markdown("### 📱 扫码率概览")
                df_scan = _filter_common(df_scan_raw)
                # Re-use out_base from above or re-calc
                if not df_scan.empty and not df_out.empty:
                    # Ensure Date Cols
                    for c in ['年份', '月份', '日']:
                        if c in df_scan.columns: df_scan[c] = pd.to_numeric(df_scan[c], errors='coerce').fillna(0).astype(int)
                    
                    # Use same oy, om, od from Outbound
                    scan_today = len(df_scan[(df_scan['年份'] == oy) & (df_scan['月份'] == om) & (df_scan['日'] == od)]) / 6.0
                    scan_month = len(df_scan[(df_scan['年份'] == oy) & (df_scan['月份'] == om)]) / 6.0
                    scan_year = len(df_scan[df_scan['年份'] == oy]) / 6.0
                    
                    l_scan_today = len(df_scan[(df_scan['年份'] == ly) & (df_scan['月份'] == om) & (df_scan['日'] == od)]) / 6.0
                    l_scan_month = len(df_scan[(df_scan['年份'] == ly) & (df_scan['月份'] == om)]) / 6.0
                    l_scan_year = len(df_scan[df_scan['年份'] == ly]) / 6.0

                    rate_today = scan_today / today_boxes if today_boxes > 0 else 0
                    rate_month = scan_month / month_boxes if month_boxes > 0 else 0
                    rate_year = scan_year / year_boxes if year_boxes > 0 else 0
                    
                    yoy_rate_d = rate_today - (l_scan_today / l_today_boxes if l_today_boxes > 0 else 0)
                    yoy_rate_m = rate_month - (l_scan_month / l_month_boxes if l_month_boxes > 0 else 0)
                    yoy_rate_y = rate_year - (l_scan_year / l_year_boxes if l_year_boxes > 0 else 0)

                    s1, s2, s3 = st.columns(3)
                    with s1:
                        trend = _trend_cls(yoy_rate_d)
                        arr = _arrow(yoy_rate_d)
                        _render_general_card("本日扫码率", "📱", fmt_pct_ratio(rate_today), [
                            ("扫码 / 出库", f"{fmt_num(scan_today)} / {fmt_num(today_boxes)}"),
                            ("同比增减", f'<span class="{trend}">{arr} {fmt_pct_value(yoy_rate_d*100)}</span>')
                        ])
                    with s2:
                        trend = _trend_cls(yoy_rate_m)
                        arr = _arrow(yoy_rate_m)
                        _render_general_card("本月扫码率", "🗓️", fmt_pct_ratio(rate_month), [
                            ("扫码 / 出库", f"{fmt_num(scan_month)} / {fmt_num(month_boxes)}"),
                            ("同比增减", f'<span class="{trend}">{arr} {fmt_pct_value(yoy_rate_m*100)}</span>')
                        ])
                    with s3:
                        trend = _trend_cls(yoy_rate_y)
                        arr = _arrow(yoy_rate_y)
                        _render_general_card("本年扫码率", "📈", fmt_pct_ratio(rate_year), [
                            ("扫码 / 出库", f"{fmt_num(scan_year)} / {fmt_num(year_boxes)}"),
                            ("同比增减", f'<span class="{trend}">{arr} {fmt_pct_value(yoy_rate_y*100)}</span>')
                        ])
                else:
                    st.info("扫码数据为空")

            # === TAB SCAN: SCAN ANALYSIS ===
            if main_tab == "📱 扫码分析":
                if df_scan_raw is not None and not df_scan_raw.empty:
                    st.subheader("📱 扫码分析")
                    
                    # 1. Date Calculation
                    # Today: Max date in max month of 2026
                    max_scan_date = None
                    df_scan_2026 = df_scan_raw[df_scan_raw['年份'] == 2026]
                    if not df_scan_2026.empty:
                        max_month = df_scan_2026['月份'].max()
                        max_day = df_scan_2026[df_scan_2026['月份'] == max_month]['日'].max()
                        max_scan_date = pd.Timestamp(year=2026, month=max_month, day=max_day)
                    
                    if max_scan_date:
                        cur_year = max_scan_date.year
                        cur_month = max_scan_date.month
                        cur_day = max_scan_date.day
                        st.info(f"📅 当前统计日期：{cur_year}年{cur_month}月{cur_day}日")
                    else:
                        st.warning("⚠️ 未找到2026年扫码数据，无法计算当日/当月指标")
                        cur_year, cur_month, cur_day = 2026, 1, 1

                    # 2. Filter Area
                    with st.expander("🔎 扫码筛选", expanded=True):
                        c_s1, c_s2, c_s3 = st.columns(3)
                        # Province
                        prov_opts_s = ['全部'] + sorted(df_scan_raw['省区'].unique().tolist())
                        sel_prov_s = c_s1.selectbox("省区", prov_opts_s, key="scan_prov")
                        
                        # Distributor
                        if sel_prov_s != '全部':
                            dist_opts_s = ['全部'] + sorted(df_scan_raw[df_scan_raw['省区'] == sel_prov_s]['经销商名称'].unique().tolist())
                        else:
                            dist_opts_s = ['全部'] + sorted(df_scan_raw['经销商名称'].unique().tolist())
                        sel_dist_s = c_s2.selectbox("经销商", dist_opts_s, key="scan_dist")
                        
                        # Category
                        cat_opts_s = ['全部'] + sorted(df_scan_raw['产品大类'].unique().tolist())
                        sel_cat_s = c_s3.selectbox("产品大类", cat_opts_s, key="scan_cat")

                    # Apply Filters
                    df_s_flt = df_scan_raw.copy()
                    last_year = cur_year - 1
                    out_base_df = None
                    out_day_df = None
                    out_day_last_df = None
                    if df_q4_raw is not None and not getattr(df_q4_raw, "empty", True):
                        tmp = df_q4_raw.copy()
                        for c in ['年份', '月份']:
                            if c in tmp.columns:
                                tmp[c] = pd.to_numeric(tmp[c], errors='coerce').fillna(0).astype(int)
                        day_col_out = None
                        if '日' in tmp.columns:
                            day_col_out = '日'
                            tmp['日'] = pd.to_numeric(tmp['日'], errors='coerce').fillna(0).astype(int)
                        else:
                            cand = next((c for c in tmp.columns if '日期' in str(c)), None)
                            if cand:
                                dt = pd.to_datetime(tmp[cand], errors='coerce')
                                tmp['年份'] = dt.dt.year
                                tmp['月份'] = dt.dt.month
                                tmp['日'] = dt.dt.day
                                day_col_out = '日'
                        qty_col_out = '数量(箱)' if '数量(箱)' in tmp.columns else next((c for c in tmp.columns if '数量' in str(c) or '箱' in str(c)), None)
                        if qty_col_out:
                            tmp['数量(箱)'] = pd.to_numeric(tmp[qty_col_out], errors='coerce').fillna(0)
                            if all(k in tmp.columns for k in ['年份', '月份', '日']):
                                out_base_df = tmp.copy()
                                for c in ['省区', '经销商名称', '产品大类', '大分类']:
                                    if c in out_base_df.columns:
                                        out_base_df[c] = out_base_df[c].fillna('').astype(str).str.strip()

                    if sel_prov_s != '全部':
                        df_s_flt = df_s_flt[df_s_flt['省区'] == sel_prov_s]
                        if out_base_df is not None and '省区' in out_base_df.columns:
                            out_base_df = out_base_df[out_base_df['省区'] == sel_prov_s]
                    if sel_dist_s != '全部':
                        df_s_flt = df_s_flt[df_s_flt['经销商名称'] == sel_dist_s]
                        if out_base_df is not None and '经销商名称' in out_base_df.columns:
                            out_base_df = out_base_df[out_base_df['经销商名称'] == sel_dist_s]
                    if sel_cat_s != '全部':
                        df_s_flt = df_s_flt[df_s_flt['产品大类'] == sel_cat_s]
                        if out_base_df is not None:
                            if '产品大类' in out_base_df.columns:
                                out_base_df = out_base_df[out_base_df['产品大类'] == sel_cat_s]
                            elif '大分类' in out_base_df.columns:
                                out_base_df = out_base_df[out_base_df['大分类'] == sel_cat_s]

                    if out_base_df is not None:
                        out_day_df = out_base_df[(out_base_df['年份'] == cur_year) & (out_base_df['月份'] == cur_month) & (out_base_df['日'] == cur_day)].copy()
                        out_day_last_df = out_base_df[(out_base_df['年份'] == last_year) & (out_base_df['月份'] == cur_month) & (out_base_df['日'] == cur_day)].copy()

                    # 3. Calculate Metrics (Scan vs Outbound)
                    # Unit: Box (6 tins = 1 box)
                    # Scan Count (Rows) / 6
                    
                    # --- Current Period (2026) ---
                    # Day
                    scan_day = len(df_s_flt[(df_s_flt['年份'] == cur_year) & (df_s_flt['月份'] == cur_month) & (df_s_flt['日'] == cur_day)]) / 6.0
                    out_day = 0
                    if out_day_df is not None:
                        qty_col_out = '数量(箱)' if '数量(箱)' in out_day_df.columns else next((c for c in out_day_df.columns if '数量' in str(c) or '箱' in str(c)), None)
                        if qty_col_out:
                            out_day = float(pd.to_numeric(out_day_df[qty_col_out], errors='coerce').fillna(0).sum())
                    out_day_last = 0
                    if out_day_last_df is not None:
                        qty_col_out = '数量(箱)' if '数量(箱)' in out_day_last_df.columns else next((c for c in out_day_last_df.columns if '数量' in str(c) or '箱' in str(c)), None)
                        if qty_col_out:
                            out_day_last = float(pd.to_numeric(out_day_last_df[qty_col_out], errors='coerce').fillna(0).sum())
                    
                    # Month
                    scan_month = len(df_s_flt[(df_s_flt['年份'] == cur_year) & (df_s_flt['月份'] == cur_month)]) / 6.0
                    out_month = float(pd.to_numeric(out_base_df[(out_base_df['年份'] == cur_year) & (out_base_df['月份'] == cur_month)]['数量(箱)'], errors='coerce').fillna(0).sum()) if out_base_df is not None else 0.0
                    
                    # Year
                    scan_year = len(df_s_flt[df_s_flt['年份'] == cur_year]) / 6.0
                    out_year = float(pd.to_numeric(out_base_df[out_base_df['年份'] == cur_year]['数量(箱)'], errors='coerce').fillna(0).sum()) if out_base_df is not None else 0.0

                    # --- Same Period Last Year (2025) ---
                    scan_day_last = len(df_s_flt[(df_s_flt['年份'] == last_year) & (df_s_flt['月份'] == cur_month) & (df_s_flt['日'] == cur_day)]) / 6.0
                    
                    # Month
                    scan_month_last = len(df_s_flt[(df_s_flt['年份'] == last_year) & (df_s_flt['月份'] == cur_month)]) / 6.0
                    out_month_last = float(pd.to_numeric(out_base_df[(out_base_df['年份'] == last_year) & (out_base_df['月份'] == cur_month)]['数量(箱)'], errors='coerce').fillna(0).sum()) if out_base_df is not None else 0.0
                    
                    # Year (YTD? or Full Year? Usually YTD for comparison or Full Year 2025)
                    # "同期" usually means same period. For Year, it means 2025 Full Year or YTD.
                    # Let's use Full Year 2025 for now as 2026 is incomplete.
                    scan_year_last = len(df_s_flt[df_s_flt['年份'] == last_year]) / 6.0
                    out_year_last = float(pd.to_numeric(out_base_df[out_base_df['年份'] == last_year]['数量(箱)'], errors='coerce').fillna(0).sum()) if out_base_df is not None else 0.0

                    # Rates
                    rate_month = (scan_month / out_month) if out_month > 0 else 0
                    rate_month_last = (scan_month_last / out_month_last) if out_month_last > 0 else 0
                    rate_year = (scan_year / out_year) if out_year > 0 else 0
                    rate_year_last = (scan_year_last / out_year_last) if out_year_last > 0 else 0
                    rate_day = (scan_day / out_day) if out_day > 0 else 0
                    rate_day_last = (scan_day_last / out_day_last) if out_day_last > 0 else 0

                    tab_overview, tab_s_cat, tab_s_prov, tab_s_map = st.tabs(["📊 扫码率概览", "🧩 分品类扫码率", "🗺️ 省区扫码率", "🧭 地图热力"])

                    with tab_overview:
                        st.caption(f"口径：今日 {cur_year}年{cur_month}月{cur_day}日｜本月 {cur_month}月｜本年 {cur_year}年")

                        def _trend_cls(x):
                            if x is None or (isinstance(x, float) and pd.isna(x)):
                                return "trend-neutral"
                            return "trend-up" if x > 0 else ("trend-down" if x < 0 else "trend-neutral")

                        def _arrow(x):
                            if x is None or (isinstance(x, float) and pd.isna(x)):
                                return ""
                            return "↑" if x > 0 else ("↓" if x < 0 else "")

                        yoy_rate_day = (rate_day - rate_day_last) if out_day_last > 0 else None
                        yoy_rate_month = (rate_month - rate_month_last) if out_month_last > 0 else None
                        yoy_rate_year = (rate_year - rate_year_last) if out_year_last > 0 else None
                        yoy_rate_day_pct = (yoy_rate_day * 100.0) if yoy_rate_day is not None else None
                        yoy_rate_month_pct = (yoy_rate_month * 100.0) if yoy_rate_month is not None else None
                        yoy_rate_year_pct = (yoy_rate_year * 100.0) if yoy_rate_year is not None else None

                        c1, c2, c3 = st.columns(3)
                        with c1:
                            st.markdown(f"""
                            <div class="out-kpi-card">
                                <div class="out-kpi-bar"></div>
                                <div class="out-kpi-head">
                                    <div class="out-kpi-ico">📱</div>
                                    <div class="out-kpi-title">本日扫码率</div>
                                </div>
                                <div class="out-kpi-val">{fmt_pct_ratio(rate_day)}</div>
                                <div class="out-kpi-sub"><span>出库(箱)</span><span>{fmt_num(out_day)}</span></div>
                                <div class="out-kpi-sub"><span>扫码(箱)</span><span>{fmt_num(scan_day)}</span></div>
                                <div class="out-kpi-sub2" style="margin-top:10px;"><span>同期({last_year})</span><span>{fmt_num(out_day_last)} 箱 / {fmt_num(scan_day_last)} 箱</span></div>
                                <div class="out-kpi-sub2"><span>同比（扫码率）</span><span class="{_trend_cls(yoy_rate_day)}">{_arrow(yoy_rate_day)} {fmt_pct_value(yoy_rate_day_pct) if yoy_rate_day_pct is not None else "—"}</span></div>
                            </div>
                            """, unsafe_allow_html=True)
                        with c2:
                            st.markdown(f"""
                            <div class="out-kpi-card">
                                <div class="out-kpi-bar"></div>
                                <div class="out-kpi-head">
                                    <div class="out-kpi-ico">🗓️</div>
                                    <div class="out-kpi-title">本月扫码率</div>
                                </div>
                                <div class="out-kpi-val">{fmt_pct_ratio(rate_month)}</div>
                                <div class="out-kpi-sub"><span>出库(箱)</span><span>{fmt_num(out_month)}</span></div>
                                <div class="out-kpi-sub"><span>扫码(箱)</span><span>{fmt_num(scan_month)}</span></div>
                                <div class="out-kpi-sub2" style="margin-top:10px;"><span>同期({last_year})</span><span>{fmt_num(out_month_last)} 箱 / {fmt_num(scan_month_last)} 箱</span></div>
                                <div class="out-kpi-sub2"><span>同比（扫码率）</span><span class="{_trend_cls(yoy_rate_month)}">{_arrow(yoy_rate_month)} {fmt_pct_value(yoy_rate_month_pct) if yoy_rate_month_pct is not None else "—"}</span></div>
                            </div>
                            """, unsafe_allow_html=True)
                        with c3:
                            st.markdown(f"""
                            <div class="out-kpi-card">
                                <div class="out-kpi-bar"></div>
                                <div class="out-kpi-head">
                                    <div class="out-kpi-ico">📈</div>
                                    <div class="out-kpi-title">本年扫码率</div>
                                </div>
                                <div class="out-kpi-val">{fmt_pct_ratio(rate_year)}</div>
                                <div class="out-kpi-sub"><span>出库(箱)</span><span>{fmt_num(out_year)}</span></div>
                                <div class="out-kpi-sub"><span>扫码(箱)</span><span>{fmt_num(scan_year)}</span></div>
                                <div class="out-kpi-sub2" style="margin-top:10px;"><span>同期({last_year})</span><span>{fmt_num(out_year_last)} 箱 / {fmt_num(scan_year_last)} 箱</span></div>
                                <div class="out-kpi-sub2"><span>同比（扫码率）</span><span class="{_trend_cls(yoy_rate_year)}">{_arrow(yoy_rate_year)} {fmt_pct_value(yoy_rate_year_pct) if yoy_rate_year_pct is not None else "—"}</span></div>
                            </div>
                            """, unsafe_allow_html=True)
                    
                    # --- Sub-Tab 1: Category ---
                    with tab_s_cat:
                        # Group by Big Category
                        
                        # --- Day Level (Sync) ---
                        s_cat_day = df_s_flt[(df_s_flt['年份'] == cur_year) & (df_s_flt['月份'] == cur_month) & (df_s_flt['日'] == cur_day)].groupby('产品大类').size().reset_index(name='本日扫码听数')
                        s_cat_day['本日扫码(箱)'] = s_cat_day['本日扫码听数'] / 6.0
                        o_cat_day = None
                        if out_day_df is not None:
                            if '产品大类' in out_day_df.columns:
                                group_col = '产品大类'
                            elif '大分类' in out_day_df.columns:
                                group_col = '大分类'
                            else:
                                group_col = None
                            qty_col_out = '数量(箱)' if '数量(箱)' in (out_day_df.columns if out_day_df is not None else []) else next((c for c in out_day_df.columns if '数量' in str(c) or '箱' in str(c)), None) if out_day_df is not None else None
                            if group_col and qty_col_out:
                                o_cat_day = out_day_df.groupby(group_col)[qty_col_out].sum().reset_index().rename(columns={group_col: '产品大类', qty_col_out: '今日出库(箱)'})
                        
                        # --- Month Level (Sync) ---
                        s_cat_month = df_s_flt[(df_s_flt['年份'] == cur_year) & (df_s_flt['月份'] == cur_month)].groupby('产品大类').size().reset_index(name='本月扫码听数')
                        s_cat_month['本月扫码(箱)'] = s_cat_month['本月扫码听数'] / 6.0
                        
                        o_cat_month = pd.DataFrame(columns=['产品大类', '本月出库(箱)'])
                        if out_base_df is not None:
                            group_col_m = '产品大类' if '产品大类' in out_base_df.columns else ('大分类' if '大分类' in out_base_df.columns else None)
                            if group_col_m:
                                o_cat_month = out_base_df[(out_base_df['年份'] == cur_year) & (out_base_df['月份'] == cur_month)].groupby(group_col_m)['数量(箱)'].sum().reset_index()
                                o_cat_month = o_cat_month.rename(columns={group_col_m: '产品大类', '数量(箱)': '本月出库(箱)'})

                        # --- Year Level (Sync) ---
                        s_cat_year = df_s_flt[df_s_flt['年份'] == cur_year].groupby('产品大类').size().reset_index(name='本年扫码听数')
                        s_cat_year['本年扫码(箱)'] = s_cat_year['本年扫码听数'] / 6.0
                        
                        o_cat_year = pd.DataFrame(columns=['产品大类', '本年出库(箱)'])
                        if out_base_df is not None:
                            group_col_y = '产品大类' if '产品大类' in out_base_df.columns else ('大分类' if '大分类' in out_base_df.columns else None)
                            if group_col_y:
                                o_cat_year = out_base_df[out_base_df['年份'] == cur_year].groupby(group_col_y)['数量(箱)'].sum().reset_index()
                                o_cat_year = o_cat_year.rename(columns={group_col_y: '产品大类', '数量(箱)': '本年出库(箱)'})
                            
                        # Merge All
                        cat_final = pd.merge(s_cat_day[['产品大类', '本日扫码(箱)']], s_cat_month[['产品大类', '本月扫码(箱)']], on='产品大类', how='outer')
                        if o_cat_day is not None:
                            cat_final = pd.merge(cat_final, o_cat_day, on='产品大类', how='outer')
                        cat_final = pd.merge(cat_final, o_cat_month, on='产品大类', how='outer')
                        cat_final = pd.merge(cat_final, s_cat_year[['产品大类', '本年扫码(箱)']], on='产品大类', how='outer')
                        cat_final = pd.merge(cat_final, o_cat_year, on='产品大类', how='outer').fillna(0)
                        
                        # Calculate Rates
                        # Day Rate: Outbound usually monthly, so Day Rate might not be accurate unless assumed uniform or N/A
                        # User requirement: "本日、本月的维度，也需要加到分品类和分省区". 
                        # Let's show Day Scan Qty. Day Rate is tricky without Day Outbound. We will show Day Scan Qty only or N/A for Rate.
                        # Month Rate
                        cat_final['本月扫码率'] = cat_final.apply(lambda x: x['本月扫码(箱)'] / x['本月出库(箱)'] if x['本月出库(箱)'] > 0 else 0, axis=1)
                        # Year Rate
                        cat_final['本年扫码率'] = cat_final.apply(lambda x: x['本年扫码(箱)'] / x['本年出库(箱)'] if x['本年出库(箱)'] > 0 else 0, axis=1)
                        # Day Rate
                        if '今日出库(箱)' in cat_final.columns:
                            cat_final['本日扫码率'] = cat_final.apply(lambda x: x['本日扫码(箱)'] / x['今日出库(箱)'] if x['今日出库(箱)'] > 0 else 0, axis=1)
                        else:
                            cat_final['今日出库(箱)'] = 0.0
                            cat_final['本日扫码率'] = 0.0
                        
                        cat_final = cat_final.sort_values('本月扫码(箱)', ascending=False)
                        
                        # Format for display
                        # Display
                        cat_disp = cat_final[['产品大类', '今日出库(箱)', '本日扫码(箱)', '本日扫码率', '本月出库(箱)', '本月扫码(箱)', '本月扫码率', '本年出库(箱)', '本年扫码(箱)', '本年扫码率']].copy()
                        cat_disp = cat_disp.rename(columns={'本日扫码(箱)': '今日扫码(箱)'})
                        cat_column_defs = [
                            {"headerName": "产品大类", "field": "产品大类", "pinned": "left", "minWidth": 120},
                            {"headerName": f"今日（{cur_month}月{cur_day}日）", "children": [
                                {"headerName": "出库(箱)", "field": "今日出库(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码(箱)", "field": "今日扫码(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码率", "field": "本日扫码率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO},
                            ]},
                            {"headerName": f"本月（{cur_month}月）", "children": [
                                {"headerName": "出库(箱)", "field": "本月出库(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码(箱)", "field": "本月扫码(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码率", "field": "本月扫码率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO},
                            ]},
                            {"headerName": f"本年（{cur_year}年）", "children": [
                                {"headerName": "出库(箱)", "field": "本年出库(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码(箱)", "field": "本年扫码(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码率", "field": "本年扫码率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO},
                            ]},
                        ]

                        show_aggrid_table(cat_disp, key="scan_cat_ag", column_defs=cat_column_defs)

                    # --- Sub-Tab 2: Province ---
                    with tab_s_prov:
                        # --- Day Level ---
                        s_prov_day = df_s_flt[(df_s_flt['年份'] == cur_year) & (df_s_flt['月份'] == cur_month) & (df_s_flt['日'] == cur_day)].groupby('省区').size().reset_index(name='本日扫码听数')
                        s_prov_day['本日扫码(箱)'] = s_prov_day['本日扫码听数'] / 6.0
                        o_prov_day = None
                        if out_day_df is not None:
                            o_prov_day = out_day_df.groupby('省区')['数量(箱)'].sum().reset_index().rename(columns={'数量(箱)': '今日出库(箱)'})

                        # --- Month Level (Current) ---
                        s_prov_cur = df_s_flt[(df_s_flt['年份'] == cur_year) & (df_s_flt['月份'] == cur_month)].groupby('省区').size().reset_index(name='扫码听数')
                        s_prov_cur['扫码箱数'] = s_prov_cur['扫码听数'] / 6.0
                        o_prov_cur = pd.DataFrame(columns=['省区', '本月出库(箱)'])
                        if out_base_df is not None:
                            o_prov_cur = out_base_df[(out_base_df['年份'] == cur_year) & (out_base_df['月份'] == cur_month)].groupby('省区')['数量(箱)'].sum().reset_index().rename(columns={'数量(箱)': '本月出库(箱)'})
                        prov_cur = pd.merge(s_prov_cur[['省区', '扫码箱数']], o_prov_cur, on='省区', how='outer').fillna(0)
                        prov_cur['本月扫码(箱)'] = prov_cur['扫码箱数']
                        prov_cur['本月扫码率'] = prov_cur.apply(lambda x: x['本月扫码(箱)'] / x['本月出库(箱)'] if x['本月出库(箱)'] > 0 else 0, axis=1)
                        prov_cur = prov_cur[['省区', '本月出库(箱)', '本月扫码(箱)', '本月扫码率']]

                        # --- Same Period Last Year (Month) ---
                        s_prov_last = df_s_flt[(df_s_flt['年份'] == last_year) & (df_s_flt['月份'] == cur_month)].groupby('省区').size().reset_index(name='扫码听数')
                        s_prov_last['扫码箱数'] = s_prov_last['扫码听数'] / 6.0
                        o_prov_last = pd.DataFrame(columns=['省区', '同期出库(箱)'])
                        if out_base_df is not None:
                            o_prov_last = out_base_df[(out_base_df['年份'] == last_year) & (out_base_df['月份'] == cur_month)].groupby('省区')['数量(箱)'].sum().reset_index().rename(columns={'数量(箱)': '同期出库(箱)'})
                        prov_last = pd.merge(s_prov_last[['省区', '扫码箱数']], o_prov_last, on='省区', how='outer').fillna(0)
                        prov_last['同期扫码(箱)'] = prov_last['扫码箱数']
                        prov_last['同期扫码率'] = prov_last.apply(lambda x: x['同期扫码(箱)'] / x['同期出库(箱)'] if x['同期出库(箱)'] > 0 else 0, axis=1)
                        prov_last = prov_last[['省区', '同期出库(箱)', '同期扫码(箱)', '同期扫码率']]

                        # --- Ring Period (Month) ---
                        if cur_month == 1:
                            ring_year = cur_year - 1
                            ring_month = 12
                        else:
                            ring_year = cur_year
                            ring_month = cur_month - 1

                        s_prov_ring = df_s_flt[(df_s_flt['年份'] == ring_year) & (df_s_flt['月份'] == ring_month)].groupby('省区').size().reset_index(name='扫码听数')
                        s_prov_ring['扫码箱数'] = s_prov_ring['扫码听数'] / 6.0
                        o_prov_ring = pd.DataFrame(columns=['省区', '环比出库(箱)'])
                        if out_base_df is not None:
                            o_prov_ring = out_base_df[(out_base_df['年份'] == ring_year) & (out_base_df['月份'] == ring_month)].groupby('省区')['数量(箱)'].sum().reset_index().rename(columns={'数量(箱)': '环比出库(箱)'})
                        prov_ring = pd.merge(s_prov_ring[['省区', '扫码箱数']], o_prov_ring, on='省区', how='outer').fillna(0)
                        prov_ring['环比扫码(箱)'] = prov_ring['扫码箱数']
                        prov_ring['环比扫码率'] = prov_ring.apply(lambda x: x['环比扫码(箱)'] / x['环比出库(箱)'] if x['环比出库(箱)'] > 0 else 0, axis=1)
                        prov_ring = prov_ring[['省区', '环比扫码率']]

                        # Merge All
                        prov_final = pd.merge(prov_cur, s_prov_day[['省区', '本日扫码(箱)']], on='省区', how='outer')
                        if o_prov_day is not None:
                            prov_final = pd.merge(prov_final, o_prov_day, on='省区', how='outer')
                        prov_final = pd.merge(prov_final, prov_last[['省区', '同期出库(箱)', '同期扫码(箱)', '同期扫码率']], on='省区', how='outer')
                        prov_final = pd.merge(prov_final, prov_ring[['省区', '环比扫码率']], on='省区', how='left').fillna(0)
                        prov_final['环比增长'] = prov_final['本月扫码率'] - prov_final['环比扫码率']
                        if '今日出库(箱)' not in prov_final.columns:
                            prov_final['今日出库(箱)'] = 0.0
                        prov_final['本日扫码率'] = prov_final.apply(lambda x: x['本日扫码(箱)'] / x['今日出库(箱)'] if x.get('今日出库(箱)', 0) > 0 else 0, axis=1)

                        prov_disp = prov_final[['省区', '本日扫码(箱)', '今日出库(箱)', '本日扫码率', '本月出库(箱)', '本月扫码(箱)', '本月扫码率', '同期出库(箱)', '同期扫码(箱)', '同期扫码率', '环比扫码率', '环比增长']].copy()
                        prov_disp = prov_disp.sort_values('本月扫码(箱)', ascending=False)
                        prov_disp = prov_disp.rename(columns={'本日扫码(箱)': '今日扫码(箱)'})
                        prov_column_defs = [
                            {"headerName": "省区", "field": "省区", "pinned": "left", "minWidth": 110},
                            {"headerName": f"今日（{cur_month}月{cur_day}日）", "children": [
                                {"headerName": "出库(箱)", "field": "今日出库(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码(箱)", "field": "今日扫码(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码率", "field": "本日扫码率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO},
                            ]},
                            {"headerName": f"本月（{cur_month}月）", "children": [
                                {"headerName": "出库(箱)", "field": "本月出库(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码(箱)", "field": "本月扫码(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码率", "field": "本月扫码率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO},
                            ]},
                            {"headerName": f"同期（{last_year}年{cur_month}月）", "children": [
                                {"headerName": "出库(箱)", "field": "同期出库(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码(箱)", "field": "同期扫码(箱)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM},
                                {"headerName": "扫码率", "field": "同期扫码率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO},
                            ]},
                            {"headerName": "环比", "children": [
                                {"headerName": "扫码率", "field": "环比扫码率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO},
                                {"headerName": "增长", "field": "环比增长", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO, "cellStyle": JS_COLOR_CONDITIONAL},
                            ]},
                        ]

                        show_aggrid_table(prov_disp, key="scan_prov_ag", column_defs=prov_column_defs)

                    with tab_s_map:
                        if ("经度" not in df_s_flt.columns) or ("纬度" not in df_s_flt.columns):
                            st.info("未检测到经纬度列：请确认扫码数据Sheet的M列为经纬度（形如 116.4,39.9 或 39.9,116.4）。")
                        else:
                            c_map1, c_map2, c_map3 = st.columns([1.1, 1.1, 1.2])
                            metric_mode = c_map1.radio("对比口径", ["扫码数", "扫码率"], horizontal=True, key="scan_map_metric_mode")
                            period_mode = c_map2.radio("时间范围", ["今日", "本月", "本年"], horizontal=True, key="scan_map_period")
                            style_mode = c_map3.radio("地图样式", ["详细", "简洁"], horizontal=True, key="scan_map_style_mode")

                            c_map4, c_map5 = st.columns([1.3, 1.0])
                            prov_opts_map = ["全国"] + sorted([p for p in df_s_flt["省区"].unique().tolist() if str(p).strip() != ""])
                            focus_prov = c_map4.selectbox("省区聚焦", prov_opts_map, key="scan_map_focus_prov")
                            palette_mode = c_map5.radio("配色", ["高对比", "色盲友好"], horizontal=True, key="scan_map_palette")

                            c_map6, c_map7 = st.columns([1.2, 1.1])
                            basemap_provider = c_map6.selectbox("底图来源", ["高德(国内)", "OpenStreetMap(外网)", "无底图(离线)", "自定义瓦片(内网/自建)"], key="scan_map_basemap_provider")
                            custom_tile_url = ""
                            if basemap_provider == "自定义瓦片(内网/自建)":
                                custom_tile_url = c_map7.text_input("瓦片URL模板", value="http://127.0.0.1:8080/{z}/{x}/{y}.png", key="scan_map_custom_tile_url")
                            else:
                                c_map7.write("")

                            show_cb_key = "scan_map_show_colorbar"
                            if show_cb_key not in st.session_state:
                                st.session_state[show_cb_key] = False
                            cb_label = "显示颜色刻度" if not st.session_state[show_cb_key] else "隐藏颜色刻度"
                            if st.button(cb_label, key="scan_map_toggle_colorbar"):
                                st.session_state[show_cb_key] = not bool(st.session_state[show_cb_key])
                                st.rerun()

                            df_map = df_s_flt.copy()
                            if period_mode == "今日":
                                df_map = df_map[(df_map["年份"] == cur_year) & (df_map["月份"] == cur_month) & (df_map["日"] == cur_day)]
                            elif period_mode == "本月":
                                df_map = df_map[(df_map["年份"] == cur_year) & (df_map["月份"] == cur_month)]
                            else:
                                df_map = df_map[df_map["年份"] == cur_year]

                            if focus_prov != "全国":
                                df_map = df_map[df_map["省区"] == focus_prov]

                            df_map = df_map.dropna(subset=["经度", "纬度"])
                            df_map = df_map[df_map["经度"].between(70, 140) & df_map["纬度"].between(0, 60)]

                            if df_map.empty:
                                st.info("当前筛选与口径下没有可用的经纬度数据。")
                            else:
                                center_lat = float(df_map["纬度"].mean())
                                center_lon = float(df_map["经度"].mean())
                                default_zoom = 3.1 if focus_prov == "全国" else 4.9
                                min_zoom, max_zoom = 2.2, 10.5

                                zoom_key = "scan_map_zoom"
                                if zoom_key not in st.session_state:
                                    st.session_state[zoom_key] = default_zoom
                                if st.session_state[zoom_key] < min_zoom or st.session_state[zoom_key] > max_zoom:
                                    st.session_state[zoom_key] = default_zoom

                                zc1, zc2, zc3, zc4 = st.columns([0.13, 0.13, 0.18, 0.56])
                                if zc1.button("＋", key="scan_map_zoom_in"):
                                    st.session_state[zoom_key] = min(max_zoom, float(st.session_state[zoom_key]) + 0.6)
                                    st.rerun()
                                if zc2.button("－", key="scan_map_zoom_out"):
                                    st.session_state[zoom_key] = max(min_zoom, float(st.session_state[zoom_key]) - 0.6)
                                    st.rerun()
                                if zc3.button("复位", key="scan_map_zoom_reset"):
                                    st.session_state[zoom_key] = default_zoom
                                    st.rerun()
                                zc4.slider("缩放", min_value=min_zoom, max_value=max_zoom, value=float(st.session_state[zoom_key]), step=0.1, key=zoom_key)

                                basemap_layers = None
                                if basemap_provider == "OpenStreetMap(外网)":
                                    map_style = "carto-positron" if style_mode == "简洁" else "open-street-map"
                                elif basemap_provider == "高德(国内)":
                                    map_style = "white-bg"
                                    gaode_style = "7" if style_mode == "详细" else "8"
                                    gaode_url = f"https://webrd02.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style={gaode_style}&x={{x}}&y={{y}}&z={{z}}"
                                    basemap_layers = [{"sourcetype": "raster", "source": [gaode_url], "below": "traces"}]
                                elif basemap_provider == "自定义瓦片(内网/自建)":
                                    map_style = "white-bg"
                                    _u = (custom_tile_url or "").strip()
                                    if _u:
                                        basemap_layers = [{"sourcetype": "raster", "source": [_u], "below": "traces"}]
                                else:
                                    map_style = "white-bg"
                                marker_opacity = 0.86
                                color_scale_count = "Turbo" if palette_mode == "高对比" else "Cividis"
                                color_scale_rate = "Viridis" if palette_mode == "高对比" else "Cividis"
                                point_scale = [
                                    [0.0, "#00C853"],
                                    [0.35, "#00C853"],
                                    [0.65, "#FFEB3B"],
                                    [0.82, "#FF9800"],
                                    [1.0, "#F44336"],
                                ]

                                if metric_mode == "扫码数":
                                    c_u1, c_u2 = st.columns([0.55, 0.45])
                                    unit_mode = c_u1.radio("单位", ["听", "箱"], horizontal=True, key="scan_map_unit")
                                    render_mode = c_u2.radio("渲染方式", ["热力", "标点"], horizontal=True, key="scan_map_render_mode")
                                    precision = st.slider("坐标聚合精度(小数位)", 0, 3, 2, key="scan_map_precision")
                                    df_grid = df_map[["经度", "纬度"]].copy()
                                    df_grid["经度"] = df_grid["经度"].round(int(precision))
                                    df_grid["纬度"] = df_grid["纬度"].round(int(precision))
                                    df_grid = df_grid.groupby(["经度", "纬度"]).size().reset_index(name="扫码听数")
                                    df_grid["扫码箱数"] = df_grid["扫码听数"] / 6.0
                                    val_col = "扫码听数" if unit_mode == "听" else "扫码箱数"

                                    if render_mode == "热力":
                                        fig = px.density_mapbox(
                                            df_grid,
                                            lat="纬度",
                                            lon="经度",
                                            z=val_col,
                                            radius=18 if focus_prov == "全国" else 14,
                                            zoom=float(st.session_state[zoom_key]),
                                            center={"lat": center_lat, "lon": center_lon},
                                            color_continuous_scale=color_scale_count,
                                            hover_data={"扫码听数": ":,.0f", "扫码箱数": ":,.2f"}
                                        )
                                        fig.update_traces(opacity=0.82)
                                    else:
                                        fig = px.scatter_mapbox(
                                            df_grid,
                                            lat="纬度",
                                            lon="经度",
                                            color=val_col,
                                            size=val_col,
                                            size_max=26,
                                            zoom=float(st.session_state[zoom_key]),
                                            center={"lat": center_lat, "lon": center_lon},
                                            color_continuous_scale=point_scale,
                                            hover_data={"扫码听数": ":,.0f", "扫码箱数": ":,.2f"}
                                        )
                                        fig.update_traces(marker={"opacity": marker_opacity})

                                    _layout_kwargs = {
                                        "mapbox_style": map_style,
                                        "margin": {"r": 0, "t": 0, "l": 0, "b": 0},
                                        "transition": {"duration": 260, "easing": "cubic-in-out"},
                                    }
                                    if basemap_layers is not None:
                                        _layout_kwargs["mapbox_layers"] = basemap_layers
                                    fig.update_layout(**_layout_kwargs)
                                    show_cb = bool(st.session_state.get(show_cb_key, False))
                                    cb_style = {"thickness": 10, "len": 0.55, "x": 1.0, "xpad": 0, "y": 0.5, "bgcolor": "rgba(255,255,255,0.25)", "outlinewidth": 0, "title": {"text": ""}}
                                    for _ax_name in [k for k in fig.layout if str(k).startswith("coloraxis")]:
                                        try:
                                            fig.layout[_ax_name].showscale = show_cb
                                        except Exception:
                                            pass
                                        if show_cb:
                                            try:
                                                fig.layout[_ax_name].colorbar = cb_style
                                            except Exception:
                                                pass
                                    for _t in fig.data:
                                        try:
                                            _t.update(showscale=show_cb)
                                        except Exception:
                                            pass
                                    st.plotly_chart(
                                        fig,
                                        use_container_width=True,
                                        config={"scrollZoom": True, "displayModeBar": True, "displaylogo": False, "responsive": True}
                                    )
                                else:
                                    render_mode_rate = st.radio("渲染方式", ["热力", "标点"], horizontal=True, key="scan_map_render_mode_rate")
                                    scan_by_prov = df_map.groupby("省区").size().reset_index(name="扫码听数")
                                    scan_by_prov["扫码箱数"] = scan_by_prov["扫码听数"] / 6.0
                                    cent = df_map.groupby("省区")[["经度", "纬度"]].mean().reset_index()
                                    prov_map = pd.merge(scan_by_prov, cent, on="省区", how="left")
                                    prov_map["出库(箱)"] = 0.0

                                    if out_base_df is not None and not getattr(out_base_df, "empty", True) and ("省区" in out_base_df.columns) and ("数量(箱)" in out_base_df.columns):
                                        out_map = out_base_df.copy()
                                        if period_mode == "今日":
                                            out_map = out_map[(out_map["年份"] == cur_year) & (out_map["月份"] == cur_month) & (out_map["日"] == cur_day)]
                                        elif period_mode == "本月":
                                            out_map = out_map[(out_map["年份"] == cur_year) & (out_map["月份"] == cur_month)]
                                        else:
                                            out_map = out_map[out_map["年份"] == cur_year]
                                        out_prov = out_map.groupby("省区")["数量(箱)"].sum().reset_index().rename(columns={"数量(箱)": "出库(箱)"})
                                        prov_map = pd.merge(prov_map.drop(columns=["出库(箱)"], errors="ignore"), out_prov, on="省区", how="left")
                                        prov_map["出库(箱)"] = pd.to_numeric(prov_map.get("出库(箱)"), errors="coerce").fillna(0.0)

                                    prov_map["扫码率"] = prov_map.apply(lambda x: x["扫码箱数"] / x["出库(箱)"] if x["出库(箱)"] > 0 else None, axis=1)
                                    prov_map = prov_map.dropna(subset=["经度", "纬度"])

                                    if render_mode_rate == "热力":
                                        fig = px.density_mapbox(
                                            prov_map.dropna(subset=["扫码率"]),
                                            lat="纬度",
                                            lon="经度",
                                            z="扫码率",
                                            radius=36 if focus_prov == "全国" else 24,
                                            zoom=float(st.session_state[zoom_key]),
                                            center={"lat": center_lat, "lon": center_lon},
                                            color_continuous_scale=color_scale_rate,
                                            hover_data={"省区": True, "扫码听数": ":,.0f", "扫码箱数": ":,.2f", "出库(箱)": ":,.0f", "扫码率": ":.2%"}
                                        )
                                        fig.update_traces(opacity=0.82)
                                    else:
                                        fig = px.scatter_mapbox(
                                            prov_map,
                                            lat="纬度",
                                            lon="经度",
                                            color="扫码率",
                                            size="扫码箱数",
                                            size_max=42,
                                            zoom=float(st.session_state[zoom_key]),
                                            center={"lat": center_lat, "lon": center_lon},
                                            color_continuous_scale=point_scale,
                                            hover_name="省区",
                                            hover_data={"扫码听数": ":,.0f", "扫码箱数": ":,.2f", "出库(箱)": ":,.0f", "扫码率": ":.2%"}
                                        )
                                        fig.update_traces(marker={"opacity": marker_opacity})
                                    _layout_kwargs = {
                                        "mapbox_style": map_style,
                                        "margin": {"r": 0, "t": 0, "l": 0, "b": 0},
                                        "transition": {"duration": 260, "easing": "cubic-in-out"},
                                    }
                                    if basemap_layers is not None:
                                        _layout_kwargs["mapbox_layers"] = basemap_layers
                                    fig.update_layout(**_layout_kwargs)
                                    show_cb = bool(st.session_state.get(show_cb_key, False))
                                    cb_style = {"thickness": 10, "len": 0.55, "x": 1.0, "xpad": 0, "y": 0.5, "bgcolor": "rgba(255,255,255,0.25)", "outlinewidth": 0, "title": {"text": ""}}
                                    for _ax_name in [k for k in fig.layout if str(k).startswith("coloraxis")]:
                                        try:
                                            fig.layout[_ax_name].showscale = show_cb
                                        except Exception:
                                            pass
                                        if show_cb:
                                            try:
                                                fig.layout[_ax_name].colorbar = cb_style
                                            except Exception:
                                                pass
                                    for _t in fig.data:
                                        try:
                                            _t.update(showscale=show_cb)
                                        except Exception:
                                            pass
                                    st.plotly_chart(
                                        fig,
                                        use_container_width=True,
                                        config={"scrollZoom": True, "displayModeBar": True, "displaylogo": False, "responsive": True}
                                    )

                else:
                    st.info("请在Excel中包含第6个Sheet（扫码数据）以查看此分析。")

            # === TAB 3: ABCD ANALYSIS ===
            if False and main_tab == "📈 ABCD效能分析":
                st.subheader("📊 Q3 vs Q4 门店效能对比分析")
                
                # Check for Q3/Q4 columns
                q3_cols = [c for c in month_cols if c in ['7月', '8月', '9月']]
                q4_cols = [c for c in month_cols if c in ['10月', '11月', '12月']]
                
                if not q3_cols or not q4_cols:
                    st.warning("⚠️ 数据源缺失7-12月的完整数据，无法进行Q3 vs Q4对比分析")
                else:
                    # Logic
                    # Calculate Q3 Class
                    df['Q3_Sum'] = df[q3_cols].sum(axis=1)
                    df['Q3_Avg'] = df['Q3_Sum'] / 3
                    
                    # Calculate Q4 Class
                    df['Q4_Sum'] = df[q4_cols].sum(axis=1)
                    df['Q4_Avg'] = df['Q4_Sum'] / 3
                    
                    def classify_score(x):
                        if x >= 4: return 'A'
                        elif 2 <= x < 4: return 'B'
                        elif 1 <= x < 2: return 'C'
                        else: return 'D'
                        
                    df['Class_Q3'] = df['Q3_Avg'].apply(classify_score)
                    df['Class_Q4'] = df['Q4_Avg'].apply(classify_score)
                    
                    # Comparison Metrics
                    q3_counts = df['Class_Q3'].value_counts().sort_index()
                    q4_counts = df['Class_Q4'].value_counts().sort_index()
                    
                    # Overview Cards
                    c1, c2, c3, c4 = st.columns(4)
                    
                    def render_metric(col, cls_label):
                        curr = q4_counts.get(cls_label, 0)
                        prev = q3_counts.get(cls_label, 0)
                        delta = curr - prev
                        col.metric(f"{cls_label}类门店 (Q4)", fmt_num(curr), f"{fmt_num(delta)} (环比)")
                        
                    render_metric(c1, 'A')
                    render_metric(c2, 'B')
                    render_metric(c3, 'C')
                    render_metric(c4, 'D')
                    
                    st.markdown("---")
                    
                    # Province Comparison Chart
                    st.subheader("🗺️ 各省区ABCD类门店数量对比 (Q3 vs Q4)")
                    
                    # Prepare Data for Chart
                    # Group by Province and Class for Q3
                    prov_q3 = df.groupby(['省区', 'Class_Q3']).size().reset_index(name='Count')
                    prov_q3['Period'] = 'Q3'
                    prov_q3.rename(columns={'Class_Q3': 'Class'}, inplace=True)
                    
                    # Group by Province and Class for Q4
                    prov_q4 = df.groupby(['省区', 'Class_Q4']).size().reset_index(name='Count')
                    prov_q4['Period'] = 'Q4'
                    prov_q4.rename(columns={'Class_Q4': 'Class'}, inplace=True)
                    
                    # Combine
                    prov_comp = pd.concat([prov_q3, prov_q4])
                    
                    # Interactive Selection
                    sel_period = st.radio("选择展示周期:", ["Q4 (本期)", "Q3 (上期)"], horizontal=True)
                    target_period = 'Q4' if 'Q4' in sel_period else 'Q3'
                    
                    chart_data = prov_comp[prov_comp['Period'] == target_period]
                    
                    fig_bar_prov_class = px.bar(chart_data, x='省区', y='Count', color='Class',
                                               title=f"各省区门店等级分布 ({target_period})",
                                               category_orders={"Class": ["A", "B", "C", "D"]},
                                               color_discrete_map={'A':'#FFC400', 'B':'#6A3AD0', 'C':'#B79BFF', 'D':'#8A8AA3'},
                                               text='Count')
                    fig_bar_prov_class.update_traces(textposition='inside', texttemplate='%{y:,.1~f}', hovertemplate='省区: %{x}<br>数量: %{y:,.1~f}<extra></extra>')
                    fig_bar_prov_class.update_layout(yaxis_title="门店数量", xaxis_title="省区", yaxis=dict(tickformat=",.1~f"), paper_bgcolor='rgba(255,255,255,0.25)', plot_bgcolor='rgba(255,255,255,0.25)')
                    st.plotly_chart(fig_bar_prov_class, use_container_width=True)
                    
                    st.markdown("---")
                    
                    # Migration Matrix
                    st.subheader("🔄 门店等级变动明细")
                    
                    # Define Change Type
                    def get_change_type(row):
                        order = {'A': 4, 'B': 3, 'C': 2, 'D': 1}
                        score_q3 = order[row['Class_Q3']]
                        score_q4 = order[row['Class_Q4']]
                        
                        if score_q3 == score_q4: return '持平'
                        elif score_q4 > score_q3: return '升级 ⬆️'
                        else: return '降级 ⬇️'
                        
                    df['变动类型'] = df.apply(get_change_type, axis=1)
                    
                    # Summary of Changes
                    change_counts = df['变动类型'].value_counts()
                    st.info(f"📊 变动概览: 升级 {fmt_num(change_counts.get('升级 ⬆️', 0), na='')} 家 | 降级 {fmt_num(change_counts.get('降级 ⬇️', 0), na='')} 家 | 持平 {fmt_num(change_counts.get('持平', 0), na='')} 家")
                    
                    # Detailed Table
                    # Filters
                    c_f1, c_f2, c_f3 = st.columns(3)
                    filter_prov = c_f1.selectbox("筛选省区", ['全部'] + list(df['省区'].unique()), key='abcd_prov')
                    
                    # Distributor Filter (Dependent on Province)
                    if filter_prov != '全部':
                        dist_opts = ['全部'] + sorted(list(df[df['省区'] == filter_prov]['经销商名称'].unique()))
                    else:
                        dist_opts = ['全部'] + sorted(list(df['经销商名称'].unique()))
                    filter_dist = c_f2.selectbox("筛选经销商", dist_opts, key='abcd_dist')
                    
                    filter_change = c_f3.selectbox("筛选变动类型", ['全部', '升级 ⬆️', '降级 ⬇️', '持平'], key='abcd_change')
                    
                    view_df = df.copy()
                    if filter_prov != '全部':
                        view_df = view_df[view_df['省区'] == filter_prov]
                    if filter_dist != '全部':
                        view_df = view_df[view_df['经销商名称'] == filter_dist]
                    if filter_change != '全部':
                        view_df = view_df[view_df['变动类型'] == filter_change]
                        
                    show_aggrid_table(view_df[['省区', '经销商名称', '门店名称', 'Class_Q3', 'Class_Q4', '变动类型', 'Q3_Avg', 'Q4_Avg']])

            # --- Tab 6: Inventory Analysis ---
            if main_tab == "📦 库存分析":
                if df_stock_raw is None:
                    st.warning("⚠️ 未检测到库存数据 (Sheet2)。请确保上传的 Excel 文件包含第二个 Sheet 页，且格式正确。")
                    st.info("数据格式要求：\nSheet2 需包含 A-L 列，顺序为：经销商编码、经销商名称、产品编码、产品名称、库存数量、箱数、省区名称、客户简称、产品大类、产品小类、重量、规格。")
                else:
                    st.caption(f"🕒 数据更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    if "省区名称" not in df_stock_raw.columns and "省区" in df_stock_raw.columns:
                        df_stock_raw["省区名称"] = df_stock_raw["省区"]
                    
                    with st.expander("🛠️ 库存筛选", expanded=False):
                        # Prepare filter lists
                        _prov_col = "省区名称" if "省区名称" in df_stock_raw.columns else ("省区" if "省区" in df_stock_raw.columns else None)
                        _dist_col = "经销商名称" if "经销商名称" in df_stock_raw.columns else None
                        _cat_col = "产品大类" if "产品大类" in df_stock_raw.columns else None
                        stock_provs = ['全部'] + (sorted(list(df_stock_raw[_prov_col].dropna().unique())) if _prov_col else [])
                        stock_dists = ['全部'] + (sorted(list(df_stock_raw[_dist_col].dropna().unique())) if _dist_col else [])
                        stock_cats = ['全部'] + (sorted(list(df_stock_raw[_cat_col].dropna().unique())) if _cat_col else [])
                        
                        # Helper to reset drill status
                        def reset_inv_drill():
                            st.session_state.drill_level = 1
                            st.session_state.selected_prov = None
                            st.session_state.selected_dist = None

                        # --- Subcategory Logic Adjustment ---
                        # User requirement: "Subcategory" dropdown should include 'Segment' and 'Ya Series'.
                        # When 'Segment' is selected, Specific Category options are ['1段', '2段', '3段'].
                        # When 'Ya Series' is selected, Specific Category options are ['雅赋', '雅耀', '雅舒', '雅护'].
                        
                        # 1. Base Subcategories
                        base_subcats = sorted(list(df_stock_raw['产品小类'].dropna().unique()))
                        # 2. Add Virtual Subcategories (Ensure uniqueness)
                        virtual_subcats = ['分段', '雅系列']
                        stock_subcats = ['全部'] + virtual_subcats + [s for s in base_subcats if s not in virtual_subcats]
                        
                        c1, c2, c3, c4, c5 = st.columns(5)
                        with c1: s_prov = st.selectbox("省区名称", stock_provs, key='stock_s_prov', on_change=reset_inv_drill)
                        with c2: 
                            if s_prov != '全部':
                                valid_dists = df_stock_raw[df_stock_raw['省区名称'] == s_prov]['经销商名称'].unique()
                                s_dist_opts = ['全部'] + sorted(list(valid_dists))
                            else:
                                s_dist_opts = stock_dists
                            s_dist = st.selectbox("经销商名称", s_dist_opts, key='stock_s_dist', on_change=reset_inv_drill)
                            
                        with c3: s_cat = st.selectbox("产品大类", stock_cats, key='stock_s_cat', on_change=reset_inv_drill)
                        
                        with c4: 
                            # Dynamic filter for subcat based on cat
                            # If we are using virtual subcats, we might want to show them regardless of Category?
                            # Or only if the Category allows? Assuming '美思雅段粉' allows them.
                            if s_cat != '全部':
                                valid_sub = df_stock_raw[df_stock_raw['产品大类'] == s_cat]['产品小类'].unique()
                                # Mix in virtuals if they make sense (assuming they are always available for filtering)
                                current_sub_opts = ['全部'] + virtual_subcats + sorted([s for s in valid_sub if s not in virtual_subcats])
                                s_sub_opts = current_sub_opts
                            else:
                                s_sub_opts = stock_subcats
                            if 'stock_s_sub' in st.session_state:
                                st.session_state.pop('stock_s_sub', None)
                            s_sub_selected = st.multiselect("产品小类(可多选)", s_sub_opts, default=['全部'], key='stock_s_sub_ms', on_change=reset_inv_drill)
                        
                        with c5:
                            # --- Dynamic Specific Category Options based on Subcategory Selection ---
                            if '分段' in s_sub_selected and '雅系列' in s_sub_selected:
                                stock_specs = ['1段', '2段', '3段', '雅赋', '雅耀', '雅舒', '雅护']
                            elif '分段' in s_sub_selected:
                                stock_specs = ['1段', '2段', '3段']
                            elif '雅系列' in s_sub_selected:
                                stock_specs = ['雅赋', '雅耀', '雅舒', '雅护']
                            else:
                                raw_specs = df_stock_raw['具体分类'].dropna().unique()
                                spec_opts = set(raw_specs)
                                stock_specs = sorted(list(spec_opts))
                                
                            s_spec = st.multiselect("具体分类 (支持多选)", stock_specs, default=[], placeholder="选择具体分类...", on_change=reset_inv_drill)
                        
                        # Apply Filters
                        df_s_filtered = df_stock_raw.copy()
                        if s_prov != '全部': df_s_filtered = df_s_filtered[df_s_filtered['省区名称'] == s_prov]
                        if s_dist != '全部': df_s_filtered = df_s_filtered[df_s_filtered['经销商名称'] == s_dist]
                        if s_cat != '全部': df_s_filtered = df_s_filtered[df_s_filtered['产品大类'] == s_cat]
                        
                        # --- Subcategory Filter Logic ---
                        if s_sub_selected and ('全部' not in s_sub_selected):
                            mask_sub = pd.Series(False, index=df_s_filtered.index)
                            if '分段' in s_sub_selected:
                                mask_sub = mask_sub | (
                                    (df_s_filtered['产品大类'].astype(str) == '美思雅段粉')
                                    & (df_s_filtered['具体分类'].fillna('').astype(str).isin(['1段', '2段', '3段']))
                                )
                            if '雅系列' in s_sub_selected:
                                mask_sub = mask_sub | (
                                    df_s_filtered['具体分类'].fillna('').astype(str).isin(['雅赋', '雅耀', '雅舒', '雅护'])
                                )
                            normal_subs = [x for x in s_sub_selected if x not in ['分段', '雅系列', '全部']]
                            if normal_subs:
                                mask_sub = mask_sub | df_s_filtered['产品小类'].astype(str).isin([str(x) for x in normal_subs])
                            df_s_filtered = df_s_filtered[mask_sub]
                        
                        # Apply Specific Category Filter
                        if s_spec:
                            def match_spec(row_val):
                                row_val = str(row_val)
                                for sel in s_spec:
                                    if sel in row_val: return True
                                return False
                            
                            mask = df_s_filtered['具体分类'].apply(match_spec)
                            df_s_filtered = df_s_filtered[mask]

                    st.markdown("### 导出库存（按省区ZIP）")
                    st.caption("导出范围：全部经销商；按省区拆分，每省一个Excel；表内按经销商与产品信息排序。产品筛选沿用当前选择。")

                    if "stock_zip_cache" not in st.session_state:
                        st.session_state.stock_zip_cache = {}
                    _stock_zip_cache = st.session_state.stock_zip_cache

                    _stock_sig_n = int(df_stock_raw.shape[0]) if df_stock_raw is not None else 0
                    _stock_sig_sum = 0.0
                    try:
                        if df_stock_raw is not None and "箱数" in df_stock_raw.columns:
                            _stock_sig_sum = float(pd.to_numeric(df_stock_raw["箱数"], errors="coerce").fillna(0.0).sum())
                    except Exception:
                        _stock_sig_sum = 0.0

                    _sub_key = tuple(sorted([str(x) for x in (s_sub_selected or [])]))
                    _spec_key = tuple(sorted([str(x) for x in (s_spec or [])]))
                    k_stock_zip = ("stock_zip_by_prov", str(s_cat), _sub_key, _spec_key, _stock_sig_n, round(_stock_sig_sum, 4))
                    k_stock_all = ("stock_all_excel", str(s_cat), _sub_key, _spec_key, _stock_sig_n, round(_stock_sig_sum, 4))

                    c_z1, c_z2, c_a1, c_a2, _ = st.columns([1.8, 2.2, 1.8, 2.2, 3.0])
                    with c_z1:
                        if st.button("生成各省区库存ZIP", key="stock_zip_gen"):
                            with st.spinner("正在生成各省区库存ZIP，请稍候…"):
                                df_all = df_stock_raw.copy()
                                if s_cat != "全部" and "产品大类" in df_all.columns:
                                    df_all = df_all[df_all["产品大类"] == s_cat]

                                if s_sub_selected and ("全部" not in s_sub_selected):
                                    mask_sub = pd.Series(False, index=df_all.index)
                                    if "分段" in s_sub_selected:
                                        if "产品大类" in df_all.columns and "具体分类" in df_all.columns:
                                            mask_sub = mask_sub | (
                                                (df_all["产品大类"].astype(str) == "美思雅段粉")
                                                & (df_all["具体分类"].fillna("").astype(str).isin(["1段", "2段", "3段"]))
                                            )
                                    if "雅系列" in s_sub_selected:
                                        if "具体分类" in df_all.columns:
                                            mask_sub = mask_sub | (df_all["具体分类"].fillna("").astype(str).isin(["雅赋", "雅耀", "雅舒", "雅护"]))
                                    normal_subs = [x for x in s_sub_selected if x not in ["分段", "雅系列", "全部"]]
                                    if normal_subs and "产品小类" in df_all.columns:
                                        mask_sub = mask_sub | df_all["产品小类"].astype(str).isin([str(x) for x in normal_subs])
                                    df_all = df_all[mask_sub]

                                if s_spec and "具体分类" in df_all.columns:
                                    def _match_spec(v):
                                        vv = str(v)
                                        for sel in s_spec:
                                            if str(sel) in vv:
                                                return True
                                        return False
                                    df_all = df_all[df_all["具体分类"].apply(_match_spec)]

                                if "省区名称" not in df_all.columns and "省区" in df_all.columns:
                                    df_all["省区名称"] = df_all["省区"]
                                prov_col = "省区名称" if "省区名称" in df_all.columns else ("省区" if "省区" in df_all.columns else None)

                                cols_order = [
                                    "省区名称",
                                    "省区",
                                    "经销商名称",
                                    "经销商全称",
                                    "经销商编码",
                                    "产品大类",
                                    "产品小类",
                                    "产品名称",
                                    "产品编码",
                                    "批次号",
                                    "库存数量(听/盒)",
                                    "箱数",
                                ]
                                cols_use = [c for c in cols_order if c in df_all.columns]
                                if cols_use:
                                    df_all = df_all[cols_use].copy()

                                sort_cols_all = [c for c in ["省区名称", "省区", "经销商名称", "产品大类", "产品小类", "产品名称", "产品编码", "批次号"] if c in df_all.columns]
                                if sort_cols_all:
                                    df_all = df_all.sort_values(sort_cols_all, kind="stable").reset_index(drop=True)

                                filter_parts = []
                                if s_cat != "全部":
                                    filter_parts.append(f"产品大类={s_cat}")
                                if s_sub_selected and ("全部" not in s_sub_selected):
                                    filter_parts.append(f"产品小类={','.join([str(x) for x in s_sub_selected])}")
                                if s_spec:
                                    filter_parts.append(f"具体分类={','.join([str(x) for x in s_spec])}")
                                filter_line = "筛选：" + ("；".join(filter_parts) if filter_parts else "无")

                                buf = io.BytesIO()
                                with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                                    provs = []
                                    if prov_col is not None:
                                        provs = (
                                            df_all[prov_col]
                                            .dropna()
                                            .astype(str)
                                            .str.strip()
                                            .tolist()
                                        )
                                    provs = sorted([p for p in set(provs) if p and p.lower() not in ("nan", "none", "null")])
                                    for p in provs:
                                        if prov_col is None:
                                            continue
                                        df_p = df_all[df_all[prov_col].astype(str).str.strip() == str(p).strip()].copy()
                                        if df_p.empty:
                                            continue
                                        sort_cols = [c for c in ["经销商名称", "产品大类", "产品小类", "产品名称", "产品编码", "批次号"] if c in df_p.columns]
                                        if sort_cols:
                                            df_p = df_p.sort_values(sort_cols, kind="stable").reset_index(drop=True)

                                        title_lines_p = [
                                            "库存明细 - 经销商库存",
                                            f"省区：{p}",
                                            filter_line,
                                            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                        ]
                                        number_headers_p = set([c for c in ["库存数量(听/盒)", "箱数"] if c in df_p.columns])
                                        number_formats_p = {}
                                        if "库存数量(听/盒)" in df_p.columns:
                                            number_formats_p["库存数量(听/盒)"] = "0"
                                        if "箱数" in df_p.columns:
                                            number_formats_p["箱数"] = "0.0"

                                        xls_p = _df_to_excel_bytes(
                                            df_p,
                                            sheet_name="库存明细",
                                            title_lines=title_lines_p,
                                            number_headers=number_headers_p,
                                            number_formats=number_formats_p,
                                            group_headers=False,
                                        )
                                        zf.writestr(f"{sanitize_filename(p)}.xlsx", xls_p)
                                buf.seek(0)
                                _stock_zip_cache[k_stock_zip] = {
                                    "bytes": buf.getvalue(),
                                    "name": sanitize_filename("库存明细_各省区.zip"),
                                }
                    with c_z2:
                        if k_stock_zip in _stock_zip_cache:
                            st.download_button(
                                "下载各省区库存ZIP",
                                data=_stock_zip_cache[k_stock_zip]["bytes"],
                                file_name=_stock_zip_cache[k_stock_zip]["name"],
                                mime="application/zip",
                                key="stock_zip_dl",
                            )
                    with c_a1:
                        if st.button("生成全部库存Excel", key="stock_all_gen"):
                            with st.spinner("正在生成全部库存Excel，请稍候…"):
                                df_all = df_stock_raw.copy()
                                if s_cat != "全部" and "产品大类" in df_all.columns:
                                    df_all = df_all[df_all["产品大类"] == s_cat]

                                if s_sub_selected and ("全部" not in s_sub_selected):
                                    mask_sub = pd.Series(False, index=df_all.index)
                                    if "分段" in s_sub_selected:
                                        if "产品大类" in df_all.columns and "具体分类" in df_all.columns:
                                            mask_sub = mask_sub | (
                                                (df_all["产品大类"].astype(str) == "美思雅段粉")
                                                & (df_all["具体分类"].fillna("").astype(str).isin(["1段", "2段", "3段"]))
                                            )
                                    if "雅系列" in s_sub_selected:
                                        if "具体分类" in df_all.columns:
                                            mask_sub = mask_sub | (df_all["具体分类"].fillna("").astype(str).isin(["雅赋", "雅耀", "雅舒", "雅护"]))
                                    normal_subs = [x for x in s_sub_selected if x not in ["分段", "雅系列", "全部"]]
                                    if normal_subs and "产品小类" in df_all.columns:
                                        mask_sub = mask_sub | df_all["产品小类"].astype(str).isin([str(x) for x in normal_subs])
                                    df_all = df_all[mask_sub]

                                if s_spec and "具体分类" in df_all.columns:
                                    def _match_spec(v):
                                        vv = str(v)
                                        for sel in s_spec:
                                            if str(sel) in vv:
                                                return True
                                        return False
                                    df_all = df_all[df_all["具体分类"].apply(_match_spec)]

                                if "省区名称" not in df_all.columns and "省区" in df_all.columns:
                                    df_all["省区名称"] = df_all["省区"]

                                cols_order = [
                                    "省区名称",
                                    "省区",
                                    "经销商名称",
                                    "经销商全称",
                                    "经销商编码",
                                    "产品大类",
                                    "产品小类",
                                    "产品名称",
                                    "产品编码",
                                    "批次号",
                                    "库存数量(听/盒)",
                                    "箱数",
                                ]
                                cols_use = [c for c in cols_order if c in df_all.columns]
                                if cols_use:
                                    df_all = df_all[cols_use].copy()

                                sort_cols_all = [c for c in ["省区名称", "省区", "经销商名称", "产品大类", "产品小类", "产品名称", "产品编码", "批次号"] if c in df_all.columns]
                                if sort_cols_all:
                                    df_all = df_all.sort_values(sort_cols_all, kind="stable").reset_index(drop=True)

                                filter_parts = []
                                if s_cat != "全部":
                                    filter_parts.append(f"产品大类={s_cat}")
                                if s_sub_selected and ("全部" not in s_sub_selected):
                                    filter_parts.append(f"产品小类={','.join([str(x) for x in s_sub_selected])}")
                                if s_spec:
                                    filter_parts.append(f"具体分类={','.join([str(x) for x in s_spec])}")
                                filter_line = "筛选：" + ("；".join(filter_parts) if filter_parts else "无")

                                title_lines_all = [
                                    "库存明细 - 全部省区（经销商库存）",
                                    filter_line,
                                    f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                ]
                                number_headers_all = set([c for c in ["库存数量(听/盒)", "箱数"] if c in df_all.columns])
                                number_formats_all = {}
                                if "库存数量(听/盒)" in df_all.columns:
                                    number_formats_all["库存数量(听/盒)"] = "0"
                                if "箱数" in df_all.columns:
                                    number_formats_all["箱数"] = "0.0"

                                xls_all = _df_to_excel_bytes(
                                    df_all,
                                    sheet_name="库存明细",
                                    title_lines=title_lines_all,
                                    number_headers=number_headers_all,
                                    number_formats=number_formats_all,
                                    group_headers=False,
                                )
                                _stock_zip_cache[k_stock_all] = {
                                    "bytes": xls_all,
                                    "name": sanitize_filename("库存明细_全部省区.xlsx"),
                                }
                    with c_a2:
                        if k_stock_all in _stock_zip_cache:
                            st.download_button(
                                "下载全部库存Excel",
                                data=_stock_zip_cache[k_stock_all]["bytes"],
                                file_name=_stock_zip_cache[k_stock_all]["name"],
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="stock_all_dl",
                            )
                    
                    st.markdown("---")

                    outbound_pivot = pd.DataFrame()
                    df_o_filtered = pd.DataFrame()
                    sales_agg_q4 = pd.DataFrame(columns=['经销商名称', 'Q4_Total', 'Q4_Avg'])

                    with st.expander("🚚 出库筛选", expanded=False):
                        if df_q4_raw is None or df_q4_raw.empty:
                            st.warning("⚠️ 未检测到出库底表数据 (Sheet3)。")
                        else:
                            o_raw = df_q4_raw.copy()
                            required_out_cols = ['省区', '经销商名称', '数量(箱)', '月份']
                            missing_out = [c for c in required_out_cols if c not in o_raw.columns]

                            if missing_out:
                                st.warning(f"⚠️ 出库底表缺失字段：{', '.join(missing_out)}")
                            else:
                                if '产品大类' not in o_raw.columns:
                                    o_raw['产品大类'] = '全部'
                                if '产品小类' not in o_raw.columns:
                                    o_raw['产品小类'] = '全部'
                                else:
                                    o_raw['产品小类'] = o_raw['产品小类'].astype(str).str.strip()
                                    o_raw.loc[o_raw['产品小类'].isin(['', 'nan', 'None', 'NULL', 'NaN']), '产品小类'] = pd.NA

                            out_provs = ['全部'] + sorted(o_raw['省区'].dropna().astype(str).unique().tolist())
                            out_dists_all = ['全部'] + sorted(o_raw['经销商名称'].dropna().astype(str).unique().tolist())
                            out_cats = ['全部'] + sorted(o_raw['产品大类'].dropna().astype(str).unique().tolist())
                            out_subs_clean = o_raw['产品小类'].dropna().astype(str).str.strip()
                            out_subs_clean = out_subs_clean[out_subs_clean != '']
                            out_subs_list = sorted(out_subs_clean.unique().tolist())
                            out_subs = ['全部'] + out_subs_list
                            empty_sub_cnt = int(o_raw['产品小类'].isna().sum()) if '产品小类' in o_raw.columns else 0
                            dup_sub_cnt = int(out_subs_clean.shape[0] - out_subs_clean.nunique())
                            if empty_sub_cnt > 0:
                                st.warning(f"⚠️ Sheet3 的M列(产品小类)存在空值：{empty_sub_cnt} 行")
                            if dup_sub_cnt > 0:
                                st.info(f"ℹ️ Sheet3 的M列(产品小类)存在重复值：{dup_sub_cnt} 行（下拉已自动去重）")
                            out_month_opts = list(range(1, 13))

                            oc1, oc2, oc3, oc4, oc5 = st.columns(5)
                            with oc1:
                                o_prov = st.selectbox("省区", out_provs, key='out_s_prov')
                            with oc2:
                                if o_prov != '全部':
                                    dists_in_prov = o_raw[o_raw['省区'].astype(str) == str(o_prov)]['经销商名称'].dropna().astype(str).unique().tolist()
                                    out_dists = ['全部'] + sorted(dists_in_prov)
                                else:
                                    out_dists = out_dists_all
                                o_dist = st.selectbox("经销商", out_dists, key='out_s_dist')
                            with oc3:
                                o_cat = st.selectbox("产品大类", out_cats, key='out_s_cat')
                            with oc4:
                                if o_cat != '全部':
                                    subs_in_cat = o_raw[o_raw['产品大类'].astype(str) == str(o_cat)]['产品小类'].dropna().astype(str).unique().tolist()
                                    out_subs2 = ['全部'] + sorted(subs_in_cat)
                                else:
                                    out_subs2 = out_subs
                                if 'out_s_sub' in st.session_state:
                                    st.session_state.pop('out_s_sub', None)
                                o_sub_selected = st.multiselect("产品小类(可多选)", out_subs2, default=['全部'], key='out_s_sub_ms')
                            with oc5:
                                o_months = st.multiselect("时间（月）", out_month_opts, default=[10, 11, 12], key='out_s_months')

                            df_o_filtered = o_raw.copy()
                            
                            # Filter for Year 2025 (as per Q4 definition)
                            if '年份' in df_o_filtered.columns:
                                df_o_filtered = df_o_filtered[df_o_filtered['年份'] == 2025]
                                
                            if o_prov != '全部':
                                df_o_filtered = df_o_filtered[df_o_filtered['省区'].astype(str) == str(o_prov)]
                            if o_dist != '全部':
                                df_o_filtered = df_o_filtered[df_o_filtered['经销商名称'].astype(str) == str(o_dist)]
                            if o_cat != '全部':
                                df_o_filtered = df_o_filtered[df_o_filtered['产品大类'].astype(str) == str(o_cat)]
                            if o_sub_selected and ('全部' not in o_sub_selected):
                                df_o_filtered = df_o_filtered[df_o_filtered['产品小类'].astype(str).isin([str(x) for x in o_sub_selected])]

                            def _to_month(v):
                                if pd.isna(v):
                                    return None
                                if isinstance(v, (int, float)) and not pd.isna(v):
                                    m = int(v)
                                    return m if 1 <= m <= 12 else None
                                s = str(v).strip()
                                if s.isdigit():
                                    m = int(s)
                                    return m if 1 <= m <= 12 else None
                                if '月' in s:
                                    digits = ''.join([ch for ch in s if ch.isdigit()])
                                    if digits:
                                        for k in (2, 1):
                                            if len(digits) >= k:
                                                m = int(digits[-k:])
                                                if 1 <= m <= 12:
                                                    return m
                                    return None
                                dt = pd.to_datetime(s, errors='coerce')
                                if pd.isna(dt):
                                    return None
                                m = int(dt.month)
                                return m if 1 <= m <= 12 else None

                            df_o_filtered['月'] = df_o_filtered['月份'].apply(_to_month)
                            df_o_filtered = df_o_filtered[df_o_filtered['月'].notna()].copy()
                            df_o_filtered['月'] = df_o_filtered['月'].astype(int)

                            if o_months:
                                df_o_filtered = df_o_filtered[df_o_filtered['月'].isin(o_months)].copy()

                            df_o_filtered['月列'] = df_o_filtered['月'].astype(str) + '月'

                            idx_cols = ['省区', '经销商名称', '产品大类', '产品小类']
                            outbound_pivot = (
                                df_o_filtered
                                .pivot_table(index=idx_cols, columns='月列', values='数量(箱)', aggfunc='sum', fill_value=0)
                                .reset_index()
                            )

                            month_cols_full = [f"{i}月" for i in range(1, 13)]
                            for mc in month_cols_full:
                                if mc not in outbound_pivot.columns:
                                    outbound_pivot[mc] = 0

                            outbound_pivot['Q4月均销'] = (outbound_pivot['10月'] + outbound_pivot['11月'] + outbound_pivot['12月']) / 3
                            outbound_pivot = outbound_pivot[idx_cols + month_cols_full + ['Q4月均销']]

                            with st.expander("📄 出库分析底表（Sheet3）", expanded=False):
                                show_aggrid_table(outbound_pivot, height=520, key="outbound_pivot_table")

                            if not outbound_pivot.empty:
                                dist_q4 = outbound_pivot.groupby('经销商名称')[['10月', '11月', '12月']].sum().reset_index()
                                dist_q4['Q4_Total'] = dist_q4['10月'] + dist_q4['11月'] + dist_q4['12月']
                                dist_q4['Q4_Avg'] = dist_q4['Q4_Total'] / 3
                                sales_agg_q4 = dist_q4[['经销商名称', 'Q4_Total', 'Q4_Avg']].copy()

                            out_xlsx = io.BytesIO()
                            try:
                                with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
                                    outbound_pivot.to_excel(writer, index=False, sheet_name='Sheet3')
                            except Exception:
                                with pd.ExcelWriter(out_xlsx, engine='xlsxwriter') as writer:
                                    outbound_pivot.to_excel(writer, index=False, sheet_name='Sheet3')
                                st.download_button(
                                    "📥 下载出库分析底表 (Excel)",
                                    data=out_xlsx.getvalue(),
                                    file_name="出库分析底表_Sheet3.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                )
                    
                    # --- Drill-down State Management ---
                    # (Initialized at top of script)
                    
                    # Threshold Config
                    with st.expander("⚙️ 阈值配置", expanded=False):
                        c_th1, c_th2 = st.columns(2)
                        high_th = c_th1.number_input("库存过高阈值 (DOS >)", value=2.0, step=0.1)
                        low_th = c_th2.number_input("库存过低阈值 (DOS <)", value=0.5, step=0.1)

                    # Logic:
                    # 1. Sum Stock '箱数' by Distributor (from filtered stock df_s_filtered)
                    # 2. Match with Sheet3 Sales 'Q4_Avg' by Distributor
                    
                    # Note: df_s_filtered '经销商名称' is now '客户简称' (H column) due to load_data mapping
                    stock_agg = df_s_filtered.groupby(['省区名称', '经销商名称'])['箱数'].sum().reset_index()
                    stock_agg.rename(columns={'箱数': '当前库存_箱'}, inplace=True)
                    stock_agg['经销商名称'] = stock_agg['经销商名称'].astype(str).str.strip()
                    
                    # Merge with Q4 sales data from Sheet3
                    # LEFT JOIN ensures we only keep distributors present in the STOCK file (filtered by top filters)
                    # However, if we filter by province in top filter, df_s_filtered only has that province.
                    # sales_agg_q4 has ALL distributors from Sheet3.
                    # Merging them attaches Sales info to the Stock info.
                    analysis_df = pd.merge(stock_agg, sales_agg_q4[['经销商名称', 'Q4_Avg']], on='经销商名称', how='left')
                    analysis_df['Q4_Avg'] = analysis_df['Q4_Avg'].fillna(0)
                    
                    # 3. Calc DOS & Status
                    analysis_df['近三月未出库'] = (analysis_df['Q4_Avg'] <= 0) & (analysis_df['当前库存_箱'] > 0)

                    # Calculate DOS
                    # Optimized: Vectorized
                    q4_avg_series = analysis_df['Q4_Avg']
                    stock_series = analysis_df['当前库存_箱']
                    mask_no_outbound = analysis_df.get('近三月未出库', pd.Series(False, index=analysis_df.index)).astype(bool)
                    
                    analysis_df['可销月(DOS)'] = np.where(
                        mask_no_outbound, np.nan,
                        np.where(
                            q4_avg_series <= 0, 0.0,
                            (stock_series / q4_avg_series)
                        )
                    )
                    
                    # Ensure thresholds are defined before use
                    if 'high_th' not in locals(): high_th = 2.0
                    if 'low_th' not in locals(): low_th = 0.5

                    # Optimized: Vectorized select
                    # Pre-calculate boolean mask for '近三月未出库'
                    mask_no_outbound = analysis_df.get('近三月未出库', pd.Series(False, index=analysis_df.index)).astype(bool)
                    
                    dos_series = analysis_df['可销月(DOS)']
                    
                    conditions = [
                        mask_no_outbound,
                        pd.isna(dos_series),
                        dos_series > high_th,
                        dos_series < low_th
                    ]
                    choices = [
                        '⚫ 近三月未出库',
                        '🟢 正常',
                        '🔴 库存过高',
                        '🟠 库存不足'
                    ]
                    analysis_df['库存状态'] = np.select(conditions, choices, default='🟢 正常')

                    # --- OVERVIEW METRICS (Moved Back & Enhanced) ---
                    # Calculate metrics based on the CURRENT context (filtered data analysis_df)
                    # If drill level is 1 (All Provs), it shows total.
                    # If drill level is 2 (One Prov), we should filter analysis_df to that prov for metrics?
                    # Or should metrics always reflect the TOP filters (df_s_filtered)?
                    # User request: "When I select a specific province (in filter), real-time update."
                    # df_s_filtered IS filtered by the top dropdowns. analysis_df is derived from it.
                    # So calculating from analysis_df is correct for the top filters.
                    # However, if user clicks "Drill Down" to level 2, should the metrics update to that province?
                    # User said "When I select to specific province". 
                    # If the user uses the *Sidebar/Top Filter*, df_s_filtered updates, so analysis_df updates.
                    # If the user uses *Drill Down*, st.session_state.selected_prov is set.
                    # Usually Overview Metrics reflect the *Global Context* of the current view.
                    # Let's support both: If Drill Level > 1, filter metrics to selected scope.
                    
                    metrics_df = analysis_df.copy()
                    if st.session_state.drill_level == 2 and st.session_state.selected_prov:
                        metrics_df = metrics_df[metrics_df['省区名称'] == st.session_state.selected_prov]
                    elif st.session_state.drill_level == 3 and st.session_state.selected_dist:
                        # For level 3, it's single distributor
                         metrics_df = metrics_df[metrics_df['经销商名称'] == st.session_state.selected_dist]

                    # Calc Metrics
                    total_stock_show = metrics_df['当前库存_箱'].sum()
                    if sales_agg_q4 is not None and not sales_agg_q4.empty and 'Q4_Total' in sales_agg_q4.columns:
                        dist_scope = (
                            metrics_df['经销商名称']
                            .dropna()
                            .astype(str)
                            .str.strip()
                            .unique()
                            .tolist()
                        )
                        sales_scope = sales_agg_q4[sales_agg_q4['经销商名称'].isin(dist_scope)].copy()
                        total_q4_avg_show = float(sales_scope['Q4_Total'].sum()) / 3 if not sales_scope.empty else 0.0
                    else:
                        total_q4_avg_show = 0.0
                    
                    # DOS = Total Stock / Total Sales
                    if total_q4_avg_show > 0:
                        dos_show = total_stock_show / total_q4_avg_show
                    else:
                        dos_show = 0.0
                    
                    if metrics_df is None or metrics_df.empty or '库存状态' not in metrics_df.columns:
                        abnormal_count_show = 0
                    else:
                        abnormal_count_show = int(
                            metrics_df['库存状态']
                            .fillna('')
                            .astype(str)
                            .str.contains('🔴|🟠|⚫', na=False)
                            .sum()
                        )
                    
                    st.markdown("### 📊 关键指标概览")
                    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
                    col_m1.metric("📦 总库存 (箱)", fmt_num(total_stock_show))
                    col_m2.metric("📉 Q4月均销", fmt_num(total_q4_avg_show))
                    col_m3.metric("📅 整体可销月", fmt_num(dos_show))
                    col_m4.metric("🚨 异常客户数", f"{abnormal_count_show} 家")
                    st.markdown("---")

                    rank_stock = (
                        metrics_df.groupby('经销商名称', as_index=False)['当前库存_箱']
                        .sum()
                        .rename(columns={'当前库存_箱': '库存数(箱)'})
                    )
                    rank_stock['经销商名称'] = rank_stock['经销商名称'].astype(str).str.strip()
                    rank_stock = pd.merge(
                        rank_stock,
                        sales_agg_q4[['经销商名称', 'Q4_Avg']] if (sales_agg_q4 is not None and 'Q4_Avg' in sales_agg_q4.columns) else pd.DataFrame(columns=['经销商名称', 'Q4_Avg']),
                        on='经销商名称',
                        how='left'
                    )
                    rank_stock['Q4_Avg'] = pd.to_numeric(rank_stock.get('Q4_Avg', 0), errors='coerce').fillna(0)
                    rank_stock['近三月未出库'] = (rank_stock['Q4_Avg'] <= 0) & (rank_stock['库存数(箱)'] > 0)

                    def _rank_dos(row):
                        q4 = float(row.get('Q4_Avg', 0) or 0)
                        stk = float(row.get('库存数(箱)', 0) or 0)
                        if q4 <= 0:
                            return float('nan') if stk > 0 else 0.0
                        return stk / q4

                    rank_stock['可销月'] = rank_stock.apply(_rank_dos, axis=1)
                    rank_stock['过高差值'] = (rank_stock['可销月'] - float(high_th))
                    rank_stock['过低差值'] = (float(low_th) - rank_stock['可销月'])

                    rank_stock_rankable = rank_stock[~rank_stock['近三月未出库']].copy()
                    high_top = rank_stock_rankable[rank_stock_rankable['过高差值'] > 0].copy().sort_values('过高差值', ascending=False).head(10)
                    low_top = rank_stock_rankable[rank_stock_rankable['过低差值'] > 0].copy().sort_values('过低差值', ascending=False).head(10)

                    st.markdown("### 🏆 异常库存TOP10经销商")
                    r1, r2 = st.columns(2)
                    with r1:
                        st.subheader("🔴 库存过高 TOP10")
                        if high_top.empty:
                            st.info("当前范围无库存过高经销商")
                        else:
                            high_chart = high_top.sort_values('过高差值', ascending=True).copy()
                            high_chart['标注'] = high_chart['过高差值'].map(lambda x: f"+{fmt_num(x, na='')}")
                            high_chart['_库存数_fmt'] = high_chart['库存数(箱)'].map(lambda x: fmt_num(x, na=''))
                            high_chart['_q4_fmt'] = high_chart['Q4_Avg'].map(lambda x: fmt_num(x, na=''))
                            high_chart['_dos_fmt'] = high_chart['可销月'].map(lambda x: fmt_num(x, na=''))
                            high_chart['_diff_fmt'] = high_chart['过高差值'].map(lambda x: fmt_num(x, na=''))
                            fig_high = px.bar(
                                high_chart,
                                x='过高差值',
                                y='经销商名称',
                                orientation='h',
                                text='标注',
                                title="超出过高阈值的差值（可销月 - 阈值）",
                                color_discrete_sequence=['#E5484D'],
                                custom_data=['_库存数_fmt', '_q4_fmt', '_dos_fmt', '_diff_fmt']
                            )
                            fig_high.update_traces(
                                textposition='outside',
                                hovertemplate=(
                                    "经销商: %{y}<br>"
                                    "库存数(箱): %{customdata[0]}<br>"
                                    "Q4月均销: %{customdata[1]}<br>"
                                    "可销月: %{customdata[2]}<br>"
                                    "超阈值差值: +%{customdata[3]}<extra></extra>"
                                )
                            )
                            fig_high.update_layout(height=420, xaxis_title="差值", yaxis_title="")
                            st.plotly_chart(fig_high, use_container_width=True)
                            show_aggrid_table(high_top[['经销商名称', '库存数(箱)', 'Q4_Avg', '可销月', '过高差值']], height=250, key='high_stock_ag')

                    with r2:
                        st.subheader("🟠 库存过低 TOP10")
                        if low_top.empty:
                            st.info("当前范围无库存过低经销商")
                        else:
                            low_chart = low_top.sort_values('过低差值', ascending=True).copy()
                            low_chart['标注'] = low_chart['过低差值'].map(lambda x: f"+{fmt_num(x, na='')}")
                            low_chart['_库存数_fmt'] = low_chart['库存数(箱)'].map(lambda x: fmt_num(x, na=''))
                            low_chart['_q4_fmt'] = low_chart['Q4_Avg'].map(lambda x: fmt_num(x, na=''))
                            low_chart['_dos_fmt'] = low_chart['可销月'].map(lambda x: fmt_num(x, na=''))
                            low_chart['_diff_fmt'] = low_chart['过低差值'].map(lambda x: fmt_num(x, na=''))
                            fig_low = px.bar(
                                low_chart,
                                x='过低差值',
                                y='经销商名称',
                                orientation='h',
                                text='标注',
                                title="低于过低阈值的差值（阈值 - 可销月）",
                                color_discrete_sequence=['#FFB000'],
                                custom_data=['_库存数_fmt', '_q4_fmt', '_dos_fmt', '_diff_fmt']
                            )
                            fig_low.update_traces(
                                textposition='outside',
                                hovertemplate=(
                                    "经销商: %{y}<br>"
                                    "库存数(箱): %{customdata[0]}<br>"
                                    "Q4月均销: %{customdata[1]}<br>"
                                    "可销月: %{customdata[2]}<br>"
                                    "低于阈值差值: +%{customdata[3]}<extra></extra>"
                                )
                            )
                            fig_low.update_layout(height=420, xaxis_title="差值", yaxis_title="")
                            st.plotly_chart(fig_low, use_container_width=True)
                            show_aggrid_table(low_top[['经销商名称', '库存数(箱)', 'Q4_Avg', '可销月', '过低差值']], height=250, key='low_stock_ag')

                    with st.expander("🔍 对账信息", expanded=False):
                        if df_o_filtered is None or df_o_filtered.empty or '月' not in df_o_filtered.columns:
                            st.warning("当前筛选下无出库明细可对账。")
                        else:
                            s10 = float(df_o_filtered[df_o_filtered['月'] == 10]['数量(箱)'].sum()) if '数量(箱)' in df_o_filtered.columns else 0.0
                            s11 = float(df_o_filtered[df_o_filtered['月'] == 11]['数量(箱)'].sum()) if '数量(箱)' in df_o_filtered.columns else 0.0
                            s12 = float(df_o_filtered[df_o_filtered['月'] == 12]['数量(箱)'].sum()) if '数量(箱)' in df_o_filtered.columns else 0.0
                            st.write(f"当前筛选下Sheet3合计：10月={fmt_num(s10)}，11月={fmt_num(s11)}，12月={fmt_num(s12)}")
                            st.write(f"当前筛选下Q4月均销=(10+11+12)/3 = {fmt_num((s10+s11+s12)/3)}")
                            if sales_agg_q4 is not None and 'Q4_Total' in sales_agg_q4.columns:
                                dist_scope_dbg = (
                                    metrics_df['经销商名称']
                                    .dropna()
                                    .astype(str)
                                    .str.strip()
                                    .unique()
                                    .tolist()
                                )
                                matched = sales_agg_q4[sales_agg_q4['经销商名称'].isin(dist_scope_dbg)]
                                st.write(f"当前范围经销商数(去重)：{len(dist_scope_dbg)}，Sheet3匹配到：{len(matched)}")
                                st.write(f"当前范围Q4月均销=(sum(Q4_Total))/3 = {fmt_num(float(matched['Q4_Total'].sum())/3)}")

                    # --- Navigation & Breadcrumbs ---
                    cols_nav = st.columns([1, 8])
                    if st.session_state.drill_level > 1:
                        if cols_nav[0].button("⬅️ 返回"):
                            st.session_state.drill_level -= 1
                            st.rerun()
                    
                    breadcrumbs = "🏠 全部省区"
                    if st.session_state.drill_level >= 2:
                        breadcrumbs += f" > 📍 {st.session_state.selected_prov}"
                    if st.session_state.drill_level >= 3:
                        breadcrumbs += f" > 🏢 {st.session_state.selected_dist}"
                    cols_nav[1].markdown(f"**当前位置**: {breadcrumbs}")

                    # --- Level 1: Province View ---
                    if st.session_state.drill_level == 1:
                        
                        # Agg by Prov
                        prov_agg = analysis_df.groupby('省区名称').agg({
                            '当前库存_箱': 'sum',
                            'Q4_Avg': 'sum',
                            '经销商名称': 'count' # Count of distributors
                        }).reset_index()
                        
                        # Calc Prov DOS
                        prov_agg['可销月(DOS)'] = prov_agg.apply(lambda x: (x['当前库存_箱'] / x['Q4_Avg']) if x['Q4_Avg'] > 0 else (float('nan') if x['当前库存_箱'] > 0 else 0.0), axis=1)
                        
                        # Count Abnormal Distributors per Prov
                        abnormal_counts = analysis_df.groupby('省区名称')['库存状态'].value_counts().unstack(fill_value=0)
                        if '🔴 库存过高' not in abnormal_counts.columns: abnormal_counts['🔴 库存过高'] = 0
                        if '🟠 库存不足' not in abnormal_counts.columns: abnormal_counts['🟠 库存不足'] = 0
                        if '⚫ 近三月未出库' not in abnormal_counts.columns: abnormal_counts['⚫ 近三月未出库'] = 0
                        
                        prov_view = pd.merge(prov_agg, abnormal_counts[['🔴 库存过高', '🟠 库存不足', '⚫ 近三月未出库']], on='省区名称', how='left').fillna(0)
                        
                        # New Logic: Calculate Total Abnormal Count and Sort
                        prov_view['合计异常数'] = prov_view['🔴 库存过高'] + prov_view['🟠 库存不足'] + prov_view['⚫ 近三月未出库']
                        prov_view['经销商总数'] = prov_view['经销商名称'] # Rename for clarity
                        
                        # Filter slider
                        max_abnormal = int(prov_view['合计异常数'].max()) if not prov_view.empty else 10
                        c_filter, _ = st.columns([1, 2])
                        min_abnormal_filter = c_filter.slider("🔎 异常数过滤 (≥)", 0, max_abnormal, 0)
                        
                        prov_view_filtered = prov_view[prov_view['合计异常数'] >= min_abnormal_filter].copy()
                        
                        # Sort Descending by Total Abnormal Count
                        prov_view_filtered = prov_view_filtered.sort_values('合计异常数', ascending=False)
                        
                        st.markdown("### 📋 省区库存异常详情列表")
                        st.caption("💡 提示：**直接点击表格中的某一行**，即可下钻查看该省区的经销商详情。")
                        
                        # Prepare DF for display
                        display_df = prov_view_filtered[["省区名称", "合计异常数", "🔴 库存过高", "🟠 库存不足", "⚫ 近三月未出库", "当前库存_箱", "Q4_Avg", "可销月(DOS)"]].reset_index(drop=True)
                        
                        # Use interactive dataframe with selection
                        # Dynamic height to show all rows
                        n_rows = len(display_df)
                        # Estimate height: 35px per row + 35px header + buffer
                        calc_height = (n_rows + 1) * 35 + 10
                        # Ensure a minimum height and reasonable max height (e.g., 2000px)
                        final_height = max(150, min(calc_height, 2000))

                        ag_inv = show_aggrid_table(
                            display_df,
                            height=final_height,
                            columns_props={'合计异常数': {'type': 'bar_count'}, '可销月(DOS)': {'type': 'number'}},
                            on_row_selected='single',
                            key='inv_prov_ag'
                        )
                        
                        # Show all province names as tags below for quick view
                        with st.expander("查看所有省区名称列表 (点击展开)", expanded=False):
                            st.markdown("  ".join([f"`{p}`" for p in display_df['省区名称'].tolist()]))
                        
                        # Handle Selection
                        selected_rows = ag_inv.get('selected_rows') if ag_inv else None
                        if selected_rows is not None and len(selected_rows) > 0:
                            if isinstance(selected_rows, pd.DataFrame):
                                first_row = selected_rows.iloc[0]
                            else:
                                first_row = selected_rows[0]
                            
                            selected_prov_name = first_row.get("省区名称") if isinstance(first_row, dict) else first_row["省区名称"]
                            st.session_state.selected_prov = selected_prov_name
                            st.session_state.drill_level = 2
                            st.rerun()

                        # Visualization: Stacked Bar Chart of Abnormalities
                        if not prov_view_filtered.empty:
                            fig_abnormal = px.bar(
                                prov_view_filtered,
                                x='省区名称',
                                y=['🔴 库存过高', '🟠 库存不足'],
                                title='各省异常库存分布',
                                labels={'value': '经销商数量', 'variable': '异常类型'},
                                color_discrete_map={'🔴 库存过高': '#E5484D', '🟠 库存不足': '#FFB000'}
                            )
                            fig_abnormal.update_layout(height=350, margin=dict(l=20, r=20, t=40, b=20))
                            st.plotly_chart(fig_abnormal, use_container_width=True)

                    # --- Level 2: Distributor View ---
                    elif st.session_state.drill_level == 2:
                        prov = st.session_state.selected_prov
                        st.caption("💡 提示：**点击表格行** 可查看该经销商的 SKU 库存明细。")
                        
                        # Filter by Prov
                        dist_view = analysis_df[analysis_df['省区名称'] == prov].copy().reset_index(drop=True)
                        
                        # Interactive Table
                        ag_dist_inv = show_aggrid_table(
                            dist_view[['经销商名称', '当前库存_箱', 'Q4_Avg', '可销月(DOS)', '库存状态']],
                            height=520,
                            columns_props={'可销月(DOS)': {'type': 'number'}},
                            on_row_selected='single',
                            key='inv_dist_ag'
                        )
                        
                        # Handle Selection
                        selected_rows_d = ag_dist_inv.get('selected_rows') if ag_dist_inv else None
                        if selected_rows_d is not None and len(selected_rows_d) > 0:
                            if isinstance(selected_rows_d, pd.DataFrame):
                                first_row_d = selected_rows_d.iloc[0]
                            else:
                                first_row_d = selected_rows_d[0]
                            
                            selected_dist_name = first_row_d.get("经销商名称") if isinstance(first_row_d, dict) else first_row_d["经销商名称"]
                            st.session_state.selected_dist = selected_dist_name
                            st.session_state.drill_level = 3
                            st.rerun()

                    # --- Level 3: SKU/Store View ---
                    elif st.session_state.drill_level == 3:
                        dist = st.session_state.selected_dist
                        
                        # Get SKU details for this distributor from filtered stock data
                        # Note: We don't have store-level sales in Sheet3 (only Dist level), 
                        # so we can only show Stock Details here, potentially calculating SKU-level DOS if we had SKU-level sales (which we don't from Sheet3).
                        # We will show SKU stock details.
                        
                        sku_view = df_s_filtered[df_s_filtered['经销商名称'] == dist][['产品名称', '产品编码', '箱数', '规格', '重量']].copy()
                        
                        show_aggrid_table(sku_view, height=520, key='inv_sku_ag')
                        st.caption("注：因Q4出库数据仅精确到经销商层级，此处仅展示SKU库存明细，不计算单品DOS。")

            if main_tab == "🚚 出库分析":
                if df_q4_raw is None or df_q4_raw.empty:
                    st.warning("⚠️ 未检测到出库数据 (Sheet3)。请确认Excel包含Sheet3且数据完整。")
                    with st.expander("🛠️ 调试信息", expanded=False):
                        for log in debug_logs:
                            st.text(log)
                else:
                    st.caption(f"🕒 数据更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

                    o_raw = df_q4_raw.copy()

                    if '产品大类' not in o_raw.columns:
                        o_raw['产品大类'] = '全部'
                    if '产品小类' not in o_raw.columns:
                        o_raw['产品小类'] = '全部'
                    for _c in ['省区', '经销商名称', '产品大类', '产品小类']:
                        if _c in o_raw.columns:
                            o_raw[_c] = o_raw[_c].fillna('').astype(str).str.strip()
                    if '经销商名称' in o_raw.columns:
                        o_raw['经销商名称'] = o_raw['经销商名称'].str.replace(r'\s+', '', regex=True)

                    big_cat_src = '透视' if '透视' in o_raw.columns else (o_raw.columns[19] if len(o_raw.columns) > 19 else None)
                    small_cat_src = '重量' if '重量' in o_raw.columns else (o_raw.columns[20] if len(o_raw.columns) > 20 else None)
                    out_prod_src = '出库产品' if '出库产品' in o_raw.columns else (o_raw.columns[8] if len(o_raw.columns) > 8 else None)
                    o_raw['_模块大类'] = o_raw[big_cat_src] if big_cat_src is not None else '全部'
                    o_raw['_模块小类'] = o_raw[small_cat_src] if small_cat_src is not None else '全部'
                    o_raw['_模块出库产品'] = o_raw[out_prod_src] if out_prod_src is not None else '全部'
                    for _c in ['_模块大类', '_模块小类', '_模块出库产品']:
                        if _c in o_raw.columns:
                            o_raw[_c] = o_raw[_c].fillna('').astype(str).str.strip()

                    day_col = next((c for c in o_raw.columns if str(c).strip() == '日'), None)
                    if day_col is None:
                        day_col = next((c for c in o_raw.columns if ('日期' in str(c)) or (str(c).strip().endswith('日') and '月' not in str(c))), None)
                    if day_col is None and len(o_raw.columns) > 14:
                        day_col = o_raw.columns[14]

                    store_name_col = o_raw.columns[5] if len(o_raw.columns) > 5 else None

                    if '数量(箱)' in o_raw.columns:
                        o_raw['数量(箱)'] = pd.to_numeric(o_raw['数量(箱)'], errors='coerce').fillna(0.0)
                    else:
                        o_raw['数量(箱)'] = 0.0

                    if store_name_col is not None and store_name_col in o_raw.columns:
                        o_raw['_门店名'] = (
                            o_raw[store_name_col]
                            .fillna('')
                            .astype(str)
                            .str.replace(r'\s+', '', regex=True)
                        )
                        o_raw.loc[o_raw['_门店名'].isin(['', 'nan', 'None', 'NULL', 'NaN']), '_门店名'] = pd.NA
                    else:
                        o_raw['_门店名'] = pd.NA

                    def _to_month(v):
                        if pd.isna(v):
                            return None
                        if isinstance(v, (int, float)) and not pd.isna(v):
                            m = int(v)
                            return m if 1 <= m <= 12 else None
                        s = str(v).strip()
                        if s.isdigit():
                            m = int(s)
                            return m if 1 <= m <= 12 else None
                        if '月' in s:
                            digits = ''.join([ch for ch in s if ch.isdigit()])
                            if digits:
                                for k in (2, 1):
                                    if len(digits) >= k:
                                        m = int(digits[-k:])
                                        if 1 <= m <= 12:
                                            return m
                            return None
                        dt = pd.to_datetime(s, errors='coerce')
                        if pd.isna(dt):
                            return None
                        m = int(dt.month)
                        return m if 1 <= m <= 12 else None

                    def _to_day(v):
                        if pd.isna(v):
                            return None
                        if isinstance(v, (int, float)) and not pd.isna(v):
                            d = int(v)
                            return d if 1 <= d <= 31 else None
                        s = str(v).strip()
                        digits = ''.join([ch for ch in s if ch.isdigit()])
                        if not digits:
                            return None
                        d = int(digits[-2:]) if len(digits) >= 2 else int(digits)
                        return d if 1 <= d <= 31 else None

                    if '年份' in o_raw.columns:
                        o_raw['_年'] = pd.to_numeric(o_raw['年份'], errors='coerce').fillna(0).astype(int)
                    else:
                        o_raw['_年'] = 0
                    if '月份' in o_raw.columns:
                        o_raw['_月'] = o_raw['月份'].apply(_to_month)
                    else:
                        o_raw['_月'] = None

                    if day_col is not None and day_col in o_raw.columns:
                        if '日期' in str(day_col):
                            dt_series = pd.to_datetime(o_raw[day_col], errors='coerce')
                            o_raw['_年'] = np.where(dt_series.notna(), dt_series.dt.year, o_raw['_年']).astype(int)
                            o_raw['_月'] = np.where(dt_series.notna(), dt_series.dt.month, o_raw['_月'])
                            o_raw['_日'] = np.where(dt_series.notna(), dt_series.dt.day, None)
                        else:
                            o_raw['_日'] = o_raw[day_col].apply(_to_day)
                    else:
                        o_raw['_日'] = None

                    if '月份' in o_raw.columns:
                        _ms = o_raw['月份'].fillna('').astype(str).str.strip()
                        _ym4 = _ms.str.extract(r'(20\d{2})\D{0,3}(0?[1-9]|1[0-2])')
                        _mask_y = pd.to_numeric(o_raw.get('_年', 0), errors='coerce').fillna(0).astype(int) <= 0
                        if _mask_y.any():
                            _y4 = pd.to_numeric(_ym4[0], errors='coerce')
                            o_raw.loc[_mask_y & _y4.notna(), '_年'] = _y4[_mask_y & _y4.notna()].astype(int)
                        _mask_m = o_raw.get('_月', pd.Series([None] * len(o_raw))).isna()
                        if _mask_m.any():
                            _m4 = pd.to_numeric(_ym4[1], errors='coerce')
                            o_raw.loc[_mask_m & _m4.notna(), '_月'] = _m4[_mask_m & _m4.notna()]

                        _ym2 = _ms.str.extract(r'(?<!\d)(\d{2})\D{0,3}(0?[1-9]|1[0-2])')
                        _mask_y2 = pd.to_numeric(o_raw.get('_年', 0), errors='coerce').fillna(0).astype(int) <= 0
                        if _mask_y2.any():
                            _y2 = pd.to_numeric(_ym2[0], errors='coerce')
                            o_raw.loc[_mask_y2 & _y2.notna(), '_年'] = (2000 + _y2[_mask_y2 & _y2.notna()]).astype(int)
                        _mask_m2 = o_raw.get('_月', pd.Series([None] * len(o_raw))).isna()
                        if _mask_m2.any():
                            _m2 = pd.to_numeric(_ym2[1], errors='coerce')
                            o_raw.loc[_mask_m2 & _m2.notna(), '_月'] = _m2[_mask_m2 & _m2.notna()]

                    o_raw = o_raw[o_raw['_年'] > 0].copy()
                    o_raw = o_raw[o_raw['_月'].notna()].copy()
                    o_raw['_月'] = o_raw['_月'].astype(int)
                    o_raw['_日'] = pd.to_numeric(o_raw['_日'], errors='coerce')

                    with st.expander("🛠️ 出库筛选", expanded=False):
                        out_provs = ['全部'] + sorted(o_raw['省区'].dropna().astype(str).unique().tolist()) if '省区' in o_raw.columns else ['全部']
                        oc1, oc2, oc3, oc4, oc5 = st.columns(5)
                        with oc1:
                            o_prov = st.selectbox("省区", out_provs, key='out2_prov')
                        with oc2:
                            if '经销商名称' in o_raw.columns:
                                if o_prov != '全部' and '省区' in o_raw.columns:
                                    dists_in_prov = o_raw[o_raw['省区'].astype(str) == str(o_prov)]['经销商名称'].dropna().astype(str).unique().tolist()
                                    out_dists = ['全部'] + sorted(dists_in_prov)
                                else:
                                    out_dists = ['全部'] + sorted(o_raw['经销商名称'].dropna().astype(str).unique().tolist())
                            else:
                                out_dists = ['全部']
                            o_dist = st.selectbox("经销商", out_dists, key='out2_dist')
                        with oc3:
                            out_cats = ['全部'] + sorted(o_raw['产品大类'].dropna().astype(str).unique().tolist())
                            o_cat = st.selectbox("产品大类", out_cats, key='out2_cat')
                        with oc4:
                            if o_cat != '全部':
                                subs_in_cat = o_raw[o_raw['产品大类'].astype(str) == str(o_cat)]['产品小类'].dropna().astype(str).unique().tolist()
                                out_subs = ['全部'] + sorted(subs_in_cat)
                            else:
                                out_subs = ['全部'] + sorted(o_raw['产品小类'].dropna().astype(str).unique().tolist())
                            o_sub = st.selectbox("产品小类", out_subs, key='out2_sub')
                        with oc5:
                            year_opts = sorted([int(y) for y in o_raw['_年'].dropna().unique().tolist() if int(y) > 0])
                            default_year = 2025 if 2025 in year_opts else (max(year_opts) if year_opts else 2025)
                            y_index = year_opts.index(default_year) if default_year in year_opts else 0
                            o_year = st.selectbox("年份", year_opts if year_opts else [2025], index=y_index, key='out2_year')
                            month_in_year = sorted([int(m) for m in o_raw[o_raw['_年'] == int(o_year)]['_月'].dropna().unique().tolist() if 1 <= int(m) <= 12])
                            month_opts = ['全部'] + month_in_year
                            o_month = st.selectbox("月份", month_opts, index=0, key='out2_month')

                    df_o = o_raw.copy()
                    if o_prov != '全部' and '省区' in df_o.columns:
                        df_o = df_o[df_o['省区'].astype(str) == str(o_prov)]
                    if o_dist != '全部' and '经销商名称' in df_o.columns:
                        df_o = df_o[df_o['经销商名称'].astype(str) == str(o_dist)]
                    if o_cat != '全部':
                        df_o = df_o[df_o['产品大类'].astype(str) == str(o_cat)]
                    if o_sub != '全部':
                        df_o = df_o[df_o['产品小类'].astype(str) == str(o_sub)]

                    df_o_prov_base = o_raw.copy()
                    if o_dist != '全部' and '经销商名称' in df_o_prov_base.columns:
                        df_o_prov_base = df_o_prov_base[df_o_prov_base['经销商名称'].astype(str) == str(o_dist)]

                    def _agg_scope(df_scope: pd.DataFrame):
                        boxes = float(df_scope.get('数量(箱)', 0).sum()) if df_scope is not None and not df_scope.empty else 0.0
                        if df_scope is None or df_scope.empty or '_门店名' not in df_scope.columns:
                            stores = 0.0
                        else:
                            df_s = df_scope[df_scope['数量(箱)'] > 0].copy()
                            stores = float(df_s['_门店名'].dropna().astype(str).nunique()) if not df_s.empty else 0.0
                        return boxes, stores

                    def _yoy(cur, last):
                        if last is None:
                            return None
                        last_v = float(last or 0)
                        if last_v <= 0:
                            return None
                        return (float(cur or 0) - last_v) / last_v

                    def _avg(boxes, stores):
                        try:
                            s = float(stores or 0)
                            return float(boxes or 0) / s if s > 0 else 0.0
                        except Exception:
                            return 0.0

                    def _fmt_num(x):
                        return fmt_num(x, na="0")

                    def _fmt_pct(x):
                        return fmt_pct_ratio(x) if x is not None else "—"

                    def _trend_cls(x):
                        if x is None or (isinstance(x, float) and pd.isna(x)):
                            return "trend-neutral"
                        return "trend-up" if x > 0 else ("trend-down" if x < 0 else "trend-neutral")

                    def _arrow(x):
                        if x is None or (isinstance(x, float) and pd.isna(x)):
                            return ""
                        return "↑" if x > 0 else ("↓" if x < 0 else "")

                    out_subtab_options = ["📊 关键指标", "📦 分品类", "🗺️ 分省区", "📈 趋势分析", "🧾 专案追踪", "门店类型滚动分析"]
                    out_subtab = st.segmented_control(
                        "出库子导航",
                        options=out_subtab_options,
                        default=st.session_state.get("out_subtab_nav", "📈 趋势分析") if isinstance(st.session_state.get("out_subtab_nav", None), str) else "📈 趋势分析",
                        key="out_subtab_nav",
                        label_visibility="collapsed",
                    )
                    
                    # Prepare Data Context (Shared)
                    sig = (st.session_state.get("_active_file_sig"), o_prov, o_dist, o_cat, o_sub, o_year, o_month)
                    if "out_subtab_cache" not in st.session_state:
                        st.session_state.out_subtab_cache = {}
                    
                    def _get_ctx():
                        ck = ("ctx", sig)
                        if ck in st.session_state.out_subtab_cache:
                            return st.session_state.out_subtab_cache[ck]
                        
                        # No spinner here to avoid flashing on every rerun, 
                        # relying on Streamlit's natural execution speed or cache if possible.
                        # If slow, we might add st.spinner inside specific heavy blocks.
                        if o_month != '全部':
                            _kpi_year = int(o_year)
                            _kpi_month = int(o_month)
                        else:
                            years_avail = sorted([int(y) for y in df_o['_年'].dropna().unique().tolist() if int(y) > 0])
                            _kpi_year = 2026 if 2026 in years_avail else (max(years_avail) if years_avail else int(o_year))
                            months_avail = sorted([int(m) for m in df_o[df_o['_年'] == int(_kpi_year)]['_月'].dropna().unique().tolist() if 1 <= int(m) <= 12])
                            _kpi_month = max(months_avail) if months_avail else 1

                        days_avail = sorted([int(d) for d in df_o[(df_o['_年'] == int(_kpi_year)) & (df_o['_月'] == int(_kpi_month))]['_日'].dropna().unique().tolist() if 1 <= int(d) <= 31])
                        _kpi_day = max(days_avail) if days_avail else None
                        _cmp_year = int(_kpi_year) - 1

                        _cur_today = (df_o[(df_o['_年'] == int(_kpi_year)) & (df_o['_月'] == int(_kpi_month)) & (df_o['_日'] == int(_kpi_day))] if _kpi_day is not None else df_o.iloc[0:0])
                        _cur_month = df_o[(df_o['_年'] == int(_kpi_year)) & (df_o['_月'] == int(_kpi_month))]
                        _cur_year = df_o[(df_o['_年'] == int(_kpi_year))]

                        _last_today = (df_o[(df_o['_年'] == int(_cmp_year)) & (df_o['_月'] == int(_kpi_month)) & (df_o['_日'] == int(_kpi_day))] if _kpi_day is not None else df_o.iloc[0:0])
                        _last_month = df_o[(df_o['_年'] == int(_cmp_year)) & (df_o['_月'] == int(_kpi_month))]
                        _last_year = df_o[(df_o['_年'] == int(_cmp_year))]

                        ctx = {
                            "kpi_year": _kpi_year,
                            "kpi_month": _kpi_month,
                            "kpi_day": _kpi_day,
                            "cmp_year": _cmp_year,
                            "cur_today": _cur_today,
                            "cur_month": _cur_month,
                            "cur_year": _cur_year,
                            "last_today": _last_today,
                            "last_month": _last_month,
                            "last_year": _last_year,
                        }
                        st.session_state.out_subtab_cache[ck] = ctx
                        return ctx

                    ctx = _get_ctx()
                    
                    # Common Caption
                    st.caption(
                        f"当前默认口径：{ctx['kpi_year']}年{int(ctx['kpi_month'])}月"
                        + (f"{int(ctx['kpi_day'])}日" if ctx["kpi_day"] is not None else "")
                    )

                    # --- Tab 1: KPI ---
                    if out_subtab == "📊 关键指标":
                        ck = ("kpi", sig)
                        if ck not in st.session_state.out_subtab_cache:
                             t_boxes, t_stores = _agg_scope(ctx["cur_today"])
                             tm_boxes, tm_stores = _agg_scope(ctx["cur_month"])
                             ty_boxes, ty_stores = _agg_scope(ctx["cur_year"])
                             lt_boxes, lt_stores = _agg_scope(ctx["last_today"])
                             ltm_boxes, ltm_stores = _agg_scope(ctx["last_month"])
                             lty_boxes, lty_stores = _agg_scope(ctx["last_year"])
                             t_yoy = _yoy(t_boxes, lt_boxes)
                             tm_yoy = _yoy(tm_boxes, ltm_boxes)
                             ty_yoy = _yoy(ty_boxes, lty_boxes)
                             t_avg = _avg(t_boxes, t_stores)
                             tm_avg = _avg(tm_boxes, tm_stores)
                             ty_avg = _avg(ty_boxes, ty_stores)
                             lt_avg = _avg(lt_boxes, lt_stores)
                             ltm_avg = _avg(ltm_boxes, ltm_stores)
                             lty_avg = _avg(lty_boxes, lty_stores)
                             st.session_state.out_subtab_cache[ck] = {
                                "t_boxes": t_boxes, "t_stores": t_stores, "t_yoy": t_yoy, "t_avg": t_avg, "lt_boxes": lt_boxes, "lt_stores": lt_stores, "lt_avg": lt_avg,
                                "tm_boxes": tm_boxes, "tm_stores": tm_stores, "tm_yoy": tm_yoy, "tm_avg": tm_avg, "ltm_boxes": ltm_boxes, "ltm_stores": ltm_stores, "ltm_avg": ltm_avg,
                                "ty_boxes": ty_boxes, "ty_stores": ty_stores, "ty_yoy": ty_yoy, "ty_avg": ty_avg, "lty_boxes": lty_boxes, "lty_stores": lty_stores, "lty_avg": lty_avg,
                             }
                        m = st.session_state.out_subtab_cache[ck]

                        k1, k2, k3 = st.columns(3)
                        with k1:
                            st.markdown(f"""
                            <div class="out-kpi-card">
                                <div class="out-kpi-bar"></div>
                                <div class="out-kpi-head">
                                    <div class="out-kpi-ico">🚚</div>
                                    <div class="out-kpi-title">本日出库</div>
                                </div>
                                <div class="out-kpi-val">{_fmt_num(m['t_boxes'])} 箱</div>
                                <div class="out-kpi-sub"><span>门店数</span><span>{_fmt_num(m['t_stores'])}</span></div>
                                <div class="out-kpi-sub2"><span>店均（箱/店）</span><span>{fmt_num(m['t_avg'])} <span style="color:rgba(27,21,48,0.55);">｜同期 {fmt_num(m['lt_avg'])}</span></span></div>
                                <div class="out-kpi-sub2" style="margin-top:10px;"><span>同期({ctx['cmp_year']})</span><span>{_fmt_num(m['lt_boxes'])} 箱 / {_fmt_num(m['lt_stores'])} 店</span></div>
                                <div class="out-kpi-sub2"><span>同比（箱）</span><span class="{_trend_cls(m['t_yoy'])}">{_arrow(m['t_yoy'])} {_fmt_pct(m['t_yoy'])}</span></div>
                            </div>
                            """, unsafe_allow_html=True)

                        with k2:
                            st.markdown(f"""
                            <div class="out-kpi-card">
                                <div class="out-kpi-bar"></div>
                                <div class="out-kpi-head">
                                    <div class="out-kpi-ico">📦</div>
                                    <div class="out-kpi-title">本月累计出库</div>
                                </div>
                                <div class="out-kpi-val">{_fmt_num(m['tm_boxes'])} 箱</div>
                                <div class="out-kpi-sub"><span>门店数</span><span>{_fmt_num(m['tm_stores'])}</span></div>
                                <div class="out-kpi-sub2"><span>店均（箱/店）</span><span>{fmt_num(m['tm_avg'])} <span style="color:rgba(27,21,48,0.55);">｜同期 {fmt_num(m['ltm_avg'])}</span></span></div>
                                <div class="out-kpi-sub2" style="margin-top:10px;"><span>同期({ctx['cmp_year']})</span><span>{_fmt_num(m['ltm_boxes'])} 箱 / {_fmt_num(m['ltm_stores'])} 店</span></div>
                                <div class="out-kpi-sub2"><span>同比（箱）</span><span class="{_trend_cls(m['tm_yoy'])}">{_arrow(m['tm_yoy'])} {_fmt_pct(m['tm_yoy'])}</span></div>
                            </div>
                            """, unsafe_allow_html=True)

                        with k3:
                            st.markdown(f"""
                            <div class="out-kpi-card">
                                <div class="out-kpi-bar"></div>
                                <div class="out-kpi-head">
                                    <div class="out-kpi-ico">🏁</div>
                                    <div class="out-kpi-title">本年累计出库</div>
                                </div>
                                <div class="out-kpi-val">{_fmt_num(m['ty_boxes'])} 箱</div>
                                <div class="out-kpi-sub"><span>门店数</span><span>{_fmt_num(m['ty_stores'])}</span></div>
                                <div class="out-kpi-sub2"><span>店均（箱/店）</span><span>{fmt_num(m['ty_avg'])} <span style="color:rgba(27,21,48,0.55);">｜同期 {fmt_num(m['lty_avg'])}</span></span></div>
                                <div class="out-kpi-sub2" style="margin-top:10px;"><span>同期({ctx['cmp_year']})</span><span>{_fmt_num(m['lty_boxes'])} 箱 / {_fmt_num(m['lty_stores'])} 店</span></div>
                                <div class="out-kpi-sub2"><span>同比（箱）</span><span class="{_trend_cls(m['ty_yoy'])}">{_arrow(m['ty_yoy'])} {_fmt_pct(m['ty_yoy'])}</span></div>
                            </div>
                            """, unsafe_allow_html=True)

                    # --- Tab 2: Category ---
                    if out_subtab == "📦 分品类":
                        ck = ("cat", sig)
                        if ck not in st.session_state.out_subtab_cache:
                            with st.spinner("正在加载分品类…"):
                                cat_dim = '产品小类' if o_cat != '全部' else '产品大类'
                                st.session_state.out_subtab_cache[ck] = {"cat_dim": cat_dim}
                        cat_dim = st.session_state.out_subtab_cache[ck]["cat_dim"]
                        dim_label = '产品小类' if cat_dim == '产品小类' else '产品大类'

                        st.caption(f"统计维度：{dim_label}（随筛选条件实时更新）")

                        def _cat_agg(df_scope: pd.DataFrame):
                            if df_scope is None or df_scope.empty:
                                return pd.DataFrame(columns=[cat_dim, '箱数', '门店数'])
                            df_t = df_scope.copy()
                            if cat_dim not in df_t.columns:
                                df_t[cat_dim] = '未知'
                            df_t[cat_dim] = df_t[cat_dim].fillna('未知').astype(str).str.strip()
                            df_t = df_t[df_t['数量(箱)'] > 0].copy()
                            if df_t.empty:
                                return pd.DataFrame(columns=[cat_dim, '箱数', '门店数'])
                            g_box = df_t.groupby(cat_dim, as_index=False)['数量(箱)'].sum().rename(columns={'数量(箱)': '箱数'})
                            if '_门店名' in df_t.columns:
                                g_store = df_t[df_t['_门店名'].notna()].groupby(cat_dim, as_index=False)['_门店名'].nunique().rename(columns={'_门店名': '门店数'})
                            else:
                                g_store = pd.DataFrame({cat_dim: g_box[cat_dim], '门店数': 0})
                            out = pd.merge(g_box, g_store, on=cat_dim, how='left').fillna(0)
                            out = out.sort_values('箱数', ascending=False).reset_index(drop=True)
                            return out

                        def _topn_with_other(df_sum: pd.DataFrame, n: int = 15):
                            if df_sum is None or df_sum.empty:
                                return df_sum
                            head = df_sum.head(n).copy()
                            tail = df_sum.iloc[n:].copy()
                            if not tail.empty:
                                other = pd.DataFrame([{
                                    cat_dim: '其他',
                                    '箱数': float(tail['箱数'].sum()),
                                    '门店数': float(tail['门店数'].sum())
                                }])
                                head = pd.concat([head, other], ignore_index=True)
                            return head

                        def _cat_table(df_cur: pd.DataFrame, df_last: pd.DataFrame):
                            cur_sum = _topn_with_other(_cat_agg(df_cur), 15)
                            last_sum = _topn_with_other(_cat_agg(df_last), 15)
                            if cur_sum is None or cur_sum.empty:
                                cur_sum = pd.DataFrame(columns=[cat_dim, '箱数', '门店数'])
                            if last_sum is None or last_sum.empty:
                                last_sum = pd.DataFrame(columns=[cat_dim, '箱数', '门店数'])
                            m = pd.merge(
                                cur_sum.rename(columns={'箱数': '箱数', '门店数': '门店数'}),
                                last_sum[[cat_dim, '箱数']].rename(columns={'箱数': '同期（箱数）'}),
                                on=cat_dim,
                                how='outer'
                            ).fillna(0)
                            m['同比'] = np.where(m['同期（箱数）'] > 0, (m['箱数'] - m['同期（箱数）']) / m['同期（箱数）'], None)
                            m = m.sort_values('箱数', ascending=False).reset_index(drop=True)
                            m = m.rename(columns={cat_dim: '品类'})
                            return m[['品类', '箱数', '门店数', '同期（箱数）', '同比']]

                        tab_cat_today, tab_cat_month, tab_cat_year = st.tabs(["本日", "本月", "本年"])
                        with tab_cat_today:
                            cat_tbl = _cat_table(ctx["cur_today"], ctx["last_today"])
                            show_aggrid_table(cat_tbl, columns_props={'同比': {'type': 'percent'}}, auto_height_limit=520)
                        with tab_cat_month:
                            cat_tbl = _cat_table(ctx["cur_month"], ctx["last_month"])
                            show_aggrid_table(cat_tbl, columns_props={'同比': {'type': 'percent'}}, auto_height_limit=520)
                        with tab_cat_year:
                            cat_tbl = _cat_table(ctx["cur_year"], ctx["last_year"])
                            show_aggrid_table(cat_tbl, columns_props={'同比': {'type': 'percent'}}, auto_height_limit=520)

                    # --- Tab 3: Province ---
                    if out_subtab == "🗺️ 分省区":

                        def _prov_agg(df_scope: pd.DataFrame):
                            if df_scope is None or df_scope.empty or '省区' not in df_scope.columns:
                                return pd.DataFrame(columns=['省区', '箱数', '门店数'])
                            g_box = (
                                df_scope
                                .groupby('省区', as_index=False)['数量(箱)']
                                .sum()
                                .rename(columns={'数量(箱)': '箱数'})
                            )

                            if '_门店名' in df_scope.columns:
                                tmp = df_scope[(df_scope['数量(箱)'] > 0) & (df_scope['_门店名'].notna())].copy()
                                g_store = (
                                    tmp
                                    .groupby('省区', as_index=False)['_门店名']
                                    .nunique()
                                    .rename(columns={'_门店名': '门店数'})
                                )
                            else:
                                g_store = pd.DataFrame(columns=['省区', '门店数'])

                            return pd.merge(g_box, g_store, on='省区', how='left').fillna(0)

                        p_cur_today = _prov_agg(ctx["cur_today"])
                        p_cur_month = _prov_agg(ctx["cur_month"])
                        p_cur_year = _prov_agg(ctx["cur_year"])
                        p_last_today = _prov_agg(ctx["last_today"])
                        p_last_month = _prov_agg(ctx["last_month"])
                        p_last_year = _prov_agg(ctx["last_year"])

                        if '省区' in df_o_prov_base.columns:
                            prov_all = sorted([x for x in df_o_prov_base['省区'].dropna().astype(str).unique().tolist() if x and x != 'nan'])
                        else:
                            prov_all = sorted(set(
                                p_cur_today['省区'].astype(str).tolist()
                                + p_cur_month['省区'].astype(str).tolist()
                                + p_cur_year['省区'].astype(str).tolist()
                            ))
                        prov_df = pd.DataFrame({'省区': prov_all})

                        def _merge(prov_base, df_left, prefix):
                            d = df_left.copy()
                            d.columns = ['省区'] + [f"{prefix}{c}" for c in d.columns if c != '省区']
                            return pd.merge(prov_base, d, on='省区', how='left').fillna(0)

                        prov_df = _merge(prov_df, p_cur_today, "今日")
                        prov_df = _merge(prov_df, p_last_today, "同期今日")
                        prov_df = _merge(prov_df, p_cur_month, "本月")
                        prov_df = _merge(prov_df, p_last_month, "同期本月")
                        prov_df = _merge(prov_df, p_cur_year, "本年")
                        prov_df = _merge(prov_df, p_last_year, "同期本年")

                        prov_df['今日同比(箱)'] = prov_df.apply(lambda r: _yoy(r.get('今日箱数', 0), r.get('同期今日箱数', 0)), axis=1)
                        prov_df['今日同比(门店)'] = prov_df.apply(lambda r: _yoy(r.get('今日门店数', 0), r.get('同期今日门店数', 0)), axis=1)
                        prov_df['本月同比(箱)'] = prov_df.apply(lambda r: _yoy(r.get('本月箱数', 0), r.get('同期本月箱数', 0)), axis=1)
                        prov_df['本月同比(门店)'] = prov_df.apply(lambda r: _yoy(r.get('本月门店数', 0), r.get('同期本月门店数', 0)), axis=1)
                        prov_df['本年同比(箱)'] = prov_df.apply(lambda r: _yoy(r.get('本年箱数', 0), r.get('同期本年箱数', 0)), axis=1)
                        prov_df['本年同比(门店)'] = prov_df.apply(lambda r: _yoy(r.get('本年门店数', 0), r.get('同期本年门店数', 0)), axis=1)

                        prov_show = pd.DataFrame({
                            '省区': prov_df['省区'],
                            '今日箱数': pd.to_numeric(prov_df.get('今日箱数', 0), errors='coerce').fillna(0),
                            '今日门店数': pd.to_numeric(prov_df.get('今日门店数', 0), errors='coerce').fillna(0),
                            '今日同期(箱数)': pd.to_numeric(prov_df.get('同期今日箱数', 0), errors='coerce').fillna(0),
                            '今日同比(箱)': pd.to_numeric(prov_df.get('今日同比(箱)', None), errors='coerce'),
                            '本月箱数': pd.to_numeric(prov_df.get('本月箱数', 0), errors='coerce').fillna(0),
                            '本月门店数': pd.to_numeric(prov_df.get('本月门店数', 0), errors='coerce').fillna(0),
                            '本月同期(箱数)': pd.to_numeric(prov_df.get('同期本月箱数', 0), errors='coerce').fillna(0),
                            '本月同比(箱)': pd.to_numeric(prov_df.get('本月同比(箱)', None), errors='coerce'),
                            '本年箱数': pd.to_numeric(prov_df.get('本年箱数', 0), errors='coerce').fillna(0),
                            '本年门店数': pd.to_numeric(prov_df.get('本年门店数', 0), errors='coerce').fillna(0),
                            '本年同期(箱数)': pd.to_numeric(prov_df.get('同期本年箱数', 0), errors='coerce').fillna(0),
                            '本年同比(箱)': pd.to_numeric(prov_df.get('本年同比(箱)', None), errors='coerce'),
                        }).fillna({'今日同比(箱)': np.nan, '本月同比(箱)': np.nan, '本年同比(箱)': np.nan})

                        day_txt = f"{int(ctx['kpi_month'])}月{int(ctx['kpi_day'])}日" if ctx["kpi_day"] is not None else f"{int(ctx['kpi_month'])}月"
                        grp_today = f"今日（{day_txt}）"
                        grp_month = f"本月（{int(ctx['kpi_month'])}月）"
                        grp_year = f"本年（{int(ctx['kpi_year'])}年）"

                        col_defs = [
                            {'headerName': '省区', 'field': '省区', 'minWidth': 110, 'headerClass': 'ag-header-center'},
                            {
                                'headerName': grp_today,
                                'children': [
                                    {'headerName': '箱数', 'field': '今日箱数', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_NUM},
                                    {'headerName': '门店数', 'field': '今日门店数', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_NUM},
                                    {'headerName': '同期（箱数）', 'field': '今日同期(箱数)', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_NUM},
                                    {'headerName': '同比（箱）', 'field': '今日同比(箱)', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_PCT_RATIO}, 
                                ],
                            },
                            {
                                'headerName': grp_month,
                                'children': [
                                    {'headerName': '箱数', 'field': '本月箱数', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_NUM},
                                    {'headerName': '门店数', 'field': '本月门店数', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_NUM},
                                    {'headerName': '同期（箱数）', 'field': '本月同期(箱数)', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_NUM},
                                    {'headerName': '同比（箱）', 'field': '本月同比(箱)', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_PCT_RATIO}, 
                                ],
                            },
                            {
                                'headerName': grp_year,
                                'children': [
                                    {'headerName': '箱数', 'field': '本年箱数', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_NUM},
                                    {'headerName': '门店数', 'field': '本年门店数', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_NUM},
                                    {'headerName': '同期（箱数）', 'field': '本年同期(箱数)', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_NUM},
                                    {'headerName': '同比（箱）', 'field': '本年同比(箱)', 'type': ['numericColumn', 'numberColumnFilter'], 'headerClass': 'ag-header-center', 'valueFormatter': JS_FMT_PCT_RATIO}, 
                                ],
                            },
                        ]

                        def _sum_col(col_name: str) -> float:
                            if col_name not in prov_show.columns:
                                return 0.0
                            return float(pd.to_numeric(prov_show[col_name], errors='coerce').fillna(0).sum())

                        _t_cur = _sum_col('今日箱数')
                        _t_last = _sum_col('今日同期(箱数)')
                        _m_cur = _sum_col('本月箱数')
                        _m_last = _sum_col('本月同期(箱数)')
                        _y_cur = _sum_col('本年箱数')
                        _y_last = _sum_col('本年同期(箱数)')

                        pinned_total = {
                            '省区': '合计',
                            '今日箱数': _t_cur,
                            '今日门店数': _sum_col('今日门店数'),
                            '今日同期(箱数)': _t_last,
                            '今日同比(箱)': ((_t_cur - _t_last) / _t_last) if _t_last > 0 else None,
                            '本月箱数': _m_cur,
                            '本月门店数': _sum_col('本月门店数'),
                            '本月同期(箱数)': _m_last,
                            '本月同比(箱)': ((_m_cur - _m_last) / _m_last) if _m_last > 0 else None,
                            '本年箱数': _y_cur,
                            '本年门店数': _sum_col('本年门店数'),
                            '本年同期(箱数)': _y_last,
                            '本年同比(箱)': ((_y_cur - _y_last) / _y_last) if _y_last > 0 else None,
                        }

                        gridOptions = {
                            'pinnedBottomRowData': [pinned_total],
                            'columnDefs': col_defs,
                            'defaultColDef': {
                                'resizable': True,
                                'sortable': True,
                                'filter': True,
                                'wrapHeaderText': True,
                                'autoHeaderHeight': True,
                                'minWidth': 70,
                                'flex': 1,
                                'cellStyle': {'textAlign': 'center', 'display': 'flex', 'justifyContent': 'center', 'alignItems': 'center'},
                                'headerClass': 'ag-header-center',
                            },
                            'rowHeight': 40,
                            'headerHeight': 60,
                            'groupHeaderHeight': 60,
                            'animateRows': True,
                            'suppressCellFocus': True,
                            'enableCellTextSelection': True,
                            'suppressDragLeaveHidesColumns': True,
                            'sideBar': {
                                "toolPanels": [
                                    {
                                        "id": "columns",
                                        "labelDefault": "列",
                                        "iconKey": "columns",
                                        "toolPanel": "agColumnsToolPanel",
                                        "toolPanelParams": {
                                            "suppressRowGroups": True,
                                            "suppressValues": True,
                                            "suppressPivots": True,
                                            "suppressPivotMode": True
                                        }
                                    }
                                ],
                                "defaultToolPanel": None
                            },
                        }

                        AgGrid(
                            prov_show,
                            gridOptions=gridOptions,
                            height=520,
                            width='100%',
                            data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                            update_mode=GridUpdateMode.NO_UPDATE,
                            fit_columns_on_grid_load=True,
                            allow_unsafe_jscode=True,
                            theme='streamlit',
                            key="outbound_prov_table"
                        )
                    
                    if out_subtab == "📈 趋势分析":
                        st.markdown("### 📈 月度出库趋势（可选多月）")

                        trend_view = st.segmented_control(
                            "趋势视图",
                            options=["多月趋势", "本月（日趋势）"],
                            default=st.session_state.get("out_m_trend_view", "多月趋势") if isinstance(st.session_state.get("out_m_trend_view", None), str) else "多月趋势",
                            key="out_m_trend_view",
                            label_visibility="collapsed",
                        )

                        if trend_view == "本月（日趋势）":
                            if "out_d_drill_level" not in st.session_state:
                                st.session_state.out_d_drill_level = 1
                            if "out_d_selected_prov" not in st.session_state:
                                st.session_state.out_d_selected_prov = None
                            if "out_d_selected_dist" not in st.session_state:
                                st.session_state.out_d_selected_dist = None
                            if "out_d_png_cache" not in st.session_state:
                                st.session_state.out_d_png_cache = {}
                            _out_d_png_cache = st.session_state.out_d_png_cache
                            if "out_d_zip_cache" not in st.session_state:
                                st.session_state.out_d_zip_cache = {}
                            _out_d_zip_cache = st.session_state.out_d_zip_cache

                            def _reset_out_d():
                                st.session_state.out_d_drill_level = 1
                                st.session_state.out_d_selected_prov = None
                                st.session_state.out_d_selected_dist = None
                                try:
                                    st.session_state.out_d_png_cache = {}
                                    st.session_state.out_d_zip_cache = {}
                                except Exception:
                                    pass

                            y0 = int(ctx.get("kpi_year") or 0)
                            m0 = 3
                            try:
                                _months = sorted([int(x) for x in df_o[df_o["_年"] == int(y0)]["_月"].dropna().unique().tolist()])
                                if m0 not in _months:
                                    m0 = int(ctx.get("kpi_month") or (max(_months) if _months else 3))
                            except Exception:
                                m0 = int(ctx.get("kpi_month") or 3)

                            df_base = df_o[df_o["_年"] == int(y0)].copy()
                            month_opts = sorted([int(x) for x in df_base["_月"].dropna().unique().tolist() if 1 <= int(x) <= 12])
                            if m0 not in month_opts and month_opts:
                                m0 = max(month_opts)

                            c_f1, c_f2, c_f3, c_f4 = st.columns([1.0, 1.25, 1.25, 1.8])
                            with c_f1:
                                m0 = st.selectbox("月份", month_opts if month_opts else [m0], index=(month_opts.index(m0) if m0 in month_opts else 0), key="out_d_month", on_change=_reset_out_d)
                            with c_f2:
                                _s_big = df_o.get("_模块大类", pd.Series(dtype=str)).dropna().astype(str).str.strip()
                                cat_big_opts = ["全部"] + sorted([x for x in _s_big.unique().tolist() if x and x.lower() not in ("nan", "none", "null")])
                                sel_big_d = st.selectbox("产品大类（T列透视）", cat_big_opts, key="out_d_cat_big", on_change=_reset_out_d)
                            with c_f3:
                                if "_模块小类" in df_o.columns:
                                    if sel_big_d != "全部" and "_模块大类" in df_o.columns:
                                        subs = df_o[df_o["_模块大类"].astype(str).str.strip() == str(sel_big_d).strip()]["_模块小类"].dropna().astype(str).str.strip().unique().tolist()
                                        cat_small_opts = ["全部"] + sorted([x for x in subs if x and str(x).strip().lower() not in ("nan", "none", "null")])
                                    else:
                                        _s_small = df_o["_模块小类"].dropna().astype(str).str.strip()
                                        cat_small_opts = ["全部"] + sorted([x for x in _s_small.unique().tolist() if x and x.lower() not in ("nan", "none", "null")])
                                else:
                                    cat_small_opts = ["全部"]
                                sel_small_d = st.selectbox("产品小类（U列重量）", cat_small_opts, key="out_d_cat_small", on_change=_reset_out_d)
                            with c_f4:
                                if "_模块出库产品" in df_o.columns:
                                    df_prod = df_o.copy()
                                    if sel_big_d != "全部" and "_模块大类" in df_prod.columns:
                                        df_prod = df_prod[df_prod["_模块大类"].astype(str).str.strip() == str(sel_big_d).strip()]
                                    if sel_small_d != "全部" and "_模块小类" in df_prod.columns:
                                        df_prod = df_prod[df_prod["_模块小类"].astype(str).str.strip() == str(sel_small_d).strip()]
                                    _s_prod = df_prod["_模块出库产品"].dropna().astype(str).str.strip()
                                    prod_opts = sorted([x for x in _s_prod.unique().tolist() if x and x.lower() not in ("nan", "none", "null")])
                                else:
                                    prod_opts = []
                                sel_prod_d = st.multiselect("出库产品（I列，可多选）", prod_opts, default=st.session_state.get("out_d_out_prod", []), key="out_d_out_prod", on_change=_reset_out_d)

                            df_m = df_o[(df_o["_年"] == int(y0)) & (df_o["_月"] == int(m0))].copy()
                            if sel_big_d != "全部" and "_模块大类" in df_m.columns:
                                df_m = df_m[df_m["_模块大类"].astype(str).str.strip() == str(sel_big_d).strip()].copy()
                            if sel_small_d != "全部" and "_模块小类" in df_m.columns:
                                df_m = df_m[df_m["_模块小类"].astype(str).str.strip() == str(sel_small_d).strip()].copy()
                            if sel_prod_d and "_模块出库产品" in df_m.columns:
                                sel_prod_norm = [str(x).strip() for x in sel_prod_d if str(x).strip()]
                                if sel_prod_norm:
                                    df_m = df_m[df_m["_模块出库产品"].astype(str).str.strip().isin(sel_prod_norm)].copy()

                            df_m["_日"] = pd.to_numeric(df_m.get("_日", None), errors="coerce")
                            df_m = df_m[df_m["_日"].notna()].copy()
                            df_m["_日"] = df_m["_日"].astype(int)
                            df_m = df_m[df_m["_日"].between(1, 31)].copy()

                            if df_m.empty:
                                st.info(f"未检测到 {y0}年{m0}月 的按日出库数据")
                                st.stop()

                            filter_parts = []
                            if sel_big_d != "全部":
                                filter_parts.append(f"产品大类={sel_big_d}")
                            if sel_small_d != "全部":
                                filter_parts.append(f"产品小类={sel_small_d}")
                            if sel_prod_d:
                                filter_parts.append(f"出库产品={','.join([str(x) for x in sel_prod_d])}")
                            filter_line = "筛选：" + ("；".join(filter_parts) if filter_parts else "无")

                            def _daily_xy(df_scope: pd.DataFrame):
                                if df_scope is None or df_scope.empty:
                                    return [], [], []
                                d2 = df_scope.copy()
                                d2["_日"] = pd.to_numeric(d2.get("_日", None), errors="coerce")
                                d2 = d2[d2["_日"].notna()].copy()
                                if d2.empty:
                                    return [], [], []
                                d2["_日"] = d2["_日"].astype(int)
                                d2 = d2[d2["_日"].between(1, 31)].copy()
                                d2["数量(箱)"] = pd.to_numeric(d2.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                g = d2.groupby("_日", as_index=False)["数量(箱)"].sum().rename(columns={"数量(箱)": "出库箱数"})
                                g["出库箱数"] = pd.to_numeric(g["出库箱数"], errors="coerce").fillna(0.0).round(1)
                                g = g.sort_values("_日").reset_index(drop=True)
                                days = g["_日"].astype(int).tolist()
                                x_labels = [f"{int(d)}日" for d in days]
                                y_vals = g["出库箱数"].astype(float).tolist()
                                return x_labels, y_vals, days

                            def _daily_line(df_scope: pd.DataFrame, title: str, export_key: str):
                                if df_scope is None or df_scope.empty:
                                    st.info("暂无可展示的数据")
                                    return
                                x_labels, y_vals, days = _daily_xy(df_scope)
                                if not days:
                                    st.info("暂无可展示的数据")
                                    return
                                g = pd.DataFrame({"_日": days, "出库箱数": y_vals})

                                fig = px.line(g, x="_日", y="出库箱数", markers=True)
                                fig.update_traces(
                                    line=dict(color="#7C3AED", width=3),
                                    marker=dict(color="#7C3AED", size=8),
                                    mode="lines+markers+text",
                                    text=[f"{float(v):.1f}" for v in y_vals],
                                    textposition="top center",
                                )
                                fig.update_xaxes(
                                    title_text="日期",
                                    tickmode="array",
                                    tickvals=days,
                                    ticktext=x_labels,
                                )
                                fig.update_yaxes(title_text="出库箱数(箱)", tickformat=".1f")
                                fig.update_layout(title=title, height=440, margin=dict(l=10, r=10, t=60, b=10))
                                st.plotly_chart(fig, use_container_width=True)

                                c_p1, c_p2, _ = st.columns([1.6, 2.0, 6.4])
                                with c_p1:
                                    if st.button("生成趋势图图片", key=f"out_d_png_gen_{export_key}"):
                                        with st.spinner("正在生成图片，请稍候…"):
                                            title_lines = [title, filter_line, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                                            _out_d_png_cache[export_key] = {
                                                "bytes": _pil_line_png(x_labels, y_vals, title_lines),
                                                "name": sanitize_filename(f"{title}.png"),
                                            }
                                with c_p2:
                                    if export_key in _out_d_png_cache:
                                        st.download_button(
                                            "下载趋势图图片",
                                            data=_out_d_png_cache[export_key]["bytes"],
                                            file_name=_out_d_png_cache[export_key]["name"],
                                            mime="image/png",
                                            key=f"out_d_png_dl_{export_key}",
                                        )

                            drill_level = int(st.session_state.get("out_d_drill_level", 1) or 1)
                            cnav = st.columns([1, 8])
                            if drill_level > 1:
                                if cnav[0].button("⬅️ 返回", key="out_d_back_btn"):
                                    drill_level -= 1
                                    st.session_state.out_d_drill_level = drill_level
                                    if drill_level == 1:
                                        st.session_state.out_d_selected_prov = None
                                        st.session_state.out_d_selected_dist = None
                                    elif drill_level == 2:
                                        st.session_state.out_d_selected_dist = None
                                    st.rerun()

                            bread = f"🏠 全国（{y0}年{int(m0)}月）"
                            if drill_level >= 2 and st.session_state.out_d_selected_prov:
                                bread += f" > 📍 {st.session_state.out_d_selected_prov}"
                            if drill_level >= 3 and st.session_state.out_d_selected_dist:
                                bread += f" > 🏢 {st.session_state.out_d_selected_dist}"
                            cnav[1].markdown(f"**当前位置**: {bread}")
                            st.caption(filter_line)

                            if drill_level == 1:
                                _daily_line(df_m, f"{y0}年{int(m0)}月 全国按日出库趋势", "nation")

                                _d_sig_sum = 0.0
                                try:
                                    _d_sig_sum = float(pd.to_numeric(df_m.get("数量(箱)", 0), errors="coerce").fillna(0.0).sum())
                                except Exception:
                                    _d_sig_sum = 0.0
                                _prod_norm_key = tuple(sorted([str(x).strip() for x in (sel_prod_d or []) if str(x).strip()]))
                                k_zip = ("out_d_prov_zip", int(y0), int(m0), str(sel_big_d), str(sel_small_d), _prod_norm_key, int(len(df_m)), round(_d_sig_sum, 4))

                                c_z1, c_z2, _ = st.columns([1.9, 2.2, 5.9])
                                with c_z1:
                                    if st.button("生成各省区日趋势图ZIP", key="out_d_zip_gen"):
                                        with st.spinner("正在生成各省区日趋势图ZIP，请稍候…"):
                                            provs = (
                                                df_m["省区"]
                                                .fillna("")
                                                .astype(str)
                                                .str.strip()
                                                .tolist()
                                                if "省区" in df_m.columns
                                                else []
                                            )
                                            provs = sorted([p for p in set(provs) if p and p.lower() not in ("nan", "none", "null")])
                                            buf = io.BytesIO()
                                            with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                                                for p in provs:
                                                    df_p = df_m[df_m["省区"].astype(str).str.strip() == str(p).strip()].copy()
                                                    if df_p.empty:
                                                        continue
                                                    x_labels, y_vals, _days = _daily_xy(df_p)
                                                    if not _days:
                                                        continue
                                                    title_p = f"{y0}年{int(m0)}月 {p} 按日出库趋势"
                                                    title_lines_p = [title_p, filter_line, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                                                    png_b = _pil_line_png(x_labels, y_vals, title_lines_p)
                                                    zf.writestr(f"{sanitize_filename(p)}.png", png_b)
                                            buf.seek(0)
                                            _out_d_zip_cache[k_zip] = {
                                                "bytes": buf.getvalue(),
                                                "name": sanitize_filename(f"{y0}年{int(m0)}月_各省区日趋势图.zip"),
                                            }
                                with c_z2:
                                    if k_zip in _out_d_zip_cache:
                                        st.download_button(
                                            "下载各省区日趋势图ZIP",
                                            data=_out_d_zip_cache[k_zip]["bytes"],
                                            file_name=_out_d_zip_cache[k_zip]["name"],
                                            mime="application/zip",
                                            key="out_d_zip_dl",
                                        )

                                pv = (
                                    df_m.groupby(["省区"], as_index=False)["数量(箱)"]
                                    .sum()
                                    .rename(columns={"数量(箱)": "本月出库箱数"})
                                )
                                pv["本月出库箱数"] = pd.to_numeric(pv["本月出库箱数"], errors="coerce").fillna(0.0).round(1)
                                pv["省区"] = pv["省区"].fillna("").astype(str).str.strip()
                                pv = pv[pv["省区"] != ""].sort_values("本月出库箱数", ascending=False).reset_index(drop=True)

                                gridOptions = {
                                    "columnDefs": [
                                        {"headerName": "省区", "field": "省区", "pinned": "left", "minWidth": 160},
                                        {"headerName": "本月出库箱数", "field": "本月出库箱数", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM_1DP, "minWidth": 160},
                                    ],
                                    "defaultColDef": {"resizable": True, "sortable": True, "filter": True},
                                    "rowSelection": "single",
                                    "pagination": False,
                                }
                                ag = AgGrid(
                                    pv,
                                    gridOptions=gridOptions,
                                    height=520,
                                    width="100%",
                                    data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                                    update_mode=GridUpdateMode.SELECTION_CHANGED,
                                    fit_columns_on_grid_load=True,
                                    allow_unsafe_jscode=True,
                                    theme="streamlit",
                                    key="out_d_prov_ag",
                                )
                                selected_rows = ag.get("selected_rows") if ag else None
                                if selected_rows is not None and len(selected_rows) > 0:
                                    first_row = selected_rows.iloc[0] if isinstance(selected_rows, pd.DataFrame) else selected_rows[0]
                                    selected_name = first_row.get("省区") if isinstance(first_row, dict) else first_row["省区"]
                                    st.session_state.out_d_selected_prov = selected_name
                                    st.session_state.out_d_drill_level = 2
                                    st.session_state.out_d_selected_dist = None
                                    st.rerun()
                            elif drill_level == 2:
                                p = str(st.session_state.get("out_d_selected_prov") or "").strip()
                                df_p = df_m[df_m["省区"].astype(str).str.strip() == p].copy()
                                _daily_line(df_p, f"{y0}年{int(m0)}月 {p} 按日出库趋势", f"prov_{sanitize_filename(p)}")
                                pv = (
                                    df_p.groupby(["经销商名称"], as_index=False)["数量(箱)"]
                                    .sum()
                                    .rename(columns={"数量(箱)": "本月出库箱数"})
                                )
                                pv["本月出库箱数"] = pd.to_numeric(pv["本月出库箱数"], errors="coerce").fillna(0.0).round(1)
                                pv["经销商名称"] = pv["经销商名称"].fillna("").astype(str).str.strip()
                                pv = pv[pv["经销商名称"] != ""].sort_values("本月出库箱数", ascending=False).reset_index(drop=True)

                                gridOptions = {
                                    "columnDefs": [
                                        {"headerName": "经销商名称", "field": "经销商名称", "pinned": "left", "minWidth": 220},
                                        {"headerName": "本月出库箱数", "field": "本月出库箱数", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM_1DP, "minWidth": 160},
                                    ],
                                    "defaultColDef": {"resizable": True, "sortable": True, "filter": True},
                                    "rowSelection": "single",
                                    "pagination": False,
                                }
                                ag = AgGrid(
                                    pv,
                                    gridOptions=gridOptions,
                                    height=520,
                                    width="100%",
                                    data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                                    update_mode=GridUpdateMode.SELECTION_CHANGED,
                                    fit_columns_on_grid_load=True,
                                    allow_unsafe_jscode=True,
                                    theme="streamlit",
                                    key="out_d_dist_ag",
                                )
                                selected_rows = ag.get("selected_rows") if ag else None
                                if selected_rows is not None and len(selected_rows) > 0:
                                    first_row = selected_rows.iloc[0] if isinstance(selected_rows, pd.DataFrame) else selected_rows[0]
                                    selected_name = first_row.get("经销商名称") if isinstance(first_row, dict) else first_row["经销商名称"]
                                    st.session_state.out_d_selected_dist = selected_name
                                    st.session_state.out_d_drill_level = 3
                                    st.rerun()
                            else:
                                p = str(st.session_state.get("out_d_selected_prov") or "").strip()
                                dname = str(st.session_state.get("out_d_selected_dist") or "").strip()
                                df_pd = df_m.copy()
                                if p:
                                    df_pd = df_pd[df_pd["省区"].astype(str).str.strip() == p].copy()
                                if dname:
                                    df_pd = df_pd[df_pd["经销商名称"].astype(str).str.strip() == dname].copy()
                                _daily_line(df_pd, f"{y0}年{int(m0)}月 {p}｜{dname} 按日出库趋势", f"dist_{sanitize_filename(p)}_{sanitize_filename(dname)}")

                            st.stop()

                        df_trend_universe = o_raw.copy()
                        if "客户简称" in df_trend_universe.columns:
                            df_trend_universe["经销商名称"] = df_trend_universe["客户简称"].fillna(df_trend_universe["经销商名称"])
                        df_trend_base = df_trend_universe.copy()
                        if df_trend_base is None or df_trend_base.empty:
                            st.info("暂无可用于月度趋势的数据")
                        else:
                            df_trend_base = df_trend_base[df_trend_base["_年"].notna() & df_trend_base["_月"].notna()].copy()
                            df_trend_base["_年"] = pd.to_numeric(df_trend_base["_年"], errors="coerce").fillna(0).astype(int)
                            df_trend_base["_月"] = pd.to_numeric(df_trend_base["_月"], errors="coerce").fillna(0).astype(int)
                            df_trend_base = df_trend_base[(df_trend_base["_年"] > 0) & df_trend_base["_月"].between(1, 12)].copy()
                            df_trend_base["_ym"] = (df_trend_base["_年"] * 100 + df_trend_base["_月"]).astype(int)
                            df_trend_base["_ym_label"] = (
                                df_trend_base["_年"].astype(str).str[-2:]
                                + "年"
                                + df_trend_base["_月"].astype(str)
                                + "月"
                            )

                            month_map_df = (
                                df_trend_base[["_ym", "_ym_label"]]
                                .drop_duplicates()
                                .sort_values("_ym")
                                .reset_index(drop=True)
                            )
                            ym_to_label = {int(r["_ym"]): str(r["_ym_label"]) for _, r in month_map_df.iterrows()}
                            label_to_ym = {str(r["_ym_label"]): int(r["_ym"]) for _, r in month_map_df.iterrows()}
                            month_labels = month_map_df["_ym_label"].astype(str).tolist()

                            month_labels_no_2512 = [x for x in month_labels if str(x).strip() != "25年12月"]
                            default_months = month_labels_no_2512[-3:] if len(month_labels_no_2512) >= 3 else month_labels_no_2512
                            if "25年12月" in month_labels:
                                default_months = ["25年12月"] + [x for x in default_months if str(x).strip() != "25年12月"]
                            if st.session_state.get("out_m_month_cols"):
                                prev = [x for x in st.session_state.out_m_month_cols if x in month_labels]
                                if prev:
                                    default_months = prev
                            c_m1, c_m2, c_m3, c_m4, c_m5 = st.columns([2.0, 1.25, 1.25, 1.4, 0.8])
                            with c_m1:
                                sel_labels = st.multiselect(
                                    "选择月份（可多选）",
                                    options=month_labels,
                                    default=default_months,
                                    key="out_m_sel_months",
                                )
                                st.session_state.out_m_month_cols = sel_labels

                            with c_m2:
                                _s_big = df_trend_universe.get('_模块大类', pd.Series(dtype=str)).dropna().astype(str).str.strip()
                                cat_big_opts = ['全部'] + sorted([x for x in _s_big.unique().tolist() if x and x.lower() not in ('nan', 'none', 'null')])
                                sel_big = st.selectbox("产品大类（T列透视）", cat_big_opts, key="out_m_cat_big")
                            with c_m3:
                                if '_模块小类' in df_trend_universe.columns:
                                    if sel_big != '全部' and '_模块大类' in df_trend_universe.columns:
                                        subs = df_trend_universe[df_trend_universe['_模块大类'].astype(str).str.strip() == str(sel_big).strip()]['_模块小类'].dropna().astype(str).str.strip().unique().tolist()
                                        cat_small_opts = ['全部'] + sorted([x for x in subs if x and str(x).strip().lower() not in ('nan', 'none', 'null')])
                                    else:
                                        _s_small = df_trend_universe['_模块小类'].dropna().astype(str).str.strip()
                                        cat_small_opts = ['全部'] + sorted([x for x in _s_small.unique().tolist() if x and x.lower() not in ('nan', 'none', 'null')])
                                else:
                                    cat_small_opts = ['全部']
                                sel_small = st.selectbox("产品小类（U列重量）", cat_small_opts, key="out_m_cat_small")

                            with c_m4:
                                if '_模块出库产品' in df_trend_universe.columns:
                                    df_prod = df_trend_universe.copy()
                                    if sel_big != '全部' and '_模块大类' in df_prod.columns:
                                        df_prod = df_prod[df_prod['_模块大类'].astype(str).str.strip() == str(sel_big).strip()]
                                    if sel_small != '全部' and '_模块小类' in df_prod.columns:
                                        df_prod = df_prod[df_prod['_模块小类'].astype(str).str.strip() == str(sel_small).strip()]
                                    _s_prod = df_prod['_模块出库产品'].dropna().astype(str).str.strip()
                                    prod_opts = sorted([x for x in _s_prod.unique().tolist() if x and x.lower() not in ('nan', 'none', 'null')])
                                else:
                                    prod_opts = []
                                sel_prod = st.multiselect("出库产品（I列，可多选）", prod_opts, default=[], key="out_m_out_prod")

                            with c_m5:
                                if st.session_state.get("out_m_drill_level", 1) > 1:
                                    if st.button("⬅️ 返回", key="out_m_back_btn"):
                                        if int(st.session_state.out_m_drill_level) == 2:
                                            st.session_state.out_m_drill_level = 1
                                            st.session_state.out_m_selected_prov = None
                                            st.session_state.out_m_selected_dist = None
                                        else:
                                            st.session_state.out_m_drill_level = 2
                                            st.session_state.out_m_selected_dist = None
                                        st.rerun()

                            sel_labels = sel_labels if sel_labels else default_months
                            sel_yms = [label_to_ym[l] for l in sel_labels if l in label_to_ym]
                            sel_yms = sorted([int(x) for x in sel_yms if x is not None])
                            _available_yms = set([int(x) for x in ym_to_label.keys()])
                            sel_yms_no_dec = [ym for ym in sel_yms if int(ym) != 202512]
                            fixed_trend_yms = [202601, 202602, 202603]
                            if all(int(ym) in _available_yms for ym in fixed_trend_yms):
                                first3_yms = fixed_trend_yms
                            else:
                                first3_yms = [ym for ym in sel_yms_no_dec if ym in _available_yms]
                                if len(first3_yms) > 3:
                                    first3_yms = first3_yms[-3:]
                            sel_yms = ([202512] if 202512 in sel_yms else []) + first3_yms
                            sel_yms = [ym for ym in sel_yms if int(ym) in _available_yms]
                            sel_month_cols = [ym_to_label.get(ym, str(ym)) for ym in sel_yms]
                            first3_cols = [ym_to_label.get(ym, str(ym)) for ym in first3_yms]
                            april_ym = 202604
                            april_col = "4月出库"
                            march_ym = april_ym
                            march_col = april_col
                            today_day = None
                            today_col = "今日出库"
                            try:
                                _ds = df_trend_base[df_trend_base["_ym"] == int(april_ym)].copy()
                                if "_日" in _ds.columns and not _ds.empty:
                                    _dmax = pd.to_numeric(_ds["_日"], errors="coerce").dropna()
                                    if not _dmax.empty:
                                        today_day = int(_dmax.max())
                                        today_col = f"{today_day}日出库"
                            except Exception:
                                today_day = None
                                today_col = "今日出库"
                            last_col = None

                            drill_level = int(st.session_state.get("out_m_drill_level", 1) or 1)
                            view_dim = "省区"
                            group_col = "省区"
                            df_level_base = df_trend_base.copy()
                            df_level = df_trend_base.copy()

                            if drill_level == 2:
                                view_dim = "经销商"
                                group_col = "经销商名称"
                                prov_name = st.session_state.get("out_m_selected_prov")
                                if prov_name:
                                    st.caption(f"当前省区：**{prov_name}**（点击经销商可下钻到门店）")
                                    df_level_base = df_level_base[df_level_base["省区"].astype(str).str.strip() == str(prov_name).strip()].copy()
                                    df_level = df_level[df_level["省区"].astype(str).str.strip() == str(prov_name).strip()].copy()
                            elif drill_level == 3:
                                view_dim = "门店"
                                group_col = "_门店名" if "_门店名" in df_level.columns else None
                                prov_name = st.session_state.get("out_m_selected_prov")
                                dist_name = st.session_state.get("out_m_selected_dist")
                                if prov_name:
                                    df_level_base = df_level_base[df_level_base["省区"].astype(str).str.strip() == str(prov_name).strip()].copy()
                                    df_level = df_level[df_level["省区"].astype(str).str.strip() == str(prov_name).strip()].copy()
                                if dist_name:
                                    df_level_base = df_level_base[df_level_base["经销商名称"].astype(str).str.strip() == str(dist_name).strip()].copy()
                                    df_level = df_level[df_level["经销商名称"].astype(str).str.strip() == str(dist_name).strip()].copy()
                                st.caption(f"当前省区：**{prov_name or '—'}** ｜ 当前经销商：**{dist_name or '—'}**")
                                if group_col is None:
                                    st.info("未检测到门店字段，无法展示门店维度趋势")
                                    df_level = df_level.iloc[0:0].copy()
                                else:
                                    df_level_base = df_level_base[df_level_base[group_col].notna()].copy()
                                    df_level = df_level[df_level[group_col].notna()].copy()

                            if sel_big != '全部' and '_模块大类' in df_level.columns:
                                df_level = df_level[df_level['_模块大类'].astype(str).str.strip() == str(sel_big).strip()].copy()
                            if sel_small != '全部' and '_模块小类' in df_level.columns:
                                df_level = df_level[df_level['_模块小类'].astype(str).str.strip() == str(sel_small).strip()].copy()
                            if sel_prod and '_模块出库产品' in df_level.columns:
                                sel_prod_norm = [str(x).strip() for x in sel_prod if str(x).strip()]
                                if sel_prod_norm:
                                    df_level = df_level[df_level['_模块出库产品'].astype(str).str.strip().isin(sel_prod_norm)].copy()

                            df_level_all = df_level.copy()

                            if (not sel_yms):
                                st.info("请选择月份")
                            else:
                                if df_level.empty:
                                    df_level = pd.DataFrame(columns=[group_col, "_ym", "数量(箱)"])
                                else:
                                    df_level = df_level[df_level["_ym"].isin(sel_yms)].copy()
                                    df_level["数量(箱)"] = pd.to_numeric(df_level.get("数量(箱)", 0), errors="coerce").fillna(0.0)

                                agg = (
                                    df_level
                                    .groupby([group_col, "_ym"], as_index=False)["数量(箱)"]
                                    .sum()
                                    .rename(columns={group_col: view_dim})
                                )
                                pv = agg.pivot(index=view_dim, columns="_ym", values="数量(箱)").fillna(0.0)

                                df_names = df_trend_universe.copy()
                                if drill_level == 2:
                                    p = st.session_state.get("out_m_selected_prov")
                                    if p and ('省区' in df_names.columns):
                                        df_names = df_names[df_names['省区'].astype(str).str.strip() == str(p).strip()].copy()
                                elif drill_level == 3:
                                    p = st.session_state.get("out_m_selected_prov")
                                    d = st.session_state.get("out_m_selected_dist")
                                    if p and ('省区' in df_names.columns):
                                        df_names = df_names[df_names['省区'].astype(str).str.strip() == str(p).strip()].copy()
                                    if d and ('经销商名称' in df_names.columns):
                                        df_names = df_names[df_names['经销商名称'].astype(str).str.strip() == str(d).strip()].copy()

                                if sel_big != '全部' and '_模块大类' in df_names.columns:
                                    df_names = df_names[df_names['_模块大类'].astype(str).str.strip() == str(sel_big).strip()].copy()
                                if sel_small != '全部' and '_模块小类' in df_names.columns:
                                    df_names = df_names[df_names['_模块小类'].astype(str).str.strip() == str(sel_small).strip()].copy()
                                if sel_prod and '_模块出库产品' in df_names.columns:
                                    sel_prod_norm = [str(x).strip() for x in sel_prod if str(x).strip()]
                                    if sel_prod_norm:
                                        df_names = df_names[df_names['_模块出库产品'].astype(str).str.strip().isin(sel_prod_norm)].copy()

                                invalid_names = {'', 'nan', 'none', 'null'}
                                if group_col == "省区":
                                    all_provs = df_names['省区'].dropna().astype(str).str.strip().unique() if '省区' in df_names.columns else []
                                    base_names = sorted([x for x in all_provs if x and x.lower() not in invalid_names])
                                elif group_col == "经销商名称":
                                    all_dists = df_names['经销商名称'].dropna().astype(str).str.strip().unique() if '经销商名称' in df_names.columns else []
                                    base_names = sorted([x for x in all_dists if x and x.lower() not in invalid_names])
                                else:
                                    tmp = df_names[group_col].dropna().astype(str).str.strip().unique() if group_col in df_names.columns else []
                                    base_names = sorted([x for x in tmp if x and x.lower() not in invalid_names])

                                if base_names:
                                    df_base_skeleton = pd.DataFrame({view_dim: base_names})
                                    pv_reset = pv.reset_index()
                                    if view_dim not in pv_reset.columns and len(pv_reset.columns) > 0:
                                        pv_reset.rename(columns={pv_reset.columns[0]: view_dim}, inplace=True)
                                    if view_dim in pv_reset.columns:
                                        pv_reset[view_dim] = pv_reset[view_dim].astype(str).str.strip()
                                        df_base_skeleton[view_dim] = df_base_skeleton[view_dim].astype(str).str.strip()
                                        pv = df_base_skeleton.merge(pv_reset, on=view_dim, how="left").fillna(0.0).set_index(view_dim)
                                    else:
                                        pv = df_base_skeleton.set_index(view_dim)

                                for ym in sel_yms:
                                    if ym not in pv.columns:
                                        pv[ym] = 0.0

                                pv = pv[sel_yms]
                                pv.columns = sel_month_cols
                                pv["_合计"] = pv.sum(axis=1)
                                pv = pv.sort_values("_合计", ascending=False).reset_index()
                                pv.drop(columns=["_合计"], inplace=True, errors="ignore")

                                avg_col = "近三月均出库"
                                if len(first3_cols) >= 1:
                                    pv[avg_col] = pv[first3_cols].mean(axis=1)
                                else:
                                    pv[avg_col] = 0.0
                                diff_col = None
                                scan_yms = [202601, 202602, 202603]
                                scan_avg_col = "近三月均扫码"
                                scan_rate_col = "近三月扫码率"
                                scan_avg_header = "近三月月均扫码（1、2、3）"

                                trend_base_cols = first3_cols if len(first3_cols) >= 1 else sel_month_cols
                                spark_vals = pv[trend_base_cols].values.tolist() if trend_base_cols else [[] for _ in range(len(pv))]
                                pv["_趋势数据"] = [json.dumps([float(x) for x in row]) for row in spark_vals]
                                pv["趋势"] = pv["_趋势数据"]

                                if df_level_all is not None and not df_level_all.empty and "_ym" in df_level_all.columns and group_col in df_level_all.columns:
                                    _ym_num = pd.to_numeric(df_level_all["_ym"], errors="coerce").fillna(0).astype(int)
                                    dm = df_level_all[_ym_num == int(april_ym)].copy()
                                    if not dm.empty:
                                        dm["数量(箱)"] = pd.to_numeric(dm.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                        gm = dm.groupby([group_col], as_index=False)["数量(箱)"].sum().rename(columns={group_col: view_dim, "数量(箱)": april_col})
                                        pv["_k_april"] = pv[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                        gm["_k_april"] = gm[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                        pv = pv.merge(gm[["_k_april", april_col]], on="_k_april", how="left")
                                        pv.drop(columns=["_k_april"], inplace=True, errors="ignore")
                                        if april_col in pv.columns:
                                            pv[april_col] = pd.to_numeric(pv[april_col], errors="coerce").fillna(0.0)
                                        else:
                                            pv[april_col] = 0.0
                                    else:
                                        pv[april_col] = 0.0
                                else:
                                    pv[april_col] = 0.0

                                if today_day is not None and (df_level_all is not None) and (not df_level_all.empty) and "_ym" in df_level_all.columns and group_col in df_level_all.columns and "_日" in df_level_all.columns:
                                    try:
                                        _ym_num2 = pd.to_numeric(df_level_all["_ym"], errors="coerce").fillna(0).astype(int)
                                        ddm = df_level_all[_ym_num2 == int(april_ym)].copy()
                                        if not ddm.empty:
                                            ddm["_日"] = pd.to_numeric(ddm["_日"], errors="coerce")
                                            ddm = ddm[ddm["_日"].notna()].copy()
                                            ddm["_日"] = ddm["_日"].astype(int)
                                            ddm = ddm[ddm["_日"] == int(today_day)].copy()
                                        if not ddm.empty:
                                            ddm["数量(箱)"] = pd.to_numeric(ddm.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                            gd = ddm.groupby([group_col], as_index=False)["数量(箱)"].sum().rename(columns={group_col: view_dim, "数量(箱)": today_col})
                                            pv["_k_today"] = pv[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            gd["_k_today"] = gd[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            pv = pv.merge(gd[["_k_today", today_col]], on="_k_today", how="left")
                                            pv.drop(columns=["_k_today"], inplace=True, errors="ignore")
                                            pv[today_col] = pd.to_numeric(pv.get(today_col, 0), errors="coerce").fillna(0.0)
                                        else:
                                            pv[today_col] = 0.0
                                    except Exception:
                                        pv[today_col] = pd.to_numeric(pv.get(today_col, 0), errors="coerce").fillna(0.0)
                                else:
                                    pv[today_col] = 0.0

                                roll_periods = [
                                    ("26年1-3月", [202601, 202602, 202603]),
                                    ("26年2-4月", [202602, 202603, 202604]),
                                ]

                                def _classify_store_abcd(x):
                                    try:
                                        v = float(x)
                                    except Exception:
                                        v = 0.0
                                    if v >= 4:
                                        return "A"
                                    if 2 <= v < 4:
                                        return "B"
                                    if 1 <= v < 2:
                                        return "C"
                                    return "D"

                                _st_order = {"A": 4, "B": 3, "C": 2, "D": 1}

                                def _store_change(prev_c: str | None, cur_c: str | None):
                                    if not prev_c or not cur_c:
                                        return ""
                                    pv0 = _st_order.get(str(prev_c).strip(), 0)
                                    cv0 = _st_order.get(str(cur_c).strip(), 0)
                                    if pv0 == cv0:
                                        return "持平"
                                    if cv0 > pv0:
                                        return "升级 ⬆️"
                                    return "降级 ⬇️"

                                def _cmp_sign(prev_c: str | None, cur_c: str | None):
                                    if not prev_c or not cur_c:
                                        return None
                                    pv0 = _st_order.get(str(prev_c).strip(), 0)
                                    cv0 = _st_order.get(str(cur_c).strip(), 0)
                                    if pv0 == cv0:
                                        return 0
                                    return 1 if cv0 > pv0 else -1

                                def _trend3_label(c1: str | None, c2: str | None, c3: str | None):
                                    s1 = _cmp_sign(c1, c2)
                                    s2 = _cmp_sign(c2, c3)
                                    if s1 is None or s2 is None:
                                        return ""
                                    if s1 > 0 and s2 > 0:
                                        return "持续升级"
                                    if s1 < 0 and s2 < 0:
                                        return "持续降级"
                                    if s1 > 0 and s2 < 0:
                                        return "先升级后降级"
                                    if s1 < 0 and s2 > 0:
                                        return "先降级后升级"
                                    if s1 == 0 and s2 == 0:
                                        return "持续持平"
                                    if s1 == 0 and s2 > 0:
                                        return "持平升级"
                                    if s1 > 0 and s2 == 0:
                                        return "升级持平"
                                    if s1 == 0 and s2 < 0:
                                        return "持平降级"
                                    if s1 < 0 and s2 == 0:
                                        return "降级持平"
                                    return "其他"

                                if drill_level == 3 and group_col and (df_level_all is not None) and (not df_level_all.empty) and (view_dim in pv.columns):
                                    need_yms = []
                                    for _, yms in roll_periods:
                                        need_yms += list(yms)
                                    need_yms = sorted(set([int(x) for x in need_yms]))
                                    d_roll = df_level_all.copy()
                                    d_roll = d_roll[d_roll["_ym"].isin(need_yms)].copy()
                                    if not d_roll.empty:
                                        d_roll["数量(箱)"] = pd.to_numeric(d_roll.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                        d_roll[group_col] = d_roll[group_col].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                        d_roll = d_roll[d_roll[group_col] != ""].copy()
                                        r_agg = (
                                            d_roll.groupby([group_col, "_ym"], as_index=False)["数量(箱)"]
                                            .sum()
                                            .rename(columns={group_col: view_dim})
                                        )
                                        r_pv = r_agg.pivot(index=view_dim, columns="_ym", values="数量(箱)").fillna(0.0)
                                        for ym in need_yms:
                                            if ym not in r_pv.columns:
                                                r_pv[ym] = 0.0
                                        r_pv = r_pv[need_yms].reset_index()

                                        for p_label, yms in roll_periods:
                                            cols = [int(x) for x in yms]
                                            r_pv[f"{p_label}月均出库"] = r_pv[cols].sum(axis=1) / 3.0
                                            r_pv[f"{p_label}门店类型"] = r_pv[f"{p_label}月均出库"].apply(_classify_store_abcd)

                                        for i in range(1, len(roll_periods)):
                                            prev_label = roll_periods[i - 1][0]
                                            cur_label = roll_periods[i][0]
                                            r_pv[f"{cur_label}变动"] = r_pv.apply(lambda r: _store_change(r.get(f"{prev_label}门店类型"), r.get(f"{cur_label}门店类型")), axis=1)

                                        if len(roll_periods) >= 3:
                                            p1, p2, p3 = roll_periods[-3][0], roll_periods[-2][0], roll_periods[-1][0]
                                            r_pv["近三周期变化"] = r_pv.apply(
                                                lambda r: _trend3_label(
                                                    r.get(f"{p1}门店类型"),
                                                    r.get(f"{p2}门店类型"),
                                                    r.get(f"{p3}门店类型"),
                                                ),
                                                axis=1,
                                            )
                                        else:
                                            r_pv["近三周期变化"] = ""

                                        r_pv[view_dim] = r_pv[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                        pv["_k_store"] = pv[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                        r_pv["_k_store"] = r_pv[view_dim]
                                        keep_cols = ["_k_store"]
                                        for p_label, _ in roll_periods:
                                            keep_cols += [f"{p_label}月均出库", f"{p_label}门店类型"]
                                        for i in range(1, len(roll_periods)):
                                            keep_cols.append(f"{roll_periods[i][0]}变动")
                                        keep_cols.append("近三周期变化")
                                        keep_cols = [c for c in keep_cols if c in r_pv.columns]
                                        r_pv = r_pv[keep_cols].copy()
                                        pv = pv.merge(r_pv, on="_k_store", how="left")
                                        pv.drop(columns=["_k_store"], inplace=True, errors="ignore")
                                        for p_label, _ in roll_periods:
                                            c_avg = f"{p_label}月均出库"
                                            if c_avg in pv.columns:
                                                pv[c_avg] = pd.to_numeric(pv[c_avg], errors="coerce").fillna(0.0).round(1)

                                avg_header = "近三月月均（1、2、3）"

                                if False and drill_level in (1, 2):
                                    pv["1月发货件数"] = 0.0
                                    pv["2月发货件数"] = 0.0
                                    pv["3月发货件数"] = 0.0
                                    pv["4月发货件数"] = 0.0
                                    if df_perf_raw is not None and not getattr(df_perf_raw, "empty", True):
                                        sp = df_perf_raw.copy()
                                        if "客户简称" in sp.columns:
                                            sp["经销商名称"] = sp["客户简称"].fillna(sp["经销商名称"])
                                        for c in ["省区", "经销商名称", "大类", "小类", "小类码", "中类", "重量"]:
                                            if c in sp.columns:
                                                if c == "经销商名称":
                                                    sp[c] = sp[c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                else:
                                                    sp[c] = sp[c].fillna("").astype(str).str.strip()
                                        sp["年份"] = pd.to_numeric(sp.get("年份", 0), errors="coerce").fillna(0).astype(int)
                                        sp["月份"] = pd.to_numeric(sp.get("月份", 0), errors="coerce").fillna(0).astype(int)
                                        sp = sp[(sp["年份"] > 0) & (sp["月份"].between(1, 12))].copy()
                                        sp["_ym"] = (sp["年份"] * 100 + sp["月份"]).astype(int)
                                        ship_yms = [202601, 202602, 202603, 202604]
                                        sp = sp[sp["_ym"].isin(ship_yms)].copy()
                                        sp["发货箱数"] = pd.to_numeric(sp.get("发货箱数", 0), errors="coerce").fillna(0.0)
                                        if sel_big != "全部" and "大类" in sp.columns:
                                            sp = sp[sp["大类"].astype(str).str.strip() == str(sel_big).strip()].copy()
                                        if sel_small != "全部":
                                            _sel_s = str(sel_small).strip()
                                            _m = re.search(r"(\d{3})", _sel_s)
                                            if _m:
                                                _code_i = int(_m.group(1))
                                                _src = "小类码" if "小类码" in sp.columns else ("小类" if "小类" in sp.columns else ("中类" if "中类" in sp.columns else ("重量" if "重量" in sp.columns else None)))
                                                if _src is not None and _src in sp.columns:
                                                    _digits = sp[_src].astype(str).str.extract(r"(\d+)")[0]
                                                    _v = pd.to_numeric(_digits, errors="coerce").fillna(-1).astype(int)
                                                    sp = sp[_v == _code_i].copy()
                                            else:
                                                if "小类码" in sp.columns:
                                                    sp = sp[sp["小类码"].astype(str).str.strip() == _sel_s].copy()
                                                elif "小类" in sp.columns:
                                                    sp = sp[sp["小类"].astype(str).str.strip() == _sel_s].copy()
                                                elif "中类" in sp.columns:
                                                    sp = sp[sp["中类"].astype(str).str.strip() == _sel_s].copy()
                                                elif "重量" in sp.columns:
                                                    sp = sp[sp["重量"].astype(str).str.strip() == _sel_s].copy()
                                        if not sp.empty:
                                            if drill_level == 1 and "省区" in sp.columns and view_dim in pv.columns:
                                                gsp = sp.groupby(["省区", "_ym"], as_index=False).agg({"发货箱数": "sum"})
                                                gsp["_k_prov"] = gsp["省区"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                pv["_k_prov"] = pv[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                ship_cols_qty = ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]
                                                for i, _ym in enumerate(ship_yms):
                                                    _cq = ship_cols_qty[i]
                                                    _sub = gsp[gsp["_ym"] == int(_ym)]
                                                    m_qty = dict(zip(_sub["_k_prov"].tolist(), _sub["发货箱数"].tolist()))
                                                    pv[_cq] = pv["_k_prov"].map(m_qty)
                                                pv.drop(columns=["_k_prov"], inplace=True, errors="ignore")
                                                for c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                                    if c not in pv.columns:
                                                        pv[c] = 0.0
                                                    pv[c] = pd.to_numeric(pv[c], errors="coerce").fillna(0.0)
                                            elif drill_level == 2 and "经销商名称" in sp.columns and view_dim in pv.columns:
                                                _p = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                                if _p and "省区" in sp.columns:
                                                    _p_norm = re.sub(r"\s+", "", str(_p))
                                                    sp = sp[sp["省区"].astype(str).str.replace(r"\s+", "", regex=True) == _p_norm].copy()
                                                dist_map2 = {}
                                                if df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                                    if "经销商全称" in df_stock_raw.columns and "经销商名称" in df_stock_raw.columns:
                                                        _m2 = (
                                                            df_stock_raw[["经销商全称", "经销商名称"]]
                                                            .dropna()
                                                            .astype(str)
                                                            .apply(lambda col: col.str.replace(r"\s+", "", regex=True))
                                                            .drop_duplicates()
                                                        )
                                                        dist_map2 = dict(zip(_m2["经销商全称"].tolist(), _m2["经销商名称"].tolist()))

                                                sp["_k_dist"] = sp["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                if dist_map2:
                                                    sp["_k_dist"] = sp["_k_dist"].map(dist_map2).fillna(sp["_k_dist"])
                                                gsp = sp.groupby(["_k_dist", "_ym"], as_index=False).agg({"发货箱数": "sum"})
                                                pv["_k_dist"] = pv[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                if dist_map2:
                                                    pv["_k_dist"] = pv["_k_dist"].map(dist_map2).fillna(pv["_k_dist"])
                                                ship_cols_qty = ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]
                                                for i, _ym in enumerate(ship_yms):
                                                    _cq = ship_cols_qty[i]
                                                    _sub = gsp[gsp["_ym"] == int(_ym)]
                                                    m_qty = dict(zip(_sub["_k_dist"].tolist(), _sub["发货箱数"].tolist()))
                                                    pv[_cq] = pv["_k_dist"].map(m_qty)
                                                pv.drop(columns=["_k_dist"], inplace=True, errors="ignore")
                                                for c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                                    if c not in pv.columns:
                                                        pv[c] = 0.0
                                                    pv[c] = pd.to_numeric(pv[c], errors="coerce").fillna(0.0)

                                if drill_level in (1, 2) and df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                    _s = df_stock_raw.copy()
                                    if "省区" not in _s.columns and "省区名称" in _s.columns:
                                        _s["省区"] = _s["省区名称"]
                                    for _c in ["省区", "经销商名称", "产品大类", "产品小类", "重量"]:
                                        if _c in _s.columns:
                                            if _c == "经销商名称":
                                                _s[_c] = _s[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            else:
                                                _s[_c] = _s[_c].fillna("").astype(str).str.strip()
                                    if "箱数" in _s.columns:
                                        _s["箱数"] = pd.to_numeric(_s["箱数"], errors="coerce").fillna(0.0)
                                    if drill_level == 2:
                                        _p = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                        if _p and "省区" in _s.columns:
                                            _s = _s[_s["省区"].astype(str).str.strip() == _p].copy()
                                    if sel_big != "全部" and "产品大类" in _s.columns:
                                        _s = _s[_s["产品大类"].astype(str).str.strip() == str(sel_big).strip()].copy()
                                    if sel_small != "全部" and "重量" in _s.columns:
                                        _sel_s = str(sel_small).strip()
                                        _m = re.search(r"(\d{3})", _sel_s)
                                        if _m:
                                            _w = _s["重量"].astype(str).str.extract(r"(\d{3})")[0].fillna("").astype(str)
                                            _s = _s[_w == _m.group(1)].copy()
                                        else:
                                            _s = _s[_s["重量"].astype(str).str.strip() == _sel_s].copy()
                                    if not _s.empty and "箱数" in _s.columns:
                                        if drill_level == 1:
                                            _inv = _s.groupby("省区", as_index=False)["箱数"].sum()
                                            _inv = _inv.rename(columns={"省区": view_dim, "箱数": "库存"})
                                        else:
                                            _inv = _s.groupby("经销商名称", as_index=False)["箱数"].sum()
                                            _inv = _inv.rename(columns={"经销商名称": view_dim, "箱数": "库存"})
                                        if view_dim in pv.columns and not _inv.empty:
                                            pv = pv.merge(_inv, on=view_dim, how="left")
                                            pv["库存"] = pd.to_numeric(pv.get("库存", 0), errors="coerce").fillna(0.0)
                                            if avg_col in pv.columns:
                                                _avg_v = pd.to_numeric(pv[avg_col], errors="coerce").fillna(0.0)
                                                pv["可销月"] = np.where(_avg_v > 0, pv["库存"] / _avg_v, 0.0)
                                                pv["可销月"] = pd.to_numeric(pv.get("可销月", 0), errors="coerce").fillna(0.0).round(1)

                                if df_newcust_raw is not None and not getattr(df_newcust_raw, "empty", True) and "_ym" in df_newcust_raw.columns:
                                    nc = df_newcust_raw.copy()
                                    if "省区" in nc.columns:
                                        nc["省区"] = nc["省区"].fillna("").astype(str).str.strip()
                                    if "经销商名称" in nc.columns:
                                        nc["经销商名称"] = nc["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                    if "门店名称" in nc.columns:
                                        nc["门店名称"] = nc["门店名称"].fillna("").astype(str).str.strip()
                                    if "新客数" in nc.columns:
                                        nc["新客数"] = pd.to_numeric(nc["新客数"], errors="coerce").fillna(0.0)
                                    nc["_ym"] = pd.to_numeric(nc["_ym"], errors="coerce").fillna(0).astype(int)
                                    nc = nc[nc["_ym"].between(200001, 209912)].copy()
                                    if drill_level == 2:
                                        _p = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                        if _p and "省区" in nc.columns:
                                            nc = nc[nc["省区"].astype(str).str.strip() == _p].copy()
                                    elif drill_level == 3:
                                        _p = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                        _d = str(st.session_state.get("out_m_selected_dist") or "").strip()
                                        if _p and "省区" in nc.columns:
                                            nc = nc[nc["省区"].astype(str).str.strip() == _p].copy()
                                        if _d and "经销商名称" in nc.columns:
                                            nc = nc[nc["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True) == _d.replace(" ", "")].copy()
                                    if not nc.empty:
                                        _yms = sorted([int(x) for x in nc["_ym"].dropna().astype(int).tolist() if 200001 <= int(x) <= 209912])
                                        _yy = int(max(_yms) // 100)
                                        latest_ym = int(_yy * 100 + 4)
                                        prev3 = [int(_yy * 100 + 1), int(_yy * 100 + 2), int(_yy * 100 + 3)]
                                        key_col = None
                                        if drill_level == 1:
                                            key_col = "省区"
                                        elif drill_level == 2:
                                            key_col = "经销商名称"
                                        else:
                                            key_col = "门店名称"
                                        if key_col in nc.columns and view_dim in pv.columns:
                                            cur = (
                                                nc[nc["_ym"] == latest_ym]
                                                .groupby(key_col, as_index=False)["新客数"]
                                                .sum()
                                                .rename(columns={key_col: view_dim, "新客数": "4月新客"})
                                            )
                                            p3 = (
                                                nc[nc["_ym"].isin(prev3)]
                                                .groupby(key_col, as_index=False)["新客数"]
                                                .sum()
                                                .rename(columns={key_col: view_dim, "新客数": "近三月新客"})
                                            )
                                            cum = (
                                                nc.groupby(key_col, as_index=False)["新客数"]
                                                .sum()
                                                .rename(columns={key_col: view_dim, "新客数": "整体新客"})
                                            )
                                            pv[view_dim] = pv[view_dim].fillna("").astype(str).str.strip()
                                            cur[view_dim] = cur[view_dim].fillna("").astype(str).str.strip()
                                            p3[view_dim] = p3[view_dim].fillna("").astype(str).str.strip()
                                            cum[view_dim] = cum[view_dim].fillna("").astype(str).str.strip()
                                            pv = pv.merge(cur, on=view_dim, how="left")
                                            pv = pv.merge(p3, on=view_dim, how="left")
                                            pv = pv.merge(cum, on=view_dim, how="left")
                                            for _c in ["4月新客", "近三月新客", "整体新客"]:
                                                if _c in pv.columns:
                                                    pv[_c] = pd.to_numeric(pv[_c], errors="coerce").fillna(0.0)
                                                else:
                                                    pv[_c] = 0.0
                                            pv["累计新客"] = pd.to_numeric(pv.get("整体新客", 0), errors="coerce").fillna(0.0)

                                if df_scan_raw is not None and not getattr(df_scan_raw, "empty", True):
                                    s = df_scan_raw.copy()
                                    for _c in ["省区", "门店名称", "经销商名称", "产品大类", "产品小类"]:
                                        if _c in s.columns:
                                            if _c == "经销商名称":
                                                s[_c] = s[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            else:
                                                s[_c] = s[_c].fillna("").astype(str).str.strip()
                                    if "年份" in s.columns:
                                        s["年份"] = pd.to_numeric(s["年份"], errors="coerce").fillna(0).astype(int)
                                    if "月份" in s.columns:
                                        s["月份"] = pd.to_numeric(s["月份"], errors="coerce").fillna(0).astype(int)
                                    s = s[(s.get("年份", 0) > 0) & (s.get("月份", 0).between(1, 12))].copy()
                                    if drill_level == 2:
                                        _p = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                        if _p and "省区" in s.columns:
                                            s = s[s["省区"].astype(str).str.strip() == _p].copy()
                                    elif drill_level == 3:
                                        _p = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                        _d = str(st.session_state.get("out_m_selected_dist") or "").strip()
                                        if _p and "省区" in s.columns:
                                            s = s[s["省区"].astype(str).str.strip() == _p].copy()
                                        if _d and "经销商名称" in s.columns:
                                            s = s[s["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True) == re.sub(r"\s+", "", _d)].copy()
                                    if sel_big != "全部":
                                        _sb = str(sel_big).strip()
                                        if _sb == "雅系列":
                                            if "产品小类" in s.columns:
                                                s = s[s["产品小类"].astype(str).str.contains(r"(雅赋|雅耀|雅舒|雅护)", regex=True)].copy()
                                            elif "产品大类" in s.columns:
                                                s = s[s["产品大类"].astype(str).str.strip() == _sb].copy()
                                        elif "产品大类" in s.columns:
                                            s = s[s["产品大类"].astype(str).str.strip() == _sb].copy()
                                    if sel_small != "全部" and "产品小类" in s.columns:
                                        _sel_s = str(sel_small).strip()
                                        _m = re.search(r"(\d{3})", _sel_s)
                                        if _m:
                                            _w = s["产品小类"].astype(str).str.extract(r"(\d{3})")[0].fillna("").astype(str)
                                            s = s[_w == _m.group(1)].copy()
                                        else:
                                            s = s[s["产品小类"].astype(str).str.strip() == _sel_s].copy()
                                    if not s.empty and view_dim in pv.columns:
                                        s["_ym"] = (s["年份"] * 100 + s["月份"]).astype(int)
                                        scan_yms = [202601, 202602, 202603]
                                        scan_avg_col = "近三月均扫码"
                                        scan_rate_col = "近三月扫码率"
                                        scan_avg_header = "近三月月均扫码（1、2、3）"
                                        if drill_level == 1:
                                            key_col = "省区"
                                        elif drill_level == 2:
                                            key_col = "经销商名称"
                                        else:
                                            key_col = "门店名称"
                                        if key_col in s.columns:
                                            pv[view_dim] = pv[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            for _ym in scan_yms:
                                                s_m = s[pd.to_numeric(s["_ym"], errors="coerce").fillna(0).astype(int) == int(_ym)].copy()
                                                c_scan = f"_scan_{int(_ym)}"
                                                if not s_m.empty:
                                                    scan_m = (
                                                        s_m.groupby(key_col, as_index=False)
                                                        .size()
                                                        .rename(columns={key_col: view_dim, "size": "_扫码听数"})
                                                    )
                                                    scan_m[c_scan] = pd.to_numeric(scan_m["_扫码听数"], errors="coerce").fillna(0.0) / 6.0
                                                    scan_m = scan_m[[view_dim, c_scan]].copy()
                                                    scan_m[view_dim] = scan_m[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                    pv = pv.merge(scan_m, on=view_dim, how="left")
                                                else:
                                                    pv[c_scan] = 0.0
                                                pv[c_scan] = pd.to_numeric(pv.get(c_scan, 0), errors="coerce").fillna(0.0)

                                            scan_cols = [f"_scan_{int(_ym)}" for _ym in scan_yms]
                                            pv[scan_avg_col] = pv[scan_cols].mean(axis=1)
                                        else:
                                            pv[scan_avg_col] = 0.0

                                        pv[scan_avg_col] = pd.to_numeric(pv.get(scan_avg_col, 0), errors="coerce").fillna(0.0)
                                        if avg_col in pv.columns:
                                            _out_avg = pd.to_numeric(pv[avg_col], errors="coerce").fillna(0.0)
                                            pv[scan_rate_col] = np.where(_out_avg > 0, pv[scan_avg_col] / _out_avg, 0.0)
                                        else:
                                            pv[scan_rate_col] = 0.0
                                        pv[scan_rate_col] = pd.to_numeric(pv.get(scan_rate_col, 0), errors="coerce").fillna(0.0)

                                if drill_level == 3 and store_geo_df is not None and not getattr(store_geo_df, "empty", True) and (view_dim in pv.columns):
                                    try:
                                        pv["_k_store_geo"] = pv[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                        pv = pv.merge(store_geo_df, on="_k_store_geo", how="left")
                                        pv.drop(columns=["_k_store_geo"], inplace=True, errors="ignore")
                                        pv["市"] = pv.get("市", "").fillna("").astype(str).str.strip()
                                        pv["区/县"] = pv.get("区/县", "").fillna("").astype(str).str.strip()
                                        pv["门店状态"] = pv.get("门店状态", "").fillna("").astype(str).str.strip()
                                    except Exception:
                                        pv["市"] = pv.get("市", "").fillna("").astype(str).str.strip()
                                        pv["区/县"] = pv.get("区/县", "").fillna("").astype(str).str.strip()
                                        pv["门店状态"] = pv.get("门店状态", "").fillna("").astype(str).str.strip()

                                region_label = "全国省区"
                                if drill_level == 2:
                                    region_label = str(st.session_state.get("out_m_selected_prov") or "").strip() or "省区"
                                elif drill_level == 3:
                                    region_label = str(st.session_state.get("out_m_selected_dist") or "").strip() or "经销商"

                                prod_txt = "全部"
                                if sel_prod:
                                    prod_norm = [str(x).strip() for x in sel_prod if str(x).strip()]
                                    if prod_norm:
                                        prod_txt = "、".join(prod_norm)

                                filter_line = f"筛选：月份={'、'.join([str(x) for x in sel_month_cols])}｜大类={sel_big}｜小类={sel_small}｜出库产品={prod_txt}"
                                area_line = f"区域：{region_label}"

                                export_cols = [view_dim]
                                for c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                    if c in pv.columns:
                                        export_cols.append(c)
                                export_cols += [c for c in sel_month_cols if c in pv.columns]
                                if avg_col in pv.columns:
                                    export_cols.append(avg_col)
                                if "趋势" in pv.columns:
                                    export_cols.append("趋势")
                                if april_col in pv.columns:
                                    export_cols.append(april_col)
                                if today_col in pv.columns:
                                    export_cols.append(today_col)
                                if drill_level in (1, 2) and "库存" in pv.columns:
                                    export_cols.append("库存")
                                if drill_level in (1, 2) and "可销月" in pv.columns:
                                    export_cols.append("可销月")
                                for _c in ["4月新客", "近三月新客", "整体新客"]:
                                    if _c in pv.columns:
                                        export_cols.append(_c)
                                if scan_avg_col in pv.columns:
                                    export_cols.append(scan_avg_col)
                                if scan_rate_col in pv.columns:
                                    export_cols.append(scan_rate_col)
                                if drill_level == 3:
                                    for p_label, _yms in roll_periods:
                                        for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                            if c in pv.columns:
                                                export_cols.append(c)
                                    for i in range(1, len(roll_periods)):
                                        c = f"{roll_periods[i][0]}变动"
                                        if c in pv.columns:
                                            export_cols.append(c)
                                    if "近三周期变化" in pv.columns:
                                        export_cols.append("近三周期变化")
                                for _c in ["市", "区/县"]:
                                    if _c in pv.columns:
                                        export_cols.append(_c)

                                df_export = pv[export_cols].copy() if export_cols else pv.copy()
                                if len(df_export) > 120:
                                    df_export = df_export.head(120).copy()
                                    area_line = f"{area_line}（仅导出前120行/共{int(len(pv))}行）"
                                try:
                                    total_export = {view_dim: "合计"}
                                    for _c in export_cols:
                                        if _c in (view_dim, "趋势", "趋势类型", "_趋势数据", "可销月", scan_rate_col, "市", "区/县"):
                                            continue
                                        if str(_c).startswith("_"):
                                            continue
                                        if _c in pv.columns:
                                            total_export[_c] = float(pd.to_numeric(pv[_c], errors="coerce").fillna(0.0).sum())
                                    if avg_col in export_cols and avg_col in pv.columns and len(first3_cols) >= 1:
                                        _t3 = [float(pd.to_numeric(pv[c], errors="coerce").fillna(0.0).sum()) for c in first3_cols if c in pv.columns]
                                        total_export[avg_col] = float(np.mean(_t3)) if _t3 else 0.0
                                    if "可销月" in export_cols:
                                        _a = float(total_export.get(avg_col, 0.0) or 0.0)
                                        _k = float(total_export.get("库存", 0.0) or 0.0)
                                        total_export["可销月"] = round((_k / _a), 1) if _a > 0 else 0.0
                                    if scan_avg_col in export_cols:
                                        _t_scan = []
                                        for _ym in scan_yms:
                                            c = f"_scan_{int(_ym)}"
                                            if c in pv.columns:
                                                _t_scan.append(float(pd.to_numeric(pv[c], errors="coerce").fillna(0.0).sum()))
                                        total_export[scan_avg_col] = float(np.mean(_t_scan)) if _t_scan else 0.0
                                    if scan_rate_col in export_cols and scan_avg_col in total_export and avg_col in total_export:
                                        denom = float(total_export.get(avg_col, 0.0) or 0.0)
                                        total_export[scan_rate_col] = (float(total_export.get(scan_avg_col, 0.0) or 0.0) / denom) if denom > 0 else 0.0
                                    if "趋势" in export_cols:
                                        _spark_cols = trend_base_cols if 'trend_base_cols' in locals() and trend_base_cols else first3_cols
                                        _spark_vals = [float(pd.to_numeric(pv[c], errors="coerce").fillna(0.0).sum()) for c in _spark_cols if c in pv.columns]
                                        _spark_json = json.dumps([float(x) for x in _spark_vals]) if _spark_vals else json.dumps([])
                                        total_export["趋势"] = _spark_json
                                    df_export = pd.concat([df_export, pd.DataFrame([total_export])], ignore_index=True)
                                except Exception:
                                    pass
                                col_types = {view_dim: "text"}
                                for c in sel_month_cols:
                                    if c in df_export.columns:
                                        col_types[c] = "num"
                                if avg_col in df_export.columns:
                                    col_types[avg_col] = "num"
                                if march_col in df_export.columns:
                                    col_types[march_col] = "num"
                                if "库存" in df_export.columns:
                                    col_types["库存"] = "num"
                                if "可销月" in df_export.columns:
                                    col_types["可销月"] = "num"
                                for _c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                    if _c in df_export.columns:
                                        col_types[_c] = "num"
                                if scan_avg_col in df_export.columns:
                                    col_types[scan_avg_col] = "num"
                                if scan_rate_col in df_export.columns:
                                    col_types[scan_rate_col] = "pct"
                                for _c in ["市", "区/县"]:
                                    if _c in df_export.columns:
                                        col_types[_c] = "text"
                                for _c in ["4月新客", "近三月新客", "整体新客"]:
                                    if _c in df_export.columns:
                                        col_types[_c] = "num"
                                if "趋势" in df_export.columns:
                                    col_types["趋势"] = "spark"
                                if drill_level == 3:
                                    for p_label, _yms in roll_periods:
                                        c_avg = f"{p_label}月均出库"
                                        c_type = f"{p_label}门店类型"
                                        if c_avg in df_export.columns:
                                            col_types[c_avg] = "num"
                                        if c_type in df_export.columns:
                                            col_types[c_type] = "tag"
                                    if "近三周期变化" in df_export.columns:
                                        col_types["近三周期变化"] = "tag"

                                export_title_lines = [
                                    f"月度出库趋势表 - {region_label}",
                                    filter_line,
                                    area_line,
                                    f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                ]
                                export_id = f"out_m_export_{drill_level}"
                                def _build_current_excel_df():
                                    cols = [view_dim]
                                    for _c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                        if _c in pv.columns:
                                            cols.append(_c)
                                    cols += [c for c in sel_month_cols if c in pv.columns]
                                    if avg_col in pv.columns:
                                        cols.append(avg_col)
                                    if drill_level in (1, 2) and "库存" in pv.columns:
                                        cols.append("库存")
                                    if drill_level in (1, 2) and "可销月" in pv.columns:
                                        cols.append("可销月")
                                    for _c in ["4月新客", "近三月新客", "整体新客"]:
                                        if _c in pv.columns:
                                            cols.append(_c)
                                    if diff_col in pv.columns:
                                        cols.append(diff_col)
                                    if march_col in pv.columns:
                                        cols.append(march_col)
                                    if today_col in pv.columns:
                                        cols.append(today_col)
                                    if last_col and (last_col in pv.columns) and (last_col not in cols):
                                        cols.append(last_col)
                                    if "完成率" in pv.columns:
                                        cols.append("完成率")
                                    if scan_avg_col in pv.columns:
                                        cols.append(scan_avg_col)
                                    if scan_rate_col in pv.columns:
                                        cols.append(scan_rate_col)
                                    for _c in ["市", "区/县"]:
                                        if _c in pv.columns:
                                            cols.append(_c)
                                    df_x = pv[cols].copy() if cols else pv.copy()
                                    if drill_level in (1, 2) and df_perf_raw is not None and not getattr(df_perf_raw, "empty", True):
                                        try:
                                            sp = df_perf_raw.copy()
                                            if "客户简称" in sp.columns:
                                                sp["经销商名称"] = sp["客户简称"].fillna(sp["经销商名称"])
                                            for c in ["省区", "经销商名称", "大类", "小类", "小类码", "中类", "重量"]:
                                                if c in sp.columns:
                                                    if c == "经销商名称":
                                                        sp[c] = sp[c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                    else:
                                                        sp[c] = sp[c].fillna("").astype(str).str.strip()
                                            sp["年份"] = pd.to_numeric(sp.get("年份", 0), errors="coerce").fillna(0).astype(int)
                                            sp["月份"] = pd.to_numeric(sp.get("月份", 0), errors="coerce").fillna(0).astype(int)
                                            sp = sp[(sp["年份"] > 0) & (sp["月份"].between(1, 12))].copy()
                                            sp["_ym"] = (sp["年份"] * 100 + sp["月份"]).astype(int)
                                            ship_yms = [202601, 202602, 202603, 202604]
                                            sp = sp[sp["_ym"].isin(ship_yms)].copy()
                                            sp["发货箱数"] = pd.to_numeric(sp.get("发货箱数", 0), errors="coerce").fillna(0.0)

                                            if sel_big != "全部" and "大类" in sp.columns:
                                                sp = sp[sp["大类"].astype(str).str.strip() == str(sel_big).strip()].copy()
                                            if sel_small != "全部":
                                                _sel_s = str(sel_small).strip()
                                                _m = re.search(r"(\d{3})", _sel_s)
                                                if _m:
                                                    _code_i = int(_m.group(1))
                                                    _src = "小类码" if "小类码" in sp.columns else ("小类" if "小类" in sp.columns else ("中类" if "中类" in sp.columns else ("重量" if "重量" in sp.columns else None)))
                                                    if _src is not None and _src in sp.columns:
                                                        _digits = sp[_src].astype(str).str.extract(r"(\d+)")[0]
                                                        _v = pd.to_numeric(_digits, errors="coerce").fillna(-1).astype(int)
                                                        sp = sp[_v == _code_i].copy()
                                                else:
                                                    if "小类码" in sp.columns:
                                                        sp = sp[sp["小类码"].astype(str).str.strip() == _sel_s].copy()
                                                    elif "小类" in sp.columns:
                                                        sp = sp[sp["小类"].astype(str).str.strip() == _sel_s].copy()
                                                    elif "中类" in sp.columns:
                                                        sp = sp[sp["中类"].astype(str).str.strip() == _sel_s].copy()
                                                    elif "重量" in sp.columns:
                                                        sp = sp[sp["重量"].astype(str).str.strip() == _sel_s].copy()

                                            if drill_level == 2:
                                                _p = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                                if _p and "省区" in sp.columns:
                                                    _p_norm = re.sub(r"\s+", "", str(_p))
                                                    sp = sp[sp["省区"].astype(str).str.replace(r"\s+", "", regex=True) == _p_norm].copy()

                                            if not sp.empty and view_dim in df_x.columns:
                                                if drill_level == 1 and "省区" in sp.columns:
                                                    sp["_k"] = sp["省区"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                    df_x["_k"] = df_x[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                else:
                                                    sp["_k"] = sp["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                    df_x["_k"] = df_x[view_dim].fillna("").astype(str).str.replace(r"\s+", "", regex=True)

                                                g = sp.groupby(["_k", "_ym"], as_index=False).agg({"发货箱数": "sum"})
                                                ship = g.pivot(index="_k", columns="_ym", values="发货箱数").fillna(0.0)
                                                ship_cols_qty = ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]
                                                for _ym in ship_yms:
                                                    if int(_ym) not in ship.columns:
                                                        ship[int(_ym)] = 0.0
                                                ship_df = ship[ship_yms].rename(columns={int(_ym): ship_cols_qty[i] for i, _ym in enumerate(ship_yms)}).reset_index().fillna(0.0)

                                                df_x = df_x.merge(ship_df, on="_k", how="left", suffixes=("", "_y"))
                                                for _c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                                    _cy = f"{_c}_y"
                                                    if _cy in df_x.columns:
                                                        if _c in df_x.columns:
                                                            _right = pd.to_numeric(df_x[_cy], errors="coerce")
                                                            _left = pd.to_numeric(df_x[_c], errors="coerce")
                                                            _mask = _right.notna() & (_right != 0)
                                                            df_x[_c] = _left.where(~_mask, _right)
                                                            df_x.drop(columns=[_cy], inplace=True, errors="ignore")
                                                        else:
                                                            df_x.rename(columns={_cy: _c}, inplace=True)
                                                    if _c in df_x.columns:
                                                        df_x[_c] = pd.to_numeric(df_x[_c], errors="coerce").fillna(0.0)
                                                df_x.drop(columns=["_k"], inplace=True, errors="ignore")
                                        except Exception:
                                            df_x.drop(columns=["_k"], inplace=True, errors="ignore")
                                    if avg_col in df_x.columns:
                                        df_x.rename(columns={avg_col: avg_header}, inplace=True)
                                    if diff_col in df_x.columns:
                                        feb = ym_to_label.get(202602)
                                        if feb and feb in df_x.columns:
                                            ordered_cols = [c for c in df_x.columns if c != diff_col]
                                            idx = ordered_cols.index(feb) + 1
                                            ordered_cols.insert(idx, diff_col)
                                            df_x = df_x[ordered_cols].copy()

                                    months = [c for c in first3_cols if c in df_x.columns]
                                    if last_col and last_col in df_x.columns and last_col not in months:
                                        months.append(last_col)

                                    anchor = months[-1] if months else None
                                    if months and first3_cols:
                                        m3 = [c for c in first3_cols if c in df_x.columns]
                                        if m3:
                                            anchor = m3[-1]

                                    before = []
                                    after = []
                                    if anchor and anchor in months:
                                        idx = months.index(anchor)
                                        before = months[: idx + 1]
                                        after = months[idx + 1 :]
                                    else:
                                        before = months

                                    ordered = [view_dim]
                                    for _c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                        if _c in df_x.columns:
                                            ordered.append(_c)
                                    ordered += before
                                    if avg_header in df_x.columns:
                                        ordered.append(avg_header)
                                    if "库存" in df_x.columns:
                                        ordered.append("库存")
                                    if "可销月" in df_x.columns:
                                        ordered.append("可销月")
                                    for _c in ["4月新客", "近三月新客", "整体新客"]:
                                        if _c in df_x.columns:
                                            ordered.append(_c)
                                    if march_col in df_x.columns:
                                        ordered.append(march_col)
                                    if today_col in df_x.columns:
                                        ordered.append(today_col)
                                    ordered += after
                                    if scan_avg_col in df_x.columns:
                                        ordered.append(scan_avg_col)
                                    if scan_rate_col in df_x.columns:
                                        ordered.append(scan_rate_col)
                                    ordered = [c for c in ordered if c in df_x.columns]
                                    if ordered:
                                        df_x = df_x[ordered].copy()
                                    if diff_col in df_x.columns:
                                        feb = ym_to_label.get(202602)
                                        if feb and feb in df_x.columns:
                                            ordered_cols = [c for c in df_x.columns if c != diff_col]
                                            idx = ordered_cols.index(feb) + 1
                                            ordered_cols.insert(idx, diff_col)
                                            df_x = df_x[ordered_cols].copy()
                                    return df_x

                                def _build_dist_detail_df(all_provinces: bool):
                                    d = df_trend_base.copy()
                                    prov_sel = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                    if not all_provinces and prov_sel:
                                        d = d[d["省区"].astype(str).str.strip() == prov_sel].copy()
                                    if sel_big != '全部' and '_模块大类' in d.columns:
                                        d = d[d['_模块大类'].astype(str).str.strip() == str(sel_big).strip()].copy()
                                    if sel_small != '全部' and '_模块小类' in d.columns:
                                        d = d[d['_模块小类'].astype(str).str.strip() == str(sel_small).strip()].copy()
                                    if sel_prod and '_模块出库产品' in d.columns:
                                        sel_prod_norm = [str(x).strip() for x in sel_prod if str(x).strip()]
                                        if sel_prod_norm:
                                            d = d[d['_模块出库产品'].astype(str).str.strip().isin(sel_prod_norm)].copy()
                                    d_all = d.copy()
                                    d = d[d["_ym"].isin(sel_yms)].copy()
                                    if d.empty:
                                        dn = df_trend_universe.copy()
                                        if not all_provinces and prov_sel and "省区" in dn.columns:
                                            dn = dn[dn["省区"].astype(str).str.strip() == prov_sel].copy()
                                        if sel_big != '全部' and '_模块大类' in dn.columns:
                                            dn = dn[dn['_模块大类'].astype(str).str.strip() == str(sel_big).strip()].copy()
                                        if sel_small != '全部' and '_模块小类' in dn.columns:
                                            dn = dn[dn['_模块小类'].astype(str).str.strip() == str(sel_small).strip()].copy()
                                        if sel_prod and '_模块出库产品' in dn.columns:
                                            sel_prod_norm = [str(x).strip() for x in sel_prod if str(x).strip()]
                                            if sel_prod_norm:
                                                dn = dn[dn['_模块出库产品'].astype(str).str.strip().isin(sel_prod_norm)].copy()
                                        if ("省区" in dn.columns) and ("经销商名称" in dn.columns):
                                            dn["省区"] = dn["省区"].fillna("").astype(str).str.strip()
                                            dn["经销商名称"] = dn["经销商名称"].fillna("").astype(str).str.strip()
                                            dn = dn[(dn["省区"] != "") & (dn["经销商名称"] != "")].copy()
                                            pv_s = (
                                                dn[["省区", "经销商名称"]]
                                                .drop_duplicates()
                                                .rename(columns={"经销商名称": "经销商"})
                                                .sort_values(["省区", "经销商"], ascending=[True, True])
                                                .reset_index(drop=True)
                                            )
                                        else:
                                            pv_s = pd.DataFrame(columns=["省区", "经销商"])
                                        for c in sel_month_cols:
                                            if c not in pv_s.columns:
                                                pv_s[c] = 0.0
                                    else:
                                        d["数量(箱)"] = pd.to_numeric(d.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                        d["省区"] = d["省区"].fillna("").astype(str).str.strip()
                                        d["经销商名称"] = d["经销商名称"].fillna("").astype(str).str.strip()
                                        d = d[(d["省区"] != "") & (d["经销商名称"] != "")].copy()
                                        g = d.groupby(["省区", "经销商名称", "_ym"], as_index=False)["数量(箱)"].sum()
                                        pv_s = g.pivot(index=["省区", "经销商名称"], columns="_ym", values="数量(箱)").fillna(0.0)
                                        for ym in sel_yms:
                                            if ym not in pv_s.columns:
                                                pv_s[ym] = 0.0
                                        pv_s = pv_s[sel_yms]
                                        pv_s.columns = sel_month_cols
                                        pv_s["_合计"] = pv_s.sum(axis=1)
                                        pv_s = pv_s.reset_index().rename(columns={"经销商名称": "经销商"})
                                        pv_s = pv_s.sort_values("_合计", ascending=False).reset_index(drop=True)
                                        pv_s.drop(columns=["_合计"], inplace=True, errors="ignore")

                                    # 不再提前初始化发货列，避免merge产生_x/_y后缀导致数据丢失
                                    if df_perf_raw is not None and not getattr(df_perf_raw, "empty", True):
                                        sp = df_perf_raw.copy()
                                        if "客户简称" in sp.columns:
                                            sp["经销商名称"] = sp["客户简称"].fillna(sp["经销商名称"])
                                        for c in ["省区", "经销商名称", "大类", "小类", "小类码", "中类", "重量"]:
                                            if c in sp.columns:
                                                if c == "经销商名称":
                                                    sp[c] = sp[c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                else:
                                                    sp[c] = sp[c].fillna("").astype(str).str.strip()
                                        sp["年份"] = pd.to_numeric(sp.get("年份", 0), errors="coerce").fillna(0).astype(int)
                                        sp["月份"] = pd.to_numeric(sp.get("月份", 0), errors="coerce").fillna(0).astype(int)
                                        sp = sp[(sp["年份"] > 0) & (sp["月份"].between(1, 12))].copy()
                                        sp["_ym"] = (sp["年份"] * 100 + sp["月份"]).astype(int)
                                        ship_yms = [202601, 202602, 202603, 202604]
                                        sp = sp[sp["_ym"].isin(ship_yms)].copy()
                                        sp["发货箱数"] = pd.to_numeric(sp.get("发货箱数", 0), errors="coerce").fillna(0.0)
                                        if not all_provinces and prov_sel and "省区" in sp.columns:
                                            sp = sp[sp["省区"].astype(str).str.strip() == prov_sel].copy()
                                        if sel_big != "全部" and "大类" in sp.columns:
                                            sp = sp[sp["大类"].astype(str).str.strip() == str(sel_big).strip()].copy()
                                        if sel_small != "全部":
                                            _sel_s = str(sel_small).strip()
                                            _m = re.search(r"(\d{3})", _sel_s)
                                            if _m:
                                                _code_i = int(_m.group(1))
                                                _src = "小类码" if "小类码" in sp.columns else ("小类" if "小类" in sp.columns else ("中类" if "中类" in sp.columns else ("重量" if "重量" in sp.columns else None)))
                                                if _src is not None and _src in sp.columns:
                                                    _digits = sp[_src].astype(str).str.extract(r"(\d+)")[0]
                                                    _v = pd.to_numeric(_digits, errors="coerce").fillna(-1).astype(int)
                                                    sp = sp[_v == _code_i].copy()
                                            else:
                                                if "小类码" in sp.columns:
                                                    sp = sp[sp["小类码"].astype(str).str.strip() == _sel_s].copy()
                                                elif "小类" in sp.columns:
                                                    sp = sp[sp["小类"].astype(str).str.strip() == _sel_s].copy()
                                                elif "中类" in sp.columns:
                                                    sp = sp[sp["中类"].astype(str).str.strip() == _sel_s].copy()
                                                elif "重量" in sp.columns:
                                                    sp = sp[sp["重量"].astype(str).str.strip() == _sel_s].copy()
                                        if not sp.empty and ("省区" in sp.columns) and ("经销商名称" in sp.columns):
                                            dist_map2 = {}
                                            if df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                                if "经销商全称" in df_stock_raw.columns and "经销商名称" in df_stock_raw.columns:
                                                    _m2 = (
                                                        df_stock_raw[["经销商全称", "经销商名称"]]
                                                        .dropna()
                                                        .astype(str)
                                                        .apply(lambda col: col.str.replace(r"\s+", "", regex=True))
                                                        .drop_duplicates()
                                                    )
                                                    dist_map2 = dict(zip(_m2["经销商全称"].tolist(), _m2["经销商名称"].tolist()))
                                            sp["_k_dist"] = sp["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map2:
                                                sp["_k_dist"] = sp["_k_dist"].map(dist_map2).fillna(sp["_k_dist"])
                                            # 统一匹配口径：不再依赖省区匹配（避免“广东”vs“广东省”不一致），仅使用经销商名称匹配
                                            gsp = sp.groupby(["_k_dist", "_ym"], as_index=False).agg({"发货箱数": "sum"})
                                            pv_s["_k_dist"] = pv_s["经销商"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map2:
                                                pv_s["_k_dist"] = pv_s["_k_dist"].map(dist_map2).fillna(pv_s["_k_dist"])
                                            ship_cols_qty = ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]
                                            qty = gsp.pivot(index=["_k_dist"], columns="_ym", values="发货箱数").fillna(0.0)
                                            for _ym in ship_yms:
                                                if int(_ym) not in qty.columns:
                                                    qty[int(_ym)] = 0.0
                                            qty = qty[ship_yms].rename(columns={int(_ym): ship_cols_qty[i] for i, _ym in enumerate(ship_yms)}).reset_index()
                                            pv_s = pv_s.merge(qty, on=["_k_dist"], how="left")
                                            pv_s.drop(columns=["_k_dist"], inplace=True, errors="ignore")
                                    
                                    if False:
                                        for c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                            if c not in pv_s.columns:
                                                pv_s[c] = 0.0
                                            pv_s[c] = pd.to_numeric(pv_s[c], errors="coerce").fillna(0.0)

                                    tcols = [c for c in first3_cols if c in pv_s.columns]
                                    if tcols:
                                        pv_s[avg_header] = pv_s[tcols].mean(axis=1)
                                    else:
                                        pv_s[avg_header] = 0.0
                                    pv_s[march_col] = 0.0
                                    try:
                                        dm = d_all[d_all["_ym"] == int(march_ym)].copy()
                                        if not dm.empty:
                                            dm["数量(箱)"] = pd.to_numeric(dm.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                            dm["省区"] = dm["省区"].fillna("").astype(str).str.strip()
                                            dm["经销商名称"] = dm["经销商名称"].fillna("").astype(str).str.strip()
                                            dm[store_col] = dm[store_col].fillna("").astype(str).str.strip()
                                            gm = (
                                                dm.groupby(["省区", "经销商名称", store_col], as_index=False)["数量(箱)"]
                                                .sum()
                                                .rename(columns={"经销商名称": "经销商", store_col: "门店", "数量(箱)": march_col})
                                            )
                                            pv_s = pv_s.merge(gm, on=["省区", "经销商", "门店"], how="left", suffixes=("", "_y"))
                                            if f"{march_col}_y" in pv_s.columns:
                                                pv_s.drop(columns=[march_col], inplace=True, errors="ignore")
                                                pv_s.rename(columns={f"{march_col}_y": march_col}, inplace=True)
                                            pv_s[march_col] = pd.to_numeric(pv_s.get(march_col, 0), errors="coerce").fillna(0.0)
                                    except Exception:
                                        pv_s[march_col] = pd.to_numeric(pv_s.get(march_col, 0), errors="coerce").fillna(0.0)

                                    pv_s[march_col] = 0.0
                                    try:
                                        dm = d_all[d_all["_ym"] == int(march_ym)].copy()
                                        if not dm.empty:
                                            dm["数量(箱)"] = pd.to_numeric(dm.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                            dm["省区"] = dm["省区"].fillna("").astype(str).str.strip()
                                            dm["经销商名称"] = dm["经销商名称"].fillna("").astype(str).str.strip()
                                            gm = dm.groupby(["省区", "经销商名称"], as_index=False)["数量(箱)"].sum().rename(columns={"经销商名称": "经销商", "数量(箱)": march_col})
                                            pv_s = pv_s.merge(gm, on=["省区", "经销商"], how="left", suffixes=("", "_y"))
                                            if f"{march_col}_y" in pv_s.columns:
                                                pv_s.drop(columns=[march_col], inplace=True, errors="ignore")
                                                pv_s.rename(columns={f"{march_col}_y": march_col}, inplace=True)
                                            pv_s[march_col] = pd.to_numeric(pv_s.get(march_col, 0), errors="coerce").fillna(0.0)
                                    except Exception:
                                        pv_s[march_col] = pd.to_numeric(pv_s.get(march_col, 0), errors="coerce").fillna(0.0)

                                    pv_s[today_col] = 0.0
                                    if today_day is not None:
                                        try:
                                            ddm = d_all[d_all["_ym"] == int(march_ym)].copy()
                                            if not ddm.empty and "_日" in ddm.columns:
                                                ddm["_日"] = pd.to_numeric(ddm["_日"], errors="coerce")
                                                ddm = ddm[ddm["_日"].notna()].copy()
                                                ddm["_日"] = ddm["_日"].astype(int)
                                                ddm = ddm[ddm["_日"] == int(today_day)].copy()
                                            if not ddm.empty:
                                                ddm["数量(箱)"] = pd.to_numeric(ddm.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                                ddm["省区"] = ddm["省区"].fillna("").astype(str).str.strip()
                                                ddm["经销商名称"] = ddm["经销商名称"].fillna("").astype(str).str.strip()
                                                gd = ddm.groupby(["省区", "经销商名称"], as_index=False)["数量(箱)"].sum().rename(columns={"经销商名称": "经销商", "数量(箱)": today_col})
                                                pv_s = pv_s.merge(gd, on=["省区", "经销商"], how="left", suffixes=("", "_y"))
                                                if f"{today_col}_y" in pv_s.columns:
                                                    pv_s.drop(columns=[today_col], inplace=True, errors="ignore")
                                                    pv_s.rename(columns={f"{today_col}_y": today_col}, inplace=True)
                                                pv_s[today_col] = pd.to_numeric(pv_s.get(today_col, 0), errors="coerce").fillna(0.0)
                                        except Exception:
                                            pv_s[today_col] = pd.to_numeric(pv_s.get(today_col, 0), errors="coerce").fillna(0.0)

                                    pv_s["库存"] = 0.0
                                    if df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                        _s = df_stock_raw.copy()
                                        if "省区" not in _s.columns and "省区名称" in _s.columns:
                                            _s["省区"] = _s["省区名称"]
                                        for _c in ["省区", "经销商名称", "产品大类", "产品小类", "重量"]:
                                            if _c in _s.columns:
                                                if _c == "经销商名称":
                                                    _s[_c] = _s[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                else:
                                                    _s[_c] = _s[_c].fillna("").astype(str).str.strip()
                                        if "箱数" in _s.columns:
                                            _s["箱数"] = pd.to_numeric(_s["箱数"], errors="coerce").fillna(0.0)
                                        if not all_provinces and prov_sel and "省区" in _s.columns:
                                            _s = _s[_s["省区"].astype(str).str.strip() == prov_sel].copy()
                                        if sel_big != "全部" and "产品大类" in _s.columns:
                                            _s = _s[_s["产品大类"].astype(str).str.strip() == str(sel_big).strip()].copy()
                                        if sel_small != "全部" and "重量" in _s.columns:
                                            _sel_s = str(sel_small).strip()
                                            _m = re.search(r"(\d{3})", _sel_s)
                                            if _m:
                                                _w = _s["重量"].astype(str).str.extract(r"(\d{3})")[0].fillna("").astype(str)
                                                _s = _s[_w == _m.group(1)].copy()
                                            else:
                                                _s = _s[_s["重量"].astype(str).str.strip() == _sel_s].copy()
                                        if not _s.empty and "箱数" in _s.columns:
                                            _inv = _s.groupby(["省区", "经销商名称"], as_index=False)["箱数"].sum()
                                            _inv = _inv.rename(columns={"经销商名称": "经销商", "箱数": "库存"})
                                            pv_s = pv_s.merge(_inv, on=["省区", "经销商"], how="left", suffixes=("", "_y"))
                                            if "库存_y" in pv_s.columns:
                                                pv_s.drop(columns=["库存"], inplace=True, errors="ignore")
                                                pv_s.rename(columns={"库存_y": "库存"}, inplace=True)
                                            pv_s["库存"] = pd.to_numeric(pv_s.get("库存", 0), errors="coerce").fillna(0.0)

                                    pv_s["可销月"] = 0.0
                                    if avg_header in pv_s.columns and "库存" in pv_s.columns:
                                        _avg_v = pd.to_numeric(pv_s[avg_header], errors="coerce").fillna(0.0)
                                        pv_s["可销月"] = np.where(_avg_v > 0, pv_s["库存"] / _avg_v, 0.0)
                                        pv_s["可销月"] = pd.to_numeric(pv_s.get("可销月", 0), errors="coerce").fillna(0.0).round(1)

                                    if df_newcust_raw is not None and not getattr(df_newcust_raw, "empty", True) and "_ym" in df_newcust_raw.columns:
                                        nc = df_newcust_raw.copy()
                                        if "省区" in nc.columns:
                                            nc["省区"] = nc["省区"].fillna("").astype(str).str.strip()
                                        if "经销商名称" in nc.columns:
                                            nc["经销商名称"] = nc["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)

                                        dist_map = {}
                                        if df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                            if "经销商全称" in df_stock_raw.columns and "经销商名称" in df_stock_raw.columns:
                                                _m = (
                                                    df_stock_raw[["经销商全称", "经销商名称"]]
                                                    .dropna()
                                                    .astype(str)
                                                    .apply(lambda col: col.str.replace(r"\s+", "", regex=True))
                                                    .drop_duplicates()
                                                )
                                                dist_map = dict(zip(_m["经销商全称"].tolist(), _m["经销商名称"].tolist()))

                                        if "新客数" in nc.columns:
                                            nc["新客数"] = pd.to_numeric(nc["新客数"], errors="coerce").fillna(0.0)
                                        nc["_ym"] = pd.to_numeric(nc["_ym"], errors="coerce").fillna(0).astype(int)
                                        nc = nc[nc["_ym"].between(200001, 209912)].copy()
                                        if not all_provinces and prov_sel and "省区" in nc.columns:
                                            nc = nc[nc["省区"].astype(str).str.strip() == prov_sel].copy()
                                        if not nc.empty:
                                            _yms = sorted([int(x) for x in nc["_ym"].dropna().astype(int).tolist() if 200001 <= int(x) <= 209912])
                                            _apr_yms = [x for x in _yms if int(x) % 100 == 4]
                                            latest_ym = int(_apr_yms[-1] if _apr_yms else _yms[-1])
                                            _yy = int(latest_ym // 100)
                                            prev3 = [int(_yy * 100 + 1), int(_yy * 100 + 2), int(_yy * 100 + 3)]

                                            pv_s["省区"] = pv_s["省区"].fillna("").astype(str).str.strip()
                                            pv_s["经销商"] = pv_s["经销商"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            pv_s["_经销商_key"] = pv_s["经销商"]
                                            if dist_map:
                                                pv_s["_经销商_key"] = pv_s["_经销商_key"].map(dist_map).fillna(pv_s["_经销商_key"])

                                            nc["_经销商_key"] = nc["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map:
                                                nc["_经销商_key"] = nc["_经销商_key"].map(dist_map).fillna(nc["_经销商_key"])

                                            _yms = sorted([int(x) for x in nc["_ym"].dropna().astype(int).tolist() if 200001 <= int(x) <= 209912])
                                            _yy = int(max(_yms) // 100)
                                            latest_ym = int(_yy * 100 + 4)
                                            prev3 = [int(_yy * 100 + 1), int(_yy * 100 + 2), int(_yy * 100 + 3)]

                                            cur = (
                                                nc[nc["_ym"] == latest_ym]
                                                .groupby(["省区", "_经销商_key"], as_index=False)["新客数"]
                                                .sum()
                                                .rename(columns={"新客数": "4月新客"})
                                            )
                                            p3 = (
                                                nc[nc["_ym"].isin(prev3)]
                                                .groupby(["省区", "_经销商_key"], as_index=False)["新客数"]
                                                .sum()
                                                .rename(columns={"新客数": "近三月新客"})
                                            )
                                            cum = (
                                                nc.groupby(["省区", "_经销商_key"], as_index=False)["新客数"]
                                                .sum()
                                                .rename(columns={"新客数": "整体新客"})
                                            )
                                            for _df in (cur, p3, cum):
                                                _df["省区"] = _df["省区"].fillna("").astype(str).str.strip()

                                            pv_s = pv_s.merge(cur, on=["省区", "_经销商_key"], how="left")
                                            pv_s = pv_s.merge(p3, on=["省区", "_经销商_key"], how="left")
                                            pv_s = pv_s.merge(cum, on=["省区", "_经销商_key"], how="left")
                                            for _c in ["4月新客", "近三月新客", "整体新客"]:
                                                if _c in pv_s.columns:
                                                    pv_s[_c] = pd.to_numeric(pv_s[_c], errors="coerce").fillna(0.0)
                                                else:
                                                    pv_s[_c] = 0.0
                                            pv_s["累计新客"] = pd.to_numeric(pv_s.get("整体新客", 0), errors="coerce").fillna(0.0)
                                            pv_s.drop(columns=["_经销商_key"], inplace=True, errors="ignore")

                                    pv_s[scan_avg_col] = 0.0
                                    pv_s[scan_rate_col] = 0.0
                                    if df_scan_raw is not None and not getattr(df_scan_raw, "empty", True):
                                        s = df_scan_raw.copy()
                                        for _c in ["省区", "经销商名称", "产品大类", "产品小类"]:
                                            if _c in s.columns:
                                                if _c == "经销商名称":
                                                    s[_c] = s[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                else:
                                                    s[_c] = s[_c].fillna("").astype(str).str.strip()
                                        s["年份"] = pd.to_numeric(s.get("年份", 0), errors="coerce").fillna(0).astype(int)
                                        s["月份"] = pd.to_numeric(s.get("月份", 0), errors="coerce").fillna(0).astype(int)
                                        s = s[(s["年份"] > 0) & (s["月份"].between(1, 12))].copy()
                                        if sel_big != "全部":
                                            _sb = str(sel_big).strip()
                                            if _sb == "雅系列":
                                                if "产品小类" in s.columns:
                                                    s = s[s["产品小类"].astype(str).str.contains(r"(雅赋|雅耀|雅舒|雅护)", regex=True)].copy()
                                                elif "产品大类" in s.columns:
                                                    s = s[s["产品大类"].astype(str).str.strip() == _sb].copy()
                                            elif "产品大类" in s.columns:
                                                s = s[s["产品大类"].astype(str).str.strip() == _sb].copy()
                                        if sel_small != "全部" and "产品小类" in s.columns:
                                            _sel_s = str(sel_small).strip()
                                            _m = re.search(r"(\d{3})", _sel_s)
                                            if _m:
                                                _w = s["产品小类"].astype(str).str.extract(r"(\d{3})")[0].fillna("").astype(str)
                                                s = s[_w == _m.group(1)].copy()
                                            else:
                                                s = s[s["产品小类"].astype(str).str.strip() == _sel_s].copy()
                                        s["_ym"] = (s["年份"] * 100 + s["月份"]).astype(int)
                                        _ym_num = pd.to_numeric(s["_ym"], errors="coerce").fillna(0).astype(int)
                                        s = s[_ym_num.isin([int(x) for x in scan_yms])].copy()
                                        if not s.empty:
                                            dist_map2 = {}
                                            if df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                                if "经销商全称" in df_stock_raw.columns and "经销商名称" in df_stock_raw.columns:
                                                    _m2 = (
                                                        df_stock_raw[["经销商全称", "经销商名称"]]
                                                        .dropna()
                                                        .astype(str)
                                                        .apply(lambda col: col.str.replace(r"\s+", "", regex=True))
                                                        .drop_duplicates()
                                                    )
                                                    dist_map2 = dict(zip(_m2["经销商全称"].tolist(), _m2["经销商名称"].tolist()))
                                            pv_s["_经销商_key_scan"] = pv_s["经销商"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map2:
                                                pv_s["_经销商_key_scan"] = pv_s["_经销商_key_scan"].map(dist_map2).fillna(pv_s["_经销商_key_scan"])
                                            s["_经销商_key_scan"] = s["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map2:
                                                s["_经销商_key_scan"] = s["_经销商_key_scan"].map(dist_map2).fillna(s["_经销商_key_scan"])
                                            for _ym in scan_yms:
                                                _ym_i = int(_ym)
                                                sf = s[pd.to_numeric(s["_ym"], errors="coerce").fillna(0).astype(int) == _ym_i].copy()
                                                c_scan = f"_scan_{_ym_i}"
                                                if not sf.empty:
                                                    agg_f = sf.groupby(["省区", "_经销商_key_scan"], as_index=False).size().rename(columns={"size": "_扫码听数"})
                                                    agg_f[c_scan] = pd.to_numeric(agg_f["_扫码听数"], errors="coerce").fillna(0.0) / 6.0
                                                    agg_f.drop(columns=["_扫码听数"], inplace=True, errors="ignore")
                                                    pv_s = pv_s.merge(agg_f[["省区", "_经销商_key_scan", c_scan]], on=["省区", "_经销商_key_scan"], how="left")
                                                else:
                                                    pv_s[c_scan] = 0.0
                                                pv_s[c_scan] = pd.to_numeric(pv_s.get(c_scan, 0), errors="coerce").fillna(0.0)
                                            _scan_cols = [f"_scan_{int(_ym)}" for _ym in scan_yms]
                                            pv_s[scan_avg_col] = pv_s[_scan_cols].mean(axis=1)
                                            pv_s.drop(columns=["_经销商_key_scan"], inplace=True, errors="ignore")
                                    pv_s[scan_avg_col] = pd.to_numeric(pv_s.get(scan_avg_col, 0), errors="coerce").fillna(0.0)
                                    if avg_header in pv_s.columns:
                                        _out_m = pd.to_numeric(pv_s[avg_header], errors="coerce").fillna(0.0)
                                        pv_s[scan_rate_col] = np.where(_out_m > 0, pv_s[scan_avg_col] / _out_m, 0.0)
                                    else:
                                        pv_s[scan_rate_col] = 0.0
                                    pv_s[scan_rate_col] = pd.to_numeric(pv_s.get(scan_rate_col, 0), errors="coerce").fillna(0.0)

                                    if store_geo_df is not None and not getattr(store_geo_df, "empty", True) and "门店" in pv_s.columns:
                                        try:
                                            pv_s["_k_store_geo"] = pv_s["门店"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            pv_s = pv_s.merge(store_geo_df, on="_k_store_geo", how="left")
                                            pv_s.drop(columns=["_k_store_geo"], inplace=True, errors="ignore")
                                            pv_s["市"] = pv_s.get("市", "").fillna("").astype(str).str.strip()
                                            pv_s["区/县"] = pv_s.get("区/县", "").fillna("").astype(str).str.strip()
                                            pv_s["门店状态"] = pv_s.get("门店状态", "").fillna("").astype(str).str.strip()
                                        except Exception:
                                            pv_s["市"] = pv_s.get("市", "").fillna("").astype(str).str.strip()
                                            pv_s["区/县"] = pv_s.get("区/县", "").fillna("").astype(str).str.strip()
                                            pv_s["门店状态"] = pv_s.get("门店状态", "").fillna("").astype(str).str.strip()

                                    months = [c for c in sel_month_cols if c in pv_s.columns]
                                    anchor = tcols[-1] if tcols else (months[-1] if months else None)
                                    before = []
                                    after = []
                                    if anchor and anchor in months:
                                        idx = months.index(anchor)
                                        before = months[: idx + 1]
                                        after = months[idx + 1 :]
                                    else:
                                        before = months

                                    ordered = ["省区", "经销商"]
                                    for _c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                        if _c in pv_s.columns:
                                            ordered.append(_c)
                                    ordered += before
                                    ordered.append(avg_header)
                                    if march_col in pv_s.columns:
                                        ordered.append(march_col)
                                    if today_col in pv_s.columns:
                                        ordered.append(today_col)
                                    ordered.append("库存")
                                    ordered.append("可销月")
                                    ordered += after
                                    for _c in ["4月新客", "近三月新客", "累计新客"]:
                                        if _c in pv_s.columns:
                                            ordered.append(_c)
                                    if scan_avg_col in pv_s.columns:
                                        ordered.append(scan_avg_col)
                                    if scan_rate_col in pv_s.columns:
                                        ordered.append(scan_rate_col)
                                    for p_label, _yms in roll_periods:
                                        for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                            if c in pv_s.columns:
                                                ordered.append(c)
                                    for i in range(1, len(roll_periods)):
                                        c = f"{roll_periods[i][0]}变动"
                                        if c in pv_s.columns:
                                            ordered.append(c)
                                    if "近三周期变化" in pv_s.columns:
                                        ordered.append("近三周期变化")
                                    ordered = [c for c in ordered if c in pv_s.columns]
                                    pv_s = pv_s[ordered].copy()
                                    return pv_s

                                def _build_store_detail_df(all_provinces: bool):
                                    d = df_trend_base.copy()
                                    store_col = "_门店名" if "_门店名" in d.columns else ("门店名称" if "门店名称" in d.columns else None)
                                    if store_col is None:
                                        cols0 = ["省区", "经销商", "门店"]
                                        if sel_month_cols:
                                            cols0.append(sel_month_cols[0])
                                        cols0 += [avg_header]
                                        cols0 += [c for c in sel_month_cols if sel_month_cols and c != sel_month_cols[0]]
                                        cols0.append("完成率")
                                        cols0 = [c for c in cols0 if c]
                                        return pd.DataFrame(columns=cols0)

                                    prov_sel = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                    dist_sel = str(st.session_state.get("out_m_selected_dist") or "").strip()
                                    if not all_provinces and prov_sel:
                                        d = d[d["省区"].astype(str).str.strip() == prov_sel].copy()
                                    if drill_level == 3 and dist_sel:
                                        d = d[d["经销商名称"].astype(str).str.strip() == dist_sel].copy()

                                    if sel_big != '全部' and '_模块大类' in d.columns:
                                        d = d[d['_模块大类'].astype(str).str.strip() == str(sel_big).strip()].copy()
                                    if sel_small != '全部' and '_模块小类' in d.columns:
                                        d = d[d['_模块小类'].astype(str).str.strip() == str(sel_small).strip()].copy()
                                    if sel_prod and '_模块出库产品' in d.columns:
                                        sel_prod_norm = [str(x).strip() for x in sel_prod if str(x).strip()]
                                        if sel_prod_norm:
                                            d = d[d['_模块出库产品'].astype(str).str.strip().isin(sel_prod_norm)].copy()

                                    d_all = d.copy()
                                    d = d[d["_ym"].isin(sel_yms)].copy()
                                    if d.empty:
                                        return pd.DataFrame(columns=["省区", "经销商", "门店"] + sel_month_cols + [avg_header])

                                    meta_cols = [c for c in ["门店编号"] if c in d_all.columns]
                                    meta_agg = None
                                    if meta_cols:
                                        meta = d_all[["省区", "经销商名称", store_col] + meta_cols].copy()
                                        meta["省区"] = meta["省区"].fillna("").astype(str).str.strip()
                                        meta["经销商名称"] = meta["经销商名称"].fillna("").astype(str).str.strip()
                                        meta[store_col] = meta[store_col].fillna("").astype(str).str.strip()
                                        for _c in meta_cols:
                                            meta[_c] = meta[_c].fillna("").astype(str).str.strip()

                                        def _first_non_empty(vs):
                                            for x in vs.tolist():
                                                s = str(x or "").strip()
                                                if s and s.lower() not in ("nan", "none", "null"):
                                                    return s
                                            return ""

                                        meta_agg = (
                                            meta.groupby(["省区", "经销商名称", store_col], as_index=False)
                                            .agg({c: _first_non_empty for c in meta_cols})
                                            .rename(columns={"经销商名称": "经销商", store_col: "门店"})
                                        )

                                    d["数量(箱)"] = pd.to_numeric(d.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                    d["省区"] = d["省区"].fillna("").astype(str).str.strip()
                                    d["经销商名称"] = d["经销商名称"].fillna("").astype(str).str.strip()
                                    d[store_col] = d[store_col].fillna("").astype(str).str.strip()
                                    d = d[(d[store_col] != "") & (d["经销商名称"] != "")].copy()

                                    g = d.groupby(["省区", "经销商名称", store_col, "_ym"], as_index=False)["数量(箱)"].sum()
                                    pv_s = g.pivot(index=["省区", "经销商名称", store_col], columns="_ym", values="数量(箱)").fillna(0.0)
                                    for ym in sel_yms:
                                        if ym not in pv_s.columns:
                                            pv_s[ym] = 0.0
                                    pv_s = pv_s[sel_yms]
                                    pv_s.columns = sel_month_cols
                                    pv_s["_合计"] = pv_s.sum(axis=1)
                                    pv_s = pv_s.reset_index().rename(columns={"经销商名称": "经销商", store_col: "门店"})
                                    if meta_agg is not None and not meta_agg.empty:
                                        pv_s = pv_s.merge(meta_agg, on=["省区", "经销商", "门店"], how="left")
                                    pv_s = pv_s.sort_values("_合计", ascending=False).reset_index(drop=True)
                                    pv_s.drop(columns=["_合计"], inplace=True, errors="ignore")

                                    try:
                                        need_yms = []
                                        for _, yms in roll_periods:
                                            need_yms += list(yms)
                                        need_yms = sorted(set([int(x) for x in need_yms]))
                                        d_roll = d_all.copy()
                                        d_roll = d_roll[d_roll["_ym"].isin(need_yms)].copy()
                                        if not d_roll.empty:
                                            d_roll["数量(箱)"] = pd.to_numeric(d_roll.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                            d_roll["省区"] = d_roll["省区"].fillna("").astype(str).str.strip()
                                            d_roll["经销商名称"] = d_roll["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            d_roll[store_col] = d_roll[store_col].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            d_roll = d_roll[(d_roll["省区"] != "") & (d_roll["经销商名称"] != "") & (d_roll[store_col] != "")].copy()

                                            g2 = d_roll.groupby(["省区", "经销商名称", store_col, "_ym"], as_index=False)["数量(箱)"].sum()
                                            rp = g2.pivot(index=["省区", "经销商名称", store_col], columns="_ym", values="数量(箱)").fillna(0.0)
                                            for ym in need_yms:
                                                if ym not in rp.columns:
                                                    rp[ym] = 0.0
                                            rp = rp[need_yms].reset_index().rename(columns={"经销商名称": "经销商", store_col: "门店"})

                                            for p_label, yms in roll_periods:
                                                cols = [int(x) for x in yms]
                                                rp[f"{p_label}月均出库"] = rp[cols].sum(axis=1) / 3.0
                                                rp[f"{p_label}门店类型"] = rp[f"{p_label}月均出库"].apply(_classify_store_abcd)
                                            for i in range(1, len(roll_periods)):
                                                prev_label = roll_periods[i - 1][0]
                                                cur_label = roll_periods[i][0]
                                                rp[f"{cur_label}变动"] = rp.apply(lambda r: _store_change(r.get(f"{prev_label}门店类型"), r.get(f"{cur_label}门店类型")), axis=1)
                                            if len(roll_periods) >= 3:
                                                p1, p2, p3 = roll_periods[-3][0], roll_periods[-2][0], roll_periods[-1][0]
                                                rp["近三周期变化"] = rp.apply(
                                                    lambda r: _trend3_label(
                                                        r.get(f"{p1}门店类型"),
                                                        r.get(f"{p2}门店类型"),
                                                        r.get(f"{p3}门店类型"),
                                                    ),
                                                    axis=1,
                                                )
                                            else:
                                                rp["近三周期变化"] = ""

                                            for p_label, _yms in roll_periods:
                                                c_avg = f"{p_label}月均出库"
                                                if c_avg in rp.columns:
                                                    rp[c_avg] = pd.to_numeric(rp[c_avg], errors="coerce").fillna(0.0).round(1)

                                            pv_s["省区"] = pv_s["省区"].fillna("").astype(str).str.strip()
                                            pv_s["经销商"] = pv_s["经销商"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            pv_s["门店"] = pv_s["门店"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            rp["省区"] = rp["省区"].fillna("").astype(str).str.strip()
                                            rp["经销商"] = rp["经销商"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            rp["门店"] = rp["门店"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)

                                            keep_cols = ["省区", "经销商", "门店"]
                                            for p_label, _yms in roll_periods:
                                                keep_cols += [f"{p_label}月均出库", f"{p_label}门店类型"]
                                            for i in range(1, len(roll_periods)):
                                                keep_cols.append(f"{roll_periods[i][0]}变动")
                                            keep_cols.append("近三周期变化")
                                            keep_cols = [c for c in keep_cols if c in rp.columns]
                                            rp = rp[keep_cols].copy()
                                            pv_s = pv_s.merge(rp, on=["省区", "经销商", "门店"], how="left")
                                    except Exception:
                                        pass

                                    tcols = [c for c in first3_cols if c in pv_s.columns]
                                    if tcols:
                                        pv_s[avg_header] = pv_s[tcols].mean(axis=1)
                                    else:
                                        pv_s[avg_header] = 0.0
                                    pv_s[march_col] = 0.0
                                    try:
                                        dm = d_all[d_all["_ym"] == int(march_ym)].copy()
                                        if not dm.empty:
                                            dm["数量(箱)"] = pd.to_numeric(dm.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                            dm["省区"] = dm["省区"].fillna("").astype(str).str.strip()
                                            dm["经销商名称"] = dm["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            dm[store_col] = dm[store_col].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            gm = (
                                                dm.groupby(["省区", "经销商名称", store_col], as_index=False)["数量(箱)"]
                                                .sum()
                                                .rename(columns={"经销商名称": "经销商", store_col: "门店", "数量(箱)": march_col})
                                            )
                                            pv_s["省区"] = pv_s["省区"].fillna("").astype(str).str.strip()
                                            pv_s["经销商"] = pv_s["经销商"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            pv_s["门店"] = pv_s["门店"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            pv_s = pv_s.merge(gm, on=["省区", "经销商", "门店"], how="left", suffixes=("", "_y"))
                                            if f"{march_col}_y" in pv_s.columns:
                                                pv_s.drop(columns=[march_col], inplace=True, errors="ignore")
                                                pv_s.rename(columns={f"{march_col}_y": march_col}, inplace=True)
                                            pv_s[march_col] = pd.to_numeric(pv_s.get(march_col, 0), errors="coerce").fillna(0.0)
                                    except Exception:
                                        pv_s[march_col] = pd.to_numeric(pv_s.get(march_col, 0), errors="coerce").fillna(0.0)

                                    pv_s[today_col] = 0.0
                                    if today_day is not None:
                                        try:
                                            ddm = d_all[d_all["_ym"] == int(march_ym)].copy()
                                            if not ddm.empty and "_日" in ddm.columns:
                                                ddm["_日"] = pd.to_numeric(ddm["_日"], errors="coerce")
                                                ddm = ddm[ddm["_日"].notna()].copy()
                                                ddm["_日"] = ddm["_日"].astype(int)
                                                ddm = ddm[ddm["_日"] == int(today_day)].copy()
                                            if not ddm.empty:
                                                ddm["数量(箱)"] = pd.to_numeric(ddm.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                                ddm["省区"] = ddm["省区"].fillna("").astype(str).str.strip()
                                                ddm["经销商名称"] = ddm["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                ddm[store_col] = ddm[store_col].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                gd = (
                                                    ddm.groupby(["省区", "经销商名称", store_col], as_index=False)["数量(箱)"]
                                                    .sum()
                                                    .rename(columns={"经销商名称": "经销商", store_col: "门店", "数量(箱)": today_col})
                                                )
                                                pv_s = pv_s.merge(gd, on=["省区", "经销商", "门店"], how="left", suffixes=("", "_y"))
                                                if f"{today_col}_y" in pv_s.columns:
                                                    pv_s.drop(columns=[today_col], inplace=True, errors="ignore")
                                                    pv_s.rename(columns={f"{today_col}_y": today_col}, inplace=True)
                                                pv_s[today_col] = pd.to_numeric(pv_s.get(today_col, 0), errors="coerce").fillna(0.0)
                                        except Exception:
                                            pv_s[today_col] = pd.to_numeric(pv_s.get(today_col, 0), errors="coerce").fillna(0.0)

                                    pv_s[april_col] = pd.to_numeric(pv_s.get(march_col, 0), errors="coerce").fillna(0.0)

                                    if df_newcust_raw is not None and not getattr(df_newcust_raw, "empty", True) and "_ym" in df_newcust_raw.columns:
                                        nc = df_newcust_raw.copy()
                                        if "省区" in nc.columns:
                                            nc["省区"] = nc["省区"].fillna("").astype(str).str.strip()
                                        if "经销商名称" in nc.columns:
                                            nc["经销商名称"] = nc["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                        if "门店名称" in nc.columns:
                                            nc["门店名称"] = nc["门店名称"].fillna("").astype(str).str.strip()
                                        if "新客数" in nc.columns:
                                            nc["新客数"] = pd.to_numeric(nc["新客数"], errors="coerce").fillna(0.0)
                                        nc["_ym"] = pd.to_numeric(nc["_ym"], errors="coerce").fillna(0).astype(int)
                                        nc = nc[nc["_ym"].between(200001, 209912)].copy()
                                        if not all_provinces and prov_sel and "省区" in nc.columns:
                                            nc = nc[nc["省区"].astype(str).str.strip() == prov_sel].copy()
                                        if drill_level == 3 and dist_sel and "经销商名称" in nc.columns:
                                            nc = nc[nc["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True) == dist_sel.replace(" ", "")].copy()
                                        if not nc.empty:
                                            _yms = sorted([int(x) for x in nc["_ym"].dropna().astype(int).tolist() if 200001 <= int(x) <= 209912])
                                            _yy = int(max(_yms) // 100)
                                            latest_ym = int(_yy * 100 + 4)
                                            prev3 = [int(_yy * 100 + 1), int(_yy * 100 + 2), int(_yy * 100 + 3)]
                                            cur = nc[nc["_ym"] == latest_ym].groupby(["省区", "经销商名称", "门店名称"], as_index=False)["新客数"].sum().rename(columns={"经销商名称": "经销商", "门店名称": "门店", "新客数": "4月新客"})
                                            p3 = nc[nc["_ym"].isin(prev3)].groupby(["省区", "经销商名称", "门店名称"], as_index=False)["新客数"].sum().rename(columns={"经销商名称": "经销商", "门店名称": "门店", "新客数": "近三月新客"})
                                            cum = nc.groupby(["省区", "经销商名称", "门店名称"], as_index=False)["新客数"].sum().rename(columns={"经销商名称": "经销商", "门店名称": "门店", "新客数": "整体新客"})
                                            for _df in (cur, p3, cum):
                                                _df["省区"] = _df["省区"].fillna("").astype(str).str.strip()
                                                _df["经销商"] = _df["经销商"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                _df["门店"] = _df["门店"].fillna("").astype(str).str.strip()
                                            pv_s["省区"] = pv_s["省区"].fillna("").astype(str).str.strip()
                                            pv_s["经销商"] = pv_s["经销商"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            pv_s["门店"] = pv_s["门店"].fillna("").astype(str).str.strip()
                                            pv_s = pv_s.merge(cur, on=["省区", "经销商", "门店"], how="left")
                                            pv_s = pv_s.merge(p3, on=["省区", "经销商", "门店"], how="left")
                                            pv_s = pv_s.merge(cum, on=["省区", "经销商", "门店"], how="left")
                                            for _c in ["4月新客", "近三月新客", "整体新客"]:
                                                if _c in pv_s.columns:
                                                    pv_s[_c] = pd.to_numeric(pv_s[_c], errors="coerce").fillna(0.0)
                                                else:
                                                    pv_s[_c] = 0.0
                                            pv_s["累计新客"] = pd.to_numeric(pv_s.get("整体新客", 0), errors="coerce").fillna(0.0)

                                    pv_s[scan_avg_col] = 0.0
                                    pv_s[scan_rate_col] = 0.0
                                    if df_scan_raw is not None and not getattr(df_scan_raw, "empty", True):
                                        s = df_scan_raw.copy()
                                        for _c in ["省区", "经销商名称", "门店名称", "产品大类", "产品小类"]:
                                            if _c in s.columns:
                                                if _c == "经销商名称":
                                                    s[_c] = s[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                elif _c == "门店名称":
                                                    s[_c] = s[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                else:
                                                    s[_c] = s[_c].fillna("").astype(str).str.strip()
                                        s["年份"] = pd.to_numeric(s.get("年份", 0), errors="coerce").fillna(0).astype(int)
                                        s["月份"] = pd.to_numeric(s.get("月份", 0), errors="coerce").fillna(0).astype(int)
                                        s = s[(s["年份"] > 0) & (s["月份"].between(1, 12))].copy()
                                        if sel_big != "全部":
                                            _sb = str(sel_big).strip()
                                            if _sb == "雅系列":
                                                if "产品小类" in s.columns:
                                                    s = s[s["产品小类"].astype(str).str.contains(r"(雅赋|雅耀|雅舒|雅护)", regex=True)].copy()
                                                elif "产品大类" in s.columns:
                                                    s = s[s["产品大类"].astype(str).str.strip() == _sb].copy()
                                            elif "产品大类" in s.columns:
                                                s = s[s["产品大类"].astype(str).str.strip() == _sb].copy()
                                        if sel_small != "全部" and "产品小类" in s.columns:
                                            _sel_s = str(sel_small).strip()
                                            _m = re.search(r"(\d{3})", _sel_s)
                                            if _m:
                                                _w = s["产品小类"].astype(str).str.extract(r"(\d{3})")[0].fillna("").astype(str)
                                                s = s[_w == _m.group(1)].copy()
                                            else:
                                                s = s[s["产品小类"].astype(str).str.strip() == _sel_s].copy()
                                        s["_ym"] = (s["年份"] * 100 + s["月份"]).astype(int)
                                        _ym_num = pd.to_numeric(s["_ym"], errors="coerce").fillna(0).astype(int)
                                        s = s[_ym_num.isin([int(x) for x in scan_yms])].copy()
                                        if not s.empty:
                                            dist_map2 = {}
                                            if df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                                if "经销商全称" in df_stock_raw.columns and "经销商名称" in df_stock_raw.columns:
                                                    _m2 = (
                                                        df_stock_raw[["经销商全称", "经销商名称"]]
                                                        .dropna()
                                                        .astype(str)
                                                        .apply(lambda col: col.str.replace(r"\s+", "", regex=True))
                                                        .drop_duplicates()
                                                    )
                                                    dist_map2 = dict(zip(_m2["经销商全称"].tolist(), _m2["经销商名称"].tolist()))
                                            pv_s["_经销商_key_scan"] = pv_s["经销商"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map2:
                                                pv_s["_经销商_key_scan"] = pv_s["_经销商_key_scan"].map(dist_map2).fillna(pv_s["_经销商_key_scan"])
                                            pv_s["_门店_key_scan"] = pv_s["门店"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            s["_经销商_key_scan"] = s["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map2:
                                                s["_经销商_key_scan"] = s["_经销商_key_scan"].map(dist_map2).fillna(s["_经销商_key_scan"])
                                            s["_门店_key_scan"] = s["门店名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            for _ym in scan_yms:
                                                _ym_i = int(_ym)
                                                sf = s[pd.to_numeric(s["_ym"], errors="coerce").fillna(0).astype(int) == _ym_i].copy()
                                                c_scan = f"_scan_{_ym_i}"
                                                if not sf.empty:
                                                    agg_f = sf.groupby(["省区", "_经销商_key_scan", "_门店_key_scan"], as_index=False).size().rename(columns={"size": "_扫码听数"})
                                                    agg_f[c_scan] = pd.to_numeric(agg_f["_扫码听数"], errors="coerce").fillna(0.0) / 6.0
                                                    agg_f.drop(columns=["_扫码听数"], inplace=True, errors="ignore")
                                                    pv_s = pv_s.merge(agg_f[["省区", "_经销商_key_scan", "_门店_key_scan", c_scan]], on=["省区", "_经销商_key_scan", "_门店_key_scan"], how="left")
                                                else:
                                                    pv_s[c_scan] = 0.0
                                                pv_s[c_scan] = pd.to_numeric(pv_s.get(c_scan, 0), errors="coerce").fillna(0.0)
                                            _scan_cols = [f"_scan_{int(_ym)}" for _ym in scan_yms]
                                            pv_s[scan_avg_col] = pv_s[_scan_cols].mean(axis=1)
                                            pv_s.drop(columns=["_经销商_key_scan"], inplace=True, errors="ignore")
                                            pv_s.drop(columns=["_门店_key_scan"], inplace=True, errors="ignore")
                                    pv_s[scan_avg_col] = pd.to_numeric(pv_s.get(scan_avg_col, 0), errors="coerce").fillna(0.0)
                                    if avg_header in pv_s.columns:
                                        _out_m = pd.to_numeric(pv_s[avg_header], errors="coerce").fillna(0.0)
                                        pv_s[scan_rate_col] = np.where(_out_m > 0, pv_s[scan_avg_col] / _out_m, 0.0)
                                    else:
                                        pv_s[scan_rate_col] = 0.0
                                    pv_s[scan_rate_col] = pd.to_numeric(pv_s.get(scan_rate_col, 0), errors="coerce").fillna(0.0)

                                    if store_geo_df is not None and not getattr(store_geo_df, "empty", True) and "门店" in pv_s.columns:
                                        try:
                                            pv_s["_k_store_geo"] = pv_s["门店"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            pv_s = pv_s.merge(store_geo_df, on="_k_store_geo", how="left")
                                            pv_s.drop(columns=["_k_store_geo"], inplace=True, errors="ignore")
                                            pv_s["市"] = pv_s.get("市", "").fillna("").astype(str).str.strip()
                                            pv_s["区/县"] = pv_s.get("区/县", "").fillna("").astype(str).str.strip()
                                        except Exception:
                                            pv_s["市"] = pv_s.get("市", "").fillna("").astype(str).str.strip()
                                            pv_s["区/县"] = pv_s.get("区/县", "").fillna("").astype(str).str.strip()

                                    months = [c for c in sel_month_cols if c in pv_s.columns]
                                    anchor = tcols[-1] if tcols else (months[-1] if months else None)
                                    before = []
                                    after = []
                                    if anchor and anchor in months:
                                        idx = months.index(anchor)
                                        before = months[: idx + 1]
                                        after = months[idx + 1 :]
                                    else:
                                        before = months

                                    ordered = ["省区", "经销商", "门店"]
                                    for _c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                        if _c in pv_s.columns:
                                            ordered.append(_c)
                                    ordered += before
                                    ordered.append(avg_header)
                                    if march_col in pv_s.columns:
                                        ordered.append(march_col)
                                    if today_col in pv_s.columns:
                                        ordered.append(today_col)
                                    ordered += after
                                    for _c in ["4月新客", "近三月新客", "累计新客"]:
                                        if _c in pv_s.columns:
                                            ordered.append(_c)
                                    if scan_avg_col in pv_s.columns:
                                        ordered.append(scan_avg_col)
                                    if scan_rate_col in pv_s.columns:
                                        ordered.append(scan_rate_col)
                                    for _c in ["门店编号", "门店状态"]:
                                        if _c in pv_s.columns:
                                            ordered.append(_c)
                                    for _c in ["市", "区/县"]:
                                        if _c in pv_s.columns:
                                            ordered.append(_c)
                                    ordered = [c for c in ordered if c in pv_s.columns]
                                    pv_s = pv_s[ordered].copy()
                                    return pv_s

                                number_headers_current = set([c for c in first3_cols if c in pv.columns])
                                if march_col in pv.columns:
                                    number_headers_current.add(march_col)
                                if today_col in pv.columns:
                                    number_headers_current.add(today_col)
                                number_headers_current |= {avg_header, "库存", "可销月", "1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"}
                                number_headers_current.discard("")
                                for _c in ["4月新客", "近三月新客", "累计新客"]:
                                    if _c in pv.columns:
                                        number_headers_current.add(_c)
                                if scan_avg_col in pv.columns:
                                    number_headers_current.add(scan_avg_col)
                                percent_headers_current = set()
                                if scan_rate_col in pv.columns:
                                    percent_headers_current.add(scan_rate_col)

                                if "out_m_excel_cache" not in st.session_state:
                                    st.session_state.out_m_excel_cache = {}
                                _excel_cache = st.session_state.out_m_excel_cache

                                _prov_sel = str(st.session_state.get("out_m_selected_prov") or "").strip()
                                _dist_sel = str(st.session_state.get("out_m_selected_dist") or "").strip()
                                _prod_norm_key = tuple(sorted([str(x).strip() for x in (sel_prod or []) if str(x).strip()]))
                                _ship_sig = 0.0
                                try:
                                    _ship_cols = ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]
                                    _ship_cols = [c for c in _ship_cols if c in pv.columns]
                                    if _ship_cols:
                                        _ship_sig = float(pd.to_numeric(pv[_ship_cols], errors="coerce").fillna(0.0).to_numpy().sum())
                                except Exception:
                                    _ship_sig = 0.0
                                _march_sig = 0.0
                                try:
                                    if march_col in pv.columns:
                                        _march_sig = float(pd.to_numeric(pv[march_col], errors="coerce").fillna(0.0).sum())
                                except Exception:
                                    _march_sig = 0.0

                                _excel_cache_ver = 9
                                _png_cache_ver = 7

                                def _excel_key(kind: str):
                                    return (
                                        kind,
                                        _excel_cache_ver,
                                        sig,
                                        int(drill_level),
                                        tuple(sel_yms),
                                        tuple(sel_month_cols),
                                        tuple(first3_cols),
                                        str(last_col or ""),
                                        str(avg_header or ""),
                                        str(sel_big or ""),
                                        str(sel_small or ""),
                                        _prod_norm_key,
                                        _prov_sel,
                                        _dist_sel,
                                        round(_ship_sig, 4),
                                        round(_march_sig, 4),
                                    )

                                now_tag = datetime.now().strftime("%Y%m%d_%H%M%S")

                                if drill_level == 1:
                                    k_cur = _excel_key("cur")
                                    c_e1, c_e2, _ = st.columns([1.1, 1.6, 6.3])
                                    with c_e1:
                                        if st.button("生成Excel（当前表）", key=f"{export_id}_gen_cur"):
                                            with st.spinner("正在生成Excel（当前表）…"):
                                                df_x = _build_current_excel_df()
                                                ren = {}
                                                for c in df_x.columns:
                                                    if c in ("1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"):
                                                        ren[c] = f"发货分析-{c}"
                                                for c in sel_month_cols:
                                                    if c in df_x.columns:
                                                        ren[c] = f"出库分析-{c}"
                                                if avg_header in df_x.columns:
                                                    ren[avg_header] = f"出库分析-{avg_header}"
                                                if march_col in df_x.columns:
                                                    ren[march_col] = f"出库分析-{march_col}"
                                                if today_col in df_x.columns:
                                                    ren[today_col] = f"出库分析-{today_col}"
                                                if "趋势" in df_x.columns:
                                                    ren["趋势"] = "出库分析-趋势图"
                                                if "库存" in df_x.columns:
                                                    ren["库存"] = "库存分析-库存"
                                                if "可销月" in df_x.columns:
                                                    ren["可销月"] = "库存分析-可销月"
                                                for _c in ["4月新客", "近三月新客", "累计新客"]:
                                                    if _c in df_x.columns:
                                                        ren[_c] = f"新客分析-{_c}"
                                                if scan_avg_col in df_x.columns:
                                                    ren[scan_avg_col] = f"扫码分析-{scan_avg_header}"
                                                if scan_rate_col in df_x.columns:
                                                    ren[scan_rate_col] = f"扫码分析-{scan_rate_col}"
                                                df_x = df_x.rename(columns=ren)
                                                number_headers_out = set()
                                                for c in number_headers_current:
                                                    number_headers_out.add(ren.get(c, c))
                                                percent_headers_out = set()
                                                for c in percent_headers_current:
                                                    percent_headers_out.add(ren.get(c, c))
                                                xlsx_bytes = _df_to_excel_bytes(
                                                    df_x,
                                                    sheet_name="趋势分析",
                                                    title_lines=export_title_lines,
                                                    number_headers=number_headers_out,
                                                    number_formats={
                                                        ren.get("可销月", "可销月"): "0.0",
                                                        ren.get(march_col, march_col): "0.0",
                                                        ren.get(today_col, today_col): "0.0",
                                                        ren.get(scan_avg_col, scan_avg_col): "0.0",
                                                    },
                                                    percent_headers=percent_headers_out,
                                                    percent_formats={ren.get(scan_rate_col, scan_rate_col): "0.0%"},
                                                    group_headers=True,
                                                )
                                                _excel_cache[k_cur] = {
                                                    "bytes": xlsx_bytes,
                                                    "name": sanitize_filename(f"出库趋势分析_分省区_{now_tag}.xlsx"),
                                                }
                                    with c_e2:
                                        if k_cur in _excel_cache:
                                            st.download_button(
                                                "下载Excel（当前表）",
                                                data=_excel_cache[k_cur]["bytes"],
                                                file_name=_excel_cache[k_cur]["name"],
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                key=f"{export_id}_dl_cur",
                                            )

                                    k_all = _excel_key("all_store")
                                    k_all_dist = _excel_key("all_dist")
                                    k_all_store = _excel_key("all_store")
                                    k_all_zip = _excel_key("all_zip")
                                    k_dist_folder_zip = _excel_key("dist_folder_zip")
                                    k_bundle_3s = _excel_key("bundle_3sheets")
                                    c_a1, c_a2, c_a3, c_a4, c_a5, c_a6, c_a7, c_a8, _ = st.columns([1.2, 1.5, 1.2, 1.5, 1.3, 1.7, 2.0, 2.2, 0.9])
                                    with c_a1:
                                        if st.button("生成导出全部（经销商）", key=f"{export_id}_gen_all_dist"):
                                            with st.spinner("正在生成导出全部（经销商），数据量较大请稍候…"):
                                                df_all_dist = _build_dist_detail_df(all_provinces=True)
                                                title_all_dist = [
                                                    "月度出库趋势表 - 导出全部经销商",
                                                    filter_line,
                                                    "区域：全部省区（省区→经销商）",
                                                    f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                                ]
                                                ren = {}
                                                for c in df_all_dist.columns:
                                                    if c in ("1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"):
                                                        ren[c] = f"发货分析-{c}"
                                                for c in sel_month_cols:
                                                    if c in df_all_dist.columns:
                                                        ren[c] = f"出库分析-{c}"
                                                if avg_header in df_all_dist.columns:
                                                    ren[avg_header] = f"出库分析-{avg_header}"
                                                if march_col in df_all_dist.columns:
                                                    ren[march_col] = f"出库分析-{march_col}"
                                                if today_col in df_all_dist.columns:
                                                    ren[today_col] = f"出库分析-{today_col}"
                                                if "库存" in df_all_dist.columns:
                                                    ren["库存"] = "库存分析-库存"
                                                if "可销月" in df_all_dist.columns:
                                                    ren["可销月"] = "库存分析-可销月"
                                                for _c in ["4月新客", "近三月新客", "累计新客"]:
                                                    if _c in df_all_dist.columns:
                                                        ren[_c] = f"新客分析-{_c}"
                                                if scan_avg_col in df_all_dist.columns:
                                                    ren[scan_avg_col] = f"扫码分析-{scan_avg_header}"
                                                if scan_rate_col in df_all_dist.columns:
                                                    ren[scan_rate_col] = f"扫码分析-{scan_rate_col}"
                                                df_all_dist = df_all_dist.rename(columns=ren)

                                                number_headers_all_dist = set(sel_month_cols + [avg_header, march_col, today_col, "库存", "可销月", "1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"])
                                                for _c in ["4月新客", "近三月新客", "累计新客"]:
                                                    if ren.get(_c, _c) in df_all_dist.columns:
                                                        number_headers_all_dist.add(_c)
                                                if ren.get(scan_avg_col, scan_avg_col) in df_all_dist.columns:
                                                    number_headers_all_dist.add(scan_avg_col)
                                                number_headers_all_dist = set([ren.get(c, c) for c in number_headers_all_dist])
                                                xlsx_all_dist = _df_to_excel_bytes(
                                                    df_all_dist,
                                                    sheet_name="趋势分析",
                                                    title_lines=title_all_dist,
                                                    number_headers=number_headers_all_dist,
                                                    number_formats={
                                                        ren.get("可销月", "可销月"): "0.0",
                                                        ren.get(march_col, march_col): "0.0",
                                                        ren.get(today_col, today_col): "0.0",
                                                        ren.get(scan_avg_col, scan_avg_col): "0.0",
                                                    },
                                                    percent_headers=set([ren.get(scan_rate_col, scan_rate_col)] if ren.get(scan_rate_col, scan_rate_col) in df_all_dist.columns else []),
                                                    percent_formats={ren.get(scan_rate_col, scan_rate_col): "0.0%"},
                                                    group_headers=True,
                                                )
                                                _excel_cache[k_all_dist] = {
                                                    "bytes": xlsx_all_dist,
                                                    "name": sanitize_filename(f"出库趋势分析_经销商_全部省区_{now_tag}.xlsx"),
                                                }
                                    with c_a2:
                                        if k_all_dist in _excel_cache:
                                            st.download_button(
                                                "下载导出全部（经销商）",
                                                data=_excel_cache[k_all_dist]["bytes"],
                                                file_name=_excel_cache[k_all_dist]["name"],
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                key=f"{export_id}_dl_all_dist",
                                            )
                                    with c_a3:
                                        if st.button("生成导出全部（门店）", key=f"{export_id}_gen_all_store"):
                                            with st.spinner("正在生成导出全部（门店明细），数据量较大请稍候…"):
                                                df_all = _build_store_detail_df(all_provinces=True)
                                                title_all = [
                                                    "月度出库趋势表 - 导出全部门店明细",
                                                    filter_line,
                                                    "区域：全部省区（省区→经销商→门店）",
                                                    f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                                ]
                                                ren = {}
                                                for c in df_all.columns:
                                                    if c in ("1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"):
                                                        ren[c] = f"发货分析-{c}"
                                                for c in sel_month_cols:
                                                    if c in df_all.columns:
                                                        ren[c] = f"出库分析-{c}"
                                                if avg_header in df_all.columns:
                                                    ren[avg_header] = f"出库分析-{avg_header}"
                                                if march_col in df_all.columns:
                                                    ren[march_col] = f"出库分析-{march_col}"
                                                if today_col in df_all.columns:
                                                    ren[today_col] = f"出库分析-{today_col}"
                                                if "趋势类型" in df_all.columns:
                                                    ren["趋势类型"] = "出库分析-趋势类型"
                                                for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                    if _c in df_all.columns:
                                                        ren[_c] = f"新客分析-{_c}"
                                                if scan_avg_col in df_all.columns:
                                                    ren[scan_avg_col] = f"扫码分析-{scan_avg_header}"
                                                if scan_rate_col in df_all.columns:
                                                    ren[scan_rate_col] = f"扫码分析-{scan_rate_col}"
                                                for p_label, _yms in roll_periods:
                                                    for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                                        if c in df_all.columns:
                                                            ren[c] = f"门店类型分析-{c}"
                                                for i in range(1, len(roll_periods)):
                                                    c = f"{roll_periods[i][0]}变动"
                                                    if c in df_all.columns:
                                                        ren[c] = f"门店类型分析-{c}"
                                                if "近三周期变化" in df_all.columns:
                                                    ren["近三周期变化"] = "门店类型分析-近三周期变化"
                                                df_all = df_all.rename(columns=ren)

                                                number_headers_all = set(sel_month_cols + [avg_header, march_col, today_col, "1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"])
                                                number_formats_all = {
                                                    ren.get(avg_header, avg_header): "0.0",
                                                    ren.get("1月发货额-万元", "1月发货额-万元"): "0.0",
                                                    ren.get("2月发货额-万元", "2月发货额-万元"): "0.0",
                                                    ren.get("3月发货额-万元", "3月发货额-万元"): "0.0",
                                                    ren.get(march_col, march_col): "0.0",
                                                    ren.get(today_col, today_col): "0.0",
                                                    "3月出库": "0.0",
                                                    ren.get(scan_avg_col, scan_avg_col): "0.0",
                                                }
                                                for p_label, _yms in roll_periods:
                                                    c = f"{p_label}月均出库"
                                                    if ren.get(c, c) in df_all.columns:
                                                        number_headers_all.add(c)
                                                        number_formats_all[ren.get(c, c)] = "0.0"
                                                for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                    if ren.get(_c, _c) in df_all.columns:
                                                        number_headers_all.add(_c)
                                                if ren.get(scan_avg_col, scan_avg_col) in df_all.columns:
                                                    number_headers_all.add(scan_avg_col)
                                                number_headers_all = set([ren.get(c, c) for c in number_headers_all])
                                                xlsx_all = _df_to_excel_bytes(
                                                    df_all,
                                                    sheet_name="趋势分析",
                                                    title_lines=title_all,
                                                    number_headers=number_headers_all,
                                                    number_formats=number_formats_all,
                                                    percent_headers=set([ren.get(scan_rate_col, scan_rate_col)] if ren.get(scan_rate_col, scan_rate_col) in df_all.columns else []),
                                                    percent_formats={ren.get(scan_rate_col, scan_rate_col): "0.0%"},
                                                    group_headers=True,
                                                )
                                                _excel_cache[k_all_store] = {
                                                    "bytes": xlsx_all,
                                                    "name": sanitize_filename(f"出库趋势分析_门店明细_全部省区_{now_tag}.xlsx"),
                                                }
                                    with c_a4:
                                        if k_all_store in _excel_cache:
                                            st.download_button(
                                                "下载导出全部（门店）",
                                                data=_excel_cache[k_all_store]["bytes"],
                                                file_name=_excel_cache[k_all_store]["name"],
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                key=f"{export_id}_dl_all_store",
                                            )
                                    with c_a5:
                                        if st.button("生成各省门店ZIP", key=f"{export_id}_gen_all_zip"):
                                            with st.spinner("正在生成各省门店ZIP，请稍候…"):
                                                df_all_raw = _build_store_detail_df(all_provinces=True)
                                                
                                                ren = {}
                                                for c in df_all_raw.columns:
                                                    if c in ("1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"):
                                                        ren[c] = f"发货分析-{c}"
                                                for c in sel_month_cols:
                                                    if c in df_all_raw.columns:
                                                        ren[c] = f"出库分析-{c}"
                                                if avg_header in df_all_raw.columns:
                                                    ren[avg_header] = f"出库分析-{avg_header}"
                                                if march_col in df_all_raw.columns:
                                                    ren[march_col] = f"出库分析-{march_col}"
                                                if today_col in df_all_raw.columns:
                                                    ren[today_col] = f"出库分析-{today_col}"
                                                if "趋势类型" in df_all_raw.columns:
                                                    ren["趋势类型"] = "出库分析-趋势类型"
                                                for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                    if _c in df_all_raw.columns:
                                                        ren[_c] = f"新客分析-{_c}"
                                                if scan_avg_col in df_all_raw.columns:
                                                    ren[scan_avg_col] = f"扫码分析-{scan_avg_header}"
                                                if scan_rate_col in df_all_raw.columns:
                                                    ren[scan_rate_col] = f"扫码分析-{scan_rate_col}"
                                                for p_label, _yms in roll_periods:
                                                    for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                                        if c in df_all_raw.columns:
                                                            ren[c] = f"门店类型分析-{c}"
                                                for i in range(1, len(roll_periods)):
                                                    c = f"{roll_periods[i][0]}变动"
                                                    if c in df_all_raw.columns:
                                                        ren[c] = f"门店类型分析-{c}"
                                                if "近三周期变化" in df_all_raw.columns:
                                                    ren["近三周期变化"] = "门店类型分析-近三周期变化"

                                                number_headers_all = set(sel_month_cols + [avg_header, march_col, today_col, "1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"])
                                                number_formats_all = {
                                                    ren.get(avg_header, avg_header): "0.0",
                                                    ren.get("1月发货额-万元", "1月发货额-万元"): "0.0",
                                                    ren.get("2月发货额-万元", "2月发货额-万元"): "0.0",
                                                    ren.get("3月发货额-万元", "3月发货额-万元"): "0.0",
                                                    ren.get(march_col, march_col): "0.0",
                                                    ren.get(today_col, today_col): "0.0",
                                                    "3月出库": "0.0",
                                                    ren.get(scan_avg_col, scan_avg_col): "0.0",
                                                }
                                                for p_label, _yms in roll_periods:
                                                    c = f"{p_label}月均出库"
                                                    if ren.get(c, c) in df_all_raw.columns:
                                                        number_headers_all.add(c)
                                                        number_formats_all[ren.get(c, c)] = "0.0"
                                                for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                    if ren.get(_c, _c) in df_all_raw.columns:
                                                        number_headers_all.add(_c)
                                                if ren.get(scan_avg_col, scan_avg_col) in df_all_raw.columns:
                                                    number_headers_all.add(scan_avg_col)
                                                number_headers_all = set([ren.get(c, c) for c in number_headers_all])
                                                percent_headers_all = set([ren.get(scan_rate_col, scan_rate_col)] if ren.get(scan_rate_col, scan_rate_col) in df_all_raw.columns else [])
                                                percent_formats_all = {ren.get(scan_rate_col, scan_rate_col): "0.0%"}

                                                buf = io.BytesIO()
                                                with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                                                    provs = []
                                                    if "省区" in df_all_raw.columns:
                                                        provs = sorted([p for p in df_all_raw["省区"].dropna().unique() if p])
                                                    for p in provs:
                                                        df_p = df_all_raw[df_all_raw["省区"] == p].copy()
                                                        if df_p.empty: continue
                                                        sort_cols = [c for c in ["经销商", "门店"] if c in df_p.columns]
                                                        if sort_cols:
                                                            df_p = df_p.sort_values(sort_cols, kind="stable").reset_index(drop=True)
                                                        title_p = [
                                                            f"月度出库趋势表 - {p}（门店明细）",
                                                            filter_line,
                                                            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                                        ]
                                                        df_p_ren = df_p.rename(columns=ren)
                                                        xlsx_p = _df_to_excel_bytes(
                                                            df_p_ren,
                                                            sheet_name="趋势分析",
                                                            title_lines=title_p,
                                                            number_headers=number_headers_all,
                                                            number_formats=number_formats_all,
                                                            percent_headers=percent_headers_all,
                                                            percent_formats=percent_formats_all,
                                                            group_headers=True,
                                                        )
                                                        zf.writestr(f"{p}.xlsx", xlsx_p)
                                                buf.seek(0)
                                                _excel_cache[k_all_zip] = {
                                                    "bytes": buf.getvalue(),
                                                    "name": f"出库趋势分析_各省门店.zip",
                                                }
                                        if st.button("生成经销商Excel ZIP", key=f"{export_id}_gen_dist_folder_zip"):
                                            with st.spinner("正在生成经销商Excel ZIP（每个经销商一个Excel），请稍候…"):
                                                df_all_raw = _build_store_detail_df(all_provinces=True)

                                                abbr_map = {}
                                                try:
                                                    if df_trend_universe is not None and not getattr(df_trend_universe, "empty", True) and "经销商名称" in df_trend_universe.columns:
                                                        tmp = df_trend_universe.copy()
                                                        tmp["经销商名称"] = tmp["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                        if "客户简称" in tmp.columns:
                                                            tmp["客户简称"] = tmp["客户简称"].fillna("").astype(str).str.strip()
                                                            tmp = tmp[tmp["经销商名称"] != ""].copy()

                                                            def _first_non_empty(vs):
                                                                for x in vs.tolist():
                                                                    s = str(x or "").strip()
                                                                    if s and s.lower() not in ("nan", "none", "null"):
                                                                        return s
                                                                return ""

                                                            m = tmp.groupby(["经销商名称"], as_index=False).agg({"客户简称": _first_non_empty})
                                                            abbr_map = dict(zip(m["经销商名称"].tolist(), m["客户简称"].tolist()))
                                                except Exception:
                                                    abbr_map = {}

                                                ren = {}
                                                for c in df_all_raw.columns:
                                                    if c in ("1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"):
                                                        ren[c] = f"发货分析-{c}"
                                                for c in sel_month_cols:
                                                    if c in df_all_raw.columns:
                                                        ren[c] = f"出库分析-{c}"
                                                if avg_header in df_all_raw.columns:
                                                    ren[avg_header] = f"出库分析-{avg_header}"
                                                if march_col in df_all_raw.columns:
                                                    ren[march_col] = f"出库分析-{march_col}"
                                                if today_col in df_all_raw.columns:
                                                    ren[today_col] = f"出库分析-{today_col}"
                                                if "趋势类型" in df_all_raw.columns:
                                                    ren["趋势类型"] = "出库分析-趋势类型"
                                                for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                    if _c in df_all_raw.columns:
                                                        ren[_c] = f"新客分析-{_c}"
                                                if scan_avg_col in df_all_raw.columns:
                                                    ren[scan_avg_col] = f"扫码分析-{scan_avg_header}"
                                                if scan_rate_col in df_all_raw.columns:
                                                    ren[scan_rate_col] = f"扫码分析-{scan_rate_col}"
                                                for p_label, _yms in roll_periods:
                                                    for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                                        if c in df_all_raw.columns:
                                                            ren[c] = f"门店类型分析-{c}"
                                                for i in range(1, len(roll_periods)):
                                                    c = f"{roll_periods[i][0]}变动"
                                                    if c in df_all_raw.columns:
                                                        ren[c] = f"门店类型分析-{c}"
                                                if "近三周期变化" in df_all_raw.columns:
                                                    ren["近三周期变化"] = "门店类型分析-近三周期变化"

                                                number_headers_all = set(sel_month_cols + [avg_header, march_col, today_col, "1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"])
                                                number_headers_all |= set(["3月新客", "近三月新客", "累计新客", scan_avg_col])
                                                number_headers_all = set([ren.get(c, c) for c in number_headers_all])
                                                percent_headers_all = set([ren.get(scan_rate_col, scan_rate_col)] if ren.get(scan_rate_col, scan_rate_col) in [ren.get(c, c) for c in df_all_raw.columns] else [])
                                                percent_formats_all = {ren.get(scan_rate_col, scan_rate_col): "0.0%"}
                                                number_formats_all = {
                                                    ren.get(avg_header, avg_header): "0.0",
                                                    ren.get("1月发货额-万元", "1月发货额-万元"): "0.0",
                                                    ren.get("2月发货额-万元", "2月发货额-万元"): "0.0",
                                                    ren.get("3月发货额-万元", "3月发货额-万元"): "0.0",
                                                    ren.get(march_col, march_col): "0.0",
                                                    ren.get(today_col, today_col): "0.0",
                                                    "3月出库": "0.0",
                                                    ren.get(scan_avg_col, scan_avg_col): "0.0",
                                                }

                                                buf = io.BytesIO()
                                                with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                                                    dists = []
                                                    if "经销商" in df_all_raw.columns:
                                                        dists = sorted([d for d in df_all_raw["经销商"].dropna().unique() if d])
                                                    used_files = set()
                                                    for dname in dists:
                                                        df_d = df_all_raw[df_all_raw["经销商"] == dname].copy()
                                                        if df_d.empty:
                                                            continue
                                                        sort_cols = [c for c in ["省区", "门店"] if c in df_d.columns]
                                                        if sort_cols:
                                                            df_d = df_d.sort_values(sort_cols, kind="stable").reset_index(drop=True)

                                                        key_norm = re.sub(r"\s+", "", str(dname))
                                                        abbr = (abbr_map.get(key_norm) or "").strip()
                                                        if not abbr:
                                                            abbr = str(dname)
                                                        base_fname = sanitize_filename(abbr, default=str(dname)) or sanitize_filename(str(dname), default="经销商")
                                                        fname = base_fname + ".xlsx"
                                                        if fname in used_files:
                                                            _p0 = ""
                                                            if "省区" in df_d.columns:
                                                                _p0 = str(df_d["省区"].iloc[0] or "").strip()
                                                            fname = sanitize_filename(f"{base_fname}_{_p0}" if _p0 else f"{base_fname}_{sanitize_filename(str(dname), default='经销商')}", default=base_fname) + ".xlsx"
                                                        used_files.add(fname)

                                                        title_d = [
                                                            f"月度出库趋势表 - {dname}（门店明细）",
                                                            filter_line,
                                                            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                                        ]
                                                        df_d_ren = df_d.rename(columns=ren)
                                                        xlsx_d = _df_to_excel_bytes(
                                                            df_d_ren,
                                                            sheet_name="趋势分析",
                                                            title_lines=title_d,
                                                            number_headers=number_headers_all,
                                                            number_formats=number_formats_all,
                                                            percent_headers=percent_headers_all,
                                                            percent_formats=percent_formats_all,
                                                            group_headers=True,
                                                        )
                                                        zf.writestr(fname, xlsx_d)
                                                buf.seek(0)
                                                _excel_cache[k_dist_folder_zip] = {
                                                    "bytes": buf.getvalue(),
                                                    "name": sanitize_filename(f"出库趋势分析_经销商ExcelZIP_{now_tag}.zip"),
                                                }
                                    with c_a6:
                                        if k_all_zip in _excel_cache:
                                            st.download_button(
                                                "下载各省门店ZIP",
                                                data=_excel_cache[k_all_zip]["bytes"],
                                                file_name=_excel_cache[k_all_zip]["name"],
                                                mime="application/zip",
                                                key=f"{export_id}_dl_all_zip",
                                            )
                                        if k_dist_folder_zip in _excel_cache:
                                            st.download_button(
                                                "下载经销商Excel ZIP",
                                                data=_excel_cache[k_dist_folder_zip]["bytes"],
                                                file_name=_excel_cache[k_dist_folder_zip]["name"],
                                                mime="application/zip",
                                                key=f"{export_id}_dl_dist_folder_zip",
                                            )
                                    with c_a7:
                                        if st.button("生成分省区、客户、门店表格", key=f"{export_id}_gen_bundle_3s"):
                                            with st.spinner("正在生成分省区、客户、门店表格，数据量较大请稍候…"):
                                                if k_cur not in _excel_cache:
                                                    df_x = _build_current_excel_df()
                                                    ren = {}
                                                    for c in df_x.columns:
                                                        if c in ("1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"):
                                                            ren[c] = f"发货分析-{c}"
                                                    for c in sel_month_cols:
                                                        if c in df_x.columns:
                                                            ren[c] = f"出库分析-{c}"
                                                    if avg_header in df_x.columns:
                                                        ren[avg_header] = f"出库分析-{avg_header}"
                                                    if march_col in df_x.columns:
                                                        ren[march_col] = f"出库分析-{march_col}"
                                                    if today_col in df_x.columns:
                                                        ren[today_col] = f"出库分析-{today_col}"
                                                    if "趋势" in df_x.columns:
                                                        ren["趋势"] = "出库分析-趋势图"
                                                    if "趋势类型" in df_x.columns:
                                                        ren["趋势类型"] = "出库分析-趋势类型"
                                                    if "库存" in df_x.columns:
                                                        ren["库存"] = "库存分析-库存"
                                                    if "可销月" in df_x.columns:
                                                        ren["可销月"] = "库存分析-可销月"
                                                    for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                        if _c in df_x.columns:
                                                            ren[_c] = f"新客分析-{_c}"
                                                    if scan_avg_col in df_x.columns:
                                                        ren[scan_avg_col] = f"扫码分析-{scan_avg_header}"
                                                    if scan_rate_col in df_x.columns:
                                                        ren[scan_rate_col] = f"扫码分析-{scan_rate_col}"
                                                    df_x = df_x.rename(columns=ren)
                                                    number_headers_out = set()
                                                    for c in number_headers_current:
                                                        number_headers_out.add(ren.get(c, c))
                                                    percent_headers_out = set()
                                                    for c in percent_headers_current:
                                                        percent_headers_out.add(ren.get(c, c))
                                                    xlsx_bytes = _df_to_excel_bytes(
                                                        df_x,
                                                        sheet_name="趋势分析",
                                                        title_lines=export_title_lines,
                                                        number_headers=number_headers_out,
                                                        number_formats={
                                                            ren.get("可销月", "可销月"): "0.0",
                                                            ren.get("1月发货额-万元", "1月发货额-万元"): "0.0",
                                                            ren.get("2月发货额-万元", "2月发货额-万元"): "0.0",
                                                            ren.get("3月发货额-万元", "3月发货额-万元"): "0.0",
                                                            ren.get(march_col, march_col): "0.0",
                                                            ren.get(today_col, today_col): "0.0",
                                                            "3月出库": "0.0",
                                                            ren.get(scan_avg_col, scan_avg_col): "0.0",
                                                        },
                                                        percent_headers=percent_headers_out,
                                                        percent_formats={ren.get(scan_rate_col, scan_rate_col): "0.0%"},
                                                        group_headers=True,
                                                    )
                                                    _excel_cache[k_cur] = {
                                                        "bytes": xlsx_bytes,
                                                        "name": sanitize_filename(f"出库趋势分析_分省区_{now_tag}.xlsx"),
                                                    }

                                                if k_all_dist not in _excel_cache:
                                                    df_all_dist = _build_dist_detail_df(all_provinces=True)
                                                    title_all_dist = [
                                                        "月度出库趋势表 - 导出全部经销商",
                                                        filter_line,
                                                        "区域：全部省区（省区→经销商）",
                                                        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                                    ]
                                                    ren = {}
                                                    for c in df_all_dist.columns:
                                                        if c in ("1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"):
                                                            ren[c] = f"发货分析-{c}"
                                                    for c in sel_month_cols:
                                                        if c in df_all_dist.columns:
                                                            ren[c] = f"出库分析-{c}"
                                                    if avg_header in df_all_dist.columns:
                                                        ren[avg_header] = f"出库分析-{avg_header}"
                                                    if march_col in df_all_dist.columns:
                                                        ren[march_col] = f"出库分析-{march_col}"
                                                    if today_col in df_all_dist.columns:
                                                        ren[today_col] = f"出库分析-{today_col}"
                                                    if "趋势类型" in df_all_dist.columns:
                                                        ren["趋势类型"] = "出库分析-趋势类型"
                                                    if "库存" in df_all_dist.columns:
                                                        ren["库存"] = "库存分析-库存"
                                                    if "可销月" in df_all_dist.columns:
                                                        ren["可销月"] = "库存分析-可销月"
                                                    for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                        if _c in df_all_dist.columns:
                                                            ren[_c] = f"新客分析-{_c}"
                                                    if scan_avg_col in df_all_dist.columns:
                                                        ren[scan_avg_col] = f"扫码分析-{scan_avg_header}"
                                                    if scan_rate_col in df_all_dist.columns:
                                                        ren[scan_rate_col] = f"扫码分析-{scan_rate_col}"
                                                    df_all_dist = df_all_dist.rename(columns=ren)

                                                    number_headers_all_dist = set(sel_month_cols + [avg_header, march_col, today_col, "库存", "可销月", "1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"])
                                                    for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                        if ren.get(_c, _c) in df_all_dist.columns:
                                                            number_headers_all_dist.add(_c)
                                                    if ren.get(scan_avg_col, scan_avg_col) in df_all_dist.columns:
                                                        number_headers_all_dist.add(scan_avg_col)
                                                    number_headers_all_dist = set([ren.get(c, c) for c in number_headers_all_dist])
                                                    xlsx_all_dist = _df_to_excel_bytes(
                                                        df_all_dist,
                                                        sheet_name="趋势分析",
                                                        title_lines=title_all_dist,
                                                        number_headers=number_headers_all_dist,
                                                        number_formats={
                                                            ren.get("可销月", "可销月"): "0.0",
                                                            ren.get("1月发货额-万元", "1月发货额-万元"): "0.0",
                                                            ren.get("2月发货额-万元", "2月发货额-万元"): "0.0",
                                                            ren.get("3月发货额-万元", "3月发货额-万元"): "0.0",
                                                            ren.get(march_col, march_col): "0.0",
                                                            ren.get(today_col, today_col): "0.0",
                                                            "3月出库": "0.0",
                                                            ren.get(scan_avg_col, scan_avg_col): "0.0",
                                                        },
                                                        percent_headers=set([ren.get(scan_rate_col, scan_rate_col)] if ren.get(scan_rate_col, scan_rate_col) in df_all_dist.columns else []),
                                                        percent_formats={ren.get(scan_rate_col, scan_rate_col): "0.0%"},
                                                        group_headers=True,
                                                    )
                                                    _excel_cache[k_all_dist] = {
                                                        "bytes": xlsx_all_dist,
                                                        "name": sanitize_filename(f"出库趋势分析_经销商_全部省区_{now_tag}.xlsx"),
                                                    }

                                                if k_all_store not in _excel_cache:
                                                    df_all = _build_store_detail_df(all_provinces=True)
                                                    title_all = [
                                                        "月度出库趋势表 - 导出全部门店明细",
                                                        filter_line,
                                                        "区域：全部省区（省区→经销商→门店）",
                                                        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                                    ]
                                                    ren = {}
                                                    for c in df_all.columns:
                                                        if c in ("1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"):
                                                            ren[c] = f"发货分析-{c}"
                                                    for c in sel_month_cols:
                                                        if c in df_all.columns:
                                                            ren[c] = f"出库分析-{c}"
                                                    if avg_header in df_all.columns:
                                                        ren[avg_header] = f"出库分析-{avg_header}"
                                                    if march_col in df_all.columns:
                                                        ren[march_col] = f"出库分析-{march_col}"
                                                    if today_col in df_all.columns:
                                                        ren[today_col] = f"出库分析-{today_col}"
                                                    if "趋势类型" in df_all.columns:
                                                        ren["趋势类型"] = "出库分析-趋势类型"
                                                    for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                        if _c in df_all.columns:
                                                            ren[_c] = f"新客分析-{_c}"
                                                    if scan_avg_col in df_all.columns:
                                                        ren[scan_avg_col] = f"扫码分析-{scan_avg_header}"
                                                    if scan_rate_col in df_all.columns:
                                                        ren[scan_rate_col] = f"扫码分析-{scan_rate_col}"
                                                    for p_label, _yms in roll_periods:
                                                        for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                                            if c in df_all.columns:
                                                                ren[c] = f"门店类型分析-{c}"
                                                    for i in range(1, len(roll_periods)):
                                                        c = f"{roll_periods[i][0]}变动"
                                                        if c in df_all.columns:
                                                            ren[c] = f"门店类型分析-{c}"
                                                    if "近三周期变化" in df_all.columns:
                                                        ren["近三周期变化"] = "门店类型分析-近三周期变化"
                                                    df_all = df_all.rename(columns=ren)

                                                    number_headers_all = set(sel_month_cols + [avg_header, march_col, today_col, "1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"])
                                                    number_formats_all = {
                                                        ren.get(avg_header, avg_header): "0.0",
                                                        ren.get("1月发货额-万元", "1月发货额-万元"): "0.0",
                                                        ren.get("2月发货额-万元", "2月发货额-万元"): "0.0",
                                                        ren.get("3月发货额-万元", "3月发货额-万元"): "0.0",
                                                        ren.get(march_col, march_col): "0.0",
                                                        "3月出库": "0.0",
                                                        ren.get(scan_avg_col, scan_avg_col): "0.0",
                                                    }
                                                    for p_label, _yms in roll_periods:
                                                        c = f"{p_label}月均出库"
                                                        if ren.get(c, c) in df_all.columns:
                                                            number_headers_all.add(c)
                                                            number_formats_all[ren.get(c, c)] = "0.0"
                                                    for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                        if ren.get(_c, _c) in df_all.columns:
                                                            number_headers_all.add(_c)
                                                    if ren.get(scan_avg_col, scan_avg_col) in df_all.columns:
                                                        number_headers_all.add(scan_avg_col)
                                                    number_headers_all = set([ren.get(c, c) for c in number_headers_all])
                                                    xlsx_all = _df_to_excel_bytes(
                                                        df_all,
                                                        sheet_name="趋势分析",
                                                        title_lines=title_all,
                                                        number_headers=number_headers_all,
                                                        number_formats=number_formats_all,
                                                        percent_headers=set([ren.get(scan_rate_col, scan_rate_col)] if ren.get(scan_rate_col, scan_rate_col) in df_all.columns else []),
                                                        percent_formats={ren.get(scan_rate_col, scan_rate_col): "0.0%"},
                                                        group_headers=True,
                                                    )
                                                    _excel_cache[k_all_store] = {
                                                        "bytes": xlsx_all,
                                                        "name": sanitize_filename(f"出库趋势分析_门店明细_全部省区_{now_tag}.xlsx"),
                                                    }

                                                xlsx_bundle = _merge_single_sheet_workbooks(
                                                    [
                                                        (_excel_cache[k_cur]["bytes"], "分省区"),
                                                        (_excel_cache[k_all_dist]["bytes"], "分经销商"),
                                                        (_excel_cache[k_all_store]["bytes"], "分门店"),
                                                    ]
                                                )
                                                _excel_cache[k_bundle_3s] = {
                                                    "bytes": xlsx_bundle,
                                                    "name": sanitize_filename(f"出库趋势分析_分省区_分经销商_分门店_{now_tag}.xlsx"),
                                                }
                                    with c_a8:
                                        if k_bundle_3s in _excel_cache:
                                            st.download_button(
                                                "导出分省区、客户、门店表格",
                                                data=_excel_cache[k_bundle_3s]["bytes"],
                                                file_name=_excel_cache[k_bundle_3s]["name"],
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                key=f"{export_id}_dl_bundle_3s",
                                            )
                                else:
                                    k_detail = _excel_key("detail_store")
                                    c_d1, c_d2, _ = st.columns([1.3, 2.0, 5.7])
                                    with c_d1:
                                        if st.button("生成门店明细Excel", key=f"{export_id}_gen_detail"):
                                            with st.spinner("正在生成门店明细Excel…"):
                                                df_detail = _build_store_detail_df(all_provinces=False)
                                                title_detail = [
                                                    f"月度出库趋势表 - {region_label}（门店明细）",
                                                    filter_line,
                                                    area_line,
                                                    f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                                ]
                                                ren = {}
                                                for c in df_detail.columns:
                                                    if c in ("1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"):
                                                        ren[c] = f"发货分析-{c}"
                                                for c in sel_month_cols:
                                                    if c in df_detail.columns:
                                                        ren[c] = f"出库分析-{c}"
                                                if avg_header in df_detail.columns:
                                                    ren[avg_header] = f"出库分析-{avg_header}"
                                                if march_col in df_detail.columns:
                                                    ren[march_col] = f"出库分析-{march_col}"
                                                if today_col in df_detail.columns:
                                                    ren[today_col] = f"出库分析-{today_col}"
                                                if "趋势类型" in df_detail.columns:
                                                    ren["趋势类型"] = "出库分析-趋势类型"
                                                for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                    if _c in df_detail.columns:
                                                        ren[_c] = f"新客分析-{_c}"
                                                if scan_avg_col in df_detail.columns:
                                                    ren[scan_avg_col] = f"扫码分析-{scan_avg_header}"
                                                if scan_rate_col in df_detail.columns:
                                                    ren[scan_rate_col] = f"扫码分析-{scan_rate_col}"
                                                for p_label, _yms in roll_periods:
                                                    for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                                        if c in df_detail.columns:
                                                            ren[c] = f"门店类型分析-{c}"
                                                for i in range(1, len(roll_periods)):
                                                    c = f"{roll_periods[i][0]}变动"
                                                    if c in df_detail.columns:
                                                        ren[c] = f"门店类型分析-{c}"
                                                if "近三周期变化" in df_detail.columns:
                                                    ren["近三周期变化"] = "门店类型分析-近三周期变化"
                                                df_detail = df_detail.rename(columns=ren)

                                                number_headers_detail = set(sel_month_cols + [avg_header, march_col, today_col, "1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"])
                                                number_formats_detail = {
                                                    ren.get(avg_header, avg_header): "0.0",
                                                    ren.get("1月发货额-万元", "1月发货额-万元"): "0.0",
                                                    ren.get("2月发货额-万元", "2月发货额-万元"): "0.0",
                                                    ren.get("3月发货额-万元", "3月发货额-万元"): "0.0",
                                                    ren.get(march_col, march_col): "0.0",
                                                    ren.get(today_col, today_col): "0.0",
                                                    "3月出库": "0.0",
                                                    ren.get(scan_avg_col, scan_avg_col): "0.0",
                                                }
                                                for p_label, _yms in roll_periods:
                                                    c = f"{p_label}月均出库"
                                                    if ren.get(c, c) in df_detail.columns:
                                                        number_headers_detail.add(c)
                                                        number_formats_detail[ren.get(c, c)] = "0.0"
                                                for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                    if ren.get(_c, _c) in df_detail.columns:
                                                        number_headers_detail.add(_c)
                                                if ren.get(scan_avg_col, scan_avg_col) in df_detail.columns:
                                                    number_headers_detail.add(scan_avg_col)
                                                number_headers_detail = set([ren.get(c, c) for c in number_headers_detail])
                                                xlsx_detail = _df_to_excel_bytes(
                                                    df_detail,
                                                    sheet_name="趋势分析",
                                                    title_lines=title_detail,
                                                    number_headers=number_headers_detail,
                                                    number_formats=number_formats_detail,
                                                    percent_headers=set([ren.get(scan_rate_col, scan_rate_col)] if ren.get(scan_rate_col, scan_rate_col) in df_detail.columns else []),
                                                    percent_formats={ren.get(scan_rate_col, scan_rate_col): "0.0%"},
                                                    group_headers=True,
                                                )
                                                _excel_cache[k_detail] = {
                                                    "bytes": xlsx_detail,
                                                    "name": sanitize_filename(f"出库趋势分析_门店明细_{region_label}_{now_tag}.xlsx"),
                                                }
                                    with c_d2:
                                        if k_detail in _excel_cache:
                                            st.download_button(
                                                "下载门店明细Excel",
                                                data=_excel_cache[k_detail]["bytes"],
                                                file_name=_excel_cache[k_detail]["name"],
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                key=f"{export_id}_dl_detail",
                                            )
                                _png_state_key = f"{export_id}_png"
                                _png_name_key = f"{export_id}_name"
                                _png_sig_key = f"{export_id}_png_sig"
                                _png_sig_march_sum = 0.0
                                if march_col in df_export.columns:
                                    _png_sig_march_sum = float(pd.to_numeric(df_export[march_col], errors="coerce").fillna(0.0).sum())
                                _png_sig = (
                                    _png_cache_ver,
                                    tuple(df_export.columns),
                                    int(drill_level),
                                    tuple(sel_yms),
                                    str(sel_big or ""),
                                    str(sel_small or ""),
                                    _prod_norm_key,
                                    _prov_sel,
                                    _dist_sel,
                                    _png_sig_march_sum,
                                )
                                if st.session_state.get(_png_sig_key) != _png_sig:
                                    st.session_state.pop(_png_state_key, None)
                                    st.session_state.pop(_png_name_key, None)
                                    st.session_state[_png_sig_key] = _png_sig

                                df_png = df_export.copy()
                                col_types_png = dict(col_types or {})
                                if today_col in df_png.columns:
                                    col_types_png[today_col] = "num"
                                ren_png = {}
                                for c in df_png.columns:
                                    if c in ("1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"):
                                        ren_png[c] = f"发货分析\n{c}"
                                for c in sel_month_cols:
                                    if c in df_png.columns:
                                        ren_png[c] = f"出库分析\n{c}"
                                if march_col in df_png.columns:
                                    ren_png[march_col] = f"出库分析\n{march_col}"
                                if today_col in df_png.columns:
                                    ren_png[today_col] = f"出库分析\n{today_col}"
                                if avg_col in df_png.columns:
                                    ren_png[avg_col] = f"出库分析\n{avg_header}"
                                if "趋势" in df_png.columns:
                                    ren_png["趋势"] = "出库分析\n趋势图"
                                if "趋势类型" in df_png.columns:
                                    ren_png["趋势类型"] = "出库分析\n趋势类型"
                                if "库存" in df_png.columns:
                                    ren_png["库存"] = "库存分析\n库存"
                                if "可销月" in df_png.columns:
                                    ren_png["可销月"] = "库存分析\n可销月"
                                for _c in ["3月新客", "近三月新客", "累计新客"]:
                                    if _c in df_png.columns:
                                        ren_png[_c] = f"新客分析\n{_c}"
                                if scan_avg_col in df_png.columns:
                                    ren_png[scan_avg_col] = f"扫码分析\n{scan_avg_header}"
                                if scan_rate_col in df_png.columns:
                                    ren_png[scan_rate_col] = f"扫码分析\n{scan_rate_col}"
                                for p_label, _yms in roll_periods:
                                    for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                        if c in df_png.columns:
                                            ren_png[c] = f"门店类型分析\n{c}"
                                for i in range(1, len(roll_periods)):
                                    c = f"{roll_periods[i][0]}变动"
                                    if c in df_png.columns:
                                        ren_png[c] = f"门店类型分析\n{c}"
                                if "近三周期变化" in df_png.columns:
                                    ren_png["近三周期变化"] = "门店类型分析\n近三周期变化"
                                if ren_png:
                                    df_png = df_png.rename(columns=ren_png)
                                    col_types_png = {ren_png.get(k, k): v for k, v in col_types_png.items()}

                                if int(drill_level) == 2 and view_dim in df_png.columns:
                                    base_name = df_png[view_dim].fillna("").astype(str)
                                    cols_now = [str(c) for c in df_png.columns.tolist()]
                                    out_idx = next((i for i, c in enumerate(cols_now) if c.startswith("出库分析\n")), None)
                                    if out_idx is not None:
                                        df_png.insert(int(out_idx), "出库分析\n客户名", base_name)
                                        col_types_png["出库分析\n客户名"] = "text"
                                    cols_now = [str(c) for c in df_png.columns.tolist()]
                                    inv_idx = next((i for i, c in enumerate(cols_now) if c.startswith("库存分析\n")), None)
                                    if inv_idx is not None:
                                        df_png.insert(int(inv_idx), "库存分析\n客户名", base_name)
                                        col_types_png["库存分析\n客户名"] = "text"

                                if st.button("生成表格图片（含趋势/颜色）", key=f"{export_id}_btn"):
                                    st.session_state[_png_state_key] = _pil_table_png(df_png, export_title_lines, font_size=16, col_types=col_types_png)
                                    st.session_state[_png_name_key] = f"月度出库趋势_{region_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                                if st.session_state.get(_png_state_key):
                                    st.download_button(
                                        "下载PNG",
                                        data=st.session_state[_png_state_key],
                                        file_name=st.session_state.get(_png_name_key, "月度出库趋势.png"),
                                        mime="image/png",
                                        key=f"{export_id}_dl",
                                    )

                                def _compute_pv(level: int, prov: str | None = None, dist: str | None = None):
                                    scan_yms = [202601, 202602, 202603]
                                    scan_avg_col = "近三月均扫码"
                                    scan_rate_col = "近三月扫码率"
                                    scan_avg_header = "近三月月均扫码（1、2、3）"

                                    view = "省区"
                                    grp = "省区"
                                    d_base = df_trend_base.copy()
                                    d = df_trend_base.copy()
                                    if level == 2:
                                        view = "经销商"
                                        grp = "经销商名称"
                                    elif level == 3:
                                        view = "门店"
                                        grp = "_门店名" if "_门店名" in d.columns else None
                                    if prov:
                                        d_base = d_base[d_base["省区"].astype(str).str.strip() == str(prov).strip()].copy()
                                        d = d[d["省区"].astype(str).str.strip() == str(prov).strip()].copy()
                                    if dist:
                                        d_base = d_base[d_base["经销商名称"].astype(str).str.strip() == str(dist).strip()].copy()
                                        d = d[d["经销商名称"].astype(str).str.strip() == str(dist).strip()].copy()
                                    if grp is None:
                                        return None, view, grp
                                    d_base = d_base[d_base[grp].notna()].copy()
                                    d = d[d[grp].notna()].copy()
                                    if sel_big != '全部' and '_模块大类' in d.columns:
                                        d = d[d['_模块大类'].astype(str).str.strip() == str(sel_big).strip()].copy()
                                    if sel_small != '全部' and '_模块小类' in d.columns:
                                        d = d[d['_模块小类'].astype(str).str.strip() == str(sel_small).strip()].copy()
                                    if sel_prod and '_模块出库产品' in d.columns:
                                        sel_prod_norm = [str(x).strip() for x in sel_prod if str(x).strip()]
                                        if sel_prod_norm:
                                            d = d[d['_模块出库产品'].astype(str).str.strip().isin(sel_prod_norm)].copy()
                                    d_all = d.copy()
                                    if d.empty:
                                        d = pd.DataFrame(columns=[grp, "_ym", "数量(箱)"])
                                    else:
                                        d = d[d["_ym"].isin(sel_yms)].copy()
                                        d["数量(箱)"] = pd.to_numeric(d.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                    agg2 = (
                                        d.groupby([grp, "_ym"], as_index=False)["数量(箱)"]
                                        .sum()
                                        .rename(columns={grp: view})
                                    )
                                    pv2 = agg2.pivot(index=view, columns="_ym", values="数量(箱)").fillna(0.0)

                                    df_names2 = df_trend_universe.copy()
                                    if prov and '省区' in df_names2.columns:
                                        df_names2 = df_names2[df_names2['省区'].astype(str).str.strip() == str(prov).strip()].copy()
                                    if dist and '经销商名称' in df_names2.columns:
                                        df_names2 = df_names2[df_names2['经销商名称'].astype(str).str.strip() == str(dist).strip()].copy()
                                    if sel_big != '全部' and '_模块大类' in df_names2.columns:
                                        df_names2 = df_names2[df_names2['_模块大类'].astype(str).str.strip() == str(sel_big).strip()].copy()
                                    if sel_small != '全部' and '_模块小类' in df_names2.columns:
                                        df_names2 = df_names2[df_names2['_模块小类'].astype(str).str.strip() == str(sel_small).strip()].copy()
                                    if sel_prod and '_模块出库产品' in df_names2.columns:
                                        sel_prod_norm = [str(x).strip() for x in sel_prod if str(x).strip()]
                                        if sel_prod_norm:
                                            df_names2 = df_names2[df_names2['_模块出库产品'].astype(str).str.strip().isin(sel_prod_norm)].copy()

                                    invalid_names = {'', 'nan', 'none', 'null'}
                                    if grp == "省区" and '省区' in df_names2.columns:
                                        base2 = df_names2['省区'].dropna().astype(str).str.strip().unique().tolist()
                                    elif grp == "经销商名称" and '经销商名称' in df_names2.columns:
                                        base2 = df_names2['经销商名称'].dropna().astype(str).str.strip().unique().tolist()
                                    else:
                                        base2 = df_names2[grp].dropna().astype(str).str.strip().unique().tolist() if grp in df_names2.columns else []
                                    base2 = sorted([x for x in base2 if x and x.lower() not in invalid_names])
                                    if base2:
                                        df_skel = pd.DataFrame({view: base2})
                                        pv_reset2 = pv2.reset_index()
                                        if view not in pv_reset2.columns and len(pv_reset2.columns) > 0:
                                            pv_reset2.rename(columns={pv_reset2.columns[0]: view}, inplace=True)
                                        if view in pv_reset2.columns:
                                            pv_reset2[view] = pv_reset2[view].astype(str).str.strip()
                                            df_skel[view] = df_skel[view].astype(str).str.strip()
                                            pv2 = df_skel.merge(pv_reset2, on=view, how="left").fillna(0.0).set_index(view)
                                        else:
                                            pv2 = df_skel.set_index(view)
                                    for ym in sel_yms:
                                        if ym not in pv2.columns:
                                            pv2[ym] = 0.0
                                    pv2 = pv2[sel_yms]
                                    pv2.columns = sel_month_cols
                                    pv2["_合计"] = pv2.sum(axis=1)
                                    pv2 = pv2.sort_values("_合计", ascending=False).reset_index()
                                    pv2.drop(columns=["_合计"], inplace=True, errors="ignore")
                                    pv2[avg_col] = pv2[first3_cols].mean(axis=1) if len(first3_cols) >= 1 else 0.0

                                    if False and level == 2:
                                        pv2["1月发货件数"] = 0.0
                                        pv2["2月发货件数"] = 0.0
                                        pv2["3月发货件数"] = 0.0
                                        pv2["4月发货件数"] = 0.0
                                        if df_perf_raw is not None and not getattr(df_perf_raw, "empty", True):
                                            sp = df_perf_raw.copy()
                                            if "客户简称" in sp.columns:
                                                sp["经销商名称"] = sp["客户简称"].fillna(sp["经销商名称"])
                                            for c in ["省区", "经销商名称", "大类", "小类", "小类码", "中类", "重量"]:
                                                if c in sp.columns:
                                                    if c == "经销商名称":
                                                        sp[c] = sp[c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                    else:
                                                        sp[c] = sp[c].fillna("").astype(str).str.strip()
                                            sp["年份"] = pd.to_numeric(sp.get("年份", 0), errors="coerce").fillna(0).astype(int)
                                            sp["月份"] = pd.to_numeric(sp.get("月份", 0), errors="coerce").fillna(0).astype(int)
                                            sp = sp[(sp["年份"] > 0) & (sp["月份"].between(1, 12))].copy()
                                            sp["_ym"] = (sp["年份"] * 100 + sp["月份"]).astype(int)
                                            ship_yms = [202601, 202602, 202603, 202604]
                                            sp = sp[sp["_ym"].isin(ship_yms)].copy()
                                            sp["发货箱数"] = pd.to_numeric(sp.get("发货箱数", 0), errors="coerce").fillna(0.0)
                                            if prov and "省区" in sp.columns:
                                                sp = sp[sp["省区"].astype(str).str.strip() == str(prov).strip()].copy()
                                            if sel_big != "全部" and "大类" in sp.columns:
                                                sp = sp[sp["大类"].astype(str).str.strip() == str(sel_big).strip()].copy()
                                            if sel_small != "全部":
                                                _sel_s = str(sel_small).strip()
                                                _m = re.search(r"(\d{3})", _sel_s)
                                                if _m:
                                                    _code_i = int(_m.group(1))
                                                    _src = "小类码" if "小类码" in sp.columns else ("小类" if "小类" in sp.columns else ("中类" if "中类" in sp.columns else ("重量" if "重量" in sp.columns else None)))
                                                    if _src is not None and _src in sp.columns:
                                                        _digits = sp[_src].astype(str).str.extract(r"(\d+)")[0]
                                                        _v = pd.to_numeric(_digits, errors="coerce").fillna(-1).astype(int)
                                                        sp = sp[_v == _code_i].copy()
                                                else:
                                                    if "小类码" in sp.columns:
                                                        sp = sp[sp["小类码"].astype(str).str.strip() == _sel_s].copy()
                                                    elif "小类" in sp.columns:
                                                        sp = sp[sp["小类"].astype(str).str.strip() == _sel_s].copy()
                                                    elif "中类" in sp.columns:
                                                        sp = sp[sp["中类"].astype(str).str.strip() == _sel_s].copy()
                                                    elif "重量" in sp.columns:
                                                        sp = sp[sp["重量"].astype(str).str.strip() == _sel_s].copy()

                                            dist_map2 = {}
                                            if df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                                if "经销商全称" in df_stock_raw.columns and "经销商名称" in df_stock_raw.columns:
                                                    _m2 = (
                                                        df_stock_raw[["经销商全称", "经销商名称"]]
                                                        .dropna()
                                                        .astype(str)
                                                        .apply(lambda col: col.str.replace(r"\s+", "", regex=True))
                                                        .drop_duplicates()
                                                    )
                                                    dist_map2 = dict(zip(_m2["经销商全称"].tolist(), _m2["经销商名称"].tolist()))
                                            sp["_k_dist"] = sp["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map2:
                                                sp["_k_dist"] = sp["_k_dist"].map(dist_map2).fillna(sp["_k_dist"])
                                            gsp = sp.groupby(["_k_dist", "_ym"], as_index=False).agg({"发货箱数": "sum"})
                                            pv2["_k_dist"] = pv2[view].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map2:
                                                pv2["_k_dist"] = pv2["_k_dist"].map(dist_map2).fillna(pv2["_k_dist"])
                                            ship_cols_qty = ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]
                                            for i, _ym in enumerate(ship_yms):
                                                _cq = ship_cols_qty[i]
                                                _sub = gsp[gsp["_ym"] == int(_ym)]
                                                m_qty = dict(zip(_sub["_k_dist"].tolist(), _sub["发货箱数"].tolist()))
                                                pv2[_cq] = pv2["_k_dist"].map(m_qty)
                                            pv2.drop(columns=["_k_dist"], inplace=True, errors="ignore")
                                            for c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                                if c not in pv2.columns:
                                                    pv2[c] = 0.0
                                                pv2[c] = pd.to_numeric(pv2[c], errors="coerce").fillna(0.0)

                                    if level == 2 and df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                        _s = df_stock_raw.copy()
                                        if "省区" not in _s.columns and "省区名称" in _s.columns:
                                            _s["省区"] = _s["省区名称"]
                                        for _c in ["省区", "经销商名称", "产品大类", "产品小类", "重量"]:
                                            if _c in _s.columns:
                                                if _c == "经销商名称":
                                                    _s[_c] = _s[_c].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                else:
                                                    _s[_c] = _s[_c].fillna("").astype(str).str.strip()
                                        if "箱数" in _s.columns:
                                            _s["箱数"] = pd.to_numeric(_s["箱数"], errors="coerce").fillna(0.0)
                                        if prov and "省区" in _s.columns:
                                            _s = _s[_s["省区"].astype(str).str.strip() == str(prov).strip()].copy()
                                        if sel_big != "全部" and "产品大类" in _s.columns:
                                            _s = _s[_s["产品大类"].astype(str).str.strip() == str(sel_big).strip()].copy()
                                        if sel_small != "全部" and "重量" in _s.columns:
                                            _sel_s = str(sel_small).strip()
                                            _m = re.search(r"(\d{3})", _sel_s)
                                            if _m:
                                                _w = _s["重量"].astype(str).str.extract(r"(\d{3})")[0].fillna("").astype(str)
                                                _s = _s[_w == _m.group(1)].copy()
                                            else:
                                                _s = _s[_s["重量"].astype(str).str.strip() == _sel_s].copy()
                                        if not _s.empty and "箱数" in _s.columns:
                                            _inv = _s.groupby("经销商名称", as_index=False)["箱数"].sum().rename(columns={"经销商名称": view, "箱数": "库存"})
                                            if view in pv2.columns:
                                                pv2 = pv2.merge(_inv, on=view, how="left")
                                                pv2["库存"] = pd.to_numeric(pv2.get("库存", 0), errors="coerce").fillna(0.0)
                                                _avg_v = pd.to_numeric(pv2[avg_col], errors="coerce").fillna(0.0)
                                                pv2["可销月"] = np.where(_avg_v > 0, pv2["库存"] / _avg_v, 0.0)
                                                pv2["可销月"] = pd.to_numeric(pv2.get("可销月", 0), errors="coerce").fillna(0.0).round(1)
                                        else:
                                            pv2["库存"] = 0.0
                                            pv2["可销月"] = 0.0

                                    if df_newcust_raw is not None and not getattr(df_newcust_raw, "empty", True) and "_ym" in df_newcust_raw.columns:
                                        nc = df_newcust_raw.copy()
                                        if "省区" in nc.columns:
                                            nc["省区"] = nc["省区"].fillna("").astype(str).str.strip()
                                        if "经销商名称" in nc.columns:
                                            nc["经销商名称"] = nc["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                        if "门店名称" in nc.columns:
                                            nc["门店名称"] = nc["门店名称"].fillna("").astype(str).str.strip()
                                        if "新客数" in nc.columns:
                                            nc["新客数"] = pd.to_numeric(nc["新客数"], errors="coerce").fillna(0.0)
                                        nc["_ym"] = pd.to_numeric(nc["_ym"], errors="coerce").fillna(0).astype(int)
                                        nc = nc[nc["_ym"].between(200001, 209912)].copy()
                                        if prov and "省区" in nc.columns:
                                            nc = nc[nc["省区"].astype(str).str.strip() == str(prov).strip()].copy()
                                        if level == 3 and dist and "经销商名称" in nc.columns:
                                            nc = nc[nc["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True) == str(dist).strip().replace(" ", "")].copy()
                                        if not nc.empty:
                                            _yms = sorted([int(x) for x in nc["_ym"].dropna().astype(int).tolist() if 200001 <= int(x) <= 209912])
                                            _yy = int(max(_yms) // 100)
                                            latest_ym = int(_yy * 100 + 4)
                                            prev3 = [int(_yy * 100 + 1), int(_yy * 100 + 2), int(_yy * 100 + 3)]
                                            key_col = "经销商名称" if level == 2 else "门店名称"
                                            if key_col in nc.columns and view in pv2.columns:
                                                cur = (
                                                    nc[nc["_ym"] == latest_ym]
                                                    .groupby(key_col, as_index=False)["新客数"]
                                                    .sum()
                                                    .rename(columns={key_col: view, "新客数": "4月新客"})
                                                )
                                                p3 = (
                                                    nc[nc["_ym"].isin(prev3)]
                                                    .groupby(key_col, as_index=False)["新客数"]
                                                    .sum()
                                                    .rename(columns={key_col: view, "新客数": "近三月新客"})
                                                )
                                                cum = (
                                                    nc.groupby(key_col, as_index=False)["新客数"]
                                                    .sum()
                                                    .rename(columns={key_col: view, "新客数": "整体新客"})
                                                )
                                                pv2[view] = pv2[view].fillna("").astype(str).str.strip()
                                                cur[view] = cur[view].fillna("").astype(str).str.strip()
                                                p3[view] = p3[view].fillna("").astype(str).str.strip()
                                                cum[view] = cum[view].fillna("").astype(str).str.strip()
                                                pv2 = pv2.merge(cur, on=view, how="left")
                                                pv2 = pv2.merge(p3, on=view, how="left")
                                                pv2 = pv2.merge(cum, on=view, how="left")
                                                for _c in ["4月新客", "近三月新客", "整体新客"]:
                                                    if _c in pv2.columns:
                                                        pv2[_c] = pd.to_numeric(pv2[_c], errors="coerce").fillna(0.0)
                                                    else:
                                                        pv2[_c] = 0.0
                                                pv2["累计新客"] = pd.to_numeric(pv2.get("整体新客", 0), errors="coerce").fillna(0.0)
                                    if df_scan_raw is not None and not getattr(df_scan_raw, "empty", True):
                                        s = df_scan_raw.copy()
                                        if "经销商名称" in s.columns:
                                            s["经销商名称"] = s["经销商名称"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                        if "省区" in s.columns:
                                            s["省区"] = s["省区"].fillna("").astype(str).str.strip()
                                        if "门店名称" in s.columns:
                                            s["门店名称"] = s["门店名称"].fillna("").astype(str).str.strip()
                                        if "产品大类" in s.columns:
                                            s["产品大类"] = s["产品大类"].fillna("").astype(str).str.strip()
                                        if "产品小类" in s.columns:
                                            s["产品小类"] = s["产品小类"].fillna("").astype(str).str.strip()
                                        s["年份"] = pd.to_numeric(s.get("年份", 0), errors="coerce").fillna(0).astype(int)
                                        s["月份"] = pd.to_numeric(s.get("月份", 0), errors="coerce").fillna(0).astype(int)
                                        s = s[(s["年份"] > 0) & (s["月份"].between(1, 12))].copy()
                                        if prov and "省区" in s.columns:
                                            s = s[s["省区"].astype(str).str.strip() == str(prov).strip()].copy()
                                        if level == 3 and dist and "经销商名称" in s.columns:
                                            _dist_norm = re.sub(r"\s+", "", str(dist).strip())
                                            s = s[s["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True) == _dist_norm].copy()
                                        if sel_big != "全部":
                                            _sb = str(sel_big).strip()
                                            if _sb == "雅系列":
                                                if "产品小类" in s.columns:
                                                    s = s[s["产品小类"].astype(str).str.contains(r"(雅赋|雅耀|雅舒|雅护)", regex=True)].copy()
                                                elif "产品大类" in s.columns:
                                                    s = s[s["产品大类"].astype(str).str.strip() == _sb].copy()
                                            elif "产品大类" in s.columns:
                                                s = s[s["产品大类"].astype(str).str.strip() == _sb].copy()
                                        if sel_small != "全部" and "产品小类" in s.columns:
                                            _sel_s = str(sel_small).strip()
                                            _m = re.search(r"(\d{3})", _sel_s)
                                            if _m:
                                                _w = s["产品小类"].astype(str).str.extract(r"(\d{3})")[0].fillna("").astype(str)
                                                s = s[_w == _m.group(1)].copy()
                                            else:
                                                s = s[s["产品小类"].astype(str).str.strip() == _sel_s].copy()
                                        if not s.empty and view in pv2.columns:
                                            s["_ym"] = (s["年份"] * 100 + s["月份"]).astype(int)
                                            key_col = "经销商名称" if level == 2 else "门店名称"
                                            pv2[scan_avg_col] = 0.0
                                            pv2[scan_rate_col] = 0.0
                                            if key_col in s.columns:
                                                pv2[view] = pv2[view].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                for _ym in scan_yms:
                                                    _ym_i = int(_ym)
                                                    c_scan = f"_scan_{_ym_i}"
                                                    s_m = s[pd.to_numeric(s["_ym"], errors="coerce").fillna(0).astype(int) == _ym_i].copy()
                                                    if not s_m.empty:
                                                        scan_agg = (
                                                            s_m.groupby(key_col, as_index=False)
                                                            .size()
                                                            .rename(columns={key_col: view, "size": "_扫码听数"})
                                                        )
                                                        scan_agg[c_scan] = pd.to_numeric(scan_agg["_扫码听数"], errors="coerce").fillna(0.0) / 6.0
                                                        scan_agg = scan_agg[[view, c_scan]].copy()
                                                        scan_agg[view] = scan_agg[view].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                        pv2 = pv2.merge(scan_agg, on=view, how="left")
                                                    else:
                                                        pv2[c_scan] = 0.0
                                                    pv2[c_scan] = pd.to_numeric(pv2.get(c_scan, 0), errors="coerce").fillna(0.0)
                                                _scan_cols = [f"_scan_{int(_ym)}" for _ym in scan_yms]
                                                pv2[scan_avg_col] = pv2[_scan_cols].mean(axis=1)
                                            pv2[scan_avg_col] = pd.to_numeric(pv2.get(scan_avg_col, 0), errors="coerce").fillna(0.0)
                                            if avg_col in pv2.columns:
                                                _out_avg = pd.to_numeric(pv2[avg_col], errors="coerce").fillna(0.0)
                                                pv2[scan_rate_col] = np.where(_out_avg > 0, pv2[scan_avg_col] / _out_avg, 0.0)
                                            else:
                                                pv2[scan_rate_col] = 0.0
                                            pv2[scan_rate_col] = pd.to_numeric(pv2.get(scan_rate_col, 0), errors="coerce").fillna(0.0)

                                    if level == 3 and store_geo_df is not None and not getattr(store_geo_df, "empty", True) and (view in pv2.columns):
                                        try:
                                            pv2["_k_store_geo"] = pv2[view].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            pv2 = pv2.merge(store_geo_df, on="_k_store_geo", how="left")
                                            pv2.drop(columns=["_k_store_geo"], inplace=True, errors="ignore")
                                            pv2["市"] = pv2.get("市", "").fillna("").astype(str).str.strip()
                                            pv2["区/县"] = pv2.get("区/县", "").fillna("").astype(str).str.strip()
                                            pv2["门店状态"] = pv2.get("门店状态", "").fillna("").astype(str).str.strip()
                                        except Exception:
                                            pv2["市"] = pv2.get("市", "").fillna("").astype(str).str.strip()
                                            pv2["区/县"] = pv2.get("区/县", "").fillna("").astype(str).str.strip()
                                            pv2["门店状态"] = pv2.get("门店状态", "").fillna("").astype(str).str.strip()

                                    try:
                                        dm = d_all[pd.to_numeric(d_all["_ym"], errors="coerce").fillna(0).astype(int) == int(march_ym)].copy()
                                        if not dm.empty:
                                            dm["数量(箱)"] = pd.to_numeric(dm.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                            dist_map_march = {}
                                            if grp == "经销商名称" and df_stock_raw is not None and not getattr(df_stock_raw, "empty", True):
                                                if "经销商全称" in df_stock_raw.columns and "经销商名称" in df_stock_raw.columns:
                                                    _m2 = (
                                                        df_stock_raw[["经销商全称", "经销商名称"]]
                                                        .dropna()
                                                        .astype(str)
                                                        .apply(lambda col: col.str.replace(r"\s+", "", regex=True))
                                                        .drop_duplicates()
                                                    )
                                                    dist_map_march = dict(zip(_m2["经销商全称"].tolist(), _m2["经销商名称"].tolist()))

                                            dm["_k_march"] = dm[grp].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map_march:
                                                dm["_k_march"] = dm["_k_march"].map(dist_map_march).fillna(dm["_k_march"])
                                            gm = dm.groupby(["_k_march"], as_index=False)["数量(箱)"].sum().rename(columns={"数量(箱)": march_col})
                                            pv2["_k_march"] = pv2[view].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                            if dist_map_march and grp == "经销商名称":
                                                pv2["_k_march"] = pv2["_k_march"].map(dist_map_march).fillna(pv2["_k_march"])
                                            pv2 = pv2.merge(gm[["_k_march", march_col]], on="_k_march", how="left", suffixes=("", "_y"))
                                            _m_y = f"{march_col}_y"
                                            if _m_y in pv2.columns:
                                                pv2.drop(columns=[march_col], inplace=True, errors="ignore")
                                                pv2.rename(columns={_m_y: march_col}, inplace=True)
                                            pv2.drop(columns=["_k_march"], inplace=True, errors="ignore")
                                            if march_col in pv2.columns:
                                                pv2[march_col] = pd.to_numeric(pv2[march_col], errors="coerce").fillna(0.0)
                                            else:
                                                pv2[march_col] = 0.0
                                        else:
                                            pv2[march_col] = 0.0
                                    except Exception:
                                        if march_col in pv2.columns:
                                            pv2[march_col] = pd.to_numeric(pv2[march_col], errors="coerce").fillna(0.0)
                                        else:
                                            pv2[march_col] = 0.0

                                    pv2[today_col] = 0.0
                                    if today_day is not None:
                                        try:
                                            ddm = d_all[pd.to_numeric(d_all["_ym"], errors="coerce").fillna(0).astype(int) == int(march_ym)].copy()
                                            if not ddm.empty and "_日" in ddm.columns:
                                                ddm["_日"] = pd.to_numeric(ddm["_日"], errors="coerce")
                                                ddm = ddm[ddm["_日"].notna()].copy()
                                                ddm["_日"] = ddm["_日"].astype(int)
                                                ddm = ddm[ddm["_日"] == int(today_day)].copy()
                                            if not ddm.empty:
                                                ddm["数量(箱)"] = pd.to_numeric(ddm.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                                ddm["_k_today"] = ddm[grp].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                if dist_map_march:
                                                    ddm["_k_today"] = ddm["_k_today"].map(dist_map_march).fillna(ddm["_k_today"])
                                                gd = ddm.groupby(["_k_today"], as_index=False)["数量(箱)"].sum().rename(columns={"数量(箱)": today_col})
                                                pv2["_k_today"] = pv2[view].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                if dist_map_march and grp == "经销商名称":
                                                    pv2["_k_today"] = pv2["_k_today"].map(dist_map_march).fillna(pv2["_k_today"])
                                                pv2 = pv2.merge(gd[["_k_today", today_col]], on="_k_today", how="left", suffixes=("", "_y"))
                                                _d_y = f"{today_col}_y"
                                                if _d_y in pv2.columns:
                                                    pv2.drop(columns=[today_col], inplace=True, errors="ignore")
                                                    pv2.rename(columns={_d_y: today_col}, inplace=True)
                                                pv2.drop(columns=["_k_today"], inplace=True, errors="ignore")
                                                pv2[today_col] = pd.to_numeric(pv2.get(today_col, 0), errors="coerce").fillna(0.0)
                                        except Exception:
                                            pv2[today_col] = pd.to_numeric(pv2.get(today_col, 0), errors="coerce").fillna(0.0)

                                    spark_vals2 = pv2[trend_base_cols].values.tolist() if trend_base_cols else [[] for _ in range(len(pv2))]
                                    pv2["_趋势数据"] = [json.dumps([float(x) for x in row]) for row in spark_vals2]
                                    pv2["趋势"] = pv2["_趋势数据"]

                                    if level == 3 and grp:
                                        try:
                                            need_yms = []
                                            for _, yms in roll_periods:
                                                need_yms += list(yms)
                                            need_yms = sorted(set([int(x) for x in need_yms]))
                                            d_roll = d_all.copy()
                                            d_roll = d_roll[d_roll["_ym"].isin(need_yms)].copy()
                                            if not d_roll.empty:
                                                d_roll["数量(箱)"] = pd.to_numeric(d_roll.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                                d_roll[grp] = d_roll[grp].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                d_roll = d_roll[d_roll[grp] != ""].copy()
                                                r_agg = (
                                                    d_roll.groupby([grp, "_ym"], as_index=False)["数量(箱)"]
                                                    .sum()
                                                    .rename(columns={grp: view})
                                                )
                                                r_pv = r_agg.pivot(index=view, columns="_ym", values="数量(箱)").fillna(0.0)
                                                for ym in need_yms:
                                                    if ym not in r_pv.columns:
                                                        r_pv[ym] = 0.0
                                                r_pv = r_pv[need_yms].reset_index()
                                                for p_label, yms in roll_periods:
                                                    cols = [int(x) for x in yms]
                                                    r_pv[f"{p_label}月均出库"] = r_pv[cols].sum(axis=1) / 3.0
                                                    r_pv[f"{p_label}门店类型"] = r_pv[f"{p_label}月均出库"].apply(_classify_store_abcd)
                                                for i in range(1, len(roll_periods)):
                                                    prev_label = roll_periods[i - 1][0]
                                                    cur_label = roll_periods[i][0]
                                                    r_pv[f"{cur_label}变动"] = r_pv.apply(lambda r: _store_change(r.get(f"{prev_label}门店类型"), r.get(f"{cur_label}门店类型")), axis=1)
                                                if len(roll_periods) >= 3:
                                                    p1, p2, p3 = roll_periods[-3][0], roll_periods[-2][0], roll_periods[-1][0]
                                                    r_pv["近三周期变化"] = r_pv.apply(
                                                        lambda r: _trend3_label(
                                                            r.get(f"{p1}门店类型"),
                                                            r.get(f"{p2}门店类型"),
                                                            r.get(f"{p3}门店类型"),
                                                        ),
                                                        axis=1,
                                                    )
                                                else:
                                                    r_pv["近三周期变化"] = ""
                                                for p_label, _yms in roll_periods:
                                                    c_avg = f"{p_label}月均出库"
                                                    if c_avg in r_pv.columns:
                                                        r_pv[c_avg] = pd.to_numeric(r_pv[c_avg], errors="coerce").fillna(0.0).round(1)

                                                pv2[view] = pv2[view].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                r_pv[view] = r_pv[view].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
                                                keep_cols = [view]
                                                for p_label, _yms in roll_periods:
                                                    keep_cols += [f"{p_label}月均出库", f"{p_label}门店类型"]
                                                for i in range(1, len(roll_periods)):
                                                    keep_cols.append(f"{roll_periods[i][0]}变动")
                                                keep_cols.append("近三周期变化")
                                                keep_cols = [c for c in keep_cols if c in r_pv.columns]
                                                r_pv = r_pv[keep_cols].copy()
                                                pv2 = pv2.merge(r_pv, on=view, how="left")
                                        except Exception:
                                            pass
                                    return pv2, view, grp

                                batch_id = f"out_m_batch_{drill_level}"
                                batch_export_ver = 3
                                if "out_m_zip_cache" not in st.session_state:
                                    st.session_state.out_m_zip_cache = {}
                                _zip_cache = st.session_state.out_m_zip_cache

                                batch_sig_key = f"{batch_id}_sig"
                                _prod_norm_key_zip = tuple(sorted([str(x).strip() for x in (sel_prod or []) if str(x).strip()]))
                                batch_sig = (
                                    batch_export_ver,
                                    int(drill_level),
                                    tuple(sel_yms),
                                    tuple(first3_cols),
                                    str(march_col or ""),
                                    str(avg_header or ""),
                                    str(sel_big or ""),
                                    str(sel_small or ""),
                                    _prod_norm_key_zip,
                                    str(st.session_state.get("out_m_selected_prov") or ""),
                                )
                                k_zip = ("zip",) + (batch_id,) + batch_sig
                                if st.session_state.get(batch_sig_key) != batch_sig:
                                    _zip_cache.pop(k_zip, None)
                                    st.session_state[batch_sig_key] = batch_sig
                                if drill_level in (1, 2):
                                    label = "全部导出省区图片ZIP" if drill_level == 1 else "全部导出经销商图片ZIP"
                                    if st.button(label, key=f"{batch_id}_btn"):
                                        scan_yms = [202601, 202602, 202603]
                                        scan_avg_col = "近三月均扫码"
                                        scan_rate_col = "近三月扫码率"
                                        scan_avg_header = "近三月月均扫码（1、2、3）"

                                        targets = []
                                        if drill_level == 1:
                                            targets = base_names
                                        else:
                                            targets = base_names
                                        zip_buf = io.BytesIO()
                                        with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                                            prog = st.progress(0)
                                            total = max(1, len(targets))
                                            for idx, name in enumerate(targets):
                                                if drill_level == 1:
                                                    pv_t, v_t, _ = _compute_pv(2, prov=name, dist=None)
                                                    region_t = str(name)
                                                else:
                                                    pv_t, v_t, _ = _compute_pv(3, prov=st.session_state.get("out_m_selected_prov"), dist=name)
                                                    region_t = str(name)
                                                if pv_t is None or pv_t.empty:
                                                    prog.progress(int((idx + 1) * 100 / total))
                                                    continue
                                                export_cols_t = [v_t]
                                                for _c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                                    if _c in pv_t.columns:
                                                        export_cols_t.append(_c)
                                                export_cols_t += [c for c in first3_cols if c in pv_t.columns]
                                                if avg_col in pv_t.columns:
                                                    export_cols_t.append(avg_col)
                                                if "趋势" in pv_t.columns:
                                                    export_cols_t.append("趋势")
                                                if march_col in pv_t.columns:
                                                    export_cols_t.append(march_col)
                                                if today_col in pv_t.columns:
                                                    export_cols_t.append(today_col)
                                                if "库存" in pv_t.columns:
                                                    export_cols_t.append("库存")
                                                if "可销月" in pv_t.columns:
                                                    export_cols_t.append("可销月")
                                                for _c in ["4月新客", "近三月新客", "累计新客"]:
                                                    if _c in pv_t.columns:
                                                        export_cols_t.append(_c)
                                                if last_col and (last_col in pv_t.columns) and (last_col not in export_cols_t):
                                                    export_cols_t.append(last_col)
                                                if "完成率" in pv_t.columns:
                                                    export_cols_t.append("完成率")
                                                if scan_avg_col in pv_t.columns:
                                                    export_cols_t.append(scan_avg_col)
                                                if scan_rate_col in pv_t.columns:
                                                    export_cols_t.append(scan_rate_col)
                                                if v_t == "门店":
                                                    for p_label, _yms in roll_periods:
                                                        for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                                            if c in pv_t.columns:
                                                                export_cols_t.append(c)
                                                    for i in range(1, len(roll_periods)):
                                                        c = f"{roll_periods[i][0]}变动"
                                                        if c in pv_t.columns:
                                                            export_cols_t.append(c)
                                                    if "近三周期变化" in pv_t.columns:
                                                        export_cols_t.append("近三周期变化")
                                                for _c in ["市", "区/县"]:
                                                    if _c in pv_t.columns:
                                                        export_cols_t.append(_c)
                                                df_t = pv_t[export_cols_t].copy()

                                                try:
                                                    total_row = {v_t: "合计"}
                                                    for _c in export_cols_t:
                                                        if _c in (v_t, "趋势", "趋势类型", "_趋势数据", "可销月", "完成率", scan_rate_col, "市", "区/县"):
                                                            continue
                                                        if str(_c).startswith("_"):
                                                            continue
                                                        total_row[_c] = float(pd.to_numeric(pv_t[_c], errors="coerce").fillna(0.0).sum()) if _c in pv_t.columns else 0.0
                                                    if avg_col in export_cols_t and len(first3_cols) >= 1:
                                                        _t3 = [float(pd.to_numeric(pv_t[c], errors="coerce").fillna(0.0).sum()) for c in first3_cols if c in pv_t.columns]
                                                        total_row[avg_col] = float(np.mean(_t3)) if _t3 else 0.0
                                                    if "可销月" in export_cols_t:
                                                        _a = float(total_row.get(avg_col, 0.0) or 0.0)
                                                        _k = float(total_row.get("库存", 0.0) or 0.0)
                                                        total_row["可销月"] = round((_k / _a), 1) if _a > 0 else 0.0
                                                    if "完成率" in export_cols_t:
                                                        _feb = ym_to_label.get(202602)
                                                        _a = float(total_row.get(avg_col, 0.0) or 0.0)
                                                        _f = float(total_row.get(_feb, 0.0) or 0.0) if _feb else 0.0
                                                        total_row["完成率"] = (_f / _a) if _a > 0 else 0.0
                                                    if scan_avg_col in export_cols_t:
                                                        _t_scan = []
                                                        for _ym in scan_yms:
                                                            c = f"_scan_{int(_ym)}"
                                                            if c in pv_t.columns:
                                                                _t_scan.append(float(pd.to_numeric(pv_t[c], errors="coerce").fillna(0.0).sum()))
                                                        total_row[scan_avg_col] = float(np.mean(_t_scan)) if _t_scan else 0.0
                                                    if scan_rate_col in export_cols_t and scan_avg_col in total_row and avg_col in total_row:
                                                        denom = float(total_row.get(avg_col, 0.0) or 0.0)
                                                        total_row[scan_rate_col] = (float(total_row.get(scan_avg_col, 0.0) or 0.0) / denom) if denom > 0 else 0.0
                                                    if "趋势" in export_cols_t:
                                                        _spark_vals = [float(pd.to_numeric(pv_t[c], errors="coerce").fillna(0.0).sum()) for c in trend_base_cols if c in pv_t.columns]
                                                        total_row["趋势"] = json.dumps([float(x) for x in _spark_vals]) if _spark_vals else json.dumps([])
                                                    df_t = pd.concat([df_t, pd.DataFrame([total_row])], ignore_index=True)
                                                except Exception:
                                                    pass

                                                col_types_t = {v_t: "text"}
                                                for c in sel_month_cols:
                                                    if c in df_t.columns:
                                                        col_types_t[c] = "num"
                                                if avg_col in df_t.columns:
                                                    col_types_t[avg_col] = "num"
                                                if march_col in df_t.columns:
                                                    col_types_t[march_col] = "num"
                                                if today_col in df_t.columns:
                                                    col_types_t[today_col] = "num"
                                                if "库存" in df_t.columns:
                                                    col_types_t["库存"] = "num"
                                                if "可销月" in df_t.columns:
                                                    col_types_t["可销月"] = "num"
                                                for _c in ["1月发货件数", "2月发货件数", "3月发货件数", "4月发货件数"]:
                                                    if _c in df_t.columns:
                                                        col_types_t[_c] = "num"
                                                for _c in ["4月新客", "近三月新客", "累计新客"]:
                                                    if _c in df_t.columns:
                                                        col_types_t[_c] = "num"
                                                if scan_avg_col in df_t.columns:
                                                    col_types_t[scan_avg_col] = "num"
                                                if last_col and last_col in df_t.columns:
                                                    col_types_t[last_col] = "num"
                                                if "完成率" in df_t.columns:
                                                    col_types_t["完成率"] = "pct"
                                                if scan_rate_col in df_t.columns:
                                                    col_types_t[scan_rate_col] = "pct"
                                                for _c in ["市", "区/县"]:
                                                    if _c in df_t.columns:
                                                        col_types_t[_c] = "text"
                                                if "趋势" in df_t.columns:
                                                    col_types_t["趋势"] = "spark"

                                                ren_t = {}
                                                if avg_col in df_t.columns:
                                                    ren_t[avg_col] = avg_header
                                                if scan_avg_col in df_t.columns:
                                                    ren_t[scan_avg_col] = scan_avg_header
                                                if ren_t:
                                                    df_t = df_t.rename(columns=ren_t)
                                                    col_types_t = {ren_t.get(k, k): v for k, v in col_types_t.items()}
                                                if v_t == "门店":
                                                    for p_label, _yms in roll_periods:
                                                        c_avg = f"{p_label}月均出库"
                                                        c_type = f"{p_label}门店类型"
                                                        if c_avg in df_t.columns:
                                                            col_types_t[c_avg] = "num"
                                                        if c_type in df_t.columns:
                                                            col_types_t[c_type] = "tag"
                                                    if "近三周期变化" in df_t.columns:
                                                        col_types_t["近三周期变化"] = "tag"

                                                ren_png_t = {}
                                                for c in df_t.columns:
                                                    if c in ("1月发货件数", "1月发货额-万元", "2月发货件数", "2月发货额-万元", "3月发货件数", "3月发货额-万元"):
                                                        ren_png_t[c] = f"发货分析\n{c}"
                                                for c in sel_month_cols:
                                                    if c in df_t.columns:
                                                        ren_png_t[c] = f"出库分析\n{c}"
                                                if march_col in df_t.columns:
                                                    ren_png_t[march_col] = f"出库分析\n{march_col}"
                                                if today_col in df_t.columns:
                                                    ren_png_t[today_col] = f"出库分析\n{today_col}"
                                                if avg_header in df_t.columns:
                                                    ren_png_t[avg_header] = f"出库分析\n{avg_header}"
                                                if "趋势" in df_t.columns:
                                                    ren_png_t["趋势"] = "出库分析\n趋势图"
                                                if "趋势类型" in df_t.columns:
                                                    ren_png_t["趋势类型"] = "出库分析\n趋势类型"
                                                if "库存" in df_t.columns:
                                                    ren_png_t["库存"] = "库存分析\n库存"
                                                if "可销月" in df_t.columns:
                                                    ren_png_t["可销月"] = "库存分析\n可销月"
                                                for _c in ["3月新客", "近三月新客", "累计新客"]:
                                                    if _c in df_t.columns:
                                                        ren_png_t[_c] = f"新客分析\n{_c}"
                                                if scan_avg_header in df_t.columns:
                                                    ren_png_t[scan_avg_header] = f"扫码分析\n{scan_avg_header}"
                                                if scan_rate_col in df_t.columns:
                                                    ren_png_t[scan_rate_col] = f"扫码分析\n{scan_rate_col}"
                                                for p_label, _yms in roll_periods:
                                                    for c in [f"{p_label}月均出库", f"{p_label}门店类型"]:
                                                        if c in df_t.columns:
                                                            ren_png_t[c] = f"门店类型分析\n{c}"
                                                for i in range(1, len(roll_periods)):
                                                    c = f"{roll_periods[i][0]}变动"
                                                    if c in df_t.columns:
                                                        ren_png_t[c] = f"门店类型分析\n{c}"
                                                if "近三周期变化" in df_t.columns:
                                                    ren_png_t["近三周期变化"] = "门店类型分析\n近三周期变化"
                                                if ren_png_t:
                                                    df_t = df_t.rename(columns=ren_png_t)
                                                    col_types_t = {ren_png_t.get(k, k): v for k, v in col_types_t.items()}

                                                if v_t == "经销商" and "经销商" in df_t.columns:
                                                    base_name = df_t["经销商"].fillna("").astype(str)
                                                    cols_now = [str(c) for c in df_t.columns.tolist()]
                                                    out_idx = next((i for i, c in enumerate(cols_now) if c.startswith("出库分析\n")), None)
                                                    if out_idx is not None:
                                                        df_t.insert(int(out_idx), "出库分析\n客户名", base_name)
                                                        col_types_t["出库分析\n客户名"] = "text"
                                                    cols_now = [str(c) for c in df_t.columns.tolist()]
                                                    inv_idx = next((i for i, c in enumerate(cols_now) if c.startswith("库存分析\n")), None)
                                                    if inv_idx is not None:
                                                        df_t.insert(int(inv_idx), "库存分析\n客户名", base_name)
                                                        col_types_t["库存分析\n客户名"] = "text"

                                                title_lines_t = [
                                                    f"月度出库趋势表 - {region_t}",
                                                    filter_line,
                                                    f"区域：{region_t}",
                                                    f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                                ]
                                                png = _pil_table_png(df_t, title_lines_t, font_size=16, col_types=col_types_t)
                                                fname = sanitize_filename(region_t, default="export") + ".png"
                                                zf.writestr(fname, png)
                                                prog.progress(int((idx + 1) * 100 / total))
                                        _zip_cache[k_zip] = {
                                            "bytes": zip_buf.getvalue(),
                                            "name": f"{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                                        }
                                    if k_zip in _zip_cache:
                                        st.download_button(
                                            "下载ZIP",
                                            data=_zip_cache[k_zip]["bytes"],
                                            file_name=_zip_cache[k_zip]["name"],
                                            mime="application/zip",
                                            key=f"{batch_id}_dl_{batch_export_ver}",
                                        )

                                _ship_drop = [c for c in pv.columns if ("发货件数" in str(c)) or ("发货额" in str(c))]
                                if _ship_drop:
                                    pv.drop(columns=_ship_drop, inplace=True, errors="ignore")
                                ship_defs = []

                                out_defs = []
                                for m in sel_month_cols:
                                    if m in pv.columns:
                                        out_defs.append(
                                            {
                                                "headerName": m,
                                                "field": m,
                                                "type": ["numericColumn", "numberColumnFilter"],
                                                "valueFormatter": JS_FMT_NUM,
                                                "width": 110,
                                            }
                                        )
                                if avg_col in pv.columns:
                                    out_defs.append(
                                        {
                                            "headerName": avg_header,
                                            "field": avg_col,
                                            "type": ["numericColumn", "numberColumnFilter"],
                                            "valueFormatter": JS_FMT_NUM,
                                            "width": 160,
                                        }
                                    )
                                if "趋势" in pv.columns:
                                    out_defs.append(
                                        {
                                            "headerName": "趋势图",
                                            "field": "趋势",
                                            "cellRenderer": JS_SPARKLINE,
                                            "cellRendererParams": {"width": 120, "height": 28},
                                            "width": 140,
                                        }
                                    )
                                if march_col in pv.columns:
                                    out_defs.append(
                                        {
                                            "headerName": march_col,
                                            "field": march_col,
                                            "type": ["numericColumn", "numberColumnFilter"],
                                            "valueFormatter": JS_FMT_NUM,
                                            "width": 110,
                                        }
                                    )
                                if today_col in pv.columns:
                                    out_defs.append(
                                        {
                                            "headerName": today_col,
                                            "field": today_col,
                                            "type": ["numericColumn", "numberColumnFilter"],
                                            "valueFormatter": JS_FMT_NUM,
                                            "width": 110,
                                        }
                                    )

                                inv_defs = []
                                if drill_level in (1, 2) and "库存" in pv.columns:
                                    inv_defs.append(
                                        {
                                            "headerName": "库存",
                                            "field": "库存",
                                            "type": ["numericColumn", "numberColumnFilter"],
                                            "valueFormatter": JS_FMT_NUM,
                                            "width": 120,
                                        }
                                    )
                                if drill_level in (1, 2) and "可销月" in pv.columns:
                                    inv_defs.append(
                                        {
                                            "headerName": "可销月",
                                            "field": "可销月",
                                            "type": ["numericColumn", "numberColumnFilter"],
                                            "valueFormatter": JS_FMT_NUM_1DP,
                                            "width": 120,
                                        }
                                    )

                                for _c in ["4月新客", "近三月新客", "整体新客"]:
                                    if _c not in pv.columns:
                                        pv[_c] = 0.0

                                nc_defs = []
                                for _c in ["4月新客", "近三月新客", "整体新客"]:
                                    if _c in pv.columns:
                                        nc_defs.append(
                                            {
                                                "headerName": _c,
                                                "field": _c,
                                                "type": ["numericColumn", "numberColumnFilter"],
                                                "valueFormatter": JS_FMT_NUM,
                                                "width": 120,
                                            }
                                        )

                                scan_defs = []
                                if scan_avg_col in pv.columns:
                                    scan_defs.append(
                                        {"headerName": scan_avg_header, "field": scan_avg_col, "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 160}
                                    )
                                if scan_rate_col in pv.columns:
                                    scan_defs.append(
                                        {"headerName": scan_rate_col, "field": scan_rate_col, "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO_1DP, "width": 140}
                                    )

                                geo_defs = []
                                for _c in ["市", "区/县"]:
                                    if _c in pv.columns:
                                        geo_defs.append({"headerName": _c, "field": _c, "minWidth": 140})

                                col_defs = [{"headerName": view_dim, "field": view_dim, "minWidth": 220, "pinned": "left", "tooltipField": view_dim}]
                                col_defs.append({"headerName": "出库分析", "children": out_defs})
                                if inv_defs:
                                    col_defs.append({"headerName": "库存分析", "children": inv_defs})
                                if nc_defs:
                                    col_defs.append({"headerName": "新客分析", "children": nc_defs})
                                if scan_defs:
                                    col_defs.append({"headerName": "扫码分析", "children": scan_defs})

                                if drill_level == 3:
                                    roll_children = []
                                    for p_label, _yms in roll_periods:
                                        c_avg = f"{p_label}月均出库"
                                        c_type = f"{p_label}门店类型"
                                        if c_avg in pv.columns:
                                            roll_children.append(
                                                {
                                                    "headerName": c_avg,
                                                    "field": c_avg,
                                                    "type": ["numericColumn", "numberColumnFilter"],
                                                    "valueFormatter": JS_FMT_NUM_1DP,
                                                    "width": 150,
                                                }
                                            )
                                        if c_type in pv.columns:
                                            roll_children.append({"headerName": c_type, "field": c_type, "cellRenderer": JS_STORE_TAG, "width": 120})
                                    for i in range(1, len(roll_periods)):
                                        c_chg = f"{roll_periods[i][0]}变动"
                                        if c_chg in pv.columns:
                                            roll_children.append({"headerName": c_chg, "field": c_chg, "width": 120})
                                    if "近三周期变化" in pv.columns:
                                        roll_children.append({"headerName": "近三周期变化", "field": "近三周期变化", "cellRenderer": JS_TREND_TAG, "width": 150})
                                    if roll_children:
                                        col_defs.append({"headerName": "门店类型分析", "children": roll_children})
                                    if geo_defs:
                                        col_defs.append({"headerName": "门店映射", "children": geo_defs})

                                col_defs.append({"headerName": "_趋势数据", "field": "_趋势数据", "hide": True})

                                pinned_total = {view_dim: "合计"}
                                try:
                                    for _c in sel_month_cols:
                                        if _c in pv.columns:
                                            pinned_total[_c] = float(pd.to_numeric(pv[_c], errors="coerce").fillna(0.0).sum())
                                    if march_col in pv.columns:
                                        pinned_total[march_col] = float(pd.to_numeric(pv[march_col], errors="coerce").fillna(0.0).sum())
                                    if today_col in pv.columns:
                                        pinned_total[today_col] = float(pd.to_numeric(pv[today_col], errors="coerce").fillna(0.0).sum())
                                    if avg_col in pv.columns and len(first3_cols) >= 1:
                                        _t3 = [float(pd.to_numeric(pv[c], errors="coerce").fillna(0.0).sum()) for c in first3_cols if c in pv.columns]
                                        pinned_total[avg_col] = float(np.mean(_t3)) if _t3 else 0.0
                                    if drill_level in (1, 2) and "库存" in pv.columns:
                                        pinned_total["库存"] = float(pd.to_numeric(pv["库存"], errors="coerce").fillna(0.0).sum())
                                    if drill_level in (1, 2) and "库存" in pinned_total and avg_col in pinned_total:
                                        _a = float(pinned_total.get(avg_col, 0.0) or 0.0)
                                        _k = float(pinned_total.get("库存", 0.0) or 0.0)
                                        pinned_total["可销月"] = round((_k / _a), 1) if _a > 0 else 0.0
                                    for _c in ["4月新客", "近三月新客", "整体新客"]:
                                        if _c in pv.columns:
                                            pinned_total[_c] = float(pd.to_numeric(pv[_c], errors="coerce").fillna(0.0).sum())
                                    if scan_avg_col in pv.columns:
                                        _t_scan = []
                                        for _ym in scan_yms:
                                            c = f"_scan_{int(_ym)}"
                                            if c in pv.columns:
                                                _t_scan.append(float(pd.to_numeric(pv[c], errors="coerce").fillna(0.0).sum()))
                                        pinned_total[scan_avg_col] = float(np.mean(_t_scan)) if _t_scan else 0.0
                                    if scan_rate_col in pv.columns and scan_avg_col in pinned_total and avg_col in pinned_total:
                                        denom = float(pinned_total.get(avg_col, 0.0) or 0.0)
                                        pinned_total[scan_rate_col] = (float(pinned_total.get(scan_avg_col, 0.0) or 0.0) / denom) if denom > 0 else 0.0
                                    _spark_cols = trend_base_cols if 'trend_base_cols' in locals() and trend_base_cols else first3_cols
                                    _spark_vals = [float(pd.to_numeric(pv[c], errors="coerce").fillna(0.0).sum()) for c in _spark_cols if c in pv.columns]
                                    _spark_json = json.dumps([float(x) for x in _spark_vals]) if _spark_vals else json.dumps([])
                                    pinned_total["_趋势数据"] = _spark_json
                                    pinned_total["趋势"] = _spark_json
                                except Exception:
                                    pinned_total["_趋势数据"] = json.dumps([])
                                    pinned_total["趋势"] = json.dumps([])

                                gridOptions = {
                                    "pinnedBottomRowData": [pinned_total],
                                    "columnDefs": col_defs,
                                    "defaultColDef": {
                                        "resizable": True,
                                        "sortable": True,
                                        "filter": True,
                                        "wrapHeaderText": True,
                                        "autoHeaderHeight": True,
                                        "cellStyle": {"textAlign": "center", "display": "flex", "justifyContent": "center", "alignItems": "center"},
                                        "headerClass": "ag-header-center",
                                    },
                                    "rowHeight": 40,
                                    "headerHeight": 56,
                                    "animateRows": True,
                                    "suppressCellFocus": True,
                                    "enableCellTextSelection": True,
                                    "suppressDragLeaveHidesColumns": True,
                                    "alwaysShowHorizontalScroll": True,
                                }
                                if drill_level in (1, 2):
                                    gridOptions["rowSelection"] = "single"

                                n_rows = int(len(pv))
                                row_h = 40
                                header_h = 56
                                padding_h = 140
                                min_h = 520
                                max_h = 1400
                                full_h = header_h + (n_rows * row_h) + padding_h
                                final_h = int(min(max_h, max(min_h, full_h)))
                                gridOptions["pagination"] = False

                                ag_key = "out_m_prov_ag" if drill_level == 1 else ("out_m_dist_ag" if drill_level == 2 else "out_m_store_ag")
                                st.caption(f"{view_dim}数量：{int(len(pv))}（可滚动查看全部）")
                                ag = AgGrid(
                                    pv,
                                    gridOptions=gridOptions,
                                    height=int(final_h),
                                    width="100%",
                                    data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                                    update_mode=GridUpdateMode.SELECTION_CHANGED,
                                    fit_columns_on_grid_load=False,
                                    allow_unsafe_jscode=True,
                                    theme="streamlit",
                                    key=ag_key,
                                )
                                selected_rows = ag.get("selected_rows") if ag else None
                                if selected_rows is not None and len(selected_rows) > 0 and drill_level in (1, 2):
                                    if isinstance(selected_rows, pd.DataFrame):
                                        first_row = selected_rows.iloc[0]
                                    else:
                                        first_row = selected_rows[0]

                                    selected_name = first_row.get(view_dim) if isinstance(first_row, dict) else first_row[view_dim]

                                    if drill_level == 1:
                                        st.session_state.out_m_selected_prov = selected_name
                                        st.session_state.out_m_drill_level = 2
                                        st.session_state.out_m_selected_dist = None
                                        st.rerun()
                                    elif drill_level == 2:
                                        st.session_state.out_m_selected_dist = selected_name
                                        st.session_state.out_m_drill_level = 3
                                        st.rerun()

                    if out_subtab == "🧾 专案追踪":
                        st.markdown("### 🧾 专案经销商出库追踪（段粉 / 雅系列 / 中老年）")
                        st.caption("口径：只统计2026年；出库值取出库底表S列“箱”；中老年按 提=箱×3；目标取第7个sheet（专案数据）门店目标；门店类型取专案数据E列。")

                        if "proj_view_mode" not in st.session_state:
                            st.session_state.proj_view_mode = "经销商列表"
                        if "proj_drill_level" not in st.session_state:
                            st.session_state.proj_drill_level = 1
                        if "proj_selected_prov" not in st.session_state:
                            st.session_state.proj_selected_prov = None
                        if "proj_selected_dist" not in st.session_state:
                            st.session_state.proj_selected_dist = None

                        mode = st.segmented_control(
                            "专案追踪视图",
                            options=["经销商列表", "省区汇总"],
                            key="proj_view_mode",
                            label_visibility="collapsed",
                        )
                        prev_mode = st.session_state.get("_proj_view_mode_prev")
                        if prev_mode != mode:
                            if mode == "经销商列表":
                                st.session_state.proj_drill_level = 2
                                st.session_state.proj_selected_prov = None
                                st.session_state.proj_selected_dist = None
                            else:
                                st.session_state.proj_drill_level = 1
                                st.session_state.proj_selected_prov = None
                                st.session_state.proj_selected_dist = None
                            st.session_state["_proj_view_mode_prev"] = mode

                        df_proj_raw = load_project_targets_sheet(cached_bytes, uploaded_name)
                        tgt_month = None if str(o_month) == "全部" else int(o_month)
                        proj_year = 2026
                        store_df, proj_logs = _build_project_tracking_store_df(o_raw, df_proj_raw, int(proj_year), tgt_month)
                        today_mmdd = ""
                        for x in (proj_logs or []):
                            s = str(x).strip()
                            if s.startswith("今日出库日期："):
                                today_mmdd = s.replace("今日出库日期：", "").strip()
                                break

                        if store_df is None or store_df.empty:
                            st.warning("暂无可展示的专案追踪数据")
                            with st.expander("🛠️ 调试信息", expanded=False):
                                for x in (proj_logs or []):
                                    st.text(str(x))
                        else:
                            cnav = st.columns([1, 8])
                            if int(st.session_state.proj_drill_level) > 1:
                                if cnav[0].button("⬅️ 返回", key="proj_back_btn"):
                                    st.session_state.proj_drill_level = int(st.session_state.proj_drill_level) - 1
                                    if mode == "经销商列表":
                                        if int(st.session_state.proj_drill_level) <= 2:
                                            st.session_state.proj_drill_level = 2
                                            st.session_state.proj_selected_prov = None
                                            st.session_state.proj_selected_dist = None
                                    else:
                                        if int(st.session_state.proj_drill_level) == 1:
                                            st.session_state.proj_selected_prov = None
                                            st.session_state.proj_selected_dist = None
                                        elif int(st.session_state.proj_drill_level) == 2:
                                            st.session_state.proj_selected_dist = None
                                    st.rerun()

                            bread = "🏠 全部省区"
                            if int(st.session_state.proj_drill_level) >= 2 and st.session_state.proj_selected_prov:
                                bread += f" > 📍 {st.session_state.proj_selected_prov}"
                            if int(st.session_state.proj_drill_level) >= 3 and st.session_state.proj_selected_dist:
                                bread += f" > 🏢 {st.session_state.proj_selected_dist}"
                            cnav[1].markdown(f"**当前位置**: {bread}")

                            def _agg(df_in: pd.DataFrame, group_col: str):
                                seg_cols = [
                                    "段粉-目标值",
                                    "段粉-出库值",
                                    "段粉-今日出库",
                                    "雅系列-目标值",
                                    "雅系列-出库值",
                                    "雅系列-今日出库",
                                    "中老年-目标值(提)",
                                    "中老年-出库值(提)",
                                    "中老年-今日出库(提)",
                                    "本月新客",
                                    "近三月新客",
                                    "累计新客",
                                    "本月扫码",
                                ]
                                seg_cols = [c for c in seg_cols if c in df_in.columns]
                                g = df_in.groupby(group_col, as_index=False)[seg_cols].sum()
                                if "库存" in df_in.columns:
                                    if group_col == "省区" and "经销商名称" in df_in.columns:
                                        _inv = (
                                            df_in.groupby(["省区", "经销商名称"], as_index=False)["库存"]
                                            .max()
                                            .groupby("省区", as_index=False)["库存"]
                                            .sum()
                                        )
                                        g = g.merge(_inv, on="省区", how="left")
                                    else:
                                        _inv = df_in.groupby(group_col, as_index=False)["库存"].max()
                                        g = g.merge(_inv, on=group_col, how="left")
                                    g["库存"] = pd.to_numeric(g.get("库存", 0), errors="coerce").fillna(0.0)
                                if "段粉-目标值" in g.columns and "段粉-出库值" in g.columns:
                                    g["段粉-完成率"] = np.where(g["段粉-目标值"] > 0, g["段粉-出库值"] / g["段粉-目标值"], None)
                                if "雅系列-目标值" in g.columns and "雅系列-出库值" in g.columns:
                                    g["雅系列-完成率"] = np.where(g["雅系列-目标值"] > 0, g["雅系列-出库值"] / g["雅系列-目标值"], None)
                                if "中老年-目标值(提)" in g.columns and "中老年-出库值(提)" in g.columns:
                                    g["中老年-完成率"] = np.where(g["中老年-目标值(提)"] > 0, g["中老年-出库值(提)"] / g["中老年-目标值(提)"], None)
                                if "本月扫码" in g.columns:
                                    out_box = (
                                        pd.to_numeric(g.get("段粉-出库值", 0), errors="coerce").fillna(0.0)
                                        + pd.to_numeric(g.get("雅系列-出库值", 0), errors="coerce").fillna(0.0)
                                        + (pd.to_numeric(g.get("中老年-出库值(提)", 0), errors="coerce").fillna(0.0) / 3.0)
                                    )
                                    scan_box = pd.to_numeric(g.get("本月扫码", 0), errors="coerce").fillna(0.0)
                                    g["本月扫码率"] = np.where(out_box > 0, scan_box / out_box, 0.0)
                                    g["本月扫码率"] = pd.to_numeric(g.get("本月扫码率", 0), errors="coerce").fillna(0.0)
                                ordered = [
                                    group_col,
                                    "段粉-目标值",
                                    "段粉-出库值",
                                    "段粉-完成率",
                                    "段粉-今日出库",
                                    "雅系列-目标值",
                                    "雅系列-出库值",
                                    "雅系列-完成率",
                                    "雅系列-今日出库",
                                    "中老年-目标值(提)",
                                    "中老年-出库值(提)",
                                    "中老年-完成率",
                                    "中老年-今日出库(提)",
                                ]
                                ordered = [c for c in ordered if c in g.columns]
                                rest = [c for c in g.columns if c not in ordered]
                                return g[ordered + rest]

                            drill_level = int(st.session_state.proj_drill_level)
                            view_df = None
                            view_dim = "省区"
                            if mode == "经销商列表":
                                if drill_level < 2:
                                    drill_level = 2
                                    st.session_state.proj_drill_level = 2
                                if drill_level == 2:
                                    view_df = _agg(store_df, "经销商名称").sort_values(["段粉-完成率", "雅系列-完成率", "中老年-完成率"], ascending=False)
                                    prov_map = (
                                        store_df[["经销商名称", "省区"]]
                                        .dropna()
                                        .astype(str)
                                        .drop_duplicates()
                                        .groupby("经销商名称")["省区"]
                                        .agg(lambda x: ",".join(sorted([t for t in x if t.strip()])))
                                        .reset_index()
                                    )
                                    view_df = view_df.merge(prov_map, on="经销商名称", how="left")
                                    cols_front = ["省区", "经销商名称"]
                                    cols_rest = [c for c in view_df.columns if c not in cols_front]
                                    view_df = view_df[cols_front + cols_rest]
                                    view_dim = "经销商名称"
                                else:
                                    d = str(st.session_state.proj_selected_dist or "").strip().replace(" ", "")
                                    p = str(st.session_state.proj_selected_prov or "").strip()
                                    df_s = store_df.copy()
                                    if d:
                                        df_s = df_s[df_s["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True) == d].copy()
                                    if p:
                                        df_s = df_s[df_s["省区"].astype(str).str.strip() == p].copy()
                                    view_df = df_s.sort_values(["段粉-完成率", "雅系列-完成率", "中老年-完成率"], ascending=False).reset_index(drop=True)
                                    view_dim = "门店名称"
                            else:
                                if drill_level == 1:
                                    view_df = _agg(store_df, "省区").sort_values(["段粉-完成率", "雅系列-完成率", "中老年-完成率"], ascending=False)
                                    view_dim = "省区"
                                elif drill_level == 2:
                                    p = str(st.session_state.proj_selected_prov or "").strip()
                                    df_p = store_df[store_df["省区"].astype(str) == p].copy() if p else store_df.copy()
                                    view_df = _agg(df_p, "经销商名称").sort_values(["段粉-完成率", "雅系列-完成率", "中老年-完成率"], ascending=False)
                                    view_dim = "经销商名称"
                                else:
                                    p = str(st.session_state.proj_selected_prov or "").strip()
                                    d = str(st.session_state.proj_selected_dist or "").strip().replace(" ", "")
                                    df_s = store_df.copy()
                                    if p:
                                        df_s = df_s[df_s["省区"].astype(str).str.strip() == p].copy()
                                    if d:
                                        df_s = df_s[df_s["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True) == d].copy()
                                    view_df = df_s.sort_values(["段粉-完成率", "雅系列-完成率", "中老年-完成率"], ascending=False).reset_index(drop=True)
                                    view_dim = "门店名称"

                            view_df = view_df.replace({np.nan: None})

                            def _total_row_from_df(df_in: pd.DataFrame, first_label: str):
                                def _sum_col(col: str) -> float:
                                    if col not in df_in.columns:
                                        return 0.0
                                    return float(pd.to_numeric(df_in[col], errors="coerce").fillna(0).sum())

                                t1 = _sum_col("段粉-目标值")
                                o1 = _sum_col("段粉-出库值")
                                td1 = _sum_col("段粉-今日出库")
                                t2 = _sum_col("雅系列-目标值")
                                o2 = _sum_col("雅系列-出库值")
                                td2 = _sum_col("雅系列-今日出库")
                                t3 = _sum_col("中老年-目标值(提)")
                                o3 = _sum_col("中老年-出库值(提)")
                                td3 = _sum_col("中老年-今日出库(提)")
                                r1 = (o1 / t1) if t1 > 0 else None
                                r2 = (o2 / t2) if t2 > 0 else None
                                r3 = (o3 / t3) if t3 > 0 else None
                                inv = _sum_col("库存") if view_dim == "省区" else float(pd.to_numeric(df_in.get("库存", 0), errors="coerce").fillna(0).max()) if "库存" in df_in.columns else 0.0
                                sc = _sum_col("本月扫码")
                                nc1 = _sum_col("本月新客")
                                nc3 = _sum_col("近三月新客")
                                ncc = _sum_col("累计新客")
                                out_box_total = o1 + o2 + (o3 / 3.0 if o3 else 0.0)
                                sc_rate = (sc / out_box_total) if out_box_total > 0 else 0.0
                                row = {
                                    view_dim: first_label,
                                    "段粉-目标值": t1,
                                    "段粉-出库值": o1,
                                    "段粉-完成率": r1,
                                    "段粉-今日出库": td1,
                                    "雅系列-目标值": t2,
                                    "雅系列-出库值": o2,
                                    "雅系列-完成率": r2,
                                    "雅系列-今日出库": td2,
                                    "中老年-目标值(提)": t3,
                                    "中老年-出库值(提)": o3,
                                    "中老年-完成率": r3,
                                    "中老年-今日出库(提)": td3,
                                    "库存": inv,
                                    "本月新客": nc1,
                                    "近三月新客": nc3,
                                    "累计新客": ncc,
                                    "本月扫码": sc,
                                    "本月扫码率": sc_rate,
                                }
                                if "省区" in df_in.columns and view_dim != "省区":
                                    row["省区"] = ""
                                if "门店类型" in df_in.columns:
                                    row["门店类型"] = ""
                                return row

                            col_defs = [
                                {"headerName": view_dim, "field": view_dim, "pinned": "left", "minWidth": 220, "tooltipField": view_dim},
                            ]
                            if mode == "经销商列表" and drill_level == 2:
                                col_defs = [
                                    {"headerName": "省区", "field": "省区", "pinned": "left", "minWidth": 140, "tooltipField": "省区"},
                                    {"headerName": "经销商名称", "field": "经销商名称", "pinned": "left", "minWidth": 220, "tooltipField": "经销商名称"},
                                ]
                            if drill_level == 3:
                                front = []
                                if "省区" in view_df.columns:
                                    front.append({"headerName": "省区", "field": "省区", "pinned": "left", "minWidth": 140, "tooltipField": "省区"})
                                front.append({"headerName": "门店名称", "field": "门店名称", "pinned": "left", "minWidth": 240, "tooltipField": "门店名称"})
                                if "门店类型" in view_df.columns:
                                    front.append({"headerName": "门店类型", "field": "门店类型", "pinned": "left", "minWidth": 120, "tooltipField": "门店类型", "cellRenderer": JS_STORE_TAG})
                                col_defs = front
                            col_defs += [
                                {
                                    "headerName": "段粉",
                                    "children": [
                                        {"headerName": "目标值", "field": "段粉-目标值", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 110},
                                        {"headerName": "出库值", "field": "段粉-出库值", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 110},
                                        {"headerName": "完成率", "field": "段粉-完成率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO, "width": 110},
                                        {"headerName": ("今日出库\\n" + today_mmdd) if today_mmdd else "今日出库", "field": "段粉-今日出库", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 120},
                                    ],
                                },
                                {
                                    "headerName": "雅系列",
                                    "children": [
                                        {"headerName": "目标值", "field": "雅系列-目标值", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 110},
                                        {"headerName": "出库值", "field": "雅系列-出库值", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 110},
                                        {"headerName": "完成率", "field": "雅系列-完成率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO, "width": 110},
                                        {"headerName": ("今日出库\\n" + today_mmdd) if today_mmdd else "今日出库", "field": "雅系列-今日出库", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 120},
                                    ],
                                },
                                {
                                    "headerName": "中老年（提）",
                                    "children": [
                                        {"headerName": "目标值", "field": "中老年-目标值(提)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 130},
                                        {"headerName": "出库值", "field": "中老年-出库值(提)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 130},
                                        {"headerName": "完成率", "field": "中老年-完成率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO, "width": 110},
                                        {"headerName": ("今日出库\\n" + today_mmdd) if today_mmdd else "今日出库", "field": "中老年-今日出库(提)", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 120},
                                    ],
                                },
                                {
                                    "headerName": "趋势分析指标",
                                    "children": [
                                        {"headerName": "库存", "field": "库存", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 110},
                                        {"headerName": "本月新客", "field": "本月新客", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 110},
                                        {"headerName": "近三月新客", "field": "近三月新客", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 120},
                                        {"headerName": "累计新客", "field": "累计新客", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 110},
                                        {"headerName": "本月扫码", "field": "本月扫码", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_NUM, "width": 110},
                                        {"headerName": "本月扫码率", "field": "本月扫码率", "type": ["numericColumn", "numberColumnFilter"], "valueFormatter": JS_FMT_PCT_RATIO, "width": 110},
                                    ],
                                },
                            ]

                            total_row = _total_row_from_df(view_df, "合计")
                            grid_overrides = {
                                "pinnedBottomRowData": [total_row],
                                "getRowStyle": JsCode("""
                                    function(params) {
                                        if (params.node && params.node.rowPinned) {
                                            return { 'fontWeight': '800', 'color': '#A16207' };
                                        }
                                        return null;
                                    }
                                """),
                            }

                            if mode == "经销商列表" and drill_level == 2:
                                st.caption("💡 提示：点击经销商行可下钻到门店。")
                                ag_proj = show_aggrid_table(
                                    view_df,
                                    height=620,
                                    on_row_selected="single",
                                    column_defs=col_defs,
                                    grid_options_overrides=grid_overrides,
                                    key="proj_ag_dist_list",
                                )
                                sel_rows = ag_proj.get("selected_rows") if ag_proj else None
                                if sel_rows is not None and len(sel_rows) > 0:
                                    first = sel_rows.iloc[0] if isinstance(sel_rows, pd.DataFrame) else sel_rows[0]
                                    chosen = first.get("经销商名称") if isinstance(first, dict) else first["经销商名称"]
                                    prov_val = ""
                                    try:
                                        prov_val = (first.get("省区") if isinstance(first, dict) else first["省区"]) or ""
                                    except Exception:
                                        prov_val = ""
                                    st.session_state.proj_selected_dist = chosen
                                    st.session_state.proj_selected_prov = str(prov_val).strip() or None
                                    st.session_state.proj_drill_level = 3
                                    st.rerun()

                            elif mode == "省区汇总" and drill_level == 1:
                                st.caption("💡 提示：点击省区行可下钻到经销商。")
                                ag_proj = show_aggrid_table(
                                    view_df,
                                    height=620,
                                    on_row_selected="single",
                                    column_defs=col_defs,
                                    grid_options_overrides=grid_overrides,
                                    key="proj_ag_prov",
                                )
                                sel_rows = ag_proj.get("selected_rows") if ag_proj else None
                                if sel_rows is not None and len(sel_rows) > 0:
                                    first = sel_rows.iloc[0] if isinstance(sel_rows, pd.DataFrame) else sel_rows[0]
                                    chosen = first.get(view_dim) if isinstance(first, dict) else first[view_dim]
                                    st.session_state.proj_selected_prov = chosen
                                    st.session_state.proj_selected_dist = None
                                    st.session_state.proj_drill_level = 2
                                    st.rerun()
                            elif mode == "省区汇总" and drill_level == 2:
                                st.caption("💡 提示：点击经销商行可下钻到门店。")
                                ag_proj = show_aggrid_table(
                                    view_df,
                                    height=620,
                                    on_row_selected="single",
                                    column_defs=col_defs,
                                    grid_options_overrides=grid_overrides,
                                    key="proj_ag_dist_in_prov",
                                )
                                sel_rows = ag_proj.get("selected_rows") if ag_proj else None
                                if sel_rows is not None and len(sel_rows) > 0:
                                    first = sel_rows.iloc[0] if isinstance(sel_rows, pd.DataFrame) else sel_rows[0]
                                    chosen = first.get("经销商名称") if isinstance(first, dict) else first["经销商名称"]
                                    st.session_state.proj_selected_dist = chosen
                                    st.session_state.proj_drill_level = 3
                                    st.rerun()
                            else:
                                show_aggrid_table(
                                    view_df,
                                    height=720,
                                    column_defs=col_defs,
                                    grid_options_overrides=grid_overrides,
                                    key=f"proj_ag_{mode}_{drill_level}",
                                )

                            export_df = view_df.copy()
                            export_df = export_df.loc[:, [c for c in export_df.columns if not str(c).startswith("::")]]
                            export_df = export_df.replace({np.nan: None})
                            export_df = pd.concat([export_df, pd.DataFrame([total_row])], ignore_index=True)

                            if "proj_export_cache" not in st.session_state:
                                st.session_state.proj_export_cache = {}
                            _excel_cache = st.session_state.proj_export_cache

                            sel_prov = st.session_state.get("proj_selected_prov")
                            sel_dist = st.session_state.get("proj_selected_dist")
                            month_label = f"{o_month}月" if str(o_month) != "全部" else "全年"

                            if drill_level == 1:
                                region_label = "全部省区"
                            elif drill_level == 2 and mode == "省区汇总":
                                region_label = str(sel_prov or "省区")
                            elif drill_level == 2 and mode == "经销商列表":
                                region_label = "35客户"
                            else:
                                region_label = str(sel_dist or "经销商")

                            export_id = f"proj_{mode}_{drill_level}"

                            def _excel_key(kind: str):
                                return (kind, int(proj_year), str(o_month), str(mode), int(drill_level), str(sel_prov or ""), str(sel_dist or ""))

                            number_headers = {
                                "段粉-目标值",
                                "段粉-出库值",
                                "段粉-今日出库",
                                "雅系列-目标值",
                                "雅系列-出库值",
                                "雅系列-今日出库",
                                "中老年-目标值(提)",
                                "中老年-出库值(提)",
                                "中老年-今日出库(提)",
                            }
                            percent_headers = {"段粉-完成率", "雅系列-完成率", "中老年-完成率"}

                            title_lines = [
                                f"专案追踪 - {proj_year}年{month_label}",
                                f"视图：{mode}｜层级：{drill_level}",
                                f"区域：{region_label}",
                                f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                            ]

                            col_types = {}
                            for c in export_df.columns:
                                s = str(c)
                                if "完成率" in s:
                                    col_types[s] = "pct"
                                elif s == "门店类型":
                                    col_types[s] = "tag"
                                elif s in number_headers:
                                    col_types[s] = "num"
                                else:
                                    col_types[s] = "text"

                            c_e1, c_e2, c_e3, c_e4 = st.columns([1.4, 1.9, 1.6, 3.1])
                            k_cur = _excel_key("current")
                            with c_e1:
                                if st.button("生成当前视图Excel", key=f"{export_id}_gen_cur"):
                                    with st.spinner("正在生成Excel…"):
                                        xlsx_bytes = _df_to_excel_bytes(
                                            export_df,
                                            sheet_name="专案追踪",
                                            title_lines=title_lines,
                                            number_headers=number_headers,
                                            percent_headers=percent_headers,
                                            store_type_header="门店类型",
                                            group_headers=True,
                                        )
                                        _excel_cache[k_cur] = {
                                            "bytes": xlsx_bytes,
                                            "name": sanitize_filename(f"专案追踪_{region_label}_{proj_year}_{month_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
                                        }
                            with c_e2:
                                if k_cur in _excel_cache:
                                    st.download_button(
                                        "下载当前视图Excel",
                                        data=_excel_cache[k_cur]["bytes"],
                                        file_name=_excel_cache[k_cur]["name"],
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"{export_id}_dl_cur",
                                    )

                            def _detail_store_df(all_scope: bool):
                                d = store_df.copy()
                                if not all_scope:
                                    if mode == "省区汇总" and drill_level >= 2 and sel_prov:
                                        d = d[d["省区"].astype(str).str.strip() == str(sel_prov).strip()].copy()
                                    if drill_level == 3 and sel_dist:
                                        d = d[d["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True) == str(sel_dist).strip().replace(" ", "")].copy()
                                cols = [
                                    "省区",
                                    "经销商名称",
                                    "门店名称",
                                    "门店类型",
                                    "段粉-目标值",
                                    "段粉-出库值",
                                    "段粉-完成率",
                                    "段粉-今日出库",
                                    "雅系列-目标值",
                                    "雅系列-出库值",
                                    "雅系列-完成率",
                                    "雅系列-今日出库",
                                    "中老年-目标值(提)",
                                    "中老年-出库值(提)",
                                    "中老年-完成率",
                                    "中老年-今日出库(提)",
                                ]
                                d = d[cols].copy()
                                d = d.loc[:, [c for c in d.columns if not str(c).startswith("::")]]
                                d = d.replace({np.nan: None})
                                return d

                            k_detail = _excel_key("detail_store")
                            with c_e3:
                                if st.button("生成门店明细Excel", key=f"{export_id}_gen_detail"):
                                    with st.spinner("正在生成门店明细Excel…"):
                                        df_detail = _detail_store_df(all_scope=False)
                                        total_detail = _total_row_from_df(df_detail, "合计")
                                        df_detail = pd.concat([df_detail, pd.DataFrame([total_detail])], ignore_index=True)
                                        xlsx_detail = _df_to_excel_bytes(
                                            df_detail,
                                            sheet_name="专案追踪",
                                            title_lines=[
                                                f"专案追踪 - {proj_year}年{month_label}（门店明细）",
                                                f"区域：{region_label}",
                                                f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                            ],
                                            number_headers=number_headers,
                                            percent_headers=percent_headers,
                                            store_type_header="门店类型",
                                            group_headers=True,
                                        )
                                        _excel_cache[k_detail] = {
                                            "bytes": xlsx_detail,
                                            "name": sanitize_filename(f"专案追踪_门店明细_{region_label}_{proj_year}_{month_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
                                        }
                            with c_e4:
                                if k_detail in _excel_cache:
                                    st.download_button(
                                        "下载门店明细Excel",
                                        data=_excel_cache[k_detail]["bytes"],
                                        file_name=_excel_cache[k_detail]["name"],
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"{export_id}_dl_detail",
                                    )

                            k_all = _excel_key("all_store")
                            c_all_1, c_all_2, _ = st.columns([1.6, 2.0, 6.4])
                            with c_all_1:
                                if st.button("生成导出全部Excel（门店明细）", key=f"{export_id}_gen_all"):
                                    with st.spinner("正在生成导出全部Excel…"):
                                        df_all = _detail_store_df(all_scope=True)
                                        df_all = df_all.loc[:, [c for c in df_all.columns if not str(c).startswith("::")]]
                                        total_all = _total_row_from_df(df_all, "合计")
                                        df_all = pd.concat([df_all, pd.DataFrame([total_all])], ignore_index=True)
                                        xlsx_all = _df_to_excel_bytes(
                                            df_all,
                                            sheet_name="专案追踪",
                                            title_lines=[
                                                f"专案追踪 - {proj_year}年{month_label}（导出全部门店明细）",
                                                "范围：所有省区｜所有经销商｜所有门店",
                                                f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                            ],
                                            number_headers=number_headers,
                                            percent_headers=percent_headers,
                                            store_type_header="门店类型",
                                            group_headers=True,
                                        )
                                        _excel_cache[k_all] = {
                                            "bytes": xlsx_all,
                                            "name": sanitize_filename(f"专案追踪_导出全部_门店明细_{proj_year}_{month_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
                                        }
                            with c_all_2:
                                if k_all in _excel_cache:
                                    st.download_button(
                                        "下载导出全部Excel（门店明细）",
                                        data=_excel_cache[k_all]["bytes"],
                                        file_name=_excel_cache[k_all]["name"],
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"{export_id}_dl_all",
                                    )

                            if st.button("生成表格图片（含颜色）", key=f"{export_id}_gen_png"):
                                st.session_state[f"{export_id}_png"] = _pil_table_png(export_df, title_lines, font_size=16, col_types=col_types)
                                st.session_state[f"{export_id}_png_name"] = sanitize_filename(f"专案追踪_{region_label}_{proj_year}_{month_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
                            if st.session_state.get(f"{export_id}_png"):
                                st.download_button(
                                    "下载PNG",
                                    data=st.session_state[f"{export_id}_png"],
                                    file_name=st.session_state.get(f"{export_id}_png_name", "专案追踪.png"),
                                    mime="image/png",
                                    key=f"{export_id}_dl_png",
                                )

                            if mode == "经销商列表" and drill_level == 2:
                                zip_id = f"{export_id}_zip_{proj_year}_{month_label}"
                                if st.button("批量生成门店明细PNG（ZIP）", key=f"{zip_id}_btn"):
                                    dists = sorted([x for x in store_df["经销商名称"].dropna().astype(str).unique().tolist() if x and x.lower() not in ("nan", "none", "null")])
                                    zip_buf = io.BytesIO()
                                    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                                        prog = st.progress(0)
                                        total = max(1, len(dists))
                                        for i, dist_name in enumerate(dists):
                                            df_d = store_df[store_df["经销商名称"].astype(str).str.replace(r"\s+", "", regex=True) == str(dist_name).strip().replace(" ", "")].copy()
                                            if df_d.empty:
                                                prog.progress(int((i + 1) * 100 / total))
                                                continue
                                            df_d = df_d.loc[:, [c for c in df_d.columns if not str(c).startswith("::")]]
                                            df_d = df_d.replace({np.nan: None})
                                            total_d = _total_row_from_df(df_d, "合计")
                                            df_d = pd.concat([df_d, pd.DataFrame([total_d])], ignore_index=True)
                                            col_types_d = {c: ("pct" if "完成率" in c else ("tag" if c == "门店类型" else ("num" if c in number_headers else "text"))) for c in df_d.columns}
                                            png = _pil_table_png(
                                                df_d,
                                                [
                                                    f"专案追踪 - {proj_year}年{month_label}（门店明细）",
                                                    f"经销商：{dist_name}",
                                                    f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                                ],
                                                font_size=16,
                                                col_types=col_types_d,
                                            )
                                            zf.writestr(sanitize_filename(dist_name, default="export") + ".png", png)
                                            prog.progress(int((i + 1) * 100 / total))
                                    st.session_state[f"{zip_id}_bytes"] = zip_buf.getvalue()
                                    st.session_state[f"{zip_id}_name"] = sanitize_filename(f"专案追踪_门店明细PNG_{proj_year}_{month_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")
                                if st.session_state.get(f"{zip_id}_bytes"):
                                    st.download_button(
                                        "下载ZIP",
                                        data=st.session_state[f"{zip_id}_bytes"],
                                        file_name=st.session_state.get(f"{zip_id}_name", "专案追踪.zip"),
                                        mime="application/zip",
                                        key=f"{zip_id}_dl",
                                    )

                    if out_subtab == "门店类型滚动分析":
                        st.markdown("### 门店类型滚动分析")
                        st.caption("口径：按月出库数据计算三个月滚动周期的月均出库数，并按A/B/C/D阈值分级；支持省区→经销商→门店下钻。")

                        if "roll_drill_level" not in st.session_state:
                            st.session_state.roll_drill_level = 1
                        if "roll_selected_prov" not in st.session_state:
                            st.session_state.roll_selected_prov = None
                        if "roll_selected_dist" not in st.session_state:
                            st.session_state.roll_selected_dist = None
                        if "roll_cat_big" not in st.session_state:
                            st.session_state.roll_cat_big = "全部"
                        if "roll_cat_small" not in st.session_state:
                            st.session_state.roll_cat_small = "全部"

                        c_f1, c_f2, c_f3 = st.columns([1.25, 1.25, 2.5])
                        df_roll_universe = o_raw.copy()
                        with c_f1:
                            _s_big = df_roll_universe.get("_模块大类", pd.Series(dtype=str)).dropna().astype(str).str.strip()
                            roll_cat_big_opts = ["全部"] + sorted([x for x in _s_big.unique().tolist() if x and x.lower() not in ("nan", "none", "null")])
                            roll_sel_big = st.selectbox("产品大类（T列透视）", roll_cat_big_opts, key="roll_cat_big")
                        with c_f2:
                            if "_模块小类" in df_roll_universe.columns:
                                if roll_sel_big != "全部" and "_模块大类" in df_roll_universe.columns:
                                    subs = (
                                        df_roll_universe[df_roll_universe["_模块大类"].astype(str).str.strip() == str(roll_sel_big).strip()][
                                            "_模块小类"
                                        ]
                                        .dropna()
                                        .astype(str)
                                        .str.strip()
                                        .unique()
                                        .tolist()
                                    )
                                    roll_cat_small_opts = ["全部"] + sorted([x for x in subs if x and str(x).strip().lower() not in ("nan", "none", "null")])
                                else:
                                    _s_small = df_roll_universe["_模块小类"].dropna().astype(str).str.strip()
                                    roll_cat_small_opts = ["全部"] + sorted([x for x in _s_small.unique().tolist() if x and x.lower() not in ("nan", "none", "null")])
                            else:
                                roll_cat_small_opts = ["全部"]
                            roll_sel_small = st.selectbox("产品小类（U列重量）", roll_cat_small_opts, key="roll_cat_small")
                        with c_f3:
                            if "_模块出库产品" in df_roll_universe.columns:
                                df_prod = df_roll_universe.copy()
                                if roll_sel_big != "全部" and "_模块大类" in df_prod.columns:
                                    df_prod = df_prod[df_prod["_模块大类"].astype(str).str.strip() == str(roll_sel_big).strip()]
                                if roll_sel_small != "全部" and "_模块小类" in df_prod.columns:
                                    df_prod = df_prod[df_prod["_模块小类"].astype(str).str.strip() == str(roll_sel_small).strip()]
                                _s_prod = df_prod["_模块出库产品"].dropna().astype(str).str.strip()
                                roll_prod_opts = sorted([x for x in _s_prod.unique().tolist() if x and x.lower() not in ("nan", "none", "null")])
                            else:
                                roll_prod_opts = []
                            roll_sel_prod = st.multiselect("出库产品（I列，可多选）", roll_prod_opts, key="roll_out_prod")

                        def _parse_ym_from_col(col_name: str):
                            s = str(col_name or "").strip().replace(" ", "")
                            m = re.search(r"(\d{2,4})年(\d{1,2})月", s)
                            if m:
                                y = int(m.group(1))
                                if y < 100:
                                    y = 2000 + y
                                mm = int(m.group(2))
                                return y, mm
                            m = re.search(r"(\d{2})年(\d{1,2})月", s)
                            if m:
                                y = 2000 + int(m.group(1))
                                mm = int(m.group(2))
                                return y, mm
                            m = re.search(r"(\d{1,2})月", s)
                            if m:
                                mm = int(m.group(1))
                                y = 2025 if mm >= 9 else 2026
                                return y, mm
                            return None

                        def _classify_abcd(x):
                            try:
                                v = float(x)
                            except Exception:
                                v = 0.0
                            if v >= 4:
                                return "A"
                            if 2 <= v < 4:
                                return "B"
                            if 1 <= v < 2:
                                return "C"
                            return "D"

                        _order = {"A": 4, "B": 3, "C": 2, "D": 1}

                        def _change(prev_c: str | None, cur_c: str | None):
                            if not prev_c or not cur_c:
                                return ""
                            pv = _order.get(str(prev_c).strip(), 0)
                            cv = _order.get(str(cur_c).strip(), 0)
                            if pv == cv:
                                return "持平"
                            if cv > pv:
                                return "升级 ⬆️"
                            return "降级 ⬇️"

                        periods = [
                            ("25年10-12月", [(2025, 10), (2025, 11), (2025, 12)]),
                            ("25年11-26年1月", [(2025, 11), (2025, 12), (2026, 1)]),
                            ("25年12-26年2月", [(2025, 12), (2026, 1), (2026, 2)]),
                        ]

                        def _cmp_sign(prev_c: str | None, cur_c: str | None):
                            if not prev_c or not cur_c:
                                return None
                            pv = _order.get(str(prev_c).strip(), 0)
                            cv = _order.get(str(cur_c).strip(), 0)
                            if pv == cv:
                                return 0
                            return 1 if cv > pv else -1

                        def _trend3_label(c1: str | None, c2: str | None, c3: str | None):
                            s1 = _cmp_sign(c1, c2)
                            s2 = _cmp_sign(c2, c3)
                            if s1 is None or s2 is None:
                                return ""
                            if s1 > 0 and s2 > 0:
                                return "持续升级"
                            if s1 < 0 and s2 < 0:
                                return "持续降级"
                            if s1 > 0 and s2 < 0:
                                return "先升级后降级"
                            if s1 < 0 and s2 > 0:
                                return "先降级后升级"
                            if s1 == 0 and s2 == 0:
                                return "持续持平"
                            if s1 == 0 and s2 > 0:
                                return "持平升级"
                            if s1 > 0 and s2 == 0:
                                return "升级持平"
                            if s1 == 0 and s2 < 0:
                                return "持平降级"
                            if s1 < 0 and s2 == 0:
                                return "降级持平"
                            return "其他"

                        _prod_norm_key_roll = tuple(sorted([str(x).strip() for x in (roll_sel_prod or []) if str(x).strip()]))
                        ck = (
                            "roll_store_df_v3",
                            st.session_state.get("_active_file_sig"),
                            sel_prov,
                            sel_dist,
                            str(roll_sel_big or ""),
                            str(roll_sel_small or ""),
                            _prod_norm_key_roll,
                        )
                        if "out_subtab_cache" not in st.session_state:
                            st.session_state.out_subtab_cache = {}

                        if ck in st.session_state.out_subtab_cache:
                            _cached = st.session_state.out_subtab_cache[ck]
                            if isinstance(_cached, tuple) and len(_cached) >= 3:
                                store_roll_df, roll_missing, roll_missing_note = _cached[0], _cached[1], _cached[2]
                            else:
                                store_roll_df, roll_missing = _cached
                                roll_missing_note = None
                        else:
                            d = df_roll_universe.copy()
                            if sel_prov != "全部" and "省区" in d.columns:
                                d = d[d["省区"].astype(str).str.strip() == str(sel_prov).strip()].copy()
                            if sel_dist != "全部" and "经销商名称" in d.columns:
                                d = d[d["经销商名称"].astype(str).str.strip() == str(sel_dist).strip()].copy()
                            if roll_sel_big != "全部" and "_模块大类" in d.columns:
                                d = d[d["_模块大类"].astype(str).str.strip() == str(roll_sel_big).strip()].copy()
                            if roll_sel_small != "全部" and "_模块小类" in d.columns:
                                d = d[d["_模块小类"].astype(str).str.strip() == str(roll_sel_small).strip()].copy()
                            if roll_sel_prod and "_模块出库产品" in d.columns:
                                _p = [str(x).strip() for x in roll_sel_prod if str(x).strip()]
                                if _p:
                                    d = d[d["_模块出库产品"].astype(str).str.strip().isin(_p)].copy()

                            d = d[d.get("_年").notna() & d.get("_月").notna()].copy()
                            d["_年"] = pd.to_numeric(d.get("_年", 0), errors="coerce").fillna(0).astype(int)
                            d["_月"] = pd.to_numeric(d.get("_月", 0), errors="coerce").fillna(0).astype(int)
                            d = d[(d["_年"] > 0) & (d["_月"].between(1, 12))].copy()
                            d["_ym"] = (d["_年"] * 100 + d["_月"]).astype(int)

                            store_col = "_门店名" if "_门店名" in d.columns else ("门店名称" if "门店名称" in d.columns else None)
                            if store_col is None:
                                store_roll_df = pd.DataFrame(columns=["省区", "经销商名称", "门店名称"])
                                roll_missing = []
                                roll_missing_note = None
                            else:
                                required_yms = []
                                for _, yms in periods:
                                    for (y, m) in yms:
                                        required_yms.append(int(y * 100 + m))
                                required_yms = list(dict.fromkeys(required_yms))
                                avail_yms = set(pd.to_numeric(d.get("_ym", pd.Series(dtype=int)), errors="coerce").fillna(0).astype(int).unique().tolist())
                                roll_missing = [f"{int(ym)//100}-{str(int(ym)%100).zfill(2)}" for ym in required_yms if int(ym) not in avail_yms]
                                roll_missing_note = None
                                if roll_missing:
                                    d0 = df_roll_universe.copy()
                                    if sel_prov != "全部" and "省区" in d0.columns:
                                        d0 = d0[d0["省区"].astype(str).str.strip() == str(sel_prov).strip()].copy()
                                    if sel_dist != "全部" and "经销商名称" in d0.columns:
                                        d0 = d0[d0["经销商名称"].astype(str).str.strip() == str(sel_dist).strip()].copy()
                                    d0 = d0[d0.get("_年").notna() & d0.get("_月").notna()].copy()
                                    d0["_年"] = pd.to_numeric(d0.get("_年", 0), errors="coerce").fillna(0).astype(int)
                                    d0["_月"] = pd.to_numeric(d0.get("_月", 0), errors="coerce").fillna(0).astype(int)
                                    d0 = d0[(d0["_年"] > 0) & (d0["_月"].between(1, 12))].copy()
                                    d0["_ym"] = (d0["_年"] * 100 + d0["_月"]).astype(int)
                                    avail_yms0 = set(pd.to_numeric(d0.get("_ym", pd.Series(dtype=int)), errors="coerce").fillna(0).astype(int).unique().tolist())
                                    notes = []
                                    for ym in required_yms:
                                        if int(ym) in avail_yms:
                                            continue
                                        label = f"{int(ym)//100}-{str(int(ym)%100).zfill(2)}"
                                        if int(ym) in avail_yms0:
                                            notes.append(f"{label}：该月有出库记录，但在当前产品筛选下未匹配到（按0处理）")
                                        else:
                                            notes.append(f"{label}：该月未找到出库记录（按0处理）")
                                    roll_missing_note = "；".join(notes) if notes else None

                                d["数量(箱)"] = pd.to_numeric(d.get("数量(箱)", 0), errors="coerce").fillna(0.0)
                                d["省区"] = d.get("省区", "").fillna("").astype(str).str.strip()
                                d["经销商名称"] = d.get("经销商名称", "").fillna("").astype(str).str.strip()
                                d[store_col] = d[store_col].fillna("").astype(str).str.strip()
                                d = d[(d["省区"] != "") & (d["经销商名称"] != "") & (d[store_col] != "")].copy()

                                g = d.groupby(["省区", "经销商名称", store_col, "_ym"], as_index=False)["数量(箱)"].sum()
                                pv_m = g.pivot(index=["省区", "经销商名称", store_col], columns="_ym", values="数量(箱)").fillna(0.0)
                                for ym in required_yms:
                                    if ym not in pv_m.columns:
                                        pv_m[int(ym)] = 0.0
                                pv_m = pv_m.reset_index().rename(columns={store_col: "门店名称"})

                                work = pv_m[["省区", "经销商名称", "门店名称"]].copy()
                                for p_label, yms in periods:
                                    cols = [int(y * 100 + m) for (y, m) in yms]
                                    avg_col = f"{p_label}月均出库"
                                    cls_col = f"{p_label}门店类型"
                                    work[avg_col] = pv_m[cols].sum(axis=1) / 3.0
                                    work[cls_col] = work[avg_col].apply(_classify_abcd)

                                for i in range(2, len(periods) + 1):
                                    prev_label = periods[i - 2][0]
                                    cur_label = periods[i - 1][0]
                                    work[f"{cur_label}变动"] = work.apply(
                                        lambda r: _change(r.get(f"{prev_label}门店类型"), r.get(f"{cur_label}门店类型")),
                                        axis=1,
                                    )

                                if len(periods) >= 3:
                                    p1, p2, p3 = periods[-3][0], periods[-2][0], periods[-1][0]
                                    work["近三周期变化"] = work.apply(
                                        lambda r: _trend3_label(
                                            r.get(f"{p1}门店类型"),
                                            r.get(f"{p2}门店类型"),
                                            r.get(f"{p3}门店类型"),
                                        ),
                                        axis=1,
                                    )
                                else:
                                    work["近三周期变化"] = ""

                                store_roll_df = work.copy()
                            st.session_state.out_subtab_cache[ck] = (store_roll_df, roll_missing, roll_missing_note)

                        if roll_missing:
                            st.warning("当前筛选条件下未找到以下月份数据，将按0处理：" + "、".join(roll_missing))
                            if roll_missing_note:
                                st.caption(roll_missing_note)

                        cnav = st.columns([1, 8])
                        if int(st.session_state.roll_drill_level) > 1:
                            if cnav[0].button("⬅️ 返回", key="roll_back_btn"):
                                st.session_state.roll_drill_level = int(st.session_state.roll_drill_level) - 1
                                if int(st.session_state.roll_drill_level) == 1:
                                    st.session_state.roll_selected_prov = None
                                    st.session_state.roll_selected_dist = None
                                elif int(st.session_state.roll_drill_level) == 2:
                                    st.session_state.roll_selected_dist = None
                                st.rerun()

                        bread = "🏠 全部省区"
                        if int(st.session_state.roll_drill_level) >= 2 and st.session_state.roll_selected_prov:
                            bread += f" > 📍 {st.session_state.roll_selected_prov}"
                        if int(st.session_state.roll_drill_level) >= 3 and st.session_state.roll_selected_dist:
                            bread += f" > 🏢 {st.session_state.roll_selected_dist}"
                        cnav[1].markdown(f"**当前位置**: {bread}")

                        def _agg_counts(df_in: pd.DataFrame, group_col: str):
                            out = pd.DataFrame({group_col: sorted([x for x in df_in[group_col].dropna().astype(str).unique().tolist() if x and x.lower() not in ("nan", "none", "null")])})
                            for p_label, _yms in periods:
                                cls_col = f"{p_label}门店类型"
                                g = (
                                    df_in.groupby([group_col, cls_col])
                                    .size()
                                    .reset_index(name="数量")
                                    .pivot(index=group_col, columns=cls_col, values="数量")
                                    .reset_index()
                                )
                                for k in ["A", "B", "C", "D"]:
                                    if k not in g.columns:
                                        g[k] = 0
                                g = g[[group_col, "A", "B", "C", "D"]].copy()
                                for k in ["A", "B", "C", "D"]:
                                    g[f"{p_label}-{k}"] = pd.to_numeric(g[k], errors="coerce").fillna(0).astype(int)
                                g = g.drop(columns=["A", "B", "C", "D"], errors="ignore")
                                out = out.merge(g, on=group_col, how="left")

                            trend_col = "近三周期变化"
                            trend_kinds = ["持续升级", "先升级后降级", "先降级后升级", "持续降级", "持续持平", "持平升级", "升级持平", "持平降级", "降级持平"]
                            if trend_col in df_in.columns:
                                tg = (
                                    df_in.groupby([group_col, trend_col])
                                    .size()
                                    .reset_index(name="数量")
                                    .pivot(index=group_col, columns=trend_col, values="数量")
                                    .reset_index()
                                )
                                for k in trend_kinds:
                                    if k not in tg.columns:
                                        tg[k] = 0
                                tg = tg[[group_col] + trend_kinds].copy()
                                for k in trend_kinds:
                                    tg[f"近三周期-{k}"] = pd.to_numeric(tg[k], errors="coerce").fillna(0).astype(int)
                                tg = tg.drop(columns=trend_kinds, errors="ignore")
                                out = out.merge(tg, on=group_col, how="left")
                            for c in out.columns:
                                if c == group_col:
                                    continue
                                out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).astype(int)
                            return out

                        drill_level = int(st.session_state.roll_drill_level)
                        view_dim = "省区"
                        view_df = None
                        if drill_level == 1:
                            view_dim = "省区"
                            view_df = _agg_counts(store_roll_df, "省区")
                        elif drill_level == 2:
                            view_dim = "经销商名称"
                            s = store_roll_df.copy()
                            if st.session_state.roll_selected_prov:
                                s = s[s["省区"].astype(str) == str(st.session_state.roll_selected_prov)]
                            view_df = _agg_counts(s, "经销商名称")
                        else:
                            s = store_roll_df.copy()
                            if st.session_state.roll_selected_prov:
                                s = s[s["省区"].astype(str) == str(st.session_state.roll_selected_prov)]
                            if st.session_state.roll_selected_dist:
                                s = s[s["经销商名称"].astype(str) == str(st.session_state.roll_selected_dist)]

                            cols = ["省区", "经销商名称", "门店名称"]
                            for p_label, _yms in periods:
                                cols += [f"{p_label}月均出库", f"{p_label}门店类型"]
                                if f"{p_label}变动" in s.columns:
                                    cols += [f"{p_label}变动"]
                            cols += ["近三周期变化"]
                            cols = [c for c in cols if c in s.columns]
                            view_df = s[cols].copy()
                            for p_label, _yms in periods:
                                avg_col = f"{p_label}月均出库"
                                if avg_col in view_df.columns:
                                    view_df[avg_col] = pd.to_numeric(view_df[avg_col], errors="coerce").fillna(0.0).round(1)

                        export_scope = "省区汇总" if drill_level == 1 else ("经销商汇总" if drill_level == 2 else "门店明细")
                        export_id = f"roll_export_{drill_level}_{sanitize_filename(st.session_state.get('roll_selected_prov') or 'all')}_{sanitize_filename(st.session_state.get('roll_selected_dist') or 'all')}"
                        cexp = st.columns([1, 3, 6])
                        if cexp[0].button("生成Excel", key=f"{export_id}_btn"):
                            title_lines = [
                                f"门店类型滚动分析 - {export_scope}",
                                f"范围：{sel_prov}/{sel_dist}",
                                f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                            ]
                            exp_df = view_df.copy()
                            number_formats = {}
                            if drill_level in (1, 2):
                                trend_kinds = ["持续升级", "先升级后降级", "先降级后升级", "持续降级", "持续持平", "持平升级", "升级持平", "持平降级", "降级持平"]
                                ren = {}
                                for k in trend_kinds:
                                    src = f"近三周期-{k}"
                                    if src in exp_df.columns:
                                        ren[src] = f"近三周期变化-{k}"
                                if ren:
                                    exp_df = exp_df.rename(columns=ren)
                                for c in exp_df.columns:
                                    if c != view_dim:
                                        number_formats[str(c)] = "0"
                            else:
                                ren = {}
                                ren["经销商名称"] = "经销商"
                                ren["门店名称"] = "门店"
                                for p_label, _yms in periods:
                                    ren[f"{p_label}月均出库"] = f"{p_label}-月均出库"
                                    ren[f"{p_label}门店类型"] = f"{p_label}-门店类型"
                                    if f"{p_label}变动" in exp_df.columns:
                                        ren[f"{p_label}变动"] = f"{p_label}-升级/降级"
                                if "近三周期变化" in exp_df.columns:
                                    ren["近三周期变化"] = "近三周期变化-变化类型"
                                exp_df = exp_df.rename(columns=ren)
                                for c in exp_df.columns:
                                    if str(c).endswith("-月均出库"):
                                        number_formats[str(c)] = "0.0"
                            xls = _df_to_excel_bytes(
                                exp_df,
                                sheet_name="门店类型滚动分析",
                                title_lines=title_lines,
                                number_formats=number_formats,
                                group_headers=True,
                            )
                            st.session_state[f"{export_id}_bytes"] = xls
                            st.session_state[f"{export_id}_name"] = sanitize_filename(f"门店类型滚动分析_{export_scope}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                        if st.session_state.get(f"{export_id}_bytes"):
                            cexp[1].download_button(
                                "下载Excel",
                                data=st.session_state[f"{export_id}_bytes"],
                                file_name=st.session_state.get(f"{export_id}_name", "门店类型滚动分析.xlsx"),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"{export_id}_dl",
                            )

                        export_all_id = f"roll_export_all_{sanitize_filename(sel_prov)}_{sanitize_filename(sel_dist)}"
                        export_all_btn_label = "导出各省门店ZIP" if drill_level == 1 else "导出全部门店"
                        if cexp[0].button(export_all_btn_label, key=f"{export_all_id}_btn"):
                            cols_all = ["省区", "经销商名称", "门店名称"]
                            for p_label, _yms in periods:
                                cols_all += [f"{p_label}月均出库", f"{p_label}门店类型"]
                                if f"{p_label}变动" in store_roll_df.columns:
                                    cols_all += [f"{p_label}变动"]
                            cols_all += ["近三周期变化"]
                            cols_all = [c for c in cols_all if c in store_roll_df.columns]

                            ren_all = {}
                            ren_all["经销商名称"] = "经销商"
                            ren_all["门店名称"] = "门店"
                            for p_label, _yms in periods:
                                ren_all[f"{p_label}月均出库"] = f"{p_label}-月均出库"
                                ren_all[f"{p_label}门店类型"] = f"{p_label}-门店类型"
                                if f"{p_label}变动" in cols_all:
                                    ren_all[f"{p_label}变动"] = f"{p_label}-升级/降级"
                            ren_all["近三周期变化"] = "近三周期变化-变化类型"

                            if drill_level == 1:
                                buf = io.BytesIO()
                                with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                                    provs = []
                                    if "省区" in store_roll_df.columns:
                                        provs = (
                                            store_roll_df["省区"]
                                            .dropna()
                                            .astype(str)
                                            .str.strip()
                                            .tolist()
                                        )
                                    provs = sorted([p for p in set(provs) if p and p.lower() not in ("nan", "none", "null")])
                                    for p in provs:
                                        df_p = store_roll_df[store_roll_df["省区"].astype(str).str.strip() == str(p).strip()][cols_all].copy()
                                        if df_p.empty:
                                            continue
                                        df_p = df_p.sort_values(["经销商名称", "门店名称"], kind="stable").reset_index(drop=True)
                                        for p_label, _yms in periods:
                                            avg_col = f"{p_label}月均出库"
                                            if avg_col in df_p.columns:
                                                df_p[avg_col] = pd.to_numeric(df_p[avg_col], errors="coerce").fillna(0.0).round(1)
                                        df_p = df_p.rename(columns=ren_all)
                                        number_formats_p = {str(c): "0.0" for c in df_p.columns if str(c).endswith("-月均出库")}
                                        title_lines_p = [
                                            "门店类型滚动分析 - 门店明细",
                                            f"省区：{p}",
                                            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                        ]
                                        xls_p = _df_to_excel_bytes(
                                            df_p,
                                            sheet_name="门店明细",
                                            title_lines=title_lines_p,
                                            number_formats=number_formats_p,
                                            group_headers=True,
                                        )
                                        zf.writestr(sanitize_filename(f"{p}.xlsx"), xls_p)
                                buf.seek(0)
                                st.session_state[f"{export_all_id}_bytes"] = buf.getvalue()
                                st.session_state[f"{export_all_id}_name"] = sanitize_filename(f"门店类型滚动分析_各省门店_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")
                                st.session_state[f"{export_all_id}_mime"] = "application/zip"
                            else:
                                df_all = store_roll_df[cols_all].copy()
                                for p_label, _yms in periods:
                                    avg_col = f"{p_label}月均出库"
                                    if avg_col in df_all.columns:
                                        df_all[avg_col] = pd.to_numeric(df_all[avg_col], errors="coerce").fillna(0.0).round(1)
                                df_all = df_all.rename(columns=ren_all)
                                number_formats_all = {str(c): "0.0" for c in df_all.columns if str(c).endswith("-月均出库")}
                                title_lines = [
                                    "门店类型滚动分析 - 全部门店",
                                    f"范围：{sel_prov}/{sel_dist}",
                                    f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                ]
                                xls_all = _df_to_excel_bytes(
                                    df_all,
                                    sheet_name="全部门店",
                                    title_lines=title_lines,
                                    number_formats=number_formats_all,
                                    group_headers=True,
                                )
                                st.session_state[f"{export_all_id}_bytes"] = xls_all
                                st.session_state[f"{export_all_id}_name"] = sanitize_filename(f"门店类型滚动分析_全部门店_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                                st.session_state[f"{export_all_id}_mime"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        if st.session_state.get(f"{export_all_id}_bytes"):
                            dl_label = "下载各省门店ZIP" if drill_level == 1 else "下载全部门店Excel"
                            cexp[2].download_button(
                                dl_label,
                                data=st.session_state[f"{export_all_id}_bytes"],
                                file_name=st.session_state.get(f"{export_all_id}_name", "门店类型滚动分析_全部门店.xlsx"),
                                mime=st.session_state.get(f"{export_all_id}_mime", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                                key=f"{export_all_id}_dl",
                            )

                        ag_key = "roll_prov_ag" if drill_level == 1 else ("roll_dist_ag" if drill_level == 2 else "roll_store_ag")
                        col_defs = None
                        if drill_level in (1, 2):
                            view_key = view_dim
                            col_defs = [
                                {
                                    "headerName": view_key,
                                    "field": view_key,
                                    "pinned": "left",
                                    "lockPinned": True,
                                    "minWidth": 140,
                                }
                            ]
                            for p_label, _yms in periods:
                                children = []
                                for cls, hdr_cls in [("A", "hdr-a"), ("B", "hdr-b"), ("C", "hdr-c"), ("D", "hdr-d")]:
                                    children.append(
                                        {
                                            "headerName": f"{cls}数量",
                                            "field": f"{p_label}-{cls}",
                                            "type": ["numericColumn", "numberColumnFilter"],
                                            "headerClass": hdr_cls,
                                            "minWidth": 92,
                                        }
                                    )
                                col_defs.append({"headerName": p_label, "children": children})

                            trend_kinds = [
                                ("持续升级", "hdr-up", {"color": "#16A34A", "fontWeight": "800"}),
                                ("先升级后降级", "hdr-ud", {"color": "#B45309", "fontWeight": "800"}),
                                ("先降级后升级", "hdr-du", {"color": "#2563EB", "fontWeight": "800"}),
                                ("持续降级", "hdr-down", {"color": "#DC2626", "fontWeight": "800"}),
                                ("持续持平", "hdr-flat", {"color": "#374151", "fontWeight": "800"}),
                                ("持平升级", "hdr-up", {"color": "#16A34A", "fontWeight": "800"}),
                                ("升级持平", "hdr-up", {"color": "#16A34A", "fontWeight": "800"}),
                                ("持平降级", "hdr-down", {"color": "#DC2626", "fontWeight": "800"}),
                                ("降级持平", "hdr-down", {"color": "#DC2626", "fontWeight": "800"}),
                            ]
                            t_children = []
                            for k, hdr_cls, cell_style in trend_kinds:
                                t_children.append(
                                    {
                                        "headerName": k,
                                        "field": f"近三周期-{k}",
                                        "type": ["numericColumn", "numberColumnFilter"],
                                        "headerClass": hdr_cls,
                                        "cellStyle": cell_style,
                                        "minWidth": 120,
                                    }
                                )
                            col_defs.append({"headerName": "近三周期变化", "children": t_children})
                        else:
                            base_children = [
                                {"headerName": "省区", "field": "省区", "pinned": "left", "lockPinned": True, "minWidth": 110},
                                {"headerName": "经销商", "field": "经销商名称", "pinned": "left", "lockPinned": True, "minWidth": 140},
                                {"headerName": "门店", "field": "门店名称", "pinned": "left", "lockPinned": True, "minWidth": 160},
                            ]
                            col_defs = [{"headerName": "基本信息", "children": [c for c in base_children if c["field"] in view_df.columns]}]
                            for p_label, _yms in periods:
                                children = []
                                avg_field = f"{p_label}月均出库"
                                cls_field = f"{p_label}门店类型"
                                chg_field = f"{p_label}变动"
                                if avg_field in view_df.columns:
                                    children.append(
                                        {
                                            "headerName": "月均出库",
                                            "field": avg_field,
                                            "type": ["numericColumn", "numberColumnFilter"],
                                            "minWidth": 110,
                                        }
                                    )
                                if cls_field in view_df.columns:
                                    children.append(
                                        {
                                            "headerName": "等级",
                                            "field": cls_field,
                                            "cellRenderer": JS_STORE_TAG,
                                            "minWidth": 90,
                                        }
                                    )
                                if chg_field in view_df.columns:
                                    children.append({"headerName": "升级/降级", "field": chg_field, "minWidth": 110})
                                if children:
                                    col_defs.append({"headerName": p_label, "children": children})

                            if "近三周期变化" in view_df.columns:
                                col_defs.append(
                                    {
                                        "headerName": "近三周期变化",
                                        "children": [
                                            {
                                                "headerName": "变化类型",
                                                "field": "近三周期变化",
                                                "cellRenderer": JS_TREND_TAG,
                                                "minWidth": 140,
                                            }
                                        ],
                                    }
                                )

                        ag = show_aggrid_table(
                            view_df,
                            on_row_selected=("single" if drill_level in (1, 2) else None),
                            key=ag_key,
                            column_defs=col_defs,
                        )
                        selected_rows = ag.get("selected_rows") if ag else None
                        if selected_rows is not None and len(selected_rows) > 0 and drill_level in (1, 2):
                            first_row = selected_rows[0] if not isinstance(selected_rows, pd.DataFrame) else selected_rows.iloc[0]
                            selected_name = first_row.get(view_dim) if isinstance(first_row, dict) else first_row[view_dim]
                            if drill_level == 1:
                                st.session_state.roll_selected_prov = selected_name
                                st.session_state.roll_selected_dist = None
                                st.session_state.roll_drill_level = 2
                                st.rerun()
                            elif drill_level == 2:
                                st.session_state.roll_selected_dist = selected_name
                                st.session_state.roll_drill_level = 3
                                st.rerun()

                    st.markdown("</div>", unsafe_allow_html=True)

                    # === TAB 7: PERFORMANCE ===
            if main_tab == "🚀 业绩分析":
                st.markdown("""
                <style>
                  .perf-wrap {display:flex; flex-direction:column; gap:16px;}
                  .perf-kpis {display:grid; grid-template-columns: repeat(4, 1fr); gap:14px;}
                  .perf-card {background:#F3E5F5; border:1px solid rgba(156,39,176,0.18); border-radius:12px; padding:14px 16px; box-shadow:0 6px 20px rgba(18,12,28,0.06);}
                  .perf-k {font-size:13px; color:rgba(27,21,48,0.72);}
                  .perf-v {font-size:20px; font-weight:800; color:#9C27B0; margin-top:8px;}
                  .perf-sub {display:flex; justify-content:space-between; align-items:center; margin-top:8px; font-size:12px; color:rgba(27,21,48,0.72);}
                  .perf-up {color:#2FBF71; font-weight:700;}
                  .perf-down {color:#E5484D; font-weight:700;}
                  .perf-mid {color:#FFB000; font-weight:700;}
                  .stDataFrame td { vertical-align: middle !important; }
                  @media (max-width: 1100px) {.perf-kpis {grid-template-columns: repeat(2, 1fr);} }
                </style>
                """, unsafe_allow_html=True)

                if df_perf_raw is None or df_perf_raw.empty:
                    st.warning("⚠️ 未检测到发货业绩数据 (Sheet4)。请确认Excel包含Sheet4且数据完整。")
                    with st.expander("🛠️ 调试信息", expanded=False):
                        for log in debug_logs: st.text(log)
                else:
                    df_perf = df_perf_raw.copy()
                    
                    # --- 1. Data Prep ---
                    # Load Targets from Sheet5 (C=Prov, D=Cat, E=Month, F=Task)
                    df_target = None
                    if df_target_raw is not None and len(df_target_raw.columns) >= 6:
                        try:
                            # Use iloc to be safe about column names
                            df_target = df_target_raw.iloc[:, [2, 3, 4, 5]].copy()
                            df_target.columns = ['省区', '品类', '月份', '任务量']
                            df_target['任务量'] = pd.to_numeric(df_target['任务量'], errors='coerce').fillna(0)
                            df_target['月份'] = pd.to_numeric(df_target['月份'], errors='coerce').fillna(0).astype(int)
                            df_target['省区'] = df_target['省区'].astype(str).str.strip()
                            df_target['品类'] = df_target['品类'].astype(str).str.strip()
                        except Exception as e:
                            st.error(f"任务表解析失败: {e}")
                            df_target = None
                    
                    # Data Cleaning
                    df_track = df_perf.copy()
                    df_track['年份'] = pd.to_numeric(df_track['年份'], errors='coerce').fillna(0).astype(int)
                    df_track['月份'] = pd.to_numeric(df_track['月份'], errors='coerce').fillna(0).astype(int)
                    
                    # Fix: Check if '发货金额' exists, if not, try to use '发货箱数' or create empty
                    if '发货金额' not in df_track.columns:
                            if '发货箱数' in df_track.columns:
                                df_track['发货金额'] = df_track['发货箱数'] # Fallback
                            else:
                                df_track['发货金额'] = 0.0
                    
                    df_track['发货金额'] = pd.to_numeric(df_track['发货金额'], errors='coerce').fillna(0.0)
                    
                    for c in ['省区', '经销商名称', '归类', '发货仓', '大分类', '月分析']:
                        if c in df_track.columns:
                            df_track[c] = df_track[c].fillna('').astype(str).str.strip()
                    
                    # Determine Year
                    years = sorted([y for y in df_track['年份'].unique() if y > 2000])
                    cur_year = 2026 if 2026 in years else (max(years) if years else 2025)
                    last_year = cur_year - 1
                    
                    # --- 2. Filters ---
                    with st.expander("🎛️ 筛选控制面板", expanded=False):
                        f1, f2, f3, f4, f5 = st.columns(5)
                        
                        # Province
                        prov_opts = ['全部'] + sorted([x for x in df_track['省区'].unique() if x])
                        with f1:
                            sel_prov = st.selectbox("省区", prov_opts, key="t26_prov")
                        
                        # Filter Step 1
                        df_f = df_track if sel_prov == '全部' else df_track[df_track['省区'] == sel_prov]
                        
                        # Distributor
                        dist_opts = ['全部'] + sorted([x for x in df_f['经销商名称'].unique() if x])
                        with f2:
                            sel_dist = st.selectbox("经销商", dist_opts, key="t26_dist")
                        if sel_dist != '全部':
                            df_f = df_f[df_f['经销商名称'] == sel_dist]
                            
                        if '大分类' in df_track.columns:
                            cat_col_S = '大分类'
                        elif '月分析' in df_track.columns:
                            cat_col_S = '月分析'
                            st.warning("⚠️ 未找到'Sheet4 S列大分类'字段名“大分类”，已使用“月分析”列作为替代。请确认源数据列名。")
                        else:
                            cat_col_S = '发货仓'
                            st.error("❌ 数据源中未找到Sheet4 S列“大分类”/“月分析”列，已临时使用“发货仓”列作为大分类筛选。")

                        if cat_col_S in df_f.columns:
                            df_f[cat_col_S] = df_f[cat_col_S].fillna('').astype(str).str.strip()

                        if cat_col_S in df_track.columns:
                            df_track[cat_col_S] = df_track[cat_col_S].fillna('').astype(str).str.strip()

                        cat_check_value = "益益成人粉"
                        cat_exists_all = False
                        cat_exists_filtered = False
                        if cat_col_S in df_track.columns:
                            cat_exists_all = bool((df_track[cat_col_S] == cat_check_value).any())
                        if cat_col_S in df_f.columns:
                            cat_exists_filtered = bool((df_f[cat_col_S] == cat_check_value).any())

                        if cat_exists_all and (not cat_exists_filtered):
                            st.warning(f"⚠️ 源数据“大分类”包含“{cat_check_value}”，但在当前省区/经销商筛选下无数据。请调整筛选查看。")

                        with st.expander("🔎 大分类数据校验", expanded=False):
                            if cat_col_S not in df_track.columns:
                                st.error(f"未找到用于大分类的字段：{cat_col_S}")
                            else:
                                s_all = df_track[cat_col_S]
                                s_all_nonempty = s_all[s_all != ""]
                                st.write(f"大分类字段：{cat_col_S}")
                                st.write(f"唯一类目数：{int(s_all_nonempty.nunique())}")
                                st.write(f"空值占比：{fmt_pct_ratio(float((s_all == '').mean()))}")
                                st.write(f"是否包含“{cat_check_value}”：{'是' if cat_exists_all else '否'}")
                                top_counts = s_all_nonempty.value_counts().head(12).reset_index()
                                top_counts.columns = ["类目", "行数"]
                                show_aggrid_table(top_counts, height=300, key="verify_table")

                        wh_opts = ['全部'] + sorted([x for x in df_f.get(cat_col_S, pd.Series(dtype=str)).unique() if x])
                        with f3:
                            sel_wh = st.selectbox(f"大类 ({cat_col_S})", wh_opts, key="t26_wh")
                        
                        if sel_wh != '全部':
                            df_f = df_f[df_f.get(cat_col_S, pd.Series(dtype=str)) == sel_wh]
                            
                        # Small Category (Group) - Multi Select
                        grp_opts = sorted([x for x in df_f['归类'].unique() if x])
                        with f4:
                            sel_grp = st.multiselect("小类 (归类)", grp_opts, default=[], key="t26_grp")
                        if sel_grp:
                            df_f = df_f[df_f['归类'].isin(sel_grp)]
                            
                        # Month Selection (Single)
                        avail_months = sorted(df_f[df_f['年份'] == cur_year]['月份'].unique())
                        def_month = int(avail_months[-1]) if avail_months else 1
                        with f5:
                            sel_month = st.selectbox("统计月份", list(range(1, 13)), index=def_month-1, key="t26_month")
                    
                    # --- 3. Calculations ---
                    # Actuals
                    act_cur_year = df_f[df_f['年份'] == cur_year]['发货金额'].sum()
                    act_last_year = df_f[df_f['年份'] == last_year]['发货金额'].sum()
                    
                    act_cur_month = df_f[(df_f['年份'] == cur_year) & (df_f['月份'] == sel_month)]['发货金额'].sum()
                    act_last_month = df_f[(df_f['年份'] == last_year) & (df_f['月份'] == sel_month)]['发货金额'].sum()
                    
                    # Targets
                    target_cur_year = 0.0
                    target_cur_month = 0.0
                    if df_target is not None:
                        # Apply filters to target (Province, Category)
                        # Note: Distributor filter can't apply to Target usually, unless target is by dist. 
                        # User said Sheet5 has Province/Category.
                        df_t_f = df_target.copy()
                        if sel_prov != '全部':
                            df_t_f = df_t_f[df_t_f['省区'] == sel_prov]
                        # Category mapping? Sheet5 '品类' vs Sheet4 '归类'/'发货仓'.
                        # User said D col is Category. Assuming it matches '归类' or needs mapping.
                        # For now, we sum all if no specific match logic provided or if '全部'.
                        # If user selected specific categories, we try to filter.
                        # BUT, without exact mapping, filtering Targets by Category is risky. 
                        # We'll calculate Total Target for selected Province.
                        
                        target_cur_year = df_t_f['任务量'].sum()
                        target_cur_month = df_t_f[df_t_f['月份'] == sel_month]['任务量'].sum()
                    
                    # Rates & YoY
                    rate_year = (act_cur_year / target_cur_year) if target_cur_year > 0 else None
                    rate_month = (act_cur_month / target_cur_month) if target_cur_month > 0 else None
                    
                    yoy_year = (act_cur_year - act_last_year) / act_last_year if act_last_year > 0 else None
                    yoy_month = (act_cur_month - act_last_month) / act_last_month if act_last_month > 0 else None
                    
                    # --- 4. KPI Cards ---
                    def _fmt_wan(x): return fmt_num((x or 0) / 10000)
                    def _fmt_pct(x): return fmt_pct_ratio(x) if x is not None else "—"
                    def _color_pct(x): return "perf-up" if x and x>0 else "perf-down"
                    def _arrow(x): return "↑" if x and x>0 else ("↓" if x and x<0 else "")

                    def _render_card(title, icon, val_wan, target_wan, rate, yoy_val_wan, yoy_pct):
                        trend_cls = "trend-up" if yoy_pct and yoy_pct > 0 else ("trend-down" if yoy_pct and yoy_pct < 0 else "trend-neutral")
                        arrow = _arrow(yoy_pct)
                        rate_txt = _fmt_pct(rate)
                        yoy_txt = _fmt_pct(yoy_pct)
                        pct_val = min(max(rate * 100 if rate else 0, 0), 100)
                        prog_color = "#28A745" if rate and rate >= 1.0 else ("#FFC107" if rate and rate >= 0.8 else "#DC3545")

                        st.markdown(f"""
                        <div class="out-kpi-card">
                            <div class="out-kpi-bar"></div>
                            <div class="out-kpi-head">
                                <div class="out-kpi-ico">{icon}</div>
                                <div class="out-kpi-title">{title}</div>
                            </div>
                            <div class="out-kpi-val">¥ {val_wan}万</div>
                            <div class="out-kpi-sub2" style="margin-top:8px;">
                                <span>达成率</span>
                                <span style="font-weight:800; color:{prog_color}">{rate_txt}</span>
                            </div>
                            <div class="out-kpi-progress" style="margin-top:6px;">
                                <div class="out-kpi-progress-bar" style="background:{prog_color}; width:{pct_val}%;"></div>
                            </div>
                            <div class="out-kpi-sub2" style="margin-top:10px;">
                                <span>目标</span>
                                <span>{target_wan}万</span>
                            </div>
                            <div class="out-kpi-sub2">
                                <span>同期</span>
                                <span>{yoy_val_wan}万</span>
                            </div>
                            <div class="out-kpi-sub2">
                                <span>同比</span>
                                <span class="{trend_cls}">{arrow} {yoy_txt}</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                    # --- TABS: KPI, Category, Province ---
                    tab_perf_kpi, tab_perf_cat, tab_perf_prov = st.tabs(["📊 核心业绩指标", "📦 分品类", "🗺️ 分省区"])

                    with tab_perf_kpi:
                        k1, k2 = st.columns(2)
                        
                        with k1:
                            _render_card("本月业绩", "📅", _fmt_wan(act_cur_month), _fmt_wan(target_cur_month), rate_month, _fmt_wan(act_last_month), yoy_month)
                        with k2:
                            _render_card("年度累计业绩", "🏆", _fmt_wan(act_cur_year), _fmt_wan(target_cur_year), rate_year, _fmt_wan(act_last_year), yoy_year)
                    
                    with tab_perf_cat:
                        # --- NEW: Category Performance Cards ---
                        
                        # Prepare Category Data
                        # Using cat_col_S ('大分类' or '月分析' or '发货仓')
                        
                        # 1. Monthly Category Data
                        cat_cur_m = df_f[(df_f['年份'] == cur_year) & (df_f['月份'] == sel_month)].groupby(cat_col_S)['发货金额'].sum().reset_index().rename(columns={'发货金额': '本月'})
                        cat_last_m = df_f[(df_f['年份'] == last_year) & (df_f['月份'] == sel_month)].groupby(cat_col_S)['发货金额'].sum().reset_index().rename(columns={'发货金额': '同期'})
                        
                        cat_m_final = pd.merge(cat_cur_m, cat_last_m, on=cat_col_S, how='outer').fillna(0)
                        cat_m_final['本月(万)'] = cat_m_final['本月'] / 10000
                        cat_m_final['同期(万)'] = cat_m_final['同期'] / 10000
                        cat_m_final['同比'] = np.where(cat_m_final['本月'] > 0, (cat_m_final['本月'] - cat_m_final['同期']) / cat_m_final['本月'], None)
                        cat_m_final = cat_m_final.sort_values('本月', ascending=False)

                        # 2. Yearly Category Data
                        cat_cur_y = df_f[df_f['年份'] == cur_year].groupby(cat_col_S)['发货金额'].sum().reset_index().rename(columns={'发货金额': '本年'})
                        cat_last_y = df_f[df_f['年份'] == last_year].groupby(cat_col_S)['发货金额'].sum().reset_index().rename(columns={'发货金额': '同期'})
                        
                        cat_y_final = pd.merge(cat_cur_y, cat_last_y, on=cat_col_S, how='outer').fillna(0)
                        cat_y_final['本年(万)'] = cat_y_final['本年'] / 10000
                        cat_y_final['同期(万)'] = cat_y_final['同期'] / 10000
                        cat_y_final['同比'] = np.where(cat_y_final['本年'] > 0, (cat_y_final['本年'] - cat_y_final['同期']) / cat_y_final['本年'], None)
                        cat_y_final = cat_y_final.sort_values('本年', ascending=False)

                        # Render 2 Columns for Tables
                        c_cat_m, c_cat_y = st.columns(2)

                        with c_cat_m:
                            st.markdown(
                                """
                                <div style="background-color: #F8F9FA; border-radius: 8px; padding: 16px; border: 1px solid #E9ECEF; box-shadow: 0 2px 4px rgba(0,0,0,0.05); height: 100%;">
                                    <div style="font-size: 14px; color: #6C757D; margin-bottom: 12px; font-weight: 500;">📅 本月分品类业绩</div>
                                """, 
                                unsafe_allow_html=True
                            )
                            # Replaced with AgGrid
                            show_aggrid_table(
                                cat_m_final[[cat_col_S, '本月(万)', '同期(万)', '同比']],
                                height=250,
                                key="ag_cat_m"
                            )
                            
                            # Donut Chart for Month
                            if not cat_m_final.empty and cat_m_final['本月(万)'].sum() > 0:
                                total_m = cat_m_final['本月(万)'].sum()
                                cat_m_final['legend_label'] = cat_m_final.apply(
                                    lambda r: f"{r[cat_col_S]}   {r['本月(万)']:.1f}万   {r['本月(万)']/total_m:.1%}", axis=1
                                )
                                
                                fig_m = go.Figure(data=[go.Pie(
                                    labels=cat_m_final['legend_label'],
                                    values=cat_m_final['本月(万)'],
                                    hole=0.6,
                                    marker=dict(colors=px.colors.qualitative.Pastel),
                                    textinfo='none',
                                    domain={'x': [0.4, 1.0]}
                                )])
                                fig_m.update_layout(
                                    showlegend=True,
                                    legend=dict(
                                        yanchor="middle", y=0.5,
                                        xanchor="left", x=0,
                                        font=dict(size=12, color="#333333")
                                    ),
                                    margin=dict(t=10, b=10, l=0, r=0), 
                                    height=250
                                )
                                st.plotly_chart(fig_m, use_container_width=True, key="perf_cat_month_donut")
                            else:
                                st.info("暂无数据")
                                
                            st.markdown("</div>", unsafe_allow_html=True)

                        with c_cat_y:
                            st.markdown(
                                """
                                <div style="background-color: #F8F9FA; border-radius: 8px; padding: 16px; border: 1px solid #E9ECEF; box-shadow: 0 2px 4px rgba(0,0,0,0.05); height: 100%;">
                                    <div style="font-size: 14px; color: #6C757D; margin-bottom: 12px; font-weight: 500;">🏆 年度分品类业绩</div>
                                """, 
                                unsafe_allow_html=True
                            )
                            # Replaced with AgGrid
                            show_aggrid_table(
                                cat_y_final[[cat_col_S, '本年(万)', '同期(万)', '同比']],
                                height=250,
                                key="ag_cat_y"
                            )
                            
                            # Donut Chart for Year
                            if not cat_y_final.empty and cat_y_final['本年(万)'].sum() > 0:
                                total_y = cat_y_final['本年(万)'].sum()
                                cat_y_final['legend_label'] = cat_y_final.apply(
                                    lambda r: f"{r[cat_col_S]}   {r['本年(万)']:.1f}万   {r['本年(万)']/total_y:.1%}", axis=1
                                )
                                
                                fig_y = go.Figure(data=[go.Pie(
                                    labels=cat_y_final['legend_label'],
                                    values=cat_y_final['本年(万)'],
                                    hole=0.6,
                                    marker=dict(colors=px.colors.qualitative.Pastel),
                                    textinfo='none',
                                    domain={'x': [0.4, 1.0]}
                                )])
                                fig_y.update_layout(
                                    showlegend=True,
                                    legend=dict(
                                        yanchor="middle", y=0.5,
                                        xanchor="left", x=0,
                                        font=dict(size=12, color="#333333")
                                    ),
                                    margin=dict(t=10, b=10, l=0, r=0), 
                                    height=250
                                )
                                st.plotly_chart(fig_y, use_container_width=True, key="perf_cat_year_donut")
                            else:
                                st.info("暂无数据")
                                
                            st.markdown("</div>", unsafe_allow_html=True)

                    with tab_perf_prov:
                        # --- 5. Province Table (Detailed) ---
                        
                        # Prepare Data
                        # Group by Province
                        # 1. Actuals (Cur Month)
                        df_m_cur = df_f[(df_f['年份'] == cur_year) & (df_f['月份'] == sel_month)]
                        prov_cur = df_m_cur.groupby('省区')['发货金额'].sum().reset_index().rename(columns={'发货金额': '本月业绩'})
                        
                        # 2. Actuals (Same Period)
                        df_m_last = df_f[(df_f['年份'] == last_year) & (df_f['月份'] == sel_month)]
                        prov_last = df_m_last.groupby('省区')['发货金额'].sum().reset_index().rename(columns={'发货金额': '同期业绩'})
                        
                        # 3. Targets (Month)
                        if df_target is not None:
                            t_m = df_target[df_target['月份'] == sel_month]
                            prov_target = t_m.groupby('省区')['任务量'].sum().reset_index().rename(columns={'任务量': '本月任务'})
                        else:
                            prov_target = pd.DataFrame(columns=['省区', '本月任务'])
                            
                        # Merge All
                        prov_final = pd.merge(prov_cur, prov_target, on='省区', how='outer')
                        prov_final = pd.merge(prov_final, prov_last, on='省区', how='outer').fillna(0)
                        
                        # Filter out rows with 0
                        prov_final = prov_final[(prov_final['本月业绩']!=0) | (prov_final['本月任务']!=0) | (prov_final['同期业绩']!=0)]
                        
                        # Metrics
                        prov_final['达成率'] = prov_final.apply(lambda x: (x['本月业绩'] / x['本月任务']) if x['本月任务'] > 0 else 0, axis=1)
                        prov_final['同比增长'] = prov_final.apply(lambda x: ((x['本月业绩'] - x['同期业绩']) / x['同期业绩']) if x['同期业绩'] > 0 else 0, axis=1)
                        
                        # Sort
                        prov_final = prov_final.sort_values('本月业绩', ascending=False)
                        
                        # Format for Display
                        prov_final['本月业绩(万)'] = prov_final['本月业绩'] / 10000
                        prov_final['本月任务(万)'] = prov_final['本月任务'] / 10000
                        prov_final['同期业绩(万)'] = prov_final['同期业绩'] / 10000
                        
                        # Display Columns
                        disp_df = prov_final[['省区', '本月业绩(万)', '本月任务(万)', '达成率', '同期业绩(万)', '同比增长']].copy()
                        
                        # Interactive Table
                        st.caption("👇 点击表格行可下钻查看详细数据")
                        
                        # AgGrid for Province Performance
                        ag_prov = show_aggrid_table(
                            disp_df, 
                            key="perf_prov_ag",
                            on_row_selected=True
                        )
                        
                        # Drill Down
                        # Check if selected_rows exists and is not empty
                        selected_rows = ag_prov.get('selected_rows') if ag_prov else None
                        
                        if selected_rows is not None and len(selected_rows) > 0:
                            # AgGrid return structure might differ based on version
                            # Sometimes it returns a DataFrame, sometimes a list of dicts
                            if isinstance(selected_rows, pd.DataFrame):
                                first_row = selected_rows.iloc[0]
                            else:
                                first_row = selected_rows[0]
                                
                            # Handle if it returns a DataFrame row or dict
                            sel_prov_drill = first_row.get('省区') if isinstance(first_row, dict) else first_row['省区']
                            
                            # Drill Down Tabs
                            st.markdown("---")
                            st.subheader(f"📍 {sel_prov_drill} - 明细数据")
                            
                            tab_dist, tab_cat = st.tabs(["🏢 经销商明细", "📦 品类明细"])
                            
                            # Filter data for selected province
                            d_cur = df_f[(df_f['年份'] == cur_year) & (df_f['月份'] == sel_month) & (df_f['省区'] == sel_prov_drill)]
                            d_last = df_f[(df_f['年份'] == last_year) & (df_f['月份'] == sel_month) & (df_f['省区'] == sel_prov_drill)]

                            # --- Tab 1: Distributor Drill Down ---
                            with tab_dist:
                                st.caption(f"正在查看：{sel_prov_drill} > 经销商明细")
                                d_cur_g = d_cur.groupby('经销商名称')['发货金额'].sum().reset_index().rename(columns={'发货金额': '本月'})
                                d_last_g = d_last.groupby('经销商名称')['发货金额'].sum().reset_index().rename(columns={'发货金额': '同期'})
                                
                                d_final = pd.merge(d_cur_g, d_last_g, on='经销商名称', how='outer').fillna(0)
                                d_final['同比增长'] = d_final.apply(lambda x: ((x['本月'] - x['同期']) / x['同期']) if x['同期'] > 0 else 0, axis=1)
                                d_final = d_final.sort_values('本月', ascending=False)
                                
                                d_final['本月(万)'] = d_final['本月'] / 10000
                                d_final['同期(万)'] = d_final['同期'] / 10000
                                
                                ag_dist = show_aggrid_table(
                                    d_final[['经销商名称', '本月(万)', '同期(万)', '同比增长']],
                                    key="perf_dist_ag",
                                    on_row_selected=True
                                )
                                
                                selected_rows_dist = ag_dist.get('selected_rows') if ag_dist else None
                                
                                if selected_rows_dist is not None and len(selected_rows_dist) > 0:
                                    if isinstance(selected_rows_dist, pd.DataFrame):
                                        first_row_dist = selected_rows_dist.iloc[0]
                                    else:
                                        first_row_dist = selected_rows_dist[0]
                                        
                                    sel_dist_drill = first_row_dist.get('经销商名称') if isinstance(first_row_dist, dict) else first_row_dist['经销商名称']
                                    st.info(f"📍 正在查看 {sel_prov_drill} > {sel_dist_drill} 的大分类明细")
                                    
                                    if '大分类' in d_cur.columns:
                                        cat_col_S = '大分类'
                                    elif '月分析' in d_cur.columns:
                                        cat_col_S = '月分析'
                                    else:
                                        cat_col_S = '发货仓'
                                    
                                    # Filter data for selected dist
                                    bc_cur = d_cur[d_cur['经销商名称'] == sel_dist_drill]
                                    bc_last = d_last[d_last['经销商名称'] == sel_dist_drill]
                                    
                                    bc_cur_g = bc_cur.groupby(cat_col_S)['发货金额'].sum().reset_index().rename(columns={'发货金额': '本月'})
                                    bc_last_g = bc_last.groupby(cat_col_S)['发货金额'].sum().reset_index().rename(columns={'发货金额': '同期'})
                                    
                                    bc_final = pd.merge(bc_cur_g, bc_last_g, on=cat_col_S, how='outer').fillna(0)
                                    bc_final['同比增长'] = bc_final.apply(lambda x: ((x['本月'] - x['同期']) / x['同期']) if x['同期'] > 0 else 0, axis=1)
                                    bc_final = bc_final.sort_values('本月', ascending=False)
                                    
                                    bc_final['本月(万)'] = bc_final['本月'] / 10000
                                    bc_final['同期(万)'] = bc_final['同期'] / 10000
                                    
                                    ag_bc = show_aggrid_table(
                                        bc_final[[cat_col_S, '本月(万)', '同期(万)', '同比增长']],
                                        key="perf_bc_table_dist_ag",
                                        on_row_selected=True
                                    )
                                    
                                    selected_rows_bc = ag_bc.get('selected_rows') if ag_bc else None
                                    
                                    if selected_rows_bc is not None and len(selected_rows_bc) > 0:
                                        if isinstance(selected_rows_bc, pd.DataFrame):
                                            first_row_bc = selected_rows_bc.iloc[0]
                                        else:
                                            first_row_bc = selected_rows_bc[0]
                                            
                                        sel_bc_drill = first_row_bc.get(cat_col_S) if isinstance(first_row_bc, dict) else first_row_bc[cat_col_S]
                                        st.info(f"📍 正在查看 {sel_prov_drill} > {sel_dist_drill} > {sel_bc_drill} 的小分类(归类)明细")
                                        
                                        # Level 4: Small Category (Group) for Selected Big Cat
                                        sc_cur = bc_cur[bc_cur[cat_col_S] == sel_bc_drill]
                                        sc_last = bc_last[bc_last[cat_col_S] == sel_bc_drill]
                                        
                                        sc_cur_g = sc_cur.groupby('归类')['发货金额'].sum().reset_index().rename(columns={'发货金额': '本月'})
                                        sc_last_g = sc_last.groupby('归类')['发货金额'].sum().reset_index().rename(columns={'发货金额': '同期'})
                                        
                                        sc_final = pd.merge(sc_cur_g, sc_last_g, on='归类', how='outer').fillna(0)
                                        sc_final['同比增长'] = sc_final.apply(lambda x: ((x['本月'] - x['同期']) / x['同期']) if x['同期'] > 0 else 0, axis=1)
                                        sc_final = sc_final.sort_values('本月', ascending=False)
                                        
                                        sc_final['本月(万)'] = sc_final['本月'] / 10000
                                        sc_final['同期(万)'] = sc_final['同期'] / 10000
                                        
                                        show_aggrid_table(
                                            sc_final[['归类', '本月(万)', '同期(万)', '同比增长']],
                                            key="perf_sc_table_dist_ag"
                                        )

                            with tab_cat:
                                st.caption(f"正在查看：{sel_prov_drill} > 品类明细 (按大分类聚合)")
                                if '大分类' in d_cur.columns:
                                    agg_col = '大分类'
                                elif '月分析' in d_cur.columns:
                                    agg_col = '月分析'
                                else:
                                    agg_col = '发货仓'
                                
                                c_cur_g = d_cur.groupby(agg_col)['发货金额'].sum().reset_index().rename(columns={'发货金额': '本月'})
                                c_last_g = d_last.groupby(agg_col)['发货金额'].sum().reset_index().rename(columns={'发货金额': '同期'})
                                
                                c_final = pd.merge(c_cur_g, c_last_g, on=agg_col, how='outer').fillna(0)
                                c_final['同比增长'] = c_final.apply(lambda x: ((x['本月'] - x['同期']) / x['同期']) if x['同期'] > 0 else 0, axis=1)
                                c_final = c_final.sort_values('本月', ascending=False)
                                
                                c_final['本月(万)'] = c_final['本月'] / 10000
                                c_final['同期(万)'] = c_final['同期'] / 10000
                                
                                ag_cat = show_aggrid_table(
                                    c_final[[agg_col, '本月(万)', '同期(万)', '同比增长']],
                                    key="perf_cat_table_ag",
                                    on_row_selected=True
                                )
                                
                                selected_rows_cat = ag_cat.get('selected_rows') if ag_cat else None
                                
                                if selected_rows_cat is not None and len(selected_rows_cat) > 0:
                                    if isinstance(selected_rows_cat, pd.DataFrame):
                                        first_row_cat = selected_rows_cat.iloc[0]
                                    else:
                                        first_row_cat = selected_rows_cat[0]
                                        
                                    sel_cat_drill = first_row_cat.get(agg_col) if isinstance(first_row_cat, dict) else first_row_cat[agg_col]
                                    st.info(f"📍 正在查看 {sel_prov_drill} > {sel_cat_drill} 的小分类(归类)明细")
                                    
                                    # Level 3: Small Category (Group) for Selected Big Cat (Province Level)
                                    sc_cur = d_cur[d_cur[agg_col] == sel_cat_drill]
                                    sc_last = d_last[d_last[agg_col] == sel_cat_drill]
                                    
                                    sc_cur_g = sc_cur.groupby('归类')['发货金额'].sum().reset_index().rename(columns={'发货金额': '本月'})
                                    sc_last_g = sc_last.groupby('归类')['发货金额'].sum().reset_index().rename(columns={'发货金额': '同期'})
                                    
                                    sc_final = pd.merge(sc_cur_g, sc_last_g, on='归类', how='outer').fillna(0)
                                    sc_final['同比增长'] = sc_final.apply(lambda x: ((x['本月'] - x['同期']) / x['同期']) if x['同期'] > 0 else 0, axis=1)
                                    sc_final = sc_final.sort_values('本月', ascending=False)
                                    
                                    sc_final['本月(万)'] = sc_final['本月'] / 10000
                                    sc_final['同期(万)'] = sc_final['同期'] / 10000
                                    
                                    # Dynamic height
                                    n_rows_sc2 = len(sc_final)
                                    calc_height_sc2 = (n_rows_sc2 + 1) * 35 + 10
                                    final_height_sc2 = max(150, min(calc_height_sc2, 2000))
                                    
                                    show_aggrid_table(
                                        sc_final[['归类', '本月(万)', '同期(万)', '同比增长']],
                                        height=final_height_sc2,
                                        key="perf_sc_table_cat_ag"
                                    )

else:
    st.info("请在左侧上传数据文件以开始分析。")
